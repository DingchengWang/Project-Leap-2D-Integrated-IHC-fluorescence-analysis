# This functional source module is assembled into one shared runtime.
from __future__ import annotations

def write_measurement_workbook(
    path: Path,
    *,
    fiji_details: dict,
    measurement: str,
    best_row: dict,
    age_profile: str,
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl 3.1.5 is required for the XLSX output in the formal Python environment"
        ) from exc

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.title = "IHC 2D Fluorescence Results"
    workbook.properties.creator = PIPELINE_NAME
    headers = [
        "ROI_Index",
        "Astrocyte_ID",
        "Original_Astrocyte_ID",
        "Source_Original_Astrocyte_IDs",
        "Compartment",
        "ROI_Name",
        "Label",
        "Area",
        "Mean",
        "Median",
        "Min",
        "Max",
        "IntDen",
        "RawIntDen",
        "Measurement_Channel",
        "Projection",
        "Z_Start_1Based",
        "Z_End_1Based_Inclusive",
        "Age_Profile",
        "Manual_Review_Used",
    ]
    sheet_specs = (
        ("Whole Cell", "whole"),
        ("Processes", "processes"),
        ("Soma", "soma"),
    )
    result_sets = fiji_details.get("result_sets", {})
    for sheet_name, key in sheet_specs:
        detail = result_sets.get(key, {})
        row_data = detail.get("row_data", [])
        if len(row_data) != int(detail.get("rows", -1)):
            raise RuntimeError(f"Fiji {key} row payload is incomplete: {detail}")
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="355070")
        for source in row_data:
            enriched = dict(source)
            enriched.update(
                {
                    "Measurement_Channel": measurement,
                    "Projection": str(best_row["projection"]).upper(),
                    "Z_Start_1Based": int(best_row["z_start_1based"]),
                    "Z_End_1Based_Inclusive": int(
                        best_row["z_end_1based_inclusive"]
                    ),
                    "Age_Profile": age_profile,
                    "Manual_Review_Used": bool(
                        fiji_details.get("manual_review_used", False)
                    ),
                }
            )
            sheet.append([enriched.get(header) for header in headers])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        numeric_headers = {"Area", "Mean", "Median", "Min", "Max", "IntDen", "RawIntDen"}
        for column_index, header in enumerate(headers, start=1):
            if header not in numeric_headers:
                continue
            for cell in list(sheet.columns)[column_index - 1][1:]:
                cell.number_format = "0.000000"
        widths = {
            "A": 12,
            "B": 14,
            "C": 22,
            "D": 28,
            "E": 14,
            "F": 34,
            "G": 54,
            "H": 14,
            "I": 14,
            "J": 12,
            "K": 12,
            "L": 12,
            "M": 16,
            "N": 18,
            "O": 22,
            "P": 12,
            "Q": 17,
            "R": 25,
            "S": 14,
            "T": 22,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
    audit_sheet = workbook.create_sheet("Review Audit")
    audit_headers = ["Sequence", "Action", "Source_Original_IDs", "Result_Lineage", "Reverted"]
    audit_sheet.append(audit_headers)
    for cell in audit_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="6D597A")
    for event in fiji_details.get("review_audit", []):
        audit_sheet.append(
            [
                event.get("sequence"),
                event.get("action"),
                ",".join(str(value) for value in event.get("source_ids", [])),
                ",".join(str(value) for value in event.get("result_lineage", [])),
                bool(event.get("reverted", False)),
            ]
        )
    audit_sheet.freeze_panes = "A2"
    for column, width in {"A": 12, "B": 14, "C": 28, "D": 28, "E": 12}.items():
        audit_sheet.column_dimensions[column].width = width
    workbook.save(path)

def validate_measurement_workbook(path: Path, fiji_details: dict) -> None:
    from openpyxl import load_workbook

    expected = {
        "Whole Cell": "whole",
        "Processes": "processes",
        "Soma": "soma",
    }
    with path.open("rb") as handle:
        workbook = load_workbook(handle, read_only=True, data_only=True)
        if set(workbook.sheetnames) != set(expected) | {"Review Audit"}:
            raise RuntimeError(f"Unexpected XLSX worksheets: {workbook.sheetnames}")
        observed_sequences: dict[str, tuple[list[int], list[int]]] = {}
        for sheet_name, key in expected.items():
            sheet = workbook[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                raise RuntimeError(f"XLSX worksheet {sheet_name} is empty")
            headings = list(rows[0])
            if "Median" not in headings or "P90" in headings or "P95" in headings:
                raise RuntimeError(f"XLSX measurement columns are invalid: {headings}")
            expected_rows = int(fiji_details["result_sets"][key]["rows"])
            if len(rows) - 1 != expected_rows:
                raise RuntimeError(
                    f"XLSX {sheet_name} has {len(rows) - 1} rows; expected {expected_rows}"
                )
            astrocyte_column = headings.index("Astrocyte_ID")
            original_column = headings.index("Original_Astrocyte_ID")
            current_ids = [int(row[astrocyte_column]) for row in rows[1:]]
            original_ids = [int(row[original_column]) for row in rows[1:]]
            if current_ids != list(range(1, len(current_ids) + 1)):
                raise RuntimeError(
                    f"XLSX {sheet_name} Astrocyte IDs are not contiguous: {current_ids}"
                )
            if len(set(original_ids)) != len(original_ids):
                raise RuntimeError(
                    f"XLSX {sheet_name} repeats an Original Astrocyte ID: {original_ids}"
                )
            observed_sequences[key] = (current_ids, original_ids)
        reference_sequences = observed_sequences["whole"]
        for key in ("processes", "soma"):
            if observed_sequences[key] != reference_sequences:
                raise RuntimeError(
                    "XLSX Whole, Soma, and Processes ID sequences do not match: "
                    f"{observed_sequences}"
                )
        workbook.close()
