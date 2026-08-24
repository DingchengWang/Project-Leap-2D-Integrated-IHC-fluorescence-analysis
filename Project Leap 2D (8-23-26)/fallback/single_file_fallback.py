from __future__ import annotations

import argparse

import fcntl

import gc

import json

import math

import os

import re

import shutil

import subprocess

import threading

import time

import uuid

import warnings

import weakref

from concurrent.futures import FIRST_COMPLETED, ThreadPoolExecutor, as_completed, wait

from contextlib import contextmanager

from dataclasses import asdict, dataclass, replace

from pathlib import Path

for _thread_environment_name in (
    "OMP_NUM_THREADS",
    "OPENBLAS_NUM_THREADS",
    "VECLIB_MAXIMUM_THREADS",
    "NUMEXPR_NUM_THREADS",
):
    os.environ.setdefault(_thread_environment_name, "1")

import numpy as np

import cv2

import tifffile as tf

from PIL import Image

from scipy import ndimage as ndi

from skimage import exposure, filters, measure, morphology, segmentation, transform

warnings.filterwarnings("ignore", category=FutureWarning)


@dataclass(frozen=True)
class TestSpec:
    name: str
    z_mode: str
    projection: str
    method: str
    egfp_weight: float = 0.55
    gfap_weight: float = 0.45
    smooth_sigma: float = 1.0
    threshold_scale: float = 1.0
    min_area: int = 180
    close_radius: int = 1
    dilate_radius: int = 1
    hole_area: int = 64
    cleanup_mode: str = "basic"
    anchor_area: int = 1600
    bridge_radius: int = 8
    low_percentile: float = 83.0
    seed_percentile: float = 93.0
    seed_min_area: int = 260
    max_area_fraction: float = 0.34
    cellpose: bool = False
    cellpose_cellprob: float = -0.5
    cellpose_diameter: float = 85.0
    cellpose_max_side: int = 2048
    dapi_support_radius: int = 24
    outline_smooth_sigma: float = 1.15
    outline_epsilon: float = 1.6
    artifact_filter: bool = False
    artifact_min_area: int = 1100
    artifact_near_radius: int = 36
    process_eccentricity: float = 0.78
    process_major_axis: float = 42.0
    branch_refine: bool = False
    branch_support_percentile: float = 38.0
    branch_support_radius: int = 2
    max_process_half_width: float = 10.0
    soma_protect_radius: int = 22
    outline_hole_min_area: int = 90
    require_soma_anchor: bool = False
    soma_anchor_radius: int = 4
    soma_anchor_percentile: float = 84.0
    soma_core_radius: float = 8.0
    soma_anchor_min_pixels: int = 8
    anchor_component_min_area: int = 3000
    connection_radius: int = 3
    connection_support_percentile: float = 84.0
    fine_branch_recovery: bool = False
    fine_branch_detail_percentile: float = 92.0
    fine_branch_intensity_percentile: float = 68.0
    fine_branch_min_area: int = 16
    fine_branch_min_major_axis: float = 12.0
    fine_branch_min_eccentricity: float = 0.62
    fine_branch_gap_radius: int = 2
    fine_branch_background_sigma: float = 7.0
    fine_branch_single_channel_offset: float = 4.0
    fine_branch_evidence_mode: str = "union"
    fine_branch_consensus_radius: int = 1
    fine_branch_topology_max_gap: int = 4
    fine_branch_topology_min_skeleton: int = 10
    fine_branch_topology_max_hops: int = 3
    exclude_border_components: bool = False
    border_margin: int = 12
    edge_qc_margin: int = 48
    preserve_complete_border_components: bool = False
    border_complete_soma_margin: int = 48
    border_complete_min_area_ratio: float = 0.75
    border_complete_min_interior_fraction: float = 0.75

@dataclass(frozen=True)
class CompartmentConfig:
    """Calibration-aware 2D soma rules applied after Whole ROI selection."""

    dapi_percentile_floor: float = 85.0
    nucleus_link_um: float = 0.55
    soma_zone_min_um: float = 2.1
    soma_zone_max_um: float = 5.2
    soma_zone_scale_process_rich: float = 2.15
    soma_zone_scale_compact: float = 2.85
    thickness_fraction_process_rich: float = 0.34
    thickness_fraction_compact: float = 0.24
    structural_percentile_process_rich: float = 58.0
    structural_percentile_compact: float = 46.0
    min_soma_area_um2: float = 4.5
    fallback_soma_radius_um: float = 1.55
    max_soma_fraction: float = 0.72
    min_process_fraction: float = 0.08
    ambiguity_score_delta: float = 0.08
    primary_anchor_min_score: float = 0.62
    primary_anchor_min_thickness_support: float = 0.60
    primary_anchor_min_structural_support: float = 0.34
    primary_anchor_min_overlap_fraction: float = 0.35
    primary_anchor_min_model_support: float = 0.55
    multi_anchor_min_score: float = 0.60
    multi_anchor_max_score_delta: float = 0.16
    multi_anchor_min_thickness_support: float = 0.65
    multi_anchor_min_structural_support: float = 0.35
    multi_anchor_min_overlap_fraction: float = 0.55
    multi_anchor_min_model_support: float = 0.60
    soma_anchor_min_separation_um: float = 6.0
    max_soma_anchors_per_whole_roi: int = 4
    soma_part_min_core_radius_um: float = 0.90
    soma_part_max_axis_ratio: float = 3.50
    soma_core_shell_max_um: float = 0.75
    soma_trusted_core_radius_scale: float = 1.55
    soma_trusted_core_min_um: float = 1.80
    soma_trusted_core_max_um: float = 3.60
    soma_trusted_core_nucleus_margin_um: float = 0.65
    soma_nucleus_shape_preserving: bool = False
    dapi_extent_percentile_floor: float = 68.0
    dapi_extent_low_high_ratio: float = 0.58
    dapi_extent_max_expand_um: float = 1.45
    dapi_extent_closing_um: float = 0.18
    instance_split_enabled: bool = True
    instance_split_min_anchor_score: float = 0.68
    instance_split_min_anchor_separation_um: float = 6.0
    instance_split_min_child_area_um2: float = 12.0
    instance_split_min_child_fraction: float = 0.12
    instance_split_strict_neck_core_ratio: float = 0.72
    instance_split_max_neck_core_ratio: float = 0.72
    instance_split_max_boundary_structural_ratio: float = 0.72
    instance_split_max_markers: int = 2
    instance_split_strategy: str = "pairwise_soma_anchor_split"
    branch_gap_restore_enabled: bool = True
    branch_gap_low_percentile: float = 25.0
    branch_gap_min_depth_um: float = 0.27
    branch_gap_nucleus_protect_um: float = 2.10
    branch_gap_min_area_um2: float = 0.14
    branch_gap_min_major_axis_um: float = 0.90
    branch_gap_min_eccentricity: float = 0.62
    morphology_outlier_filter_enabled: bool = True
    morphology_outlier_min_reference_count: int = 8
    morphology_outlier_robust_z: float = 3.50
    morphology_outlier_min_consensus: int = 3
    morphology_fragment_min_axis_ratio: float = 4.25
    morphology_fragment_max_branchpoints: int = 1
    morphology_fragment_max_core_radius_um: float = 0.95

@dataclass(frozen=True)
class AgeProfileDecision:
    profile: str
    source: str
    neonatal_score: float | None
    threshold: float
    confidence_margin: float | None
    tagged_files: tuple[str, ...]
    features: dict[str, float | int]

@dataclass(frozen=True)
class Neonatal3DConfig:
    """Calibrated object/surface gate for neonatal DAPI soma anchors."""

    candidate_link_um: float = 0.75
    crop_margin_um: float = 4.60
    dapi_xy_support_margin_um: float = 0.75
    dapi_active_profile_fraction: float = 0.22
    dapi_low_fraction: float = 0.28
    dapi_high_fraction: float = 0.48
    dapi_min_z_span_um: float = 0.35
    dapi_min_volume_um3: float = 0.45
    dapi_min_projection_overlap: float = 0.30
    egfp_smooth_um: float = 0.12
    shell_inner_um: float = 0.08
    shell_outer_um: float = 0.95
    background_inner_um: float = 1.35
    background_outer_um: float = 2.20
    surface_contact_um: float = 0.55
    central_slice_area_fraction: float = 0.45
    slice_support_threshold: float = 0.42
    angular_sector_count: int = 12
    angular_sector_support_fraction: float = 0.20
    min_surface_coverage: float = 0.34
    min_angular_coverage: float = 0.75
    min_z_support_fraction: float = 0.34
    min_shell_enrichment: float = 0.04
    min_enclosure_score: float = 0.48
    canonical_envelope_closing_xy_um: float = 0.22
    canonical_envelope_closing_z_um: float = 0.28
    canonical_crop_margin_um: float = 2.60
    canonical_min_peak_radius_um: float = 0.72
    canonical_peak_h_um: float = 0.38
    canonical_min_peak_separation_um: float = 2.35
    canonical_min_child_volume_um3: float = 3.0
    canonical_min_child_z_span_um: float = 0.45
    canonical_max_neck_peak_ratio: float = 0.58
    canonical_core_radius_fraction: float = 0.46
    canonical_max_instances_per_envelope: int = 4

@dataclass(frozen=True)
class DapiFragmentWorkloadLimits:
    """Versioned resource limits applied before DAPI fragment-job submission."""

    policy_version: str
    max_parent_fragments: int
    max_total_fragments: int
    max_parent_voxel_comparisons: int
    max_total_voxel_comparisons: int
    max_parent_result_payload_bytes_lower_bound: int
    max_pending_tasks: int = 24
    heartbeat_seconds: float = 5.0

@dataclass(frozen=True)
class DapiParentFragmentWorkload:
    parent_core_id: int
    parent_index_1based: int
    bbox_yx_0based: tuple[int, int, int, int]
    crop_shape_zyx: tuple[int, int, int]
    crop_voxels: int
    crop_xy_pixels: int
    fragment_count: int
    estimated_voxel_comparisons: int
    estimated_result_payload_bytes_lower_bound: int

class DapiFragmentWorkloadLimitExceeded(RuntimeError):
    """Expected safety stop for pathological DAPI fragment workloads."""

    def __init__(
        self,
        diagnostic: dict[str, object],
        diagnostic_path: Path | None,
    ) -> None:
        self.diagnostic = diagnostic
        self.diagnostic_path = diagnostic_path
        super().__init__(
            str(
                diagnostic.get(
                    "reason_code",
                    "DAPI_FRAGMENT_WORKLOAD_LIMIT_EXCEEDED",
                )
            )
        )

    def user_message(self) -> str:
        parent = self.diagnostic["offending_parent"]
        assert isinstance(parent, dict)
        shape = parent["crop_shape_zyx"]
        assert isinstance(shape, (list, tuple))
        selected_z = self.diagnostic["selected_z_range_1based"]
        assert isinstance(selected_z, (list, tuple))
        limit = self.diagnostic["limit"]
        observed = self.diagnostic["observed"]
        details = (
            str(self.diagnostic_path)
            if self.diagnostic_path is not None
            else "Diagnostic JSON could not be written."
        )
        return "\n".join(
            [
                "ANALYSIS STOPPED SAFELY — DAPI fragment workload limit exceeded",
                "",
                "The DAPI 3D check found an abnormal nuclear-fragment workload.",
                "Continuing could keep the CPU busy for a very long time, so the run",
                "was stopped before jobs were submitted for the offending region.",
                "",
                f"Selected Z range: {selected_z[0]}–{selected_z[1]}",
                f"Offending parent: {parent['parent_core_id']}",
                f"3D fragments: {int(parent['fragment_count']):,}",
                "Parent crop: "
                f"{int(shape[0]):,} × {int(shape[1]):,} × {int(shape[2]):,} voxels",
                "Estimated voxel checks: "
                f"{int(parent['estimated_voxel_comparisons']):,}",
                f"Triggered metric: {self.diagnostic['trigger_metric']} "
                f"({int(observed):,} > {int(limit):,})",
                f"Safety policy: {self.diagnostic['policy_version']}",
                "",
                "Protected state:",
                "• No fragment jobs were submitted for this parent.",
                "• The KCNN/KCNJ measurement stack was not loaded.",
                "• Fiji was not launched.",
                "• Existing production outputs were not replaced.",
                "",
                "This does not mean that the TIFF is corrupted, and it does not by",
                "itself prove acquisition overexposure. Common causes include DAPI",
                "saturation or clipping during conversion to 8-bit.",
                "",
                "Next step: inspect the original high-bit-depth DAPI data and the",
                "acquisition histogram. Re-export linear grayscale data if clipping",
                "occurred during export; reacquire or exclude the image if the",
                "original acquisition is saturated.",
                "",
                "Diagnostic:",
                str(self.diagnostic["reason_code"]),
                f"Details: {details}",
            ]
        )

@dataclass(frozen=True)
class NucleusOwnershipConfig:
    """Conservative 3D nucleus ownership and foreign-soma conflict rules."""

    fragment_radius_sum_factor: float = 1.25
    owner_min_overlap_um2: float = 0.60
    accepted_min_extent_overlap_fraction: float = 0.45
    foreign_min_overlap_um2: float = 1.60
    foreign_min_overlap_fraction: float = 0.45
    foreign_min_owner_overlap_ratio: float = 0.15
    foreign_min_component_dominance: float = 0.80
    accepted_min_volume_um3: float = 4.0
    unowned_min_volume_um3: float = 12.0
    unowned_min_enclosure_score: float = 0.20
    unowned_barrier_radius_um: float = 1.10
    unowned_barrier_inner_width_um: float = 0.18
    multi_owner_unowned_exclusion_um: float = 0.45
    marker_search_um: float = 0.75
    marker_radius_um: float = 0.25
    minimum_child_area_um2: float = 12.0
    minimum_owner_child_fraction: float = 0.12
    minimum_accepted_child_fraction: float = 0.08
    maximum_boundary_core_ratio: float = 0.86
    maximum_boundary_structural_ratio: float = 0.88
    fragment_bridge_max_gap_um: float = 0.38
    fragment_bridge_min_z_overlap_fraction: float = 0.35
    fragment_bridge_max_axis_ratio: float = 2.25
    projection_occlusion_enabled: bool = True
    projection_foreign_halo_um: float = 1.10
    projection_contact_stop_um: float = 0.12
    projection_min_foreign_volume_um3: float = 3.0
    projection_min_foreign_z_span_um: float = 0.45
    projection_min_foreign_extent_um2: float = 1.2
    projection_min_retained_child_um2: float = 12.0

@dataclass(frozen=True)
class AxialTruncationConfig:
    """Conservative selected-Z cuboid guard for partially observed nuclei."""

    enabled: bool = True
    guard_depth_um: float = 2.0
    boundary_band_um: float = 0.55
    min_inside_volume_fraction: float = 0.62
    min_inside_z_span_um: float = 0.50
    min_outside_continuation_um3: float = 0.35
    min_outside_to_inside_ratio: float = 0.12
    min_boundary_area_ratio: float = 0.24
    min_reference_nuclei: int = 4
    minimum_relative_volume: float = 0.34

@dataclass(frozen=True)
class CanonicalIdentityConfig:
    """Frozen-baseline gates for local 3D nucleus identity reconciliation."""

    enabled: bool = True
    minimum_extent_overlap_um2: float = 0.55
    minimum_extent_overlap_fraction: float = 0.16
    merge_contact_distance_um: float = 0.30
    maximum_merge_source_count: int = 3
    split_minimum_extent_overlap_fraction: float = 0.32
    satellite_max_volume_ratio: float = 0.38
    satellite_max_radius_sum_factor: float = 1.20
    satellite_min_z_overlap_fraction: float = 0.60

@dataclass(frozen=True)
class SomaNuclearCompletionConfig:
    """Conservative completion of a Soma to its already assigned 3D nucleus."""

    enabled: bool = True
    minimum_existing_extent_coverage: float = 0.70
    maximum_added_fraction_of_existing_soma: float = 0.15
    minimum_owner_overlap_um2: float = 0.55
    minimum_foreign_overlap_um2: float = 0.55
    minimum_foreign_overlap_fraction: float = 0.16
    maximum_local_foreign_distance_um: float = 1.10

@dataclass(frozen=True)
class Neonatal3DContext:
    dapi_stack: np.ndarray
    egfp_stack: np.ndarray
    z_start_0based: int
    z_end_0based_inclusive: int
    pixel_depth_um: float
    calibration_source: str
    structural_channel: str = "eGFP"

@dataclass
class ValidatedNucleusAnchors:
    accepted_core_mask_2d: np.ndarray
    accepted_extent_mask_2d: np.ndarray
    metrics: dict
    object_core_labels_2d: np.ndarray | None = None
    object_extent_labels_2d: np.ndarray | None = None
    dapi_valid_object_ids: tuple[int, ...] = ()
    accepted_object_ids: tuple[int, ...] = ()
    object_records: tuple[dict, ...] = ()
    nucleus_instance_core_labels_2d: np.ndarray | None = None
    nucleus_instance_extent_labels_2d: np.ndarray | None = None
    accepted_instance_ids: tuple[int, ...] = ()
    ambiguous_instance_ids: tuple[int, ...] = ()
    nucleus_instance_records: tuple[dict, ...] = ()
    source_object_to_instance_ids: dict[int, tuple[int, ...]] | None = None

@dataclass
class CanonicalNucleusResolution:
    core_labels_2d: np.ndarray
    extent_labels_2d: np.ndarray
    records: tuple[dict, ...]
    accepted_ids: tuple[int, ...]
    ambiguous_ids: tuple[int, ...]
    metrics: dict

@dataclass(frozen=True)
class CandidateBaseResult:
    mask: np.ndarray
    cellpose_mask: np.ndarray
    method_used: str
    error: str

@dataclass(frozen=True)
class CandidateWindowContext:
    projection_key: tuple[int, int, str]
    structural_key: tuple
    dapi_projection: np.ndarray
    structural_projections: dict[str, np.ndarray]
    structural_map: np.ndarray
    cellpose_mask: np.ndarray
    cellpose_note: str

@dataclass(frozen=True)
class Log1pGMMThreshold:
    threshold_raw: float
    threshold_log1p: float
    background_peak_raw: float
    background_peak_log1p: float
    background_sigma_log1p: float
    means_log1p: tuple[float, float]
    variances_log1p: tuple[float, float]
    weights: tuple[float, float]
    ashman_separation: float
    posterior_at_threshold: float
    clip_bounds_raw: tuple[float, float]
    converged: bool
    iterations: int
    qc_pass: bool
    qc_reasons: tuple[str, ...]


PRODUCT_DISPLAY_NAME = "Project Leap 2D"

ANALYSIS_CORE_NAME = "Project Leap 2D Analysis Core"

PIPELINE_NAME = ANALYSIS_CORE_NAME

TERMINAL_RULE = "/" * 112

_RUN_STARTED_AT: float | None = None

DEFAULT_INPUT_DIR = Path.home() / "Desktop" / "IHC IMAGE"

DEFAULT_OUT_DIR = DEFAULT_INPUT_DIR

WHOLE_OVERLAY_FILENAME = "IHC_2D_Whole_Astrocyte_Overlay.png"

SOMA_OVERLAY_FILENAME = "IHC_2D_Astrocyte_Soma_Overlay.png"

PROCESS_OVERLAY_FILENAME = "IHC_2D_Astrocyte_Processes_Overlay.png"

REPORT_FILENAME = "IHC_2D_Analysis_Report.txt"

WORKBOOK_FILENAME = "IHC_2D_Fluorescence_Results.xlsx"

DEBUG_WHOLE_OVERLAY_FILENAME = "IHC_2D_DEBUG_Whole_Overlay.png"

DEBUG_SOMA_OVERLAY_FILENAME = "IHC_2D_DEBUG_Soma_Overlay.png"

DEBUG_PROCESS_OVERLAY_FILENAME = "IHC_2D_DEBUG_Processes_Overlay.png"

DEBUG_REPORT_FILENAME = "IHC_2D_DEBUG_Report.txt"

DEBUG_STATE_FILENAME = "IHC_2D_DEBUG_Compartment_State.npz"

STRUCTURAL_CHANNELS = ("eGFP", "GFAP")

MEASUREMENT_CHANNELS = ("KCNN1", "KCNN2", "KCNN3", "KCNJ10")

CHANNEL_PATTERNS = {
    "DAPI": re.compile(r"(?<![a-z0-9])dapi(?![a-z0-9])", re.IGNORECASE),
    "eGFP": re.compile(r"(?<![a-z0-9])(?:e?gfp)(?![a-z0-9])", re.IGNORECASE),
    "GFAP": re.compile(r"(?<![a-z0-9])gfap(?![a-z0-9])", re.IGNORECASE),
    "KCNN1": re.compile(r"(?<![a-z0-9])(?:kcnn1|sk1)(?![a-z0-9])", re.IGNORECASE),
    "KCNN2": re.compile(r"(?<![a-z0-9])(?:kcnn2|sk2)(?![a-z0-9])", re.IGNORECASE),
    "KCNN3": re.compile(r"(?<![a-z0-9])(?:kcnn3|sk3)(?![a-z0-9])", re.IGNORECASE),
    "KCNJ10": re.compile(
        r"(?<![a-z0-9])(?:kcnj10|kir4(?:[._ -]?1)?)(?![a-z0-9])",
        re.IGNORECASE,
    ),
}

AGE_PROFILE_PATTERNS = {
    "neonatal": re.compile(r"(?<![a-z0-9])neonatal(?![a-z0-9])", re.IGNORECASE),
    "mature": re.compile(r"(?<![a-z0-9])mature(?![a-z0-9])", re.IGNORECASE),
}

AGE_PROFILE_THRESHOLD = 0.55

_CELLPOSE_MODEL = None

_CELLPOSE_DEVICE = None

_CELLPOSE_WORKING_MAX_SIDE = None

_CELLPOSE_EFFECTIVE_BATCH_SIZE: int | None = None

_CELLPOSE_MASK_CACHE: dict[tuple, tuple[np.ndarray, str]] = {}

CELLPOSE_BATCH_SIZE = 32

CELLPOSE_BATCH_FALLBACKS = (16, 8)

CELLPOSE_PRIOR_MIN_AREA_PX = 55

CANDIDATE_CPU_WORKERS = 12

DAPI_INVENTORY_CPU_WORKERS = 12

REVIEW_MERGE_MAX_SOMA_GAP_UM = 1.0

_EFFECTIVE_CANDIDATE_CPU_WORKERS = CANDIDATE_CPU_WORKERS

_EFFECTIVE_DAPI_INVENTORY_CPU_WORKERS = DAPI_INVENTORY_CPU_WORKERS

NEAR_DUPLICATE_CANDIDATE_IOU = 0.995

CHALLENGER_MIN_SCORE_MARGIN = 0.01

MORPHOLOGY_BASELINE_CANDIDATE_COUNT = 30

STRUCTURAL_REFINEMENT_CANDIDATE_COUNT = 30

DISTRIBUTIONAL_THRESHOLD_CANDIDATE_COUNT = 30

PRE_DISTRIBUTION_BASELINE_CANDIDATE_COUNT = (
    MORPHOLOGY_BASELINE_CANDIDATE_COUNT
    + STRUCTURAL_REFINEMENT_CANDIDATE_COUNT
)

TOTAL_CANDIDATE_COUNT = (
    PRE_DISTRIBUTION_BASELINE_CANDIDATE_COUNT
    + DISTRIBUTIONAL_THRESHOLD_CANDIDATE_COUNT
)

EXPECTED_Z_INTERVAL_COUNT = 5

PRE_DISTRIBUTION_BASELINE_PROFILES_PER_Z = 12

EXPECTED_PROFILES_PER_Z = 18

PRE_DISTRIBUTION_BASELINE_CANDIDATE_FAMILIES = (
    "process_sensitivity",
    "balanced_adaptive",
    "precision",
    "strict_merge",
    "channel_consensus",
    "topology_continuity",
)

EXPECTED_CANDIDATE_FAMILIES = (
    *PRE_DISTRIBUTION_BASELINE_CANDIDATE_FAMILIES,
    "distributional_threshold",
)

_BRANCH_FEATURE_CACHE: dict[
    tuple,
    dict[str, tuple[np.ndarray, np.ndarray, np.ndarray]],
] = {}

_CACHE_LOCK = threading.RLock()

_DAPI_NUCLEI_CACHE: dict[tuple, np.ndarray] = {}

_FULL_PERCENTILE_CACHE: dict[
    tuple, tuple[weakref.ReferenceType[np.ndarray], float]
] = {}

_FULL_SUM_CACHE: dict[tuple, tuple[weakref.ReferenceType[np.ndarray], float]] = {}

_TOP_HAT_CACHE: dict[tuple, tuple[np.ndarray, float]] = {}

_NORMALIZED_PROJECTION_CACHE: dict[
    tuple, tuple[weakref.ReferenceType[np.ndarray], np.ndarray]
] = {}

_CANDIDATE_BASE_CACHE: dict[tuple, "CandidateBaseResult"] = {}

_CANDIDATE_BASE_LOCKS: dict[tuple, threading.Lock] = {}

_CACHE_KEY_LOCKS: dict[tuple, threading.Lock] = {}

_DISTRIBUTION_MODEL_CACHE: dict[tuple, "Log1pGMMThreshold"] = {}

_DISTRIBUTION_MODEL_FAILURES: dict[tuple, str] = {}

_DISTRIBUTION_DIAGNOSTIC_CACHE: dict[tuple, dict[str, object]] = {}

_RUNTIME_TIMINGS: dict[str, object] = {
    "cellpose_model_init_seconds": 0.0,
    "cellpose_inference_events": [],
    "candidate_postprocess_seconds": [],
    "candidate_total_seconds": [],
    "candidate_stage_wall_seconds": 0.0,
    "rank_candidates_seconds": 0.0,
    "compartment_split_seconds": 0.0,
}

DAPI_FRAGMENT_WORKLOAD_LIMITS: DapiFragmentWorkloadLimits | None = (
    DapiFragmentWorkloadLimits(
        policy_version="DAPI-Fragment-Workload-Guard-v1",
        max_parent_fragments=2_600,
        max_total_fragments=7_500,
        max_parent_voxel_comparisons=27_000_000_000,
        max_total_voxel_comparisons=60_000_000_000,
        max_parent_result_payload_bytes_lower_bound=4 * 1024**3,
    )
)


def runtime_elapsed_seconds() -> float:
    if _RUN_STARTED_AT is None:
        return 0.0
    return float(time.perf_counter() - _RUN_STARTED_AT)

def print_terminal_stage(title: str, detail: str | None = None) -> None:
    print(f"\n{TERMINAL_RULE}", flush=True)
    print(f"{title} | elapsed={runtime_elapsed_seconds():.3f} s", flush=True)
    if detail:
        print(detail, flush=True)
    print(TERMINAL_RULE, flush=True)

def print_terminal_event(message: str) -> None:
    print(f"[elapsed={runtime_elapsed_seconds():.3f} s] {message}", flush=True)

def physical_memory_bytes() -> int | None:
    try:
        pages = int(os.sysconf("SC_PHYS_PAGES"))
        page_size = int(os.sysconf("SC_PAGE_SIZE"))
    except (AttributeError, OSError, TypeError, ValueError):
        return None
    total = pages * page_size
    return total if total > 0 else None

def select_candidate_cpu_workers(requested: int = CANDIDATE_CPU_WORKERS) -> int:
    """Select a stable preflight worker count without retrying completed jobs."""
    cpu_cap = max(1, min(int(requested), 12, int(os.cpu_count() or 1)))
    total_memory = physical_memory_bytes()
    if total_memory is None:
        memory_cap = cpu_cap
    else:
        gib = total_memory / float(1024**3)
        if gib >= 32.0:
            memory_cap = 12
        elif gib >= 24.0:
            memory_cap = 8
        elif gib >= 16.0:
            memory_cap = 6
        else:
            memory_cap = 4
    return max(1, min(cpu_cap, memory_cap))

def print_runtime_timing_summary(fiji_details: dict) -> None:
    events = _RUNTIME_TIMINGS["cellpose_inference_events"]
    candidate_times = _RUNTIME_TIMINGS["candidate_postprocess_seconds"]
    assert isinstance(events, list)
    assert isinstance(candidate_times, list)
    print_terminal_stage(
        "09 | FINAL RUNTIME SUMMARY",
        "Timing is displayed in Terminal only; no timing log is retained in the analysis folder.",
    )
    print(
        f"  Cellpose model initialization: "
        f"{float(_RUNTIME_TIMINGS['cellpose_model_init_seconds']):.3f} s "
        "(excluded from inference and morphology/postprocess timings)",
        flush=True,
    )
    if events:
        for index, event in enumerate(events, start=1):
            status = "ok" if event.get("success", False) else "failed"
            print(
                f"  Cellpose inference {index:02d}: {float(event['seconds']):.3f} s "
                f"({event['device']}, max_side={event['max_side']}, {status})",
                flush=True,
            )
        print(
            f"  Cellpose inference total: {sum(float(event['seconds']) for event in events):.3f} s "
            "(model initialization excluded; cache hits excluded)",
            flush=True,
        )
    else:
        print("  Cellpose inference: no uncached model.eval call", flush=True)
    if candidate_times:
        values = np.asarray(candidate_times, dtype=np.float64)
        print(
            f"  Candidate morphology/postprocess worker compute total: "
            f"{float(values.sum()):.3f} s; "
            f"median={float(np.median(values)):.3f} s; max={float(values.max()):.3f} s",
            flush=True,
        )
        print(
            f"  Candidate evaluation elapsed wall time: "
            f"{float(_RUNTIME_TIMINGS['candidate_stage_wall_seconds']):.3f} s",
            flush=True,
        )
    print(
        f"  rank_candidates: {float(_RUNTIME_TIMINGS['rank_candidates_seconds']):.3f} s",
        flush=True,
    )
    print(
        f"  Soma/Processes split: {float(_RUNTIME_TIMINGS['compartment_split_seconds']):.3f} s",
        flush=True,
    )
    startup = fiji_details.get("fiji_startup_seconds")
    measurement = fiji_details.get("measurement_seconds")
    review_wait = fiji_details.get("review_wait_seconds")
    if startup is not None:
        print(f"  Fiji startup to six-window review-ready: {float(startup):.3f} s", flush=True)
    if measurement is not None:
        print(f"  Fiji native measurement (Whole, Processes, Soma): {float(measurement):.3f} s", flush=True)
    if review_wait is not None:
        print(
            f"  Human review/decision wait excluded from performance comparison: {float(review_wait):.3f} s",
            flush=True,
        )
    print(f"  Complete pipeline elapsed: {runtime_elapsed_seconds():.3f} s", flush=True)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Generate the best Whole Astrocyte ROI from 90 candidates, split Soma and Processes, "
            "open six compartment views in Fiji, and measure one raw KCNN/KCNJ10 projection."
        )
    )
    parser.add_argument("--input-dir", type=Path, default=DEFAULT_INPUT_DIR)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUT_DIR)
    parser.add_argument(
        "--disable-cellpose",
        action="store_true",
        help="Debug fallback: disable Cellpose-SAM and use deterministic structural processing.",
    )
    parser.add_argument(
        "--skip-fiji",
        action="store_true",
        help="Debug only: stop after ROI selection and write a Python preview overlay.",
    )
    parser.add_argument(
        "--fiji-auto-continue",
        action="store_true",
        help="Skip the pre-measurement review dialog for controlled automated validation.",
    )
    parser.add_argument(
        "--fiji-timeout-minutes",
        type=float,
        default=120.0,
        help="Maximum time to wait for the Fiji display and measurement workflow.",
    )
    parser.add_argument(
        "--fiji-launcher",
        type=Path,
        default=None,
        help="Optional explicit Fiji launcher path.",
    )
    parser.add_argument(
        "--dapi-fragment-workload-preflight-only",
        action="store_true",
        help=(
            "Validation only: reconstruct and count 3D DAPI fragments without "
            "submitting fragment jobs, loading the measurement stack, or launching Fiji."
        ),
    )
    parser.add_argument(
        "--dapi-fragment-workload-json",
        type=Path,
        default=None,
        help=(
            "Optional DAPI fragment workload diagnostic JSON path. Required with "
            "--dapi-fragment-workload-preflight-only."
        ),
    )
    return parser.parse_args(argv)


def discover_channel_paths(input_dir: Path) -> tuple[dict[str, Path], list[str]]:
    if not input_dir.is_dir():
        raise FileNotFoundError(input_dir)

    candidates: dict[str, list[Path]] = {name: [] for name in CHANNEL_PATTERNS}
    ignored: list[str] = []
    tif_paths = sorted(
        path for path in input_dir.iterdir() if path.is_file() and path.suffix.lower() in {".tif", ".tiff"}
    )
    for path in tif_paths:
        matches = [name for name, pattern in CHANNEL_PATTERNS.items() if pattern.search(path.stem)]
        if not matches:
            ignored.append(f"{path.name}: no supported channel token")
            continue
        if len(matches) != 1:
            raise ValueError(f"Ambiguous channel tokens in {path.name}: {matches}")
        try:
            with tf.TiffFile(str(path)) as tif:
                series = tif.series[0]
                shape = tuple(int(x) for x in series.shape)
                axes = str(series.axes)
        except Exception as exc:
            raise ValueError(f"Unreadable TIFF {path.name}: {exc!r}") from exc
        if len(shape) != 3 or axes != "ZYX":
            raise ValueError(
                f"{path.name} must be a split single-channel ZYX stack; found axes={axes!r}, shape={shape}"
            )
        candidates[matches[0]].append(path)

    selected: dict[str, Path] = {}
    for channel, paths in candidates.items():
        if not paths:
            continue
        if len(paths) != 1:
            names = ", ".join(path.name for path in paths)
            raise ValueError(f"Exactly one {channel} stack is allowed; found: {names}")
        selected[channel] = paths[0]

    if "DAPI" not in selected:
        raise ValueError(f"No split DAPI Z-stack found in {input_dir}")
    if not any(channel in selected for channel in STRUCTURAL_CHANNELS):
        raise ValueError(f"No split eGFP or GFAP structural Z-stack found in {input_dir}")
    measurement = [channel for channel in MEASUREMENT_CHANNELS if channel in selected]
    if len(measurement) != 1:
        raise ValueError(
            "Exactly one measurement stack is required: KCNN1, KCNN2, KCNN3, or KCNJ10; "
            f"found {measurement or 'none'}"
        )
    return selected, ignored

def detect_filename_age_profile(paths: dict[str, Path]) -> AgeProfileDecision | None:
    """Use an explicit age token from ROI-defining channel filenames only."""

    evidence: dict[str, list[str]] = {name: [] for name in AGE_PROFILE_PATTERNS}
    for channel in ("DAPI", *STRUCTURAL_CHANNELS):
        path = paths.get(channel)
        if path is None:
            continue
        matched = [
            name
            for name, pattern in AGE_PROFILE_PATTERNS.items()
            if pattern.search(path.stem)
        ]
        if len(matched) > 1:
            raise ValueError(
                f"Conflicting age-profile tokens in one input filename: {path.name}"
            )
        if matched:
            evidence[matched[0]].append(path.name)

    detected = [name for name, filenames in evidence.items() if filenames]
    if len(detected) > 1:
        details = "; ".join(
            f"{name}={','.join(evidence[name])}" for name in detected
        )
        raise ValueError(
            "Input filenames contain conflicting neonatal/mature labels: " + details
        )
    if not detected:
        return None

    profile = detected[0]
    return AgeProfileDecision(
        profile=profile,
        source="filename",
        neonatal_score=None,
        threshold=AGE_PROFILE_THRESHOLD,
        confidence_margin=None,
        tagged_files=tuple(sorted(evidence[profile])),
        features={},
    )

def channel_mode(paths: dict[str, Path]) -> str:
    structural = [channel for channel in STRUCTURAL_CHANNELS if channel in paths]
    return "+".join(["DAPI", *structural])

def measurement_channel(paths: dict[str, Path]) -> str:
    channels = [channel for channel in MEASUREMENT_CHANNELS if channel in paths]
    if len(channels) != 1:
        raise ValueError(f"Expected exactly one measurement channel, found {channels}")
    return channels[0]

def validate_shared_geometry(metadata: dict[str, dict]) -> None:
    def signature(meta: dict) -> tuple:
        return (
            meta.get("pixel_width_um"),
            meta.get("pixel_height_um"),
            meta.get("pixel_depth_um"),
        )

    def signatures_match(left: tuple, right: tuple) -> bool:
        for left_value, right_value in zip(left, right):
            if left_value is None or right_value is None:
                if left_value is not None or right_value is not None:
                    return False
                continue
            if not math.isclose(
                float(left_value),
                float(right_value),
                rel_tol=1e-5,
                abs_tol=1e-9,
            ):
                return False
        return True

    reference = signature(metadata["DAPI"])
    mismatched = {
        channel: signature(channel_meta)
        for channel, channel_meta in metadata.items()
        if not signatures_match(signature(channel_meta), reference)
    }
    if mismatched:
        raise ValueError(
            f"Channel calibration metadata do not match DAPI {reference}: {mismatched}"
        )


@contextmanager
def analysis_lock():
    cache_root = Path.home() / "Library" / "Caches" / "IHC2DAnalysis"
    cache_root.mkdir(parents=True, exist_ok=True)
    lock_path = cache_root / "analysis.lock"
    handle = lock_path.open("a+", encoding="utf-8")
    try:
        try:
            fcntl.flock(handle.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
        except BlockingIOError as exc:
            raise RuntimeError(
                "Another IHC 2D Analysis process is already running. Finish or stop it before starting a new run."
            ) from exc
        handle.seek(0)
        handle.truncate()
        handle.write(f"pid={os.getpid()} started={time.strftime('%Y-%m-%d %H:%M:%S %Z')}\n")
        handle.flush()
        yield
    finally:
        try:
            fcntl.flock(handle.fileno(), fcntl.LOCK_UN)
        finally:
            handle.close()


def resolution_um_per_pixel(resolution: object, unit: object) -> float | None:
    if resolution is None or unit is None:
        return None
    try:
        if isinstance(resolution, (tuple, list)) and len(resolution) == 2:
            pixels_per_unit = float(resolution[0]) / float(resolution[1])
        else:
            pixels_per_unit = float(resolution)
        unit_code = int(unit)
    except (TypeError, ValueError, ZeroDivisionError):
        return None
    if pixels_per_unit <= 0:
        return None
    if unit_code == 2:  # inch
        return 25400.0 / pixels_per_unit
    if unit_code == 3:  # centimeter
        return 10000.0 / pixels_per_unit
    return None

def imagej_axis_scale_um(imagej_metadata: dict, axis: str) -> tuple[float | None, str | None]:
    """Read a calibrated SCIFIO/ImageJ axis scale without assuming isotropic voxels."""

    def convert_to_um(value: object, unit: object) -> float | None:
        try:
            scale = float(value)
        except (TypeError, ValueError):
            return None
        if not np.isfinite(scale) or scale <= 0:
            return None
        normalized_unit = (
            str(unit)
            .replace("\\u00B5", "u")
            .replace("µ", "u")
            .replace("μ", "u")
            .strip()
            .lower()
        )
        factors = {
            "um": 1.0,
            "micron": 1.0,
            "microns": 1.0,
            "micrometer": 1.0,
            "micrometers": 1.0,
            "micrometre": 1.0,
            "micrometres": 1.0,
            "nm": 0.001,
            "mm": 1000.0,
        }
        factor = factors.get(normalized_unit)
        return None if factor is None else scale * factor

    axis = axis.upper()
    if axis == "Z":
        spacing_um = convert_to_um(
            imagej_metadata.get("spacing"),
            imagej_metadata.get("unit"),
        )
        if spacing_um is not None:
            return spacing_um, "ImageJ spacing/unit"

    axes = [value.strip().upper() for value in str(imagej_metadata.get("axes", "")).split(",")]
    scales = [value.strip() for value in str(imagej_metadata.get("scales", "")).split(",")]
    units = [value.strip() for value in str(imagej_metadata.get("units", "")).split(",")]
    if axis not in axes or len(scales) != len(axes):
        return None, None
    index = axes.index(axis)
    unit = units[index] if len(units) == len(axes) else str(imagej_metadata.get("unit", ""))
    scale_um = convert_to_um(scales[index], unit)
    if scale_um is None:
        return None, None
    return scale_um, "SCIFIO axes/scales/units"

def read_meta(path: Path) -> dict:
    with tf.TiffFile(str(path)) as tif:
        series = tif.series[0]
        page0 = tif.pages[0]
        tags = page0.tags
        imagej_metadata = tif.imagej_metadata or {}
        x_resolution = tags["XResolution"].value if "XResolution" in tags else None
        y_resolution = tags["YResolution"].value if "YResolution" in tags else None
        resolution_unit = tags["ResolutionUnit"].value if "ResolutionUnit" in tags else None
        pixel_depth_um, pixel_depth_source = imagej_axis_scale_um(imagej_metadata, "Z")
        return {
            "path": str(path),
            "exists": path.exists(),
            "size_mb": round(path.stat().st_size / 1024 / 1024, 3),
            "shape": tuple(int(x) for x in series.shape),
            "axes": series.axes,
            "dtype": str(series.dtype),
            "pages": len(tif.pages),
            "image_width": int(tags["ImageWidth"].value) if "ImageWidth" in tags else None,
            "image_length": int(tags["ImageLength"].value) if "ImageLength" in tags else None,
            "bits_per_sample": tags["BitsPerSample"].value if "BitsPerSample" in tags else None,
            "x_resolution": str(x_resolution) if x_resolution is not None else None,
            "y_resolution": str(y_resolution) if y_resolution is not None else None,
            "resolution_unit": str(resolution_unit) if resolution_unit is not None else None,
            "pixel_width_um": resolution_um_per_pixel(x_resolution, resolution_unit),
            "pixel_height_um": resolution_um_per_pixel(y_resolution, resolution_unit),
            "pixel_depth_um": pixel_depth_um,
            "pixel_depth_source": pixel_depth_source,
            "imagej_metadata": {
                k: v
                for k, v in imagej_metadata.items()
                if k
                in {
                    "channels",
                    "slices",
                    "frames",
                    "unit",
                    "spacing",
                    "axes",
                    "scales",
                    "units",
                }
            },
        }

def load_stack(path: Path) -> np.ndarray:
    arr = tf.imread(str(path))
    if arr.ndim != 3:
        raise ValueError(f"Expected ZYX stack for {path}, got {arr.shape}")
    return arr


def robust01(img: np.ndarray, lo: float = 0.5, hi: float = 99.8) -> np.ndarray:
    x = img.astype(np.float32, copy=False)
    p0, p1 = np.percentile(x, [lo, hi])
    if p1 <= p0:
        return np.zeros_like(x, dtype=np.float32)
    y = (x - p0) / (p1 - p0)
    return np.clip(y, 0, 1).astype(np.float32)

def normalized_projection(img: np.ndarray) -> np.ndarray:
    """Return one immutable robust normalization for a shared projection."""
    key = ("normalized_projection", array_identity_key(img))
    with _CACHE_LOCK:
        entry = _NORMALIZED_PROJECTION_CACHE.get(key)
    if entry is not None and entry[0]() is img:
        return entry[1]
    with cache_key_lock(key):
        with _CACHE_LOCK:
            entry = _NORMALIZED_PROJECTION_CACHE.get(key)
        if entry is None or entry[0]() is not img:
            normalized = robust01(img)
            normalized.setflags(write=False)
            with _CACHE_LOCK:
                _NORMALIZED_PROJECTION_CACHE[key] = (
                    weakref.ref(img),
                    normalized,
                )
            return normalized
        return entry[1]

def array_identity_key(array: np.ndarray) -> tuple:
    return (
        int(array.__array_interface__["data"][0]),
        tuple(int(value) for value in array.shape),
        tuple(int(value) for value in array.strides),
        array.dtype.str,
    )

def cache_key_lock(key: tuple) -> threading.Lock:
    with _CACHE_LOCK:
        return _CACHE_KEY_LOCKS.setdefault(key, threading.Lock())

def full_array_percentile(array: np.ndarray, percentile: float) -> float:
    key = ("percentile", id(array), float(percentile))
    with _CACHE_LOCK:
        entry = _FULL_PERCENTILE_CACHE.get(key)
    if entry is not None and entry[0]() is array:
        return entry[1]
    with cache_key_lock(key):
        with _CACHE_LOCK:
            entry = _FULL_PERCENTILE_CACHE.get(key)
        if entry is None or entry[0]() is not array:
            value = float(np.percentile(array, percentile))
            with _CACHE_LOCK:
                _FULL_PERCENTILE_CACHE[key] = (weakref.ref(array), value)
            return value
        return entry[1]

def full_array_sum(array: np.ndarray) -> float:
    key = ("sum_float64", id(array))
    with _CACHE_LOCK:
        entry = _FULL_SUM_CACHE.get(key)
    if entry is not None and entry[0]() is array:
        return entry[1]
    with cache_key_lock(key):
        with _CACHE_LOCK:
            entry = _FULL_SUM_CACHE.get(key)
        if entry is None or entry[0]() is not array:
            value = float(np.sum(array, dtype=np.float64))
            with _CACHE_LOCK:
                _FULL_SUM_CACHE[key] = (weakref.ref(array), value)
            return value
        return entry[1]

def project(stack: np.ndarray, z0: int, z1: int, mode: str) -> np.ndarray:
    sub = stack[z0 : z1 + 1]
    if mode == "max":
        return sub.max(axis=0)
    if mode == "mean":
        return sub.mean(axis=0).astype(np.float32)
    if mode == "sum":
        return sub.sum(axis=0).astype(np.float32)
    raise ValueError(mode)

def z_profile(structural_stacks: dict[str, np.ndarray]) -> np.ndarray:
    # Normalize each marker profile before averaging so a brighter channel cannot
    # dominate Z selection merely because it has a wider acquisition range.
    normalized_profiles: list[np.ndarray] = []
    for stack in structural_stacks.values():
        profile = np.asarray(
            [np.percentile(stack[z], 99.2) for z in range(stack.shape[0])],
            dtype=np.float32,
        )
        low, high = np.percentile(profile, [5, 95])
        if high > low:
            profile = np.clip((profile - low) / (high - low), 0, 1)
        elif float(profile.max()) > 0:
            profile = profile / float(profile.max())
        else:
            profile = np.ones_like(profile)
        normalized_profiles.append(profile)
    return np.mean(normalized_profiles, axis=0).astype(np.float32)

def z_range_from_mode(mode: str, prof: np.ndarray) -> tuple[int, int]:
    n = len(prof)
    auto_match = re.fullmatch(r"auto_(\d+)", mode)
    if auto_match:
        width = min(n, max(1, int(auto_match.group(1))))
        smooth_profile = ndi.gaussian_filter1d(prof.astype(np.float64), sigma=1.0)
        window_activity = np.convolve(smooth_profile, np.ones(width), mode="valid")
        z0 = int(np.argmax(window_activity))
        return z0, z0 + width - 1
    if mode == "full":
        return 0, n - 1
    if mode == "middle_70":
        return int(round(n * 0.15)), int(round(n * 0.85)) - 1
    if mode == "middle_50":
        return int(round(n * 0.25)), int(round(n * 0.75)) - 1
    if mode == "active_wide":
        thr = np.percentile(prof, 55)
        idx = np.flatnonzero(prof >= thr)
    elif mode == "active_core":
        thr = np.percentile(prof, 70)
        idx = np.flatnonzero(prof >= thr)
    elif mode == "peak_window":
        center = int(np.argmax(prof))
        half = max(10, n // 8)
        return max(0, center - half), min(n - 1, center + half)
    elif mode == "peak_narrow":
        center = int(np.argmax(prof))
        half = max(7, n // 12)
        return max(0, center - half), min(n - 1, center + half)
    elif mode == "peak_tight":
        center = int(np.argmax(prof))
        half = max(5, n // 18)
        return max(0, center - half), min(n - 1, center + half)
    else:
        raise ValueError(mode)
    if len(idx) == 0:
        return 0, n - 1
    return max(0, int(idx.min()) - 2), min(n - 1, int(idx.max()) + 2)

def active_channel_weights(
    structural_projections: dict[str, np.ndarray],
    spec: TestSpec,
) -> dict[str, float]:
    requested = {"eGFP": spec.egfp_weight, "GFAP": spec.gfap_weight}
    weights = {
        channel: max(0.0, float(requested[channel]))
        for channel in STRUCTURAL_CHANNELS
        if channel in structural_projections
    }
    total = sum(weights.values())
    if total <= 0:
        equal = 1.0 / len(weights)
        return {channel: equal for channel in weights}
    return {channel: weight / total for channel, weight in weights.items()}

def structural_map(
    structural_projections: dict[str, np.ndarray],
    spec: TestSpec,
) -> np.ndarray:
    weights = active_channel_weights(structural_projections, spec)
    combined = sum(
        weights[channel] * normalized_projection(structural_projections[channel])
        for channel in weights
    )
    combined = combined / max(float(combined.max()), 1e-6)
    if spec.smooth_sigma > 0:
        combined = filters.gaussian(combined, sigma=spec.smooth_sigma, preserve_range=True)
    return np.clip(combined, 0, 1).astype(np.float32)

def smooth_outline_source(mask: np.ndarray, spec: TestSpec) -> np.ndarray:
    if spec.outline_smooth_sigma <= 0:
        return mask.astype(bool)
    smoothed = filters.gaussian(mask.astype(np.float32), sigma=spec.outline_smooth_sigma, preserve_range=True)
    outline_mask = smoothed >= 0.42
    outline_mask = morphology.binary_closing(outline_mask, footprint=morphology.disk(1))
    outline_mask = morphology.remove_small_objects(outline_mask, min_size=max(30, spec.min_area // 2))
    return outline_mask.astype(bool)

def draw_smooth_label_contours(
    rgb: np.ndarray,
    labels: np.ndarray,
    spec: TestSpec,
    color: tuple[int, int, int],
) -> np.ndarray:
    """Draw each instance separately so touching labels retain their shared boundary."""

    overlay = (np.clip(rgb, 0, 1) * 255).astype(np.uint8)
    for label_id in (int(value) for value in np.unique(labels) if int(value) > 0):
        source = smooth_outline_source(labels == label_id, spec)
        contours, hierarchy = cv2.findContours(
            source.astype(np.uint8) * 255,
            cv2.RETR_CCOMP,
            cv2.CHAIN_APPROX_NONE,
        )
        hierarchy_rows = hierarchy[0] if hierarchy is not None else []
        drawable_contours = []
        for index, contour in enumerate(contours):
            if len(contour) < 4:
                continue
            is_internal = len(hierarchy_rows) > index and hierarchy_rows[index][3] >= 0
            if is_internal and abs(cv2.contourArea(contour)) < spec.outline_hole_min_area:
                continue
            drawable_contours.append(contour)
        if drawable_contours:
            cv2.drawContours(
                overlay,
                drawable_contours,
                -1,
                color,
                2,
                lineType=cv2.LINE_AA,
            )
    return overlay

def make_fiji_like_composite(
    dapi: np.ndarray,
    structural_projections: dict[str, np.ndarray],
    measurement_projection: np.ndarray,
) -> np.ndarray:
    blue = exposure.adjust_gamma(robust01(dapi), gamma=0.85)
    measurement_red = exposure.adjust_gamma(robust01(measurement_projection), gamma=0.85)
    if "eGFP" in structural_projections and "GFAP" in structural_projections:
        egfp = exposure.adjust_gamma(robust01(structural_projections["eGFP"]), gamma=0.85)
        gfap = exposure.adjust_gamma(robust01(structural_projections["GFAP"]), gamma=0.85)
        red = np.clip(measurement_red + gfap, 0, 1)
        green = np.clip(gfap + egfp, 0, 1)
    else:
        only_channel = next(iter(structural_projections))
        red = measurement_red
        green = exposure.adjust_gamma(
            robust01(structural_projections[only_channel]),
            gamma=0.85,
        )
    return np.stack([red, green, blue], axis=-1).astype(np.float32)

def display_range(image: np.ndarray, low: float = 0.3, high: float = 99.8) -> list[float]:
    sample = image[::4, ::4]
    minimum, maximum = np.percentile(sample, [low, high])
    if maximum <= minimum:
        minimum = float(np.min(sample))
        maximum = float(np.max(sample))
    if maximum <= minimum:
        maximum = minimum + 1.0
    return [round(float(minimum), 6), round(float(maximum), 6)]


def clear_candidate_computation_caches() -> None:
    with _CACHE_LOCK:
        _DAPI_NUCLEI_CACHE.clear()
        _FULL_PERCENTILE_CACHE.clear()
        _FULL_SUM_CACHE.clear()
        _TOP_HAT_CACHE.clear()
        _NORMALIZED_PROJECTION_CACHE.clear()
        _CANDIDATE_BASE_CACHE.clear()
        _CANDIDATE_BASE_LOCKS.clear()
        _CACHE_KEY_LOCKS.clear()
        _BRANCH_FEATURE_CACHE.clear()
        _DISTRIBUTION_MODEL_CACHE.clear()
        _DISTRIBUTION_MODEL_FAILURES.clear()
        _DISTRIBUTION_DIAGNOSTIC_CACHE.clear()

def weighted_value_quantile(
    values: np.ndarray,
    weights: np.ndarray,
    fraction: float,
) -> float:
    values = np.asarray(values, dtype=np.float64)
    weights = np.asarray(weights, dtype=np.float64)
    if values.ndim != 1 or values.shape != weights.shape or values.size == 0:
        raise ValueError("Invalid weighted quantile inputs")
    if np.any(weights < 0) or not np.any(weights > 0):
        raise ValueError("Weighted quantile requires positive total weight")
    order = np.argsort(values, kind="stable")
    ordered_values = values[order]
    ordered_weights = weights[order]
    cumulative = np.cumsum(ordered_weights, dtype=np.float64)
    target = float(np.clip(fraction, 0.0, 1.0)) * float(cumulative[-1])
    position = int(np.searchsorted(cumulative, target, side="left"))
    return float(ordered_values[min(position, ordered_values.size - 1)])

def gaussian_posterior_intersection(
    means: np.ndarray | tuple[float, float],
    variances: np.ndarray | tuple[float, float],
    weights: np.ndarray | tuple[float, float],
) -> float:
    """Return the background-to-signal equal-posterior root between the means."""
    means_array = np.asarray(means, dtype=np.float64)
    variances_array = np.asarray(variances, dtype=np.float64)
    weights_array = np.asarray(weights, dtype=np.float64)
    if means_array.shape != (2,) or variances_array.shape != (2,) or weights_array.shape != (2,):
        raise ValueError("Two Gaussian components are required")
    order = np.argsort(means_array, kind="stable")
    means_array = means_array[order]
    variances_array = variances_array[order]
    weights_array = weights_array[order]
    if (
        not np.all(np.isfinite(means_array))
        or not np.all(np.isfinite(variances_array))
        or not np.all(np.isfinite(weights_array))
        or np.any(variances_array <= 0)
        or np.any(weights_array <= 0)
        or means_array[1] <= means_array[0]
    ):
        raise ValueError("Invalid Gaussian component parameters")

    mean0, mean1 = map(float, means_array)
    var0, var1 = map(float, variances_array)
    weight0, weight1 = map(float, weights_array)
    sigma0 = math.sqrt(var0)
    sigma1 = math.sqrt(var1)
    a = 0.5 / var1 - 0.5 / var0
    b = mean0 / var0 - mean1 / var1
    c = (
        -0.5 * mean0 * mean0 / var0
        + 0.5 * mean1 * mean1 / var1
        + math.log((weight0 / sigma0) / (weight1 / sigma1))
    )
    scale = max(abs(0.5 / var0), abs(0.5 / var1), 1.0)
    if abs(a) <= 1e-12 * scale:
        if abs(b) <= np.finfo(np.float64).eps:
            raise ValueError("Gaussian posteriors do not have a unique intersection")
        roots = [-c / b]
    else:
        discriminant = b * b - 4.0 * a * c
        tolerance = 1e-12 * max(b * b, abs(4.0 * a * c), 1.0)
        if discriminant < -tolerance:
            raise ValueError("Gaussian posteriors do not intersect")
        discriminant = max(discriminant, 0.0)
        square_root = math.sqrt(discriminant)
        roots = [
            (-b - square_root) / (2.0 * a),
            (-b + square_root) / (2.0 * a),
        ]
    interval_tolerance = 1e-10 * max(abs(mean0), abs(mean1), 1.0)
    valid_roots = sorted(
        root
        for root in roots
        if np.isfinite(root)
        and mean0 + interval_tolerance < root < mean1 - interval_tolerance
    )
    if len(valid_roots) != 1:
        raise ValueError(
            "Gaussian posteriors require exactly one intersection between component means"
        )
    root = float(valid_roots[0])
    epsilon = max((mean1 - mean0) * 1e-6, 1e-12)

    def signal_log_odds(value: float) -> float:
        background = (
            math.log(weight0 / sigma0)
            - 0.5 * (value - mean0) ** 2 / var0
        )
        signal = (
            math.log(weight1 / sigma1)
            - 0.5 * (value - mean1) ** 2 / var1
        )
        return signal - background

    if not (
        signal_log_odds(root - epsilon) < 0.0
        and signal_log_odds(root + epsilon) > 0.0
    ):
        raise ValueError(
            "Posterior intersection does not switch from background to signal"
        )
    return root

def fit_weighted_two_component_gmm(
    values: np.ndarray,
    counts: np.ndarray,
    *,
    max_iterations: int = 300,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, bool, int]:
    """Fit an exact weighted univariate GMM to unique log-intensity values."""
    x = np.asarray(values, dtype=np.float64)
    sample_weights = np.asarray(counts, dtype=np.float64)
    if x.ndim != 1 or x.shape != sample_weights.shape or x.size < 4:
        raise ValueError("Too few unique values for a two-component GMM")
    total_weight = float(sample_weights.sum(dtype=np.float64))
    if total_weight < 200 or np.any(sample_weights < 0):
        raise ValueError("Too few pixels for a two-component GMM")
    global_mean = float(np.sum(sample_weights * x, dtype=np.float64) / total_weight)
    global_variance = float(
        np.sum(sample_weights * (x - global_mean) ** 2, dtype=np.float64)
        / total_weight
    )
    if not np.isfinite(global_variance) or global_variance <= 0:
        raise ValueError("GMM requires a non-constant intensity distribution")
    variance_floor = max(global_variance * 1e-6, 1e-10)
    initial_quantiles = ((0.20, 0.75), (0.35, 0.85), (0.10, 0.65))
    best: tuple[float, np.ndarray, np.ndarray, np.ndarray, bool, int] | None = None

    for lower_fraction, upper_fraction in initial_quantiles:
        means = np.asarray(
            [
                weighted_value_quantile(x, sample_weights, lower_fraction),
                weighted_value_quantile(x, sample_weights, upper_fraction),
            ],
            dtype=np.float64,
        )
        if means[1] <= means[0]:
            continue
        hard_labels = np.argmin(np.abs(x[:, None] - means[None, :]), axis=1)
        component_weights = np.asarray(
            [float(sample_weights[hard_labels == index].sum(dtype=np.float64)) for index in range(2)],
            dtype=np.float64,
        )
        if np.any(component_weights <= 0):
            component_weights = np.asarray([0.70, 0.30], dtype=np.float64) * total_weight
        mixing = component_weights / total_weight
        variances = np.full(2, global_variance, dtype=np.float64)
        previous_log_likelihood = -np.inf
        converged = False
        completed_iterations = 0

        for iteration in range(1, max_iterations + 1):
            log_probabilities = []
            for component in range(2):
                variance = max(float(variances[component]), variance_floor)
                log_probabilities.append(
                    math.log(max(float(mixing[component]), 1e-12))
                    - 0.5
                    * (
                        math.log(2.0 * math.pi * variance)
                        + (x - float(means[component])) ** 2 / variance
                    )
                )
            log_probability_0, log_probability_1 = log_probabilities
            log_normalizer = np.logaddexp(log_probability_0, log_probability_1)
            responsibility_0 = np.exp(log_probability_0 - log_normalizer)
            responsibilities = np.column_stack(
                [responsibility_0, 1.0 - responsibility_0]
            )
            weighted_responsibilities = responsibilities * sample_weights[:, None]
            effective_counts = weighted_responsibilities.sum(axis=0, dtype=np.float64)
            if np.any(effective_counts <= max(1e-8 * total_weight, 1e-6)):
                break
            mixing = effective_counts / total_weight
            means = (
                (weighted_responsibilities * x[:, None]).sum(axis=0, dtype=np.float64)
                / effective_counts
            )
            variances = (
                (
                    weighted_responsibilities
                    * (x[:, None] - means[None, :]) ** 2
                ).sum(axis=0, dtype=np.float64)
                / effective_counts
            )
            variances = np.maximum(variances, variance_floor)
            log_likelihood = float(
                np.sum(sample_weights * log_normalizer, dtype=np.float64)
            )
            completed_iterations = iteration
            if (
                iteration >= 5
                and abs(log_likelihood - previous_log_likelihood)
                <= 1e-9 * (1.0 + abs(previous_log_likelihood))
            ):
                converged = True
                previous_log_likelihood = log_likelihood
                break
            previous_log_likelihood = log_likelihood

        if not np.isfinite(previous_log_likelihood):
            continue
        candidate = (
            previous_log_likelihood,
            means.copy(),
            variances.copy(),
            mixing.copy(),
            converged,
            completed_iterations,
        )
        if best is None or candidate[0] > best[0]:
            best = candidate
    if best is None:
        raise ValueError("All deterministic GMM initializations failed")
    _likelihood, means, variances, mixing, converged, iterations = best
    order = np.argsort(means, kind="stable")
    return (
        means[order],
        variances[order],
        mixing[order],
        bool(converged),
        int(iterations),
    )

def background_peak_log_model(
    log_values: np.ndarray,
    counts: np.ndarray,
    *,
    upper_bound: float,
    bins: int = 512,
) -> tuple[float, float]:
    """Estimate the low-intensity mode and FWHM sigma using true bin centers."""
    x = np.asarray(log_values, dtype=np.float64)
    sample_weights = np.asarray(counts, dtype=np.float64)
    if x.size < 4 or x.shape != sample_weights.shape or x[-1] <= x[0]:
        raise ValueError("Background-peak modeling requires a non-constant distribution")
    hist, edges = np.histogram(
        x,
        bins=max(64, int(bins)),
        range=(float(x[0]), float(x[-1])),
        weights=sample_weights,
    )
    centers = 0.5 * (edges[:-1] + edges[1:])
    smoothed = ndi.gaussian_filter1d(hist.astype(np.float64), sigma=2.0)
    allowed = np.flatnonzero(centers < float(upper_bound))
    if allowed.size < 3:
        raise ValueError("No low-intensity histogram range below the posterior threshold")
    peak_index = int(allowed[np.argmax(smoothed[allowed])])
    baseline = float(np.min(smoothed[allowed]))
    peak_height = float(smoothed[peak_index])
    if peak_height <= baseline:
        raise ValueError("Background peak has no measurable prominence")
    half_height = baseline + 0.5 * (peak_height - baseline)
    left_index = peak_index
    while left_index > 0 and smoothed[left_index] > half_height:
        left_index -= 1
    right_index = peak_index
    while right_index < smoothed.size - 1 and smoothed[right_index] > half_height:
        right_index += 1
    left_width = float(centers[peak_index] - centers[left_index])
    right_width = float(centers[right_index] - centers[peak_index])
    if left_width > 0 and right_width > 0:
        fwhm = left_width + right_width
    elif left_width > 0:
        fwhm = 2.0 * left_width
    elif right_width > 0:
        fwhm = 2.0 * right_width
    else:
        raise ValueError("Background peak width could not be estimated")
    bin_width = float(np.mean(np.diff(centers), dtype=np.float64))
    sigma = max(fwhm / 2.354820045, bin_width)
    return float(centers[peak_index]), float(sigma)

def distribution_model_cache_key(projection: np.ndarray) -> tuple:
    return (
        "log1p_gmm_posterior_v1",
        array_identity_key(projection),
        0.01,
        99.99,
        4096,
        512,
    )

def fit_log1p_gmm_threshold(projection: np.ndarray) -> Log1pGMMThreshold:
    values = np.asarray(projection, dtype=np.float64).reshape(-1)
    values = values[np.isfinite(values)]
    if values.size < 200:
        raise ValueError("Too few finite pixels for log1p GMM thresholding")
    clip_low, clip_high = np.percentile(values, [0.01, 99.99])
    if not np.isfinite(clip_low) or not np.isfinite(clip_high) or clip_high <= clip_low:
        raise ValueError("log1p GMM requires a non-constant clipped intensity range")
    clipped_log_values = np.log1p(np.clip(values, clip_low, clip_high))
    histogram, edges = np.histogram(
        clipped_log_values,
        bins=4096,
        range=(float(np.log1p(clip_low)), float(np.log1p(clip_high))),
    )
    centers = 0.5 * (edges[:-1] + edges[1:])
    populated = histogram > 0
    unique_log_values = centers[populated]
    counts = histogram[populated].astype(np.float64, copy=False)
    means, variances, weights, converged, iterations = fit_weighted_two_component_gmm(
        unique_log_values,
        counts,
    )
    threshold_log = gaussian_posterior_intersection(means, variances, weights)
    background_peak_log, background_sigma_log = background_peak_log_model(
        unique_log_values,
        counts,
        upper_bound=threshold_log,
    )
    variance_sum = float(variances[0] + variances[1])
    ashman_separation = float(
        math.sqrt(2.0) * (means[1] - means[0]) / math.sqrt(variance_sum)
    )
    log_densities = np.asarray(
        [
            math.log(float(weights[index]))
            - 0.5
            * (
                math.log(2.0 * math.pi * float(variances[index]))
                + (threshold_log - float(means[index])) ** 2
                / float(variances[index])
            )
            for index in range(2)
        ],
        dtype=np.float64,
    )
    posterior_at_threshold = float(
        math.exp(log_densities[1] - float(np.logaddexp(*log_densities)))
    )
    reasons: list[str] = []
    if not converged:
        reasons.append("gmm_not_converged")
    if float(np.min(weights)) < 0.005:
        reasons.append("component_weight_below_0.005")
    if ashman_separation < 1.0:
        reasons.append("component_separation_below_1.0")
    if not (float(means[0]) < threshold_log < float(means[1])):
        reasons.append("posterior_intersection_outside_means")
    if threshold_log <= background_peak_log + 0.25 * background_sigma_log:
        reasons.append("posterior_intersection_not_above_background_noise")
    if abs(posterior_at_threshold - 0.5) > 1e-6:
        reasons.append("posterior_intersection_numerically_inaccurate")
    threshold_raw = float(np.expm1(threshold_log))
    if not (float(clip_low) <= threshold_raw <= float(clip_high)):
        reasons.append("posterior_threshold_outside_clipped_range")
    return Log1pGMMThreshold(
        threshold_raw=threshold_raw,
        threshold_log1p=float(threshold_log),
        background_peak_raw=float(np.expm1(background_peak_log)),
        background_peak_log1p=float(background_peak_log),
        background_sigma_log1p=float(background_sigma_log),
        means_log1p=(float(means[0]), float(means[1])),
        variances_log1p=(float(variances[0]), float(variances[1])),
        weights=(float(weights[0]), float(weights[1])),
        ashman_separation=ashman_separation,
        posterior_at_threshold=posterior_at_threshold,
        clip_bounds_raw=(float(clip_low), float(clip_high)),
        converged=bool(converged),
        iterations=int(iterations),
        qc_pass=not reasons,
        qc_reasons=tuple(reasons),
    )

def get_log1p_gmm_threshold(projection: np.ndarray) -> Log1pGMMThreshold:
    key = distribution_model_cache_key(projection)
    with _CACHE_LOCK:
        cached = _DISTRIBUTION_MODEL_CACHE.get(key)
        cached_failure = _DISTRIBUTION_MODEL_FAILURES.get(key)
    if cached is not None:
        return cached
    if cached_failure is not None:
        raise ValueError(cached_failure)
    with cache_key_lock(key):
        with _CACHE_LOCK:
            cached = _DISTRIBUTION_MODEL_CACHE.get(key)
            cached_failure = _DISTRIBUTION_MODEL_FAILURES.get(key)
        if cached is not None:
            return cached
        if cached_failure is not None:
            raise ValueError(cached_failure)
        try:
            result = fit_log1p_gmm_threshold(projection)
        except Exception as exc:
            message = f"{type(exc).__name__}: {exc}"
            with _CACHE_LOCK:
                _DISTRIBUTION_MODEL_FAILURES[key] = message
            raise ValueError(message) from exc
        with _CACHE_LOCK:
            _DISTRIBUTION_MODEL_CACHE[key] = result
        return result

def distributional_threshold_diagnostics(
    struct: np.ndarray,
    structural_projections: dict[str, np.ndarray],
    spec: TestSpec,
) -> dict[str, object]:
    key = (
        "distributional_threshold_diagnostics_v1",
        array_identity_key(struct),
        tuple(sorted(structural_projections)),
        round(float(spec.threshold_scale), 12),
    )
    with _CACHE_LOCK:
        cached = _DISTRIBUTION_DIAGNOSTIC_CACHE.get(key)
    if cached is not None:
        return dict(cached)
    model_error = ""
    result: Log1pGMMThreshold | None = None
    try:
        result = get_log1p_gmm_threshold(struct)
    except Exception as exc:
        model_error = repr(exc)
    if result is not None and not result.qc_pass:
        model_error = ",".join(result.qc_reasons)
    valid = result is not None and result.qc_pass
    posterior_threshold = result.threshold_raw if result is not None else None
    scaled_threshold = (
        result.threshold_raw * float(spec.threshold_scale)
        if valid and result is not None
        else None
    )
    diagnostics: dict[str, object] = {
        "distribution_model": "log1p_two_component_gmm_true_posterior",
        "distribution_background_peak_role": "qc_only",
        "distribution_qc_pass": bool(valid),
        "distribution_valid_channels": ",".join(sorted(structural_projections)),
        "distribution_failed_channels": model_error,
        "distribution_thresholds_raw": scaled_threshold,
        "distribution_posterior_thresholds_raw": posterior_threshold,
        "distribution_background_peaks_raw": (
            result.background_peak_raw if result is not None else None
        ),
        "distribution_background_sigmas_log1p": (
            result.background_sigma_log1p if result is not None else None
        ),
        "distribution_ashman_separations": (
            result.ashman_separation if result is not None else None
        ),
        "distribution_component_parameters": (
            json.dumps(
                {
                    "means_log1p": result.means_log1p,
                    "variances_log1p": result.variances_log1p,
                    "weights": result.weights,
                    "posterior_at_threshold": result.posterior_at_threshold,
                    "iterations": result.iterations,
                },
                sort_keys=True,
            )
            if result is not None
            else ""
        ),
    }
    with _CACHE_LOCK:
        _DISTRIBUTION_DIAGNOSTIC_CACHE[key] = dict(diagnostics)
    return diagnostics

def log1p_gmm_mask(
    struct: np.ndarray,
    structural_projections: dict[str, np.ndarray],
    spec: TestSpec,
) -> np.ndarray:
    diagnostics = distributional_threshold_diagnostics(
        struct,
        structural_projections,
        spec,
    )
    if not bool(diagnostics["distribution_qc_pass"]):
        raise ValueError(
            "No structural channel passed log1p GMM/background-peak QC: "
            f"{diagnostics['distribution_failed_channels']}"
        )
    threshold = diagnostics["distribution_thresholds_raw"]
    if threshold is None or not np.isfinite(float(threshold)):
        raise ValueError("No valid structural-composite log1p GMM threshold")
    return np.asarray(struct, dtype=np.float64) >= float(threshold)

def threshold_mask(
    struct: np.ndarray,
    structural_projections: dict[str, np.ndarray],
    spec: TestSpec,
) -> np.ndarray:
    if spec.method == "otsu":
        thr = filters.threshold_otsu(struct) * spec.threshold_scale
        mask = struct >= thr
    elif spec.method == "yen":
        thr = filters.threshold_yen(struct) * spec.threshold_scale
        mask = struct >= thr
    elif spec.method == "li":
        thr = filters.threshold_li(struct) * spec.threshold_scale
        mask = struct >= thr
    elif spec.method == "sauvola":
        # Local thresholding is useful for uneven staining and projection background.
        local = filters.threshold_sauvola(struct, window_size=121, k=0.12)
        mask = struct >= (local * spec.threshold_scale)
    elif spec.method == "hysteresis":
        hi = full_array_percentile(struct, 88) * spec.threshold_scale
        lo = hi * 0.45
        mask = filters.apply_hysteresis_threshold(struct, lo, hi)
    elif spec.method == "dual_channel_union":
        masks = []
        for channel, projection in structural_projections.items():
            normalized = normalized_projection(projection)
            threshold_factor = 0.82 if channel == "eGFP" else 0.76
            threshold = filters.threshold_otsu(filters.gaussian(normalized, sigma=1.0))
            masks.append(normalized >= threshold * threshold_factor * spec.threshold_scale)
        mask = np.logical_or.reduce(masks)
    elif spec.method == "top_hat_union":
        weights = active_channel_weights(structural_projections, spec)
        cache_key = (
            "top_hat_union",
            tuple(
                (
                    channel,
                    round(float(weights[channel]), 12),
                    array_identity_key(structural_projections[channel]),
                )
                for channel in sorted(weights)
            ),
        )
        with _CACHE_LOCK:
            cached = _TOP_HAT_CACHE.get(cache_key)
        if cached is None:
            with cache_key_lock(cache_key):
                with _CACHE_LOCK:
                    cached = _TOP_HAT_CACHE.get(cache_key)
                if cached is None:
                    fp = morphology.disk(9)
                    base_mix = np.zeros_like(struct, dtype=np.float32)
                    top_hat_mix = np.zeros_like(struct, dtype=np.float32)
                    for channel, weight in weights.items():
                        normalized = normalized_projection(
                            structural_projections[channel]
                        )
                        base_mix += weight * normalized
                        top_hat_mix += weight * morphology.white_tophat(
                            normalized,
                            footprint=fp,
                        )
                    mix = np.clip(0.45 * base_mix + 0.55 * top_hat_mix, 0, None)
                    mix = (mix / max(float(mix.max()), 1e-6)).astype(
                        np.float32,
                        copy=False,
                    )
                    threshold = float(filters.threshold_otsu(mix))
                    mix.setflags(write=False)
                    cached = (mix, threshold)
                    with _CACHE_LOCK:
                        _TOP_HAT_CACHE[cache_key] = cached
        mix, threshold = cached
        mask = mix >= (threshold * spec.threshold_scale)
    elif spec.method == "log1p_gmm":
        mask = log1p_gmm_mask(struct, structural_projections, spec)
    else:
        raise ValueError(spec.method)
    return mask

def cleanup_mask(mask: np.ndarray, spec: TestSpec) -> np.ndarray:
    mask = morphology.remove_small_objects(mask.astype(bool), min_size=spec.min_area)
    if spec.close_radius > 0:
        mask = morphology.binary_closing(mask, footprint=morphology.disk(spec.close_radius))
    if spec.hole_area > 0:
        mask = morphology.remove_small_holes(mask, area_threshold=spec.hole_area)
    if spec.dilate_radius > 0:
        mask = morphology.binary_dilation(mask, footprint=morphology.disk(spec.dilate_radius))
    mask = morphology.remove_small_objects(mask, min_size=spec.min_area)
    # Avoid accepting a biologically implausible whole-field mask.
    frac = float(mask.mean())
    if frac > 0.42:
        labels = measure.label(mask)
        props = sorted(measure.regionprops(labels), key=lambda p: p.area, reverse=True)
        keep_labels = [p.label for p in props[:12] if p.area >= spec.min_area]
        mask = np.isin(labels, keep_labels)
    return mask.astype(bool)

def anchor_mask(candidate: np.ndarray, struct: np.ndarray, spec: TestSpec) -> np.ndarray:
    labels = measure.label(candidate)
    if labels.max() == 0:
        return np.zeros_like(candidate, dtype=bool)
    candidate_values = struct[candidate]
    if candidate_values.size == 0:
        return np.zeros_like(candidate, dtype=bool)
    mean_cut = float(np.percentile(candidate_values, 72))
    anchor_labels: list[int] = []
    for prop in measure.regionprops(labels, intensity_image=struct):
        if prop.area >= spec.anchor_area:
            anchor_labels.append(int(prop.label))
        elif prop.area >= max(120, spec.anchor_area // 5) and prop.mean_intensity >= mean_cut:
            anchor_labels.append(int(prop.label))
    anchor = np.isin(labels, anchor_labels)

    bright_seed = candidate & (
        struct >= full_array_percentile(struct, spec.seed_percentile)
    )
    bright_seed = morphology.remove_small_objects(bright_seed, min_size=spec.seed_min_area)
    return anchor | bright_seed

def anchor_connected_cleanup(
    candidate: np.ndarray,
    struct: np.ndarray,
    spec: TestSpec,
    extra_anchor: np.ndarray | None = None,
) -> np.ndarray:
    candidate = cleanup_mask(candidate, spec)
    anchors = anchor_mask(candidate, struct, spec)
    if extra_anchor is not None and extra_anchor.any():
        anchors |= extra_anchor.astype(bool)
    if not anchors.any():
        return candidate

    if spec.cleanup_mode == "anchor_bridge":
        expanded = morphology.binary_dilation(candidate, footprint=morphology.disk(spec.bridge_radius))
        labels = measure.label(expanded)
        anchor_labels = np.unique(labels[morphology.binary_dilation(anchors, footprint=morphology.disk(spec.bridge_radius))])
        anchor_labels = anchor_labels[anchor_labels != 0]
        keep_region = np.isin(labels, anchor_labels)
        cleaned = candidate & keep_region
    elif spec.cleanup_mode == "seed_reconstruct":
        low_mask = struct >= full_array_percentile(struct, spec.low_percentile)
        low_mask = morphology.remove_small_objects(low_mask, min_size=max(40, spec.min_area // 2))
        cleaned = ndi.binary_propagation(anchors, mask=low_mask)
    elif spec.cleanup_mode == "hybrid_reconstruct":
        low_mask = struct >= full_array_percentile(struct, spec.low_percentile)
        bridge = morphology.binary_dilation(candidate, footprint=morphology.disk(spec.bridge_radius))
        cleaned = ndi.binary_propagation(anchors, mask=(low_mask & bridge))
        cleaned |= candidate & morphology.binary_dilation(cleaned, footprint=morphology.disk(max(2, spec.bridge_radius // 2)))
    else:
        cleaned = candidate

    cleaned = cleanup_mask(cleaned, spec)
    if float(cleaned.mean()) > spec.max_area_fraction:
        # Revert to conservative anchor-bridge if reconstruction leaks into broad background.
        expanded = morphology.binary_dilation(candidate, footprint=morphology.disk(max(3, spec.bridge_radius // 2)))
        labels = measure.label(expanded)
        anchor_labels = np.unique(labels[anchors])
        anchor_labels = anchor_labels[anchor_labels != 0]
        cleaned = cleanup_mask(candidate & np.isin(labels, anchor_labels), spec)
    return cleaned.astype(bool)


def get_cellpose_model():
    global _CELLPOSE_MODEL, _CELLPOSE_DEVICE
    from cellpose import models
    import torch

    if _CELLPOSE_MODEL is not None:
        return _CELLPOSE_MODEL, _CELLPOSE_DEVICE

    np.random.seed(0)
    torch.manual_seed(0)

    if torch.backends.mps.is_available():
        device = torch.device("mps")
    else:
        device = torch.device("cpu")
    model_init_started = time.perf_counter()
    _CELLPOSE_MODEL = models.CellposeModel(
        gpu=(device.type != "cpu"),
        pretrained_model="cpsam_v2",
        device=device,
        use_bfloat16=False,
    )
    _RUNTIME_TIMINGS["cellpose_model_init_seconds"] = float(
        time.perf_counter() - model_init_started
    )
    _CELLPOSE_DEVICE = str(device)
    return _CELLPOSE_MODEL, _CELLPOSE_DEVICE

def synchronize_mps() -> None:
    try:
        import torch

        if torch.backends.mps.is_available():
            torch.mps.synchronize()
    except Exception:
        pass

def clear_mps_cache() -> None:
    try:
        import torch

        if torch.backends.mps.is_available():
            torch.mps.synchronize()
            torch.mps.empty_cache()
    except Exception:
        pass
    gc.collect()

def is_recoverable_cellpose_batch_error(exc: Exception) -> bool:
    message = repr(exc).lower()
    resource_markers = (
        "out of memory",
        "failed to allocate",
        "allocation failed",
        "resource exhausted",
        "insufficient memory",
        "recommended max working set size",
    )
    return any(marker in message for marker in resource_markers)

def run_cellpose_mask(struct: np.ndarray, spec: TestSpec, cache_key: tuple) -> tuple[np.ndarray, str]:
    global _CELLPOSE_WORKING_MAX_SIDE
    model, device_name = get_cellpose_model()
    if cache_key in _CELLPOSE_MASK_CACHE:
        return _CELLPOSE_MASK_CACHE[cache_key]

    sizes: list[int] = []
    if _CELLPOSE_WORKING_MAX_SIDE is not None:
        sizes.append(_CELLPOSE_WORKING_MAX_SIDE)
    else:
        sizes.append(spec.cellpose_max_side)
        for fallback in [1536, 1280, 1024]:
            if fallback < spec.cellpose_max_side:
                sizes.append(fallback)

    last_error = ""
    for max_side in sizes:
        try:
            print_terminal_event(
                "Cellpose-SAM inference started | "
                f"device={device_name} | max_side={max_side} | "
                f"diameter={spec.cellpose_diameter} | cellprob={spec.cellpose_cellprob}"
            )
            synchronize_mps()
            inference_started = time.perf_counter()
            try:
                mask, note = run_cellpose_mask_at_size(struct, spec, model, device_name, max_side)
            except Exception as exc:
                synchronize_mps()
                events = _RUNTIME_TIMINGS["cellpose_inference_events"]
                assert isinstance(events, list)
                events.append(
                    {
                        "seconds": float(time.perf_counter() - inference_started),
                        "device": str(device_name),
                        "max_side": int(max_side),
                        "diameter": float(spec.cellpose_diameter),
                        "cellprob": float(spec.cellpose_cellprob),
                        "success": False,
                        "error": repr(exc),
                    }
                )
                raise
            synchronize_mps()
            inference_seconds = time.perf_counter() - inference_started
            events = _RUNTIME_TIMINGS["cellpose_inference_events"]
            assert isinstance(events, list)
            events.append(
                {
                    "seconds": float(inference_seconds),
                    "device": str(device_name),
                    "max_side": int(max_side),
                    "diameter": float(spec.cellpose_diameter),
                    "cellprob": float(spec.cellpose_cellprob),
                    "success": True,
                }
            )
            print_terminal_event(
                f"Cellpose-SAM inference {len(events):02d} completed | "
                f"device={device_name} | max_side={max_side} | "
                f"batch={_CELLPOSE_EFFECTIVE_BATCH_SIZE} | "
                f"time={float(inference_seconds):.3f} s"
            )
            _CELLPOSE_WORKING_MAX_SIDE = max_side
            result = (mask, note)
            _CELLPOSE_MASK_CACHE[cache_key] = result
            return result
        except Exception as exc:
            last_error = repr(exc)
            continue
    raise RuntimeError(f"Cellpose-SAM failed at all attempted sizes: {last_error}")

def run_cellpose_mask_at_size(
    struct: np.ndarray,
    spec: TestSpec,
    model,
    device_name: str,
    max_side: int,
) -> tuple[np.ndarray, str]:
    global _CELLPOSE_EFFECTIVE_BATCH_SIZE
    scale = min(1.0, max_side / max(struct.shape))
    if scale < 1.0:
        small = transform.resize(
            struct,
            (int(round(struct.shape[0] * scale)), int(round(struct.shape[1] * scale))),
            preserve_range=True,
            anti_aliasing=True,
        ).astype(np.float32)
    else:
        small = struct.astype(np.float32, copy=False)
    img = (np.clip(small, 0, 1) * 255).astype(np.uint8)
    eval_kwargs = {
        "channels": [0, 0],
        "diameter": max(8.0, spec.cellpose_diameter * scale),
        "cellprob_threshold": spec.cellpose_cellprob,
        "flow_threshold": 0.4,
        "min_size": max(20, int(CELLPOSE_PRIOR_MIN_AREA_PX * scale * scale)),
    }
    batch_sizes: list[int] = []
    starting_batch = (
        _CELLPOSE_EFFECTIVE_BATCH_SIZE
        if _CELLPOSE_EFFECTIVE_BATCH_SIZE is not None
        else CELLPOSE_BATCH_SIZE
    )
    for batch_size in (starting_batch, *CELLPOSE_BATCH_FALLBACKS):
        if (
            batch_size > 0
            and batch_size <= starting_batch
            and batch_size not in batch_sizes
        ):
            batch_sizes.append(batch_size)
    last_batch_error: Exception | None = None
    for attempt_index, effective_batch_size in enumerate(batch_sizes):
        try:
            masks, flows, styles = model.eval(
                img,
                batch_size=effective_batch_size,
                **eval_kwargs,
            )[:3]
            break
        except Exception as batch_error:
            last_batch_error = batch_error
            if (
                attempt_index + 1 >= len(batch_sizes)
                or not is_recoverable_cellpose_batch_error(batch_error)
            ):
                raise
            clear_mps_cache()
            print(
                f"Cellpose-SAM batch_size={effective_batch_size} failed; retrying "
                f"the same input at batch_size={batch_sizes[attempt_index + 1]}: "
                f"{batch_error!r}",
                flush=True,
            )
    else:
        raise RuntimeError(f"Cellpose-SAM batch evaluation failed: {last_batch_error!r}")
    _CELLPOSE_EFFECTIVE_BATCH_SIZE = int(effective_batch_size)
    if masks is None:
        return np.zeros_like(struct, dtype=bool), f"cellpose_cpsam_v2_{device_name}_{max_side}_empty"
    mask_small = np.asarray(masks) > 0
    if scale < 1.0:
        mask = transform.resize(
            mask_small.astype(np.uint8),
            struct.shape,
            order=0,
            preserve_range=True,
            anti_aliasing=False,
        ).astype(bool)
    else:
        mask = mask_small.astype(bool)
    return (
        mask.astype(bool),
        f"cellpose_cpsam_v2_{device_name}_{max_side}_batch{effective_batch_size}",
    )


def dapi_nuclei_mask(dapi_proj: np.ndarray, percentile_floor: float | None = None) -> np.ndarray:
    key = (
        "dapi_nuclei",
        array_identity_key(dapi_proj),
        None if percentile_floor is None else float(percentile_floor),
    )
    with _CACHE_LOCK:
        cached = _DAPI_NUCLEI_CACHE.get(key)
    if cached is not None:
        return cached
    with cache_key_lock(key):
        with _CACHE_LOCK:
            cached = _DAPI_NUCLEI_CACHE.get(key)
        if cached is None:
            d = normalized_projection(dapi_proj)
            try:
                threshold = float(filters.threshold_otsu(d))
            except ValueError:
                threshold = full_array_percentile(d, 96)
            if percentile_floor is not None:
                threshold = max(
                    threshold,
                    full_array_percentile(d, percentile_floor),
                )
            cached = d >= threshold
            cached = morphology.binary_opening(
                cached,
                footprint=morphology.disk(1),
            )
            cached = morphology.remove_small_objects(cached, min_size=40).astype(bool)
            cached.setflags(write=False)
            with _CACHE_LOCK:
                _DAPI_NUCLEI_CACHE[key] = cached
        return cached

def dapi_nuclei_core_and_extent(
    dapi_proj: np.ndarray,
    mean_pixel_um: float,
    config: CompartmentConfig,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, dict]:
    """Return strict DAPI seeds and their lower-threshold, seed-connected extents."""

    dapi_norm = normalized_projection(dapi_proj)
    try:
        otsu_threshold = float(filters.threshold_otsu(dapi_norm))
    except ValueError:
        otsu_threshold = full_array_percentile(dapi_norm, 96)
    high_threshold = max(
        otsu_threshold,
        full_array_percentile(dapi_norm, config.dapi_percentile_floor),
    )
    nuclei_core = dapi_norm >= high_threshold
    nuclei_core = morphology.binary_opening(
        nuclei_core,
        footprint=morphology.disk(1),
    )
    nuclei_core = morphology.remove_small_objects(nuclei_core, min_size=40)

    low_threshold = min(
        high_threshold,
        max(
            full_array_percentile(dapi_norm, config.dapi_extent_percentile_floor),
            high_threshold * config.dapi_extent_low_high_ratio,
        ),
    )
    extent_domain = dapi_norm >= low_threshold
    nuclei_extent = ndi.binary_propagation(
        nuclei_core,
        structure=np.ones((3, 3), dtype=bool),
        mask=extent_domain,
    ).astype(bool)
    closing_radius_px = max(
        1,
        int(round(config.dapi_extent_closing_um / max(mean_pixel_um, 1e-4))),
    )
    nuclei_extent = morphology.binary_closing(
        nuclei_extent,
        footprint=morphology.disk(closing_radius_px),
    )
    nuclei_extent = morphology.remove_small_holes(
        nuclei_extent,
        area_threshold=max(16, int(round(0.20 / max(mean_pixel_um**2, 1e-6)))),
    )
    nuclei_extent = morphology.remove_small_objects(nuclei_extent, min_size=40)
    nuclei_extent |= nuclei_core
    return nuclei_core.astype(bool), nuclei_extent.astype(bool), dapi_norm, {
        "high_threshold": round(high_threshold, 6),
        "low_threshold": round(low_threshold, 6),
        "strict_core_px": int(nuclei_core.sum()),
        "reconstructed_extent_px": int(nuclei_extent.sum()),
    }


def longest_true_run(values: np.ndarray) -> int:
    best = 0
    current = 0
    for value in np.asarray(values, dtype=bool):
        if value:
            current += 1
            best = max(best, current)
        else:
            current = 0
    return best

def neonatal_3d_slice_angular_coverage(
    boundary: np.ndarray,
    covered: np.ndarray,
    sector_count: int,
    sector_support_fraction: float,
) -> float:
    coords = np.argwhere(boundary)
    if coords.size == 0:
        return 0.0
    center_y, center_x = coords.mean(axis=0)
    angles = np.mod(
        np.arctan2(coords[:, 0] - center_y, coords[:, 1] - center_x),
        2.0 * math.pi,
    )
    sectors = np.floor(angles * sector_count / (2.0 * math.pi)).astype(int)
    supported = 0
    observed = 0
    covered_values = covered[coords[:, 0], coords[:, 1]]
    for sector in range(sector_count):
        selected = sectors == sector
        if not np.any(selected):
            continue
        observed += 1
        supported += int(float(covered_values[selected].mean()) >= sector_support_fraction)
    return float(supported) / max(observed, 1)

def evaluate_nucleus_object_3d(
    nucleus_volume: np.ndarray,
    local_core: np.ndarray,
    local_dapi: np.ndarray,
    local_structural: np.ndarray,
    voxel_sampling: tuple[float, float, float],
    dapi_dynamic: float,
    dapi_min_dynamic: float,
    low_threshold: float,
    high_threshold: float,
    cfg: Neonatal3DConfig,
    projection_overlap_denominator: int,
) -> dict:
    voxel_volume_um3 = float(np.prod(voxel_sampling))
    volume_px = int(nucleus_volume.sum())
    volume_um3 = volume_px * voxel_volume_um3
    z_present = np.any(nucleus_volume, axis=(1, 2))
    z_indices = np.flatnonzero(z_present)
    z_span_um = (
        (int(z_indices[-1]) - int(z_indices[0]) + 1) * voxel_sampling[0]
        if z_indices.size
        else 0.0
    )
    projection = np.any(nucleus_volume, axis=0)
    projection_overlap = float((projection & local_core).sum()) / max(
        int(projection_overlap_denominator),
        1,
    )
    dapi_valid = bool(
        dapi_dynamic >= dapi_min_dynamic
        and volume_um3 >= cfg.dapi_min_volume_um3
        and z_span_um >= cfg.dapi_min_z_span_um
        and projection_overlap >= cfg.dapi_min_projection_overlap
    )

    surface_coverage = 0.0
    median_boundary_coverage = 0.0
    median_angular_coverage = 0.0
    z_support_fraction = 0.0
    shell_enrichment = -1.0
    radial_band_fraction = 0.0
    enclosure_score = 0.0
    structural_threshold = math.nan
    if dapi_valid:
        smooth_sigma = tuple(
            max(0.0, cfg.egfp_smooth_um / spacing) for spacing in voxel_sampling
        )
        smooth_structural = ndi.gaussian_filter(
            local_structural,
            sigma=smooth_sigma,
            mode="nearest",
        )
        distance_from_nucleus = ndi.distance_transform_edt(
            ~nucleus_volume,
            sampling=voxel_sampling,
        )
        shell = (
            (distance_from_nucleus >= cfg.shell_inner_um)
            & (distance_from_nucleus <= cfg.shell_outer_um)
            & ~nucleus_volume
        )
        background_shell = (
            (distance_from_nucleus >= cfg.background_inner_um)
            & (distance_from_nucleus <= cfg.background_outer_um)
        )
        background_values = smooth_structural[background_shell]
        if background_values.size < 32:
            background_values = smooth_structural[~nucleus_volume]
        background_median = float(np.percentile(background_values, 50.0))
        background_mad = float(
            1.4826 * np.median(np.abs(background_values - background_median))
        )
        local_high = float(np.percentile(smooth_structural, 99.0))
        structural_numeric_floor = (
            np.finfo(np.float32).eps
            * max(
                float(np.max(np.abs(smooth_structural))),
                float(np.finfo(np.float32).tiny),
            )
            * 16.0
        )
        structural_threshold = background_median + max(
            structural_numeric_floor,
            1.5 * background_mad,
            0.12 * max(local_high - background_median, 0.0),
        )
        positive = (smooth_structural > structural_threshold) & ~nucleus_volume
        if positive.any():
            distance_to_positive = ndi.distance_transform_edt(
                ~positive,
                sampling=voxel_sampling,
            )
        else:
            distance_to_positive = np.full(
                nucleus_volume.shape,
                np.inf,
                dtype=np.float32,
            )
        surface = nucleus_volume & ~ndi.binary_erosion(
            nucleus_volume,
            structure=ndi.generate_binary_structure(3, 1),
            border_value=0,
        )
        surface_coverage = (
            float((distance_to_positive[surface] <= cfg.surface_contact_um).mean())
            if surface.any()
            else 0.0
        )

        shell_values = smooth_structural[shell]
        shell_p75 = float(np.percentile(shell_values, 75.0)) if shell_values.size else 0.0
        background_p75 = float(np.percentile(background_values, 75.0))
        local_dynamic = max(
            local_high - background_median,
            structural_numeric_floor,
        )
        shell_enrichment = float(
            np.clip((shell_p75 - background_p75) / local_dynamic, -1.0, 1.0)
        )
        radial_support: list[bool] = []
        radial_edges = np.linspace(cfg.shell_inner_um, cfg.shell_outer_um, 4)
        for inner, outer in zip(radial_edges[:-1], radial_edges[1:]):
            band = (
                (distance_from_nucleus >= inner)
                & (distance_from_nucleus < outer)
                & ~nucleus_volume
            )
            radial_support.append(
                bool(band.any())
                and float(positive[band].mean()) >= 0.08
                and float(np.percentile(smooth_structural[band], 70.0))
                >= structural_threshold
            )
        radial_band_fraction = float(np.mean(radial_support)) if radial_support else 0.0

        areas = nucleus_volume.sum(axis=(1, 2))
        central = z_present & (
            areas >= cfg.central_slice_area_fraction * max(int(areas.max()), 1)
        )
        boundary_coverages: list[float] = []
        angular_coverages: list[float] = []
        supported_slices = np.zeros(nucleus_volume.shape[0], dtype=bool)
        for z_index in np.flatnonzero(central):
            plane = nucleus_volume[z_index]
            boundary = plane & ~morphology.binary_erosion(
                plane,
                footprint=morphology.disk(1),
            )
            if not boundary.any():
                continue
            positive_plane = positive[z_index]
            if positive_plane.any():
                distance_2d = ndi.distance_transform_edt(
                    ~positive_plane,
                    sampling=voxel_sampling[1:],
                )
                covered = boundary & (distance_2d <= cfg.surface_contact_um)
            else:
                covered = np.zeros_like(boundary, dtype=bool)
            boundary_coverage = float(covered.sum()) / max(int(boundary.sum()), 1)
            angular_coverage = neonatal_3d_slice_angular_coverage(
                boundary,
                covered,
                cfg.angular_sector_count,
                cfg.angular_sector_support_fraction,
            )
            boundary_coverages.append(boundary_coverage)
            angular_coverages.append(angular_coverage)
            supported_slices[z_index] = (
                boundary_coverage >= cfg.slice_support_threshold
                and angular_coverage >= cfg.min_angular_coverage
            )
        median_boundary_coverage = (
            float(np.median(boundary_coverages)) if boundary_coverages else 0.0
        )
        median_angular_coverage = (
            float(np.median(angular_coverages)) if angular_coverages else 0.0
        )
        z_support_fraction = float(longest_true_run(supported_slices)) / max(
            int(central.sum()),
            1,
        )

        surface_score = float(np.clip((surface_coverage - 0.20) / 0.55, 0.0, 1.0))
        angular_score = float(
            np.clip((median_angular_coverage - 0.20) / 0.65, 0.0, 1.0)
        )
        z_score = float(np.clip((z_support_fraction - 0.15) / 0.70, 0.0, 1.0))
        enrichment_score = float(
            np.clip((shell_enrichment + 0.02) / 0.35, 0.0, 1.0)
        )
        enclosure_score = (
            0.30 * surface_score
            + 0.25 * angular_score
            + 0.20 * z_score
            + 0.15 * enrichment_score
            + 0.10 * radial_band_fraction
        )

    criteria = {
        "surface": surface_coverage >= cfg.min_surface_coverage,
        "angular": median_angular_coverage >= cfg.min_angular_coverage,
        "z_continuity": z_support_fraction >= cfg.min_z_support_fraction,
        "shell_enrichment": shell_enrichment >= cfg.min_shell_enrichment,
    }
    accepted = bool(
        dapi_valid
        and enclosure_score >= cfg.min_enclosure_score
        and criteria["surface"]
        and criteria["angular"]
        and criteria["z_continuity"]
        and sum(criteria.values()) >= 3
    )
    failed: list[str] = []
    if not dapi_valid:
        failed.append("invalid_3d_dapi_object")
    if enclosure_score < cfg.min_enclosure_score:
        failed.append("low_enclosure_score")
    failed.extend(name for name, passed in criteria.items() if not passed)
    coords = np.argwhere(nucleus_volume)
    center_z, center_y, center_x = (
        coords.mean(axis=0) if coords.size else np.asarray([math.nan, math.nan, math.nan])
    )
    return {
        "accepted": accepted,
        "dapi_valid": dapi_valid,
        "reason": "accepted" if accepted else ",".join(dict.fromkeys(failed)),
        "center_z_local": float(center_z),
        "center_y_local": float(center_y),
        "center_x_local": float(center_x),
        "dapi_volume_um3": float(volume_um3),
        "dapi_z_span_um": float(z_span_um),
        "dapi_projection_overlap": float(projection_overlap),
        "surface_coverage": float(surface_coverage),
        "median_xy_boundary_coverage": float(median_boundary_coverage),
        "angular_coverage": float(median_angular_coverage),
        "z_support_fraction": float(z_support_fraction),
        "shell_enrichment": float(shell_enrichment),
        "radial_band_fraction": float(radial_band_fraction),
        "enclosure_score": float(enclosure_score),
        "dapi_low_threshold": float(low_threshold),
        "dapi_high_threshold": float(high_threshold),
        "structural_threshold": (
            float(structural_threshold) if np.isfinite(structural_threshold) else None
        ),
        "projection": projection,
    }

def evaluate_raw_inventory_object_3d(
    local_volume_id: int,
    volume_labels: np.ndarray,
    volume_count: int,
    local_core: np.ndarray,
    local_dapi: np.ndarray,
    local_structural: np.ndarray,
    extent_component_labels_crop: np.ndarray,
    voxel_sampling: tuple[float, float, float],
    dapi_dynamic: float,
    dapi_min_dynamic: float,
    low_threshold: float,
    high_threshold: float,
    cfg: Neonatal3DConfig,
) -> dict:
    """Evaluate one raw 3D DAPI object without assigning global IDs or labels."""

    object_volume = volume_labels == int(local_volume_id)
    object_projection = np.any(object_volume, axis=0)
    overlap_px = int((object_projection & local_core).sum())
    denominator = (
        int(local_core.sum())
        if int(volume_count) == 1
        else min(int(local_core.sum()), int(object_projection.sum()))
    )
    evaluated = evaluate_nucleus_object_3d(
        object_volume,
        local_core,
        local_dapi,
        local_structural,
        voxel_sampling,
        dapi_dynamic,
        dapi_min_dynamic,
        low_threshold,
        high_threshold,
        cfg,
        denominator,
    )
    z_coordinates = np.flatnonzero(np.any(object_volume, axis=(1, 2)))
    extent_component_values = extent_component_labels_crop[object_projection]
    extent_component_values = extent_component_values[
        extent_component_values > 0
    ]
    extent_component_id = (
        int(np.bincount(extent_component_values).argmax())
        if extent_component_values.size
        else 0
    )
    peak_map = np.max(
        np.where(object_volume, local_dapi, -np.inf),
        axis=0,
    )
    return {
        "local_volume_id": int(local_volume_id),
        "projection": object_projection,
        "overlap_px": overlap_px,
        "evaluated": evaluated,
        "z_coordinates": z_coordinates,
        "extent_component_id": extent_component_id,
        "peak_map": peak_map,
    }

def dapi_parent_fragment_workload(
    *,
    parent_core_id: int,
    parent_index_1based: int,
    bbox_yx_0based: tuple[int, int, int, int],
    volume_labels: np.ndarray,
    fragment_count: int,
) -> DapiParentFragmentWorkload:
    crop_shape = tuple(int(value) for value in volume_labels.shape)
    crop_voxels = int(volume_labels.size)
    crop_xy_pixels = int(crop_shape[1] * crop_shape[2])
    fragments = int(fragment_count)
    return DapiParentFragmentWorkload(
        parent_core_id=int(parent_core_id),
        parent_index_1based=int(parent_index_1based),
        bbox_yx_0based=tuple(int(value) for value in bbox_yx_0based),
        crop_shape_zyx=crop_shape,
        crop_voxels=crop_voxels,
        crop_xy_pixels=crop_xy_pixels,
        fragment_count=fragments,
        estimated_voxel_comparisons=int(fragments * crop_voxels),
        estimated_result_payload_bytes_lower_bound=int(
            fragments
            * crop_xy_pixels
            * (
                np.dtype(np.bool_).itemsize
                + np.dtype(np.float32).itemsize
            )
        ),
    )

def append_dapi_parent_fragment_workload(
    summary: dict[str, object],
    workload: DapiParentFragmentWorkload,
) -> dict[str, object]:
    record = asdict(workload)
    parent_records = summary["parent_records"]
    assert isinstance(parent_records, list)
    parent_records.append(record)
    summary["parents_linked_to_whole"] = int(
        summary["parents_linked_to_whole"]
    ) + 1
    summary["total_fragments"] = int(summary["total_fragments"]) + int(
        workload.fragment_count
    )
    summary["total_voxel_comparisons"] = int(
        summary["total_voxel_comparisons"]
    ) + int(workload.estimated_voxel_comparisons)
    summary["max_parent_fragments"] = max(
        int(summary["max_parent_fragments"]),
        int(workload.fragment_count),
    )
    summary["max_parent_voxel_comparisons"] = max(
        int(summary["max_parent_voxel_comparisons"]),
        int(workload.estimated_voxel_comparisons),
    )
    summary["max_parent_result_payload_bytes_lower_bound"] = max(
        int(summary["max_parent_result_payload_bytes_lower_bound"]),
        int(workload.estimated_result_payload_bytes_lower_bound),
    )
    return record

def dapi_fragment_workload_violation(
    summary: dict[str, object],
    parent: DapiParentFragmentWorkload,
    limits: DapiFragmentWorkloadLimits,
) -> tuple[str, int, int] | None:
    checks = (
        (
            "parent_result_payload_bytes_lower_bound",
            int(parent.estimated_result_payload_bytes_lower_bound),
            int(limits.max_parent_result_payload_bytes_lower_bound),
        ),
        (
            "parent_voxel_comparisons",
            int(parent.estimated_voxel_comparisons),
            int(limits.max_parent_voxel_comparisons),
        ),
        (
            "parent_fragments",
            int(parent.fragment_count),
            int(limits.max_parent_fragments),
        ),
        (
            "total_voxel_comparisons",
            int(summary["total_voxel_comparisons"]),
            int(limits.max_total_voxel_comparisons),
        ),
        (
            "total_fragments",
            int(summary["total_fragments"]),
            int(limits.max_total_fragments),
        ),
    )
    for metric, observed, limit in checks:
        if observed > limit:
            return metric, observed, limit
    return None

def atomic_write_dapi_fragment_workload_json(
    path: Path,
    payload: dict[str, object],
) -> None:
    resolved = path.expanduser().resolve()
    resolved.parent.mkdir(parents=True, exist_ok=True)
    temporary = (
        resolved.parent
        / f"temporary_{resolved.name}.{uuid.uuid4().hex}.tmp"
    )
    try:
        temporary.write_text(
            json.dumps(payload, indent=2, sort_keys=True),
            encoding="utf-8",
        )
        os.replace(temporary, resolved)
    finally:
        temporary.unlink(missing_ok=True)

def bounded_ordered_map(
    executor: ThreadPoolExecutor,
    function,
    arguments,
    *,
    max_pending: int,
    cancel_event: threading.Event,
    heartbeat_seconds: float,
    progress_callback=None,
) -> list:
    """Evaluate a bounded number of Futures while yielding input order."""

    if max_pending <= 0:
        raise ValueError("max_pending must be positive")
    if heartbeat_seconds <= 0:
        raise ValueError("heartbeat_seconds must be positive")
    indexed_arguments = iter(enumerate(arguments))
    pending: dict[object, int] = {}
    completed: dict[int, object] = {}
    results: list[object] = []
    next_yield_index = 0
    submitted_count = 0

    def fill_window() -> None:
        nonlocal submitted_count
        while len(pending) + len(completed) < max_pending:
            if cancel_event.is_set():
                return
            try:
                index, call_arguments = next(indexed_arguments)
            except StopIteration:
                return
            future = executor.submit(function, *call_arguments)
            pending[future] = int(index)
            submitted_count += 1

    fill_window()
    try:
        while pending:
            if cancel_event.is_set():
                raise RuntimeError("DAPI fragment evaluation was cancelled")
            done, _ = wait(
                tuple(pending),
                timeout=float(heartbeat_seconds),
                return_when=FIRST_COMPLETED,
            )
            if not done:
                if progress_callback is not None:
                    progress_callback(
                        {
                            "event": "heartbeat",
                            "submitted": int(submitted_count),
                            "yielded": int(next_yield_index),
                            "pending": int(len(pending)),
                        }
                    )
                continue
            for future in done:
                index = pending.pop(future)
                completed[int(index)] = future.result()
            while next_yield_index in completed:
                results.append(completed.pop(next_yield_index))
                next_yield_index += 1
                if progress_callback is not None:
                    progress_callback(
                        {
                            "event": "completed",
                            "submitted": int(submitted_count),
                            "yielded": int(next_yield_index),
                            "pending": int(len(pending)),
                        }
                    )
            fill_window()
    except BaseException:
        cancel_event.set()
        for future in pending:
            future.cancel()
        raise
    if completed:
        raise RuntimeError(
            "DAPI fragment scheduler retained out-of-order results"
        )
    return results

def separable_physical_binary_closing(
    mask: np.ndarray,
    radii_um: tuple[float, float, float],
    sampling: tuple[float, float, float],
) -> np.ndarray:
    """Close small intranuclear gaps with three bounded one-dimensional passes."""

    output = np.asarray(mask, dtype=bool)
    for axis, (radius_um, spacing_um) in enumerate(zip(radii_um, sampling)):
        radius_px = max(1, int(math.ceil(radius_um / spacing_um)))
        shape = [1, 1, 1]
        shape[axis] = 2 * radius_px + 1
        output = ndi.binary_closing(output, structure=np.ones(shape, dtype=bool))
    return output

def resolve_connected_nuclear_envelope(
    envelope: np.ndarray,
    voxel_sampling: tuple[float, float, float],
    config: Neonatal3DConfig,
) -> tuple[list[np.ndarray], str, dict]:
    """Resolve one connected DAPI envelope by shape basins, never intensity peaks."""

    distance_um = ndi.distance_transform_edt(envelope, sampling=voxel_sampling)
    maxima = morphology.h_maxima(distance_um, config.canonical_peak_h_um)
    maxima &= envelope & (distance_um >= config.canonical_min_peak_radius_um)
    maxima_labels = measure.label(maxima, connectivity=3)
    peak_rows: list[tuple[float, tuple[int, int, int]]] = []
    for prop in measure.regionprops(maxima_labels, intensity_image=distance_um):
        coordinates = prop.coords
        values = distance_um[tuple(coordinates.T)]
        best = coordinates[int(np.argmax(values))]
        peak_rows.append(
            (
                float(values.max()),
                (int(best[0]), int(best[1]), int(best[2])),
            )
        )
    selected: list[tuple[float, tuple[int, int, int]]] = []
    for row in sorted(peak_rows, reverse=True):
        coordinate = row[1]
        if any(
            math.sqrt(
                sum(
                    ((coordinate[axis] - other[1][axis]) * voxel_sampling[axis]) ** 2
                    for axis in range(3)
                )
            )
            < config.canonical_min_peak_separation_um
            for other in selected
        ):
            continue
        selected.append(row)
        if len(selected) >= config.canonical_max_instances_per_envelope:
            break
    if len(selected) < 2:
        return [envelope], "single", {
            "shape_peak_count": len(selected),
            "split_accepted": False,
            "neck_peak_ratio": None,
        }

    markers = np.zeros(envelope.shape, dtype=np.int32)
    for marker_id, (_peak, coordinate) in enumerate(selected, start=1):
        markers[coordinate] = marker_id
    partition = segmentation.watershed(
        -distance_um,
        markers=markers,
        mask=envelope,
        connectivity=np.ones((3, 3, 3), dtype=bool),
        watershed_line=False,
    )
    voxel_volume_um3 = float(np.prod(voxel_sampling))
    children = [partition == marker_id for marker_id in range(1, len(selected) + 1)]
    child_volumes = [int(child.sum()) * voxel_volume_um3 for child in children]
    child_spans = [
        int(np.any(child, axis=(1, 2)).sum()) * voxel_sampling[0]
        for child in children
    ]
    boundary = np.zeros(envelope.shape, dtype=bool)
    for axis in range(3):
        left = [slice(None)] * 3
        right = [slice(None)] * 3
        left[axis] = slice(1, None)
        right[axis] = slice(None, -1)
        different = (
            (partition[tuple(left)] > 0)
            & (partition[tuple(right)] > 0)
            & (partition[tuple(left)] != partition[tuple(right)])
        )
        boundary[tuple(left)] |= different
        boundary[tuple(right)] |= different
    minimum_peak = min(row[0] for row in selected)
    neck_peak_ratio = (
        float(np.percentile(distance_um[boundary], 75.0)) / max(minimum_peak, 1e-9)
        if boundary.any()
        else math.inf
    )
    quality_passed = bool(
        all(
            volume >= config.canonical_min_child_volume_um3
            for volume in child_volumes
        )
        and all(
            span >= config.canonical_min_child_z_span_um for span in child_spans
        )
        and neck_peak_ratio <= config.canonical_max_neck_peak_ratio
    )
    diagnostics = {
        "shape_peak_count": len(selected),
        "split_accepted": quality_passed,
        "child_volumes_um3": child_volumes,
        "child_z_spans_um": child_spans,
        "neck_peak_ratio": float(neck_peak_ratio),
    }
    if quality_passed:
        ordered = sorted(
            children,
            key=lambda child: tuple(np.argwhere(child).mean(axis=0)),
        )
        return ordered, "connected_object_split", diagnostics
    return [envelope], "ambiguous", diagnostics

def resolve_2d_nuclear_extent_families(
    nuclei_extent: np.ndarray,
    pixel_width_um: float,
    pixel_height_um: float,
    config: Neonatal3DConfig,
) -> np.ndarray:
    """Partition projection-connected DAPI extent into bounded shape-basin families."""

    connected_labels = measure.label(nuclei_extent, connectivity=2)
    output = np.zeros(nuclei_extent.shape, dtype=np.uint32)
    next_id = 1
    minimum_area_px = max(
        12,
        int(
            math.ceil(
                math.pi * config.canonical_min_peak_radius_um**2
                / (pixel_width_um * pixel_height_um)
            )
        ),
    )
    for prop in measure.regionprops(connected_labels):
        min_row, min_col, max_row, max_col = prop.bbox
        crop = np.s_[min_row:max_row, min_col:max_col]
        component = connected_labels[crop] == int(prop.label)
        distance_um = ndi.distance_transform_edt(
            component,
            sampling=(pixel_height_um, pixel_width_um),
        )
        maxima = morphology.h_maxima(distance_um, config.canonical_peak_h_um)
        maxima &= component & (distance_um >= config.canonical_min_peak_radius_um)
        maxima_labels = measure.label(maxima, connectivity=2)
        peaks: list[tuple[float, tuple[int, int]]] = []
        for maximum in measure.regionprops(maxima_labels, intensity_image=distance_um):
            coordinates = maximum.coords
            values = distance_um[tuple(coordinates.T)]
            best = coordinates[int(np.argmax(values))]
            peaks.append((float(values.max()), (int(best[0]), int(best[1]))))
        selected: list[tuple[float, tuple[int, int]]] = []
        for peak in sorted(peaks, reverse=True):
            coordinate = peak[1]
            if any(
                math.hypot(
                    (coordinate[0] - other[1][0]) * pixel_height_um,
                    (coordinate[1] - other[1][1]) * pixel_width_um,
                )
                < config.canonical_min_peak_separation_um
                for other in selected
            ):
                continue
            selected.append(peak)
        if len(selected) <= 1:
            local_output = output[crop]
            local_output[component] = next_id
            next_id += 1
            continue
        markers = np.zeros(component.shape, dtype=np.int32)
        for marker_id, (_value, coordinate) in enumerate(selected, start=1):
            markers[coordinate] = marker_id
        partition = segmentation.watershed(
            -distance_um,
            markers=markers,
            mask=component,
            watershed_line=False,
            connectivity=np.ones((3, 3), dtype=bool),
        )
        child_ids = [
            child_id
            for child_id in range(1, len(selected) + 1)
            if int((partition == child_id).sum()) >= minimum_area_px
        ]
        if len(child_ids) <= 1:
            local_output = output[crop]
            local_output[component] = next_id
            next_id += 1
            continue
        retained_markers = np.zeros(component.shape, dtype=np.int32)
        for marker_id, child_id in enumerate(child_ids, start=1):
            coordinate = selected[child_id - 1][1]
            retained_markers[coordinate] = marker_id
        partition = segmentation.watershed(
            -distance_um,
            markers=retained_markers,
            mask=component,
            watershed_line=False,
            connectivity=np.ones((3, 3), dtype=bool),
        )
        local_output = output[crop]
        for child_id in range(1, len(child_ids) + 1):
            local_output[partition == child_id] = next_id
            next_id += 1
    return output

def evaluate_canonical_extent_family_3d(
    family_id: int,
    bbox: tuple[int, int, int, int],
    extent_labels: np.ndarray,
    nuclei_core: np.ndarray,
    distance_to_whole: np.ndarray,
    dapi_substack: np.ndarray,
    structural_substack: np.ndarray,
    pixel_width_um: float,
    pixel_height_um: float,
    voxel_sampling: tuple[float, float, float],
    pad_y: int,
    pad_x: int,
    minimum_component_voxels: int,
    config: Neonatal3DConfig,
) -> dict | None:
    """Compute one canonical DAPI family without assigning global instance IDs."""

    min_row, min_col, max_row, max_col = bbox
    row0 = max(0, min_row - pad_y)
    col0 = max(0, min_col - pad_x)
    row1 = min(extent_labels.shape[0], max_row + pad_y)
    col1 = min(extent_labels.shape[1], max_col + pad_x)
    crop = np.s_[row0:row1, col0:col1]
    local_family = extent_labels[crop] == int(family_id)
    if (
        float(distance_to_whole[crop][local_family].min(initial=math.inf))
        > config.candidate_link_um
    ):
        return None
    local_core = nuclei_core[crop] & local_family
    if not local_core.any():
        return None
    support_distance = ndi.distance_transform_edt(
        ~local_family,
        sampling=(pixel_height_um, pixel_width_um),
    )
    support = support_distance <= config.dapi_xy_support_margin_um
    local_dapi = dapi_substack[:, row0:row1, col0:col1].astype(
        np.float32,
        copy=False,
    )
    local_structural = structural_substack[:, row0:row1, col0:col1].astype(
        np.float32,
        copy=False,
    )
    ring = support & (
        support_distance >= 0.50 * config.dapi_xy_support_margin_um
    )
    if not ring.any():
        ring = ~support
    background_values = local_dapi[:, ring] if ring.any() else local_dapi.ravel()
    dapi_background = float(np.percentile(background_values, 50.0))
    dapi_peak = float(np.percentile(local_dapi[:, local_core], 99.0))
    dapi_dynamic = max(dapi_peak - dapi_background, 0.0)
    field_dynamic = max(
        float(np.percentile(local_dapi, 99.9))
        - float(np.percentile(local_dapi, 0.1)),
        0.0,
    )
    numeric_floor = (
        np.finfo(np.float32).eps
        * max(
            float(np.max(np.abs(local_dapi))),
            float(np.finfo(np.float32).tiny),
        )
        * 16.0
    )
    minimum_dynamic = max(numeric_floor, 0.04 * field_dynamic)
    low_threshold = dapi_background + config.dapi_low_fraction * dapi_dynamic
    high_threshold = dapi_background + config.dapi_high_fraction * dapi_dynamic
    envelope = (local_dapi >= low_threshold) & support[None, :, :]
    envelope = separable_physical_binary_closing(
        envelope,
        (
            config.canonical_envelope_closing_z_um,
            config.canonical_envelope_closing_xy_um,
            config.canonical_envelope_closing_xy_um,
        ),
        voxel_sampling,
    )
    envelope = ndi.binary_fill_holes(envelope)
    envelope = morphology.remove_small_objects(
        envelope,
        min_size=minimum_component_voxels,
    )
    connected_labels = measure.label(envelope, connectivity=3)
    family_instances: list[tuple[np.ndarray, str, dict]] = []
    split_envelope_count = 0
    ambiguous_count = 0
    for connected_id in range(1, int(connected_labels.max()) + 1):
        connected = connected_labels == connected_id
        if int(connected.sum()) < minimum_component_voxels:
            continue
        children, resolution, diagnostics = resolve_connected_nuclear_envelope(
            connected,
            voxel_sampling,
            config,
        )
        split_envelope_count += int(resolution == "connected_object_split")
        ambiguous_count += int(resolution == "ambiguous")
        family_instances.extend(
            (child, resolution, diagnostics) for child in children
        )

    family_instances.sort(
        key=lambda item: tuple(np.argwhere(item[0]).mean(axis=0))
    )
    evaluated_instances: list[dict] = []
    for instance_volume, resolution, diagnostics in family_instances:
        projection = np.any(instance_volume, axis=0) & local_family
        if not projection.any():
            continue
        distance_um = ndi.distance_transform_edt(
            instance_volume,
            sampling=voxel_sampling,
        )
        peak_radius = float(distance_um.max(initial=0.0))
        interior = instance_volume & (
            distance_um
            >= config.canonical_core_radius_fraction
            * max(peak_radius, 1e-9)
        )
        core_projection = np.any(interior, axis=0) & projection
        if not core_projection.any():
            projection_coordinates = np.argwhere(projection)
            projection_center = projection_coordinates.mean(axis=0)
            nearest = projection_coordinates[
                int(
                    np.argmin(
                        np.square(
                            projection_coordinates - projection_center
                        ).sum(axis=1)
                    )
                )
            ]
            core_projection[int(nearest[0]), int(nearest[1])] = True
        evaluated = evaluate_nucleus_object_3d(
            instance_volume,
            projection,
            local_dapi,
            local_structural,
            voxel_sampling,
            dapi_dynamic,
            minimum_dynamic,
            low_threshold,
            high_threshold,
            config,
            int(projection.sum()),
        )
        coordinates = np.argwhere(instance_volume)
        z_coordinates = np.flatnonzero(
            np.any(instance_volume, axis=(1, 2))
        )
        evaluated_instances.append(
            {
                "resolution": resolution,
                "diagnostics": diagnostics,
                "projection": projection,
                "core_projection": core_projection,
                "evaluated": evaluated,
                "center_z_local": float(coordinates[:, 0].mean()),
                "center_y_local": float(coordinates[:, 1].mean()),
                "center_x_local": float(coordinates[:, 2].mean()),
                "z_min_local": int(z_coordinates.min()),
                "z_max_local": int(z_coordinates.max()),
                "volume_px": int(instance_volume.sum()),
            }
        )
    return {
        "family_id": int(family_id),
        "row0": int(row0),
        "col0": int(col0),
        "row1": int(row1),
        "col1": int(col1),
        "split_envelope_count": int(split_envelope_count),
        "ambiguous_count": int(ambiguous_count),
        "instances": evaluated_instances,
    }

def resolve_canonical_nucleus_instances_3d(
    frozen_whole_mask: np.ndarray,
    nuclei_core: np.ndarray,
    nuclei_extent: np.ndarray,
    context: Neonatal3DContext,
    pixel_width_um: float,
    pixel_height_um: float,
    config: Neonatal3DConfig,
    raw_object_extent_labels: np.ndarray | None = None,
    max_workers: int = 1,
) -> CanonicalNucleusResolution:
    """Resolve heterogeneous DAPI into one canonical identity per 3D nuclear envelope."""

    z0 = int(context.z_start_0based)
    z1 = int(context.z_end_0based_inclusive)
    dapi_substack = context.dapi_stack[z0 : z1 + 1]
    structural_substack = context.egfp_stack[z0 : z1 + 1]
    voxel_sampling = (
        float(context.pixel_depth_um),
        float(pixel_height_um),
        float(pixel_width_um),
    )
    voxel_volume_um3 = float(np.prod(voxel_sampling))
    extent_labels = resolve_2d_nuclear_extent_families(
        nuclei_extent,
        pixel_width_um,
        pixel_height_um,
        config,
    )
    distance_to_whole = ndi.distance_transform_edt(
        ~frozen_whole_mask.astype(bool),
        sampling=(pixel_height_um, pixel_width_um),
    )
    pad_y = max(2, int(math.ceil(config.canonical_crop_margin_um / pixel_height_um)))
    pad_x = max(2, int(math.ceil(config.canonical_crop_margin_um / pixel_width_um)))
    core_output = np.zeros(nuclei_extent.shape, dtype=np.uint32)
    extent_output = np.zeros(nuclei_extent.shape, dtype=np.uint32)
    records: list[dict] = []
    next_instance_id = 1
    ambiguous_count = 0
    split_envelope_count = 0

    minimum_component_voxels = max(
        1,
        int(math.ceil(config.dapi_min_volume_um3 / voxel_volume_um3)),
    )

    family_arguments = [
        (
            int(prop.label),
            tuple(int(value) for value in prop.bbox),
            extent_labels,
            nuclei_core,
            distance_to_whole,
            dapi_substack,
            structural_substack,
            pixel_width_um,
            pixel_height_um,
            voxel_sampling,
            pad_y,
            pad_x,
            minimum_component_voxels,
            config,
        )
        for prop in measure.regionprops(extent_labels)
    ]
    worker_count = max(1, min(int(max_workers), 12))
    if worker_count == 1:
        family_results = [
            evaluate_canonical_extent_family_3d(*arguments)
            for arguments in family_arguments
        ]
    else:
        with ThreadPoolExecutor(
            max_workers=min(worker_count, max(len(family_arguments), 1)),
            thread_name_prefix="ihc-dapi-canonical",
        ) as canonical_executor:
            canonical_futures = [
                canonical_executor.submit(
                    evaluate_canonical_extent_family_3d,
                    *arguments,
                )
                for arguments in family_arguments
            ]
            family_results = [
                future.result() for future in canonical_futures
            ]

    for family_result in family_results:
        if family_result is None:
            continue
        family_id = int(family_result["family_id"])
        row0 = int(family_result["row0"])
        col0 = int(family_result["col0"])
        row1 = int(family_result["row1"])
        col1 = int(family_result["col1"])
        crop = np.s_[row0:row1, col0:col1]
        split_envelope_count += int(
            family_result["split_envelope_count"]
        )
        ambiguous_count += int(family_result["ambiguous_count"])
        for instance_result in family_result["instances"]:
            resolution = str(instance_result["resolution"])
            diagnostics = instance_result["diagnostics"]
            projection = instance_result["projection"]
            core_projection = instance_result["core_projection"]
            evaluated = instance_result["evaluated"]
            instance_id = next_instance_id
            next_instance_id += 1
            core_view = core_output[crop]
            extent_view = extent_output[crop]
            unclaimed_extent = projection & (extent_view == 0)
            unclaimed_core = core_projection & (core_view == 0)
            extent_view[unclaimed_extent] = instance_id
            core_view[unclaimed_core] = instance_id
            extent_view[unclaimed_core] = instance_id
            claimed_projection = unclaimed_extent | unclaimed_core
            source_object_ids: tuple[int, ...] = ()
            if raw_object_extent_labels is not None:
                source_object_ids = tuple(
                    sorted(
                        int(value)
                        for value in np.unique(
                            raw_object_extent_labels[crop][claimed_projection]
                        )
                        if int(value) > 0
                    )
                )
            identity_status = "ambiguous" if resolution == "ambiguous" else "resolved"
            accepted = bool(evaluated["accepted"] and identity_status == "resolved")
            records.append(
                {
                    "instance_id": instance_id,
                    "object_id": instance_id,
                    "nucleus_id_2d": instance_id,
                    "object_id_3d": instance_id,
                    "source_object_ids": source_object_ids,
                    "extent_component_2d_id": family_id,
                    "resolution": resolution,
                    "identity_status": identity_status,
                    "accepted": accepted,
                    "dapi_valid": bool(evaluated["dapi_valid"]),
                    "center_z": float(
                        z0 + instance_result["center_z_local"]
                    ),
                    "center_y": float(
                        row0 + instance_result["center_y_local"]
                    ),
                    "center_x": float(
                        col0 + instance_result["center_x_local"]
                    ),
                    "z_min_0based": int(
                        z0 + instance_result["z_min_local"]
                    ),
                    "z_max_0based_inclusive": int(
                        z0 + instance_result["z_max_local"]
                    ),
                    "volume_um3": float(
                        int(instance_result["volume_px"])
                        * voxel_volume_um3
                    ),
                    "projection_area_px": int(claimed_projection.sum()),
                    "extent_area_px": int(claimed_projection.sum()),
                    "enclosure_score": float(evaluated["enclosure_score"]),
                    "dapi_low_threshold": float(
                        evaluated["dapi_low_threshold"]
                    ),
                    "dapi_high_threshold": float(
                        evaluated["dapi_high_threshold"]
                    ),
                    "dapi_z_span_um": float(evaluated["dapi_z_span_um"]),
                    "dapi_projection_overlap": float(evaluated["dapi_projection_overlap"]),
                    "surface_coverage": float(evaluated["surface_coverage"]),
                    "median_xy_boundary_coverage": float(
                        evaluated["median_xy_boundary_coverage"]
                    ),
                    "angular_coverage": float(evaluated["angular_coverage"]),
                    "z_support_fraction": float(evaluated["z_support_fraction"]),
                    "shell_enrichment": float(evaluated["shell_enrichment"]),
                    "radial_band_fraction": float(evaluated["radial_band_fraction"]),
                    "reason": (
                        "ambiguous_nuclear_envelope"
                        if identity_status == "ambiguous"
                        else str(evaluated["reason"])
                    ),
                    "resolution_diagnostics": diagnostics,
                }
            )

    accepted_ids = tuple(
        int(row["instance_id"]) for row in records if bool(row["accepted"])
    )
    ambiguous_ids = tuple(
        int(row["instance_id"])
        for row in records
        if row["identity_status"] == "ambiguous"
    )
    return CanonicalNucleusResolution(
        core_labels_2d=core_output,
        extent_labels_2d=extent_output,
        records=tuple(records),
        accepted_ids=accepted_ids,
        ambiguous_ids=ambiguous_ids,
        metrics={
            "method": (
                "canonical 3D nuclear envelope resolution by anisotropic distance basins; "
                "DAPI intensity peaks are not identity markers"
            ),
            "instance_count": len(records),
            "accepted_instance_count": len(accepted_ids),
            "ambiguous_instance_count": len(ambiguous_ids),
            "connected_envelope_split_count": split_envelope_count,
            "source_extent_family_count": int(extent_labels.max()),
        },
    )

def build_dapi_object_inventory_3d(
    frozen_whole_mask: np.ndarray,
    dapi_projection: np.ndarray,
    context: Neonatal3DContext,
    pixel_width_um: float,
    pixel_height_um: float,
    compartment_config: CompartmentConfig,
    validation_config: Neonatal3DConfig | None = None,
    max_workers: int = 1,
    preflight_only: bool = False,
    workload_limits: DapiFragmentWorkloadLimits | None = None,
    workload_diagnostic_path: Path | None = None,
) -> ValidatedNucleusAnchors:
    """Preserve and independently validate every reconstructed 3D DAPI object."""

    inventory_started = time.perf_counter()
    cfg = validation_config or Neonatal3DConfig()
    if context.pixel_depth_um <= 0:
        raise ValueError("Positive Z calibration is required for 3D nucleus inventory")
    z0 = int(context.z_start_0based)
    z1 = int(context.z_end_0based_inclusive)
    if z0 < 0 or z1 < z0 or z1 >= context.dapi_stack.shape[0]:
        raise ValueError(f"Invalid 3D nucleus inventory Z range: {z0}-{z1}")
    if context.dapi_stack.shape != context.egfp_stack.shape:
        raise ValueError("DAPI and structural stacks must have identical shapes")
    limits = (
        workload_limits
        if workload_limits is not None
        else DAPI_FRAGMENT_WORKLOAD_LIMITS
    )
    workload_summary: dict[str, object] = {
        "schema_version": 1,
        "mode": "preflight_only" if preflight_only else "enforce",
        "status": "running",
        "policy_version": (
            limits.policy_version if limits is not None else "calibration_only"
        ),
        "z_start_1based": int(z0 + 1),
        "z_end_1based_inclusive": int(z1 + 1),
        "parents_seen": 0,
        "parents_linked_to_whole": 0,
        "total_fragments": 0,
        "total_voxel_comparisons": 0,
        "max_parent_fragments": 0,
        "max_parent_voxel_comparisons": 0,
        "max_parent_result_payload_bytes_lower_bound": 0,
        "guard_triggered": False,
        "guard_reason": None,
        "parent_records": [],
    }
    mean_pixel_um = math.sqrt(pixel_width_um * pixel_height_um)
    nuclei_core, nuclei_extent, _, dapi_2d_metrics = dapi_nuclei_core_and_extent(
        dapi_projection,
        mean_pixel_um,
        compartment_config,
    )
    core_labels = measure.label(nuclei_core, connectivity=2)
    extent_component_labels = measure.label(nuclei_extent, connectivity=2)
    object_core_labels = np.zeros(nuclei_core.shape, dtype=np.uint32)
    object_extent_labels = np.zeros(nuclei_core.shape, dtype=np.uint32)
    if int(core_labels.max()) == 0:
        workload_summary["status"] = (
            "preflight_completed" if preflight_only else "completed"
        )
        return ValidatedNucleusAnchors(
            np.zeros_like(nuclei_core, dtype=bool),
            np.zeros_like(nuclei_extent, dtype=bool),
            {
                "status": "completed_no_2d_dapi_candidates",
                "method": "object-preserving calibrated 3D DAPI inventory",
                "measurement_channel_used": False,
                "z_start_1based": z0 + 1,
                "z_end_1based_inclusive": z1 + 1,
                "voxel_size_um": [
                    round(context.pixel_depth_um, 9),
                    round(pixel_height_um, 9),
                    round(pixel_width_um, 9),
                ],
                "calibration_source": context.calibration_source,
                "structural_channel": context.structural_channel,
                "parent_2d_core_count": 0,
                "candidate_count": 0,
                "accepted_count": 0,
                "rejected_count": 0,
                "dapi_valid_count": 0,
                "multi_object_parent_core_count": 0,
                "per_nucleus": [],
                "config": asdict(cfg),
                "dapi_2d": dapi_2d_metrics,
                "dapi_fragment_workload": workload_summary,
            },
            object_core_labels,
            object_extent_labels,
        )

    core_distance, core_nearest_indices = ndi.distance_transform_edt(
        ~nuclei_core,
        sampling=(pixel_height_um, pixel_width_um),
        return_indices=True,
    )
    nearest_core_labels = core_labels[
        core_nearest_indices[0],
        core_nearest_indices[1],
    ]
    distance_to_whole = ndi.distance_transform_edt(
        ~frozen_whole_mask.astype(bool),
        sampling=(pixel_height_um, pixel_width_um),
    )
    dapi_substack = context.dapi_stack[z0 : z1 + 1]
    structural_substack = context.egfp_stack[z0 : z1 + 1]
    voxel_sampling = (
        float(context.pixel_depth_um),
        float(pixel_height_um),
        float(pixel_width_um),
    )
    pad_y = max(2, int(math.ceil(cfg.crop_margin_um / pixel_height_um)))
    pad_x = max(2, int(math.ceil(cfg.crop_margin_um / pixel_width_um)))
    per_object: list[dict] = []
    object_records: list[dict] = []
    next_object_id = 1
    multi_object_parent_core_count = 0
    worker_count = max(1, min(int(max_workers), 12))

    for parent_index, prop in enumerate(
        measure.regionprops(core_labels),
        start=1,
    ):
        workload_summary["parents_seen"] = int(
            workload_summary["parents_seen"]
        ) + 1
        parent_core_id = int(prop.label)
        min_row, min_col, max_row, max_col = prop.bbox
        row0 = max(0, min_row - pad_y)
        col0 = max(0, min_col - pad_x)
        row1 = min(core_labels.shape[0], max_row + pad_y)
        col1 = min(core_labels.shape[1], max_col + pad_x)
        crop = np.s_[row0:row1, col0:col1]
        local_core = core_labels[crop] == parent_core_id
        if (
            float(distance_to_whole[crop][local_core].min(initial=math.inf))
            > cfg.candidate_link_um
        ):
            continue
        local_extent = (
            nuclei_extent[crop]
            & (nearest_core_labels[crop] == parent_core_id)
            & (core_distance[crop] <= compartment_config.dapi_extent_max_expand_um)
        )
        local_extent |= local_core
        local_support_distance = ndi.distance_transform_edt(
            ~(local_core | local_extent),
            sampling=(pixel_height_um, pixel_width_um),
        )
        xy_support = local_support_distance <= cfg.dapi_xy_support_margin_um
        local_dapi = dapi_substack[:, row0:row1, col0:col1].astype(
            np.float32,
            copy=False,
        )
        local_structural = structural_substack[:, row0:row1, col0:col1].astype(
            np.float32,
            copy=False,
        )

        profile = np.percentile(local_dapi[:, local_core], 85.0, axis=1)
        profile_baseline = float(np.percentile(profile, 10.0))
        profile_peak = float(np.max(profile))
        profile_contrast = profile_peak - profile_baseline
        active = profile >= (
            profile_baseline + cfg.dapi_active_profile_fraction * max(profile_contrast, 0.0)
        )
        active[int(np.argmax(profile))] = True
        active = ndi.binary_closing(active, structure=np.ones(3, dtype=bool))

        support_ring = xy_support & (
            local_support_distance >= 0.50 * cfg.dapi_xy_support_margin_um
        )
        if not support_ring.any():
            support_ring = ~xy_support
        background_values = (
            local_dapi[:, support_ring] if support_ring.any() else local_dapi.ravel()
        )
        dapi_background = float(np.percentile(background_values, 50.0))
        dapi_peak = float(np.percentile(local_dapi[:, local_core], 99.0))
        dapi_dynamic = max(dapi_peak - dapi_background, 0.0)
        dapi_field_dynamic = max(
            float(np.percentile(local_dapi, 99.9))
            - float(np.percentile(local_dapi, 0.1)),
            0.0,
        )
        dapi_numeric_floor = (
            np.finfo(np.float32).eps
            * max(
                float(np.max(np.abs(local_dapi))),
                float(np.finfo(np.float32).tiny),
            )
            * 16.0
        )
        dapi_min_dynamic = max(dapi_numeric_floor, 0.04 * dapi_field_dynamic)
        low_threshold = dapi_background + cfg.dapi_low_fraction * dapi_dynamic
        high_threshold = dapi_background + cfg.dapi_high_fraction * dapi_dynamic
        low_domain = (
            (local_dapi >= low_threshold)
            & xy_support[None, :, :]
            & active[:, None, None]
        )
        seeds = (
            (local_dapi >= high_threshold)
            & local_core[None, :, :]
            & active[:, None, None]
        )
        if not seeds.any() and dapi_dynamic > 0:
            seed_values = np.where(
                local_core[None, :, :] & active[:, None, None],
                local_dapi,
                -np.inf,
            )
            seed = np.unravel_index(int(np.argmax(seed_values)), seed_values.shape)
            seeds[seed] = True
        nucleus_volume = (
            ndi.binary_propagation(
                seeds,
                structure=np.ones((3, 3, 3), dtype=bool),
                mask=low_domain,
            ).astype(bool)
            if seeds.any()
            else np.zeros_like(low_domain, dtype=bool)
        )
        volume_labels = measure.label(nucleus_volume, connectivity=3)
        local_entries: list[dict] = []
        volume_count = int(volume_labels.max())
        multi_object_parent_core_count += int(volume_count > 1)
        parent_workload = dapi_parent_fragment_workload(
            parent_core_id=parent_core_id,
            parent_index_1based=parent_index,
            bbox_yx_0based=(row0, col0, row1, col1),
            volume_labels=volume_labels,
            fragment_count=volume_count,
        )
        parent_workload_record = append_dapi_parent_fragment_workload(
            workload_summary,
            parent_workload,
        )
        if preflight_only:
            continue
        if limits is not None:
            violation = dapi_fragment_workload_violation(
                workload_summary,
                parent_workload,
                limits,
            )
            if violation is not None:
                trigger_metric, observed, limit = violation
                workload_summary["status"] = "blocked"
                workload_summary["guard_triggered"] = True
                workload_summary["guard_reason"] = trigger_metric
                diagnostic: dict[str, object] = {
                    "schema_version": 1,
                    "status": "blocked",
                    "guard_triggered": True,
                    "reason_code": (
                        "DAPI_FRAGMENT_WORKLOAD_LIMIT_EXCEEDED"
                    ),
                    "analysis_stage": "dapi_fragment_pre_submission",
                    "policy_version": limits.policy_version,
                    "trigger_metric": trigger_metric,
                    "observed": int(observed),
                    "limit": int(limit),
                    "selected_z_range_1based": [int(z0 + 1), int(z1 + 1)],
                    "offending_parent": parent_workload_record,
                    "jobs_submitted_for_offending_parent": False,
                    "measurement_stack_loaded": False,
                    "fiji_launched": False,
                    "production_outputs_replaced": False,
                    "workload_summary": workload_summary,
                }
                if workload_diagnostic_path is not None:
                    atomic_write_dapi_fragment_workload_json(
                        workload_diagnostic_path,
                        diagnostic,
                    )
                raise DapiFragmentWorkloadLimitExceeded(
                    diagnostic,
                    workload_diagnostic_path,
                )
        extent_component_labels_crop = extent_component_labels[crop]
        task_arguments = (
            (
                local_volume_id,
                volume_labels,
                volume_count,
                local_core,
                local_dapi,
                local_structural,
                extent_component_labels_crop,
                voxel_sampling,
                dapi_dynamic,
                dapi_min_dynamic,
                low_threshold,
                high_threshold,
                cfg,
            )
            for local_volume_id in range(1, volume_count + 1)
        )
        if worker_count == 1 or volume_count <= 1:
            local_results = [
                evaluate_raw_inventory_object_3d(*arguments)
                for arguments in task_arguments
            ]
        else:
            cancel_event = threading.Event()
            max_pending = min(
                int(volume_count),
                int(
                    limits.max_pending_tasks
                    if limits is not None
                    else max(1, 2 * worker_count)
                ),
            )
            heartbeat_seconds = float(
                limits.heartbeat_seconds if limits is not None else 5.0
            )
            progress_step = max(1, int(math.ceil(volume_count / 10)))

            def report_dapi_fragment_progress(
                progress: dict[str, object],
            ) -> None:
                yielded = int(progress["yielded"])
                if (
                    progress["event"] == "heartbeat"
                    or yielded == volume_count
                    or yielded % progress_step == 0
                ):
                    print_terminal_event(
                        "DAPI fragments | "
                        f"parent={parent_index}/{int(core_labels.max())} | "
                        f"completed={yielded}/{volume_count} | "
                        f"pending={int(progress['pending'])}"
                    )

            with ThreadPoolExecutor(
                max_workers=min(worker_count, volume_count),
                thread_name_prefix="ihc-dapi-object",
            ) as object_executor:
                local_results = bounded_ordered_map(
                    object_executor,
                    evaluate_raw_inventory_object_3d,
                    task_arguments,
                    max_pending=max_pending,
                    cancel_event=cancel_event,
                    heartbeat_seconds=heartbeat_seconds,
                    progress_callback=report_dapi_fragment_progress,
                )
        for object_result in local_results:
            object_projection = object_result["projection"]
            overlap_px = int(object_result["overlap_px"])
            evaluated = object_result["evaluated"]
            z_coordinates = object_result["z_coordinates"]
            extent_component_id = int(object_result["extent_component_id"])
            peak_map = object_result["peak_map"]
            object_id = next_object_id
            next_object_id += 1
            local_entries.append(
                {
                    "object_id": object_id,
                    "projection": object_projection,
                    "peak_map": peak_map,
                    "evaluated": evaluated,
                }
            )
            per_object.append(
                {
                    "nucleus_id_2d": object_id,
                    "object_id_3d": object_id,
                    "parent_core_2d_id": parent_core_id,
                    "parent_core_object_count": volume_count,
                    "accepted": bool(evaluated["accepted"]),
                    "dapi_valid": bool(evaluated["dapi_valid"]),
                    "reason": str(evaluated["reason"]),
                    "center_z": round(z0 + evaluated["center_z_local"], 3),
                    "center_y": round(row0 + evaluated["center_y_local"], 3),
                    "center_x": round(col0 + evaluated["center_x_local"], 3),
                    "dapi_volume_um3": round(evaluated["dapi_volume_um3"], 6),
                    "dapi_z_span_um": round(evaluated["dapi_z_span_um"], 6),
                    "dapi_projection_overlap": round(
                        evaluated["dapi_projection_overlap"],
                        6,
                    ),
                    "surface_coverage": round(evaluated["surface_coverage"], 6),
                    "median_xy_boundary_coverage": round(
                        evaluated["median_xy_boundary_coverage"],
                        6,
                    ),
                    "angular_coverage": round(evaluated["angular_coverage"], 6),
                    "z_support_fraction": round(
                        evaluated["z_support_fraction"],
                        6,
                    ),
                    "shell_enrichment": round(evaluated["shell_enrichment"], 6),
                    "radial_band_fraction": round(
                        evaluated["radial_band_fraction"],
                        6,
                    ),
                    "enclosure_score": round(evaluated["enclosure_score"], 6),
                    "dapi_low_threshold": round(evaluated["dapi_low_threshold"], 6),
                    "dapi_high_threshold": round(evaluated["dapi_high_threshold"], 6),
                    "egfp_threshold": (
                        round(float(evaluated["structural_threshold"]), 6)
                        if evaluated["structural_threshold"] is not None
                        else None
                    ),
                    "structural_channel": context.structural_channel,
                    "raw_projection_overlap_px": overlap_px,
                    "z_min_1based": int(z0 + z_coordinates.min() + 1),
                    "z_max_1based_inclusive": int(z0 + z_coordinates.max() + 1),
                    "projection_area_px": int(object_projection.sum()),
                    "extent_component_2d_id": extent_component_id,
                }
            )
            object_records.append(
                {
                    "object_id": object_id,
                    "parent_core_2d_id": parent_core_id,
                    "extent_component_2d_id": extent_component_id,
                    "accepted": bool(evaluated["accepted"]),
                    "dapi_valid": bool(evaluated["dapi_valid"]),
                    "center_z": float(z0 + evaluated["center_z_local"]),
                    "center_y": float(row0 + evaluated["center_y_local"]),
                    "center_x": float(col0 + evaluated["center_x_local"]),
                    "z_min_0based": int(z0 + z_coordinates.min()),
                    "z_max_0based_inclusive": int(z0 + z_coordinates.max()),
                    "volume_um3": float(evaluated["dapi_volume_um3"]),
                    "projection_area_px": int(object_projection.sum()),
                    "enclosure_score": float(evaluated["enclosure_score"]),
                }
            )

        label_entries = [
            entry for entry in local_entries if bool(entry["evaluated"]["dapi_valid"])
        ]
        if not label_entries:
            continue
        strengths = np.stack(
            [np.asarray(entry["peak_map"], dtype=np.float32) for entry in label_entries],
            axis=0,
        )
        supported = np.isfinite(strengths)
        assigned_index = np.argmax(strengths, axis=0)
        any_supported = np.any(supported, axis=0)
        projection_stack = np.stack(
            [np.asarray(entry["projection"], dtype=bool) for entry in label_entries],
            axis=0,
        )
        distance_stack = np.stack(
            [ndi.distance_transform_edt(~projection) for projection in projection_stack],
            axis=0,
        )
        assigned_index[~any_supported] = np.argmin(distance_stack[:, ~any_supported], axis=0)
        local_object_core_labels = np.zeros(local_core.shape, dtype=np.uint32)
        for index, entry in enumerate(label_entries):
            local_object_core_labels[local_core & (assigned_index == index)] = int(
                entry["object_id"]
            )

        used_seed_pixels: set[tuple[int, int]] = set()
        core_coords = np.argwhere(local_core)
        for entry in label_entries:
            object_id = int(entry["object_id"])
            object_core = local_object_core_labels == object_id
            if object_core.any():
                seed_coord = tuple(np.argwhere(object_core)[0])
                used_seed_pixels.add((int(seed_coord[0]), int(seed_coord[1])))
                continue
            center = np.asarray(
                [
                    float(entry["evaluated"]["center_y_local"]),
                    float(entry["evaluated"]["center_x_local"]),
                ]
            )
            order = np.argsort(np.square(core_coords - center).sum(axis=1))
            for coordinate_index in order:
                seed_y, seed_x = map(int, core_coords[int(coordinate_index)])
                if (seed_y, seed_x) in used_seed_pixels:
                    continue
                local_object_core_labels[seed_y, seed_x] = object_id
                used_seed_pixels.add((seed_y, seed_x))
                break

        object_ids = [int(entry["object_id"]) for entry in label_entries]
        object_seed_masks = [local_object_core_labels == object_id for object_id in object_ids]
        seed_distance_stack = np.stack(
            [ndi.distance_transform_edt(~seed_mask) for seed_mask in object_seed_masks],
            axis=0,
        )
        extent_assignment = np.argmin(seed_distance_stack, axis=0)
        local_object_extent_labels = np.zeros(local_extent.shape, dtype=np.uint32)
        for index, object_id in enumerate(object_ids):
            local_object_extent_labels[local_extent & (extent_assignment == index)] = object_id
        local_object_extent_labels[local_object_core_labels > 0] = local_object_core_labels[
            local_object_core_labels > 0
        ]
        core_view = object_core_labels[crop]
        extent_view = object_extent_labels[crop]
        if np.any((core_view > 0) & (local_object_core_labels > 0)):
            raise RuntimeError("3D nucleus inventory produced overlapping parent-core labels")
        if np.any((extent_view > 0) & (local_object_extent_labels > 0)):
            raise RuntimeError("3D nucleus inventory produced overlapping extent labels")
        core_view[local_object_core_labels > 0] = local_object_core_labels[
            local_object_core_labels > 0
        ]
        extent_view[local_object_extent_labels > 0] = local_object_extent_labels[
            local_object_extent_labels > 0
        ]

    if preflight_only:
        workload_summary["status"] = "preflight_completed"
        return ValidatedNucleusAnchors(
            accepted_core_mask_2d=np.zeros_like(nuclei_core, dtype=bool),
            accepted_extent_mask_2d=np.zeros_like(nuclei_extent, dtype=bool),
            metrics={
                "status": "preflight_completed",
                "method": (
                    "DAPI fragment workload preflight using the production DAPI "
                    "parent reconstruction path without fragment evaluation"
                ),
                "measurement_channel_used": False,
                "z_start_1based": z0 + 1,
                "z_end_1based_inclusive": z1 + 1,
                "parent_2d_core_count": int(core_labels.max()),
                "candidate_count": int(workload_summary["total_fragments"]),
                "accepted_count": 0,
                "rejected_count": 0,
                "dapi_valid_count": 0,
                "per_nucleus": [],
                "config": asdict(cfg),
                "dapi_2d": dapi_2d_metrics,
                "dapi_fragment_workload": workload_summary,
            },
            object_core_labels_2d=object_core_labels,
            object_extent_labels_2d=object_extent_labels,
        )

    workload_summary["status"] = "completed"

    raw_dapi_valid_ids = tuple(
        int(row["object_id_3d"]) for row in per_object if bool(row["dapi_valid"])
    )
    raw_accepted_ids = tuple(
        int(row["object_id_3d"]) for row in per_object if bool(row["accepted"])
    )
    per_object_by_id = {
        int(row["object_id_3d"]): row for row in per_object
    }
    extent_area_counts = np.bincount(
        object_extent_labels.ravel(),
        minlength=next_object_id,
    )
    for record in object_records:
        object_id = int(record["object_id"])
        extent_area_px = int(extent_area_counts[object_id])
        record["extent_area_px"] = extent_area_px
        per_object_by_id[object_id]["extent_area_px"] = extent_area_px
    print(
        "3D DAPI inventory complete | "
        f"objects={len(object_records)} | "
        f"elapsed={time.perf_counter() - inventory_started:.3f} s; "
        "resolving canonical nuclei...",
        flush=True,
    )
    canonical_started = time.perf_counter()
    canonical = resolve_canonical_nucleus_instances_3d(
        frozen_whole_mask,
        nuclei_core,
        nuclei_extent,
        context,
        pixel_width_um,
        pixel_height_um,
        cfg,
        raw_object_extent_labels=object_extent_labels,
        max_workers=worker_count,
    )
    print(
        "Canonical nucleus resolution complete | "
        f"instances={len(canonical.records)} | "
        f"elapsed={time.perf_counter() - canonical_started:.3f} s",
        flush=True,
    )
    canonical_records = [dict(row) for row in canonical.records]
    dapi_valid_ids = raw_dapi_valid_ids
    accepted_ids = raw_accepted_ids
    accepted_core = np.isin(object_core_labels, accepted_ids)
    accepted_extent = np.isin(object_extent_labels, accepted_ids) | accepted_core
    source_object_to_instance_ids: dict[int, list[int]] = {}
    for row in canonical_records:
        for source_object_id in row["source_object_ids"]:
            source_object_to_instance_ids.setdefault(int(source_object_id), []).append(
                int(row["instance_id"])
            )
    return ValidatedNucleusAnchors(
        accepted_core_mask_2d=accepted_core,
        accepted_extent_mask_2d=accepted_extent,
        metrics={
            "status": "completed",
            "method": (
                "object-preserving calibrated 3D DAPI inventory + "
                "independent canonical nuclear-envelope audit layer"
            ),
            "measurement_channel_used": False,
            "z_start_1based": z0 + 1,
            "z_end_1based_inclusive": z1 + 1,
            "voxel_size_um": [
                round(context.pixel_depth_um, 9),
                round(pixel_height_um, 9),
                round(pixel_width_um, 9),
            ],
            "calibration_source": context.calibration_source,
            "structural_channel": context.structural_channel,
            "parent_2d_core_count": int(core_labels.max()),
            "candidate_count": len(per_object),
            "dapi_valid_count": len(dapi_valid_ids),
            "accepted_count": len(accepted_ids),
            "rejected_count": len(per_object) - len(accepted_ids),
            "unowned_dapi_valid_count": len(set(dapi_valid_ids) - set(accepted_ids)),
            "multi_object_parent_core_count": multi_object_parent_core_count,
            "accepted_2d_nucleus_ids": list(accepted_ids),
            "rejected_2d_nucleus_ids": [
                int(row["object_id_3d"])
                for row in per_object
                if not bool(row["accepted"])
            ],
            "config": asdict(cfg),
            "dapi_2d": dapi_2d_metrics,
            "dapi_fragment_workload": workload_summary,
            "canonical_resolution": canonical.metrics,
            "raw_object_qc": {
                "candidate_count": len(per_object),
                "dapi_valid_count": len(raw_dapi_valid_ids),
                "accepted_count": len(raw_accepted_ids),
                "multi_object_parent_core_count": multi_object_parent_core_count,
            },
            "canonical_per_nucleus": canonical_records,
            "per_nucleus": per_object,
        },
        object_core_labels_2d=object_core_labels,
        object_extent_labels_2d=object_extent_labels,
        dapi_valid_object_ids=dapi_valid_ids,
        accepted_object_ids=accepted_ids,
        object_records=tuple(object_records),
        nucleus_instance_core_labels_2d=canonical.core_labels_2d,
        nucleus_instance_extent_labels_2d=canonical.extent_labels_2d,
        accepted_instance_ids=tuple(int(value) for value in canonical.accepted_ids),
        ambiguous_instance_ids=canonical.ambiguous_ids,
        nucleus_instance_records=tuple(canonical_records),
        source_object_to_instance_ids={
            object_id: tuple(sorted(instance_ids))
            for object_id, instance_ids in source_object_to_instance_ids.items()
        },
    )


def dapi_supported_anchor(dapi_proj: np.ndarray, struct: np.ndarray, candidate: np.ndarray, spec: TestSpec) -> np.ndarray:
    nuclei = dapi_nuclei_mask(dapi_proj)
    nuclei = morphology.binary_dilation(nuclei, footprint=morphology.disk(spec.dapi_support_radius))
    structural_high = struct >= full_array_percentile(struct, 74)
    support = candidate & nuclei & structural_high
    support = morphology.remove_small_objects(support, min_size=max(40, spec.min_area // 2))
    return support

def strict_soma_anchor(
    mask: np.ndarray,
    dapi_proj: np.ndarray,
    struct: np.ndarray,
    cellpose_mask: np.ndarray,
    spec: TestSpec,
    distance: np.ndarray | None = None,
) -> np.ndarray:
    if not mask.any():
        return np.zeros_like(mask, dtype=bool)
    nuclei = dapi_nuclei_mask(dapi_proj, percentile_floor=85.0)
    near_nuclei = morphology.binary_dilation(
        nuclei,
        footprint=morphology.disk(spec.soma_anchor_radius),
    )
    if distance is None:
        distance = ndi.distance_transform_edt(mask)
    soma_core = mask & (distance >= spec.soma_core_radius)
    structural_core = struct >= full_array_percentile(
        struct,
        spec.soma_anchor_percentile,
    )
    model_or_structure = structural_core | (cellpose_mask & mask)
    return (soma_core & near_nuclei & model_or_structure).astype(bool)

def retain_soma_connected_components(
    mask: np.ndarray,
    dapi_proj: np.ndarray,
    struct: np.ndarray,
    cellpose_mask: np.ndarray,
    spec: TestSpec,
) -> np.ndarray:
    labels = measure.label(mask)
    if labels.max() == 0:
        return mask.astype(bool)

    soma_anchor = strict_soma_anchor(mask, dapi_proj, struct, cellpose_mask, spec)
    component_count = int(labels.max())
    component_areas = np.bincount(labels.ravel(), minlength=component_count + 1)
    anchor_counts = np.bincount(labels[soma_anchor], minlength=component_count + 1)
    primary_labels = np.flatnonzero(
        (component_areas >= spec.anchor_component_min_area)
        & (anchor_counts >= spec.soma_anchor_min_pixels)
    )
    primary_labels = primary_labels[primary_labels > 0]
    primary = np.isin(labels, primary_labels)
    if not primary.any():
        return np.zeros_like(mask, dtype=bool)

    if spec.connection_radius <= 0:
        return primary.astype(bool)

    support_cut = full_array_percentile(
        struct,
        spec.connection_support_percentile,
    )
    bridge_band = morphology.binary_dilation(
        primary,
        footprint=morphology.disk(spec.connection_radius),
    ) & (struct >= support_cut)
    connection_domain = mask | bridge_band
    domain_labels = measure.label(connection_domain)
    touching_labels = np.unique(domain_labels[primary])
    touching_labels = touching_labels[touching_labels > 0]
    connected = np.isin(domain_labels, touching_labels)
    connected = morphology.remove_small_objects(connected, min_size=spec.min_area)
    return connected.astype(bool)

def empty_branch_recovery_metrics() -> dict:
    return {
        "fine_branch_evidence_px": 0,
        "fine_branch_added_px": 0,
        "fine_branch_added_fraction": 0.0,
        "fine_branch_added_structural_mean": 0.0,
        "fine_branch_consensus_evidence_px": 0,
        "fine_branch_single_channel_retained_px": 0,
        "fine_branch_topology_bridge_px": 0,
        "fine_branch_topology_skeleton_px": 0,
        "fine_branch_topology_rejected_px": 0,
    }

def fine_branch_features(
    structural_projections: dict[str, np.ndarray],
    cache_key: tuple,
    background_sigma: float,
) -> dict[str, tuple[np.ndarray, np.ndarray, np.ndarray]]:
    projection_identity = tuple(
        (channel, array_identity_key(projection))
        for channel, projection in sorted(structural_projections.items())
    )
    feature_key = (
        *cache_key,
        projection_identity,
        round(float(background_sigma), 3),
    )
    with _CACHE_LOCK:
        cached = _BRANCH_FEATURE_CACHE.get(feature_key)
    if cached is not None:
        return cached
    with cache_key_lock(("branch_features", *feature_key)):
        with _CACHE_LOCK:
            cached = _BRANCH_FEATURE_CACHE.get(feature_key)
        if cached is not None:
            return cached
        features: dict[str, tuple[np.ndarray, np.ndarray, np.ndarray]] = {}
        for channel, projection in structural_projections.items():
            normalized = normalized_projection(projection)
            smoothed = filters.gaussian(normalized, sigma=0.8, preserve_range=True)
            local_background = filters.gaussian(
                smoothed,
                sigma=background_sigma,
                preserve_range=True,
            )
            local_detail = np.clip(smoothed - local_background, 0, None).astype(
                np.float32
            )
            ridge = filters.sato(
                smoothed,
                sigmas=(1, 2, 3),
                black_ridges=False,
            ).astype(np.float32)
            local_detail.setflags(write=False)
            ridge.setflags(write=False)
            features[channel] = (normalized, local_detail, ridge)

        with _CACHE_LOCK:
            _BRANCH_FEATURE_CACHE[feature_key] = features
        return features

def channel_consensus_branch_evidence(
    mask: np.ndarray,
    channel_evidence: list[np.ndarray],
    channel_support: list[np.ndarray],
    radius: int,
) -> tuple[np.ndarray, np.ndarray, dict[str, int]]:
    evidence = np.logical_or.reduce(channel_evidence)
    low_support = np.logical_or.reduce(channel_support)
    if len(channel_evidence) < 2:
        footprint = morphology.disk(max(1, int(radius)))
        evidence_labels = measure.label(evidence)
        directly_connected = morphology.binary_dilation(
            mask,
            footprint=footprint,
        )
        retained_labels = np.unique(evidence_labels[directly_connected])
        retained_labels = retained_labels[retained_labels > 0]
        guarded_evidence = np.isin(evidence_labels, retained_labels)
        guarded_support = low_support & morphology.binary_dilation(
            mask | guarded_evidence,
            footprint=footprint,
        )
        return guarded_evidence, guarded_support, {
            "fine_branch_consensus_evidence_px": 0,
            "fine_branch_single_channel_retained_px": int(guarded_evidence.sum()),
        }

    footprint = morphology.disk(max(1, int(radius)))
    consensus = np.zeros_like(mask, dtype=bool)
    consensus_support = np.zeros_like(mask, dtype=bool)
    for left in range(len(channel_evidence)):
        for right in range(left + 1, len(channel_evidence)):
            left_evidence = channel_evidence[left]
            right_evidence = channel_evidence[right]
            consensus |= left_evidence & morphology.binary_dilation(
                right_evidence,
                footprint=footprint,
            )
            consensus |= right_evidence & morphology.binary_dilation(
                left_evidence,
                footprint=footprint,
            )
            left_support = channel_support[left]
            right_support = channel_support[right]
            consensus_support |= left_support & morphology.binary_dilation(
                right_support,
                footprint=footprint,
            )
            consensus_support |= right_support & morphology.binary_dilation(
                left_support,
                footprint=footprint,
            )

    single_channel = evidence & ~consensus
    single_labels = measure.label(single_channel)
    trusted_seed = mask | consensus
    trusted_neighborhood = morphology.binary_dilation(
        trusted_seed,
        footprint=morphology.disk(max(1, int(radius) + 1)),
    )
    retained_labels = np.unique(single_labels[trusted_neighborhood])
    retained_labels = retained_labels[retained_labels > 0]
    retained_single = np.isin(single_labels, retained_labels)
    guarded_evidence = consensus | retained_single
    guarded_support = consensus_support | (
        low_support
        & morphology.binary_dilation(
            mask | guarded_evidence,
            footprint=footprint,
        )
    )
    return guarded_evidence, guarded_support, {
        "fine_branch_consensus_evidence_px": int(consensus.sum()),
        "fine_branch_single_channel_retained_px": int(retained_single.sum()),
    }

def topology_continuity_branch_evidence(
    mask: np.ndarray,
    evidence: np.ndarray,
    low_support: np.ndarray,
    *,
    max_gap: int,
    min_skeleton: int,
    max_hops: int,
) -> tuple[np.ndarray, np.ndarray, dict[str, int]]:
    evidence_labels = measure.label(evidence)
    if evidence_labels.max() == 0:
        return evidence, np.zeros_like(mask, dtype=bool), {
            "fine_branch_topology_bridge_px": 0,
            "fine_branch_topology_skeleton_px": 0,
            "fine_branch_topology_rejected_px": 0,
        }

    skeleton = morphology.skeletonize(evidence)
    skeleton_counts = np.bincount(
        evidence_labels[skeleton],
        minlength=int(evidence_labels.max()) + 1,
    )
    valid_labels = np.flatnonzero(skeleton_counts >= max(1, int(min_skeleton)))
    valid_labels = valid_labels[valid_labels > 0]
    line_evidence = np.isin(evidence_labels, valid_labels)
    line_labels = measure.label(line_evidence)
    accepted = np.zeros_like(mask, dtype=bool)
    bridges = np.zeros_like(mask, dtype=bool)
    frontier = mask.astype(bool, copy=True)
    one_pixel = morphology.disk(1)
    max_gap = max(1, int(max_gap))

    for _ in range(max(1, int(max_hops))):
        reachable = frontier.copy()
        for _step in range(max_gap):
            reachable |= morphology.binary_dilation(
                reachable,
                footprint=one_pixel,
            ) & low_support
        touching = morphology.binary_dilation(
            reachable,
            footprint=one_pixel,
        )
        touching_labels = np.unique(line_labels[touching])
        touching_labels = touching_labels[touching_labels > 0]
        newly_accepted = np.isin(line_labels, touching_labels) & ~accepted
        if not newly_accepted.any():
            break
        bridge_target = morphology.binary_dilation(
            newly_accepted,
            footprint=morphology.disk(max_gap),
        )
        bridges |= reachable & bridge_target & low_support
        accepted |= newly_accepted
        frontier = mask | accepted | bridges

    accepted_skeleton_px = int(morphology.skeletonize(accepted).sum())
    return accepted, bridges, {
        "fine_branch_topology_bridge_px": int(bridges.sum()),
        "fine_branch_topology_skeleton_px": accepted_skeleton_px,
        "fine_branch_topology_rejected_px": int((evidence & ~accepted).sum()),
    }

def recover_anchor_connected_fine_processes(
    mask: np.ndarray,
    structural_projections: dict[str, np.ndarray],
    dapi_proj: np.ndarray,
    struct: np.ndarray,
    cellpose_mask: np.ndarray,
    spec: TestSpec,
    cache_key: tuple,
) -> tuple[np.ndarray, dict]:
    if not spec.fine_branch_recovery or not mask.any():
        return mask.astype(bool), empty_branch_recovery_metrics()

    features = fine_branch_features(
        structural_projections,
        cache_key=cache_key,
        background_sigma=spec.fine_branch_background_sigma,
    )
    evidence = np.zeros_like(mask, dtype=bool)
    low_support = np.zeros_like(mask, dtype=bool)
    channel_evidence_masks: list[np.ndarray] = []
    channel_support_masks: list[np.ndarray] = []
    has_gfap = "GFAP" in structural_projections

    for channel, (normalized, local_detail, ridge) in features.items():
        if channel == "GFAP":
            channel_offset = 0.0
        elif has_gfap:
            channel_offset = spec.fine_branch_single_channel_offset * 0.5
        else:
            channel_offset = spec.fine_branch_single_channel_offset

        detail_percentile = min(99.5, spec.fine_branch_detail_percentile + channel_offset)
        intensity_percentile = min(
            99.0,
            spec.fine_branch_intensity_percentile + 1.5 * channel_offset,
        )
        detail_cut = full_array_percentile(local_detail, detail_percentile)
        ridge_cut = full_array_percentile(ridge, detail_percentile)
        intensity_cut = full_array_percentile(normalized, intensity_percentile)
        channel_evidence = (
            (normalized >= intensity_cut)
            & (local_detail >= detail_cut)
            & (ridge >= ridge_cut)
        )
        channel_min_area = max(
            4,
            int(round(spec.fine_branch_min_area + 1.5 * channel_offset)),
        )
        channel_evidence = morphology.remove_small_objects(
            channel_evidence,
            min_size=channel_min_area,
        )

        labels = measure.label(channel_evidence)
        shaped_labels: list[int] = []
        min_major_axis = spec.fine_branch_min_major_axis + 1.5 * channel_offset
        min_eccentricity = min(
            0.95,
            spec.fine_branch_min_eccentricity + 0.025 * channel_offset,
        )
        for prop in measure.regionprops(labels):
            major_axis = float(
                prop.axis_major_length
                if hasattr(prop, "axis_major_length")
                else prop.major_axis_length
            )
            line_like = (
                prop.area >= channel_min_area
                and major_axis >= min_major_axis
                and (
                    prop.eccentricity >= min_eccentricity
                    or major_axis >= 1.8 * min_major_axis
                )
            )
            if line_like:
                shaped_labels.append(int(prop.label))
        shaped = np.isin(labels, shaped_labels)
        channel_evidence = morphology.binary_closing(
            shaped,
            footprint=morphology.disk(1),
        )
        evidence |= channel_evidence
        channel_evidence_masks.append(channel_evidence)

        support_detail_cut = full_array_percentile(
            local_detail,
            max(50.0, detail_percentile - 20.0),
        )
        support_intensity_cut = full_array_percentile(
            normalized,
            max(45.0, intensity_percentile - 10.0),
        )
        channel_support = (
            (normalized >= support_intensity_cut)
            & (local_detail >= support_detail_cut)
        )
        low_support |= channel_support
        channel_support_masks.append(channel_support)

    mode_metrics: dict[str, int] = {}
    topology_bridge = np.zeros_like(mask, dtype=bool)
    if spec.fine_branch_evidence_mode == "channel_consensus":
        evidence, low_support, mode_metrics = channel_consensus_branch_evidence(
            mask,
            channel_evidence_masks,
            channel_support_masks,
            spec.fine_branch_consensus_radius,
        )
    elif spec.fine_branch_evidence_mode == "topology_continuity":
        evidence, topology_bridge, mode_metrics = topology_continuity_branch_evidence(
            mask,
            evidence,
            low_support,
            max_gap=spec.fine_branch_topology_max_gap,
            min_skeleton=spec.fine_branch_topology_min_skeleton,
            max_hops=spec.fine_branch_topology_max_hops,
        )
    elif spec.fine_branch_evidence_mode != "union":
        raise ValueError(
            f"Unknown fine-branch evidence mode: {spec.fine_branch_evidence_mode}"
        )

    if not evidence.any():
        return mask.astype(bool), empty_branch_recovery_metrics()

    if spec.fine_branch_evidence_mode == "topology_continuity":
        bridge = topology_bridge
    else:
        bridge = morphology.binary_dilation(
            mask,
            footprint=morphology.disk(spec.fine_branch_gap_radius),
        ) & low_support
    domain = mask | evidence | bridge
    base_labels = measure.label(mask)
    distance_from_base = ndi.distance_transform_edt(~mask)
    grown_labels = segmentation.watershed(
        distance_from_base,
        markers=base_labels,
        mask=domain,
    )

    # Keep neighboring astrocytes separated if one permissive line component
    # reaches two pre-existing soma-anchored components.
    seam = np.zeros_like(mask, dtype=bool)
    vertical = (
        (grown_labels[1:] > 0)
        & (grown_labels[:-1] > 0)
        & (grown_labels[1:] != grown_labels[:-1])
    )
    horizontal = (
        (grown_labels[:, 1:] > 0)
        & (grown_labels[:, :-1] > 0)
        & (grown_labels[:, 1:] != grown_labels[:, :-1])
    )
    diagonal_down = (
        (grown_labels[1:, 1:] > 0)
        & (grown_labels[:-1, :-1] > 0)
        & (grown_labels[1:, 1:] != grown_labels[:-1, :-1])
    )
    diagonal_up = (
        (grown_labels[1:, :-1] > 0)
        & (grown_labels[:-1, 1:] > 0)
        & (grown_labels[1:, :-1] != grown_labels[:-1, 1:])
    )
    seam[1:] |= vertical
    seam[:-1] |= vertical
    seam[:, 1:] |= horizontal
    seam[:, :-1] |= horizontal
    seam[1:, 1:] |= diagonal_down
    seam[:-1, :-1] |= diagonal_down
    seam[1:, :-1] |= diagonal_up
    seam[:-1, 1:] |= diagonal_up
    grown = grown_labels > 0
    grown[seam & ~mask] = False
    grown |= mask

    if spec.require_soma_anchor:
        grown = retain_soma_connected_components(
            grown,
            dapi_proj,
            struct,
            cellpose_mask,
            spec,
        )
    added = grown & ~mask
    metrics = {
        "fine_branch_evidence_px": int(evidence.sum()),
        "fine_branch_added_px": int(added.sum()),
        "fine_branch_added_fraction": round(
            float(added.sum()) / max(int(grown.sum()), 1),
            6,
        ),
        "fine_branch_added_structural_mean": round(
            float(struct[added].mean()) if added.any() else 0.0,
            6,
        ),
        **{
            key: int(value)
            for key, value in mode_metrics.items()
        },
    }
    for key, value in empty_branch_recovery_metrics().items():
        metrics.setdefault(key, value)
    return grown.astype(bool), metrics

def empty_border_exclusion_metrics() -> dict:
    return {
        "border_candidate_components": 0,
        "border_preserved_complete_components": 0,
        "border_preserved_complete_area_px": 0,
        "border_preserved_complete_area_fraction": 0.0,
        "border_reference_median_area_px": 0.0,
        "border_removed_components": 0,
        "border_removed_area_px": 0,
        "border_removed_area_fraction": 0.0,
    }

def edge_zone_mask(shape: tuple[int, int], margin: int) -> np.ndarray:
    height, width = shape
    bounded_margin = min(max(1, int(margin)), height // 2, width // 2)
    zone = np.zeros(shape, dtype=bool)
    zone[:bounded_margin] = True
    zone[-bounded_margin:] = True
    zone[:, :bounded_margin] = True
    zone[:, -bounded_margin:] = True
    return zone

def complete_soma_component_labels(
    mask: np.ndarray,
    dapi_proj: np.ndarray,
    struct: np.ndarray,
    spec: TestSpec,
    labels: np.ndarray | None = None,
    distance: np.ndarray | None = None,
) -> tuple[np.ndarray, np.ndarray, float]:
    if labels is None:
        labels = measure.label(mask)
    component_count = int(labels.max())
    if component_count == 0:
        return labels, np.empty(0, dtype=np.int32), 0.0

    # A process may leave the field of view while the biological cell remains
    # usable. Require its DAPI-supported structural soma core to be well inside
    # the image, then compare the component with complete interior cells.
    soma_anchor = strict_soma_anchor(
        mask,
        dapi_proj,
        struct,
        np.zeros_like(mask, dtype=bool),
        spec,
        distance=distance,
    )
    soma_interior = ~edge_zone_mask(mask.shape, spec.border_complete_soma_margin)
    soma_anchor &= soma_interior

    component_areas = np.bincount(labels.ravel(), minlength=component_count + 1)
    anchor_counts = np.bincount(labels[soma_anchor], minlength=component_count + 1)
    interior_counts = np.bincount(
        labels[soma_interior],
        minlength=component_count + 1,
    )
    supported = (
        (component_areas >= spec.anchor_component_min_area)
        & (anchor_counts >= spec.soma_anchor_min_pixels)
    )
    supported[0] = False

    border_zone = edge_zone_mask(mask.shape, spec.border_margin)
    border_labels = np.unique(labels[border_zone])
    border_labels = border_labels[border_labels > 0]
    interior_supported = np.flatnonzero(supported & ~np.isin(np.arange(component_count + 1), border_labels))
    if interior_supported.size:
        reference_area = float(np.median(component_areas[interior_supported]))
    else:
        all_supported = np.flatnonzero(supported)
        reference_area = float(np.median(component_areas[all_supported])) if all_supported.size else 0.0

    area_floor = reference_area * spec.border_complete_min_area_ratio
    interior_fraction = interior_counts / np.maximum(component_areas, 1)
    complete = supported & (component_areas >= area_floor)
    complete &= interior_fraction >= spec.border_complete_min_interior_fraction
    complete[0] = False
    return labels, np.flatnonzero(complete).astype(np.int32), reference_area

def exclude_incomplete_border_components(
    mask: np.ndarray,
    dapi_proj: np.ndarray,
    struct: np.ndarray,
    spec: TestSpec,
) -> tuple[np.ndarray, dict]:
    if not spec.exclude_border_components or not mask.any():
        return mask.astype(bool), empty_border_exclusion_metrics()

    labels, complete_labels, reference_area = complete_soma_component_labels(
        mask,
        dapi_proj,
        struct,
        spec,
    )
    border_zone = edge_zone_mask(mask.shape, spec.border_margin)
    border_labels = np.unique(labels[border_zone])
    border_labels = border_labels[border_labels > 0]
    if spec.preserve_complete_border_components:
        preserved_labels = np.intersect1d(border_labels, complete_labels)
    else:
        preserved_labels = np.empty(0, dtype=border_labels.dtype)
    removed_labels = np.setdiff1d(border_labels, preserved_labels)
    removed = np.isin(labels, removed_labels)
    preserved = np.isin(labels, preserved_labels)
    kept = mask & ~removed
    preserved_fraction = float(preserved.sum()) / max(int(kept.sum()), 1)
    removed_fraction = float(removed.sum()) / max(int(mask.sum()), 1)
    metrics = {
        "border_candidate_components": int(len(border_labels)),
        "border_preserved_complete_components": int(len(preserved_labels)),
        "border_preserved_complete_area_px": int(preserved.sum()),
        "border_preserved_complete_area_fraction": round(preserved_fraction, 6),
        "_raw_border_preserved_complete_area_fraction": preserved_fraction,
        "border_reference_median_area_px": round(reference_area, 3),
        "border_removed_components": int(len(removed_labels)),
        "border_removed_area_px": int(removed.sum()),
        "border_removed_area_fraction": round(removed_fraction, 6),
        "_raw_border_removed_area_fraction": removed_fraction,
    }
    return kept.astype(bool), metrics

def remove_isolated_artifact_fragments(
    mask: np.ndarray,
    dapi_proj: np.ndarray,
    struct: np.ndarray,
    cellpose_mask: np.ndarray,
    spec: TestSpec,
) -> np.ndarray:
    if spec.require_soma_anchor:
        return retain_soma_connected_components(mask, dapi_proj, struct, cellpose_mask, spec)

    labels = measure.label(mask)
    if labels.max() == 0:
        return mask
    props = measure.regionprops(labels, intensity_image=struct)
    large_labels = [int(prop.label) for prop in props if prop.area >= spec.anchor_area]
    large = np.isin(labels, large_labels)
    near_large = morphology.binary_dilation(large, footprint=morphology.disk(spec.artifact_near_radius))
    dapi_support = dapi_supported_anchor(dapi_proj, struct, mask, spec)
    near_dapi_soma = morphology.binary_dilation(
        dapi_support,
        footprint=morphology.disk(spec.artifact_near_radius),
    )
    soma_min_area = max(180, spec.min_area * 2)
    process_min_area = max(120, spec.min_area)

    keep = np.zeros_like(mask, dtype=bool)
    for prop in props:
        component_slice = prop.slice
        comp = prop.image
        area = prop.area
        near_main = bool((comp & near_large[component_slice]).any())
        has_dapi_support = bool((comp & dapi_support[component_slice]).any())
        process_like = (
            prop.eccentricity >= spec.process_eccentricity
            and prop.major_axis_length >= spec.process_major_axis
        )
        large_enough = area >= spec.artifact_min_area
        near_supported_soma = bool((comp & near_dapi_soma[component_slice]).any())
        supported_soma = has_dapi_support and area >= soma_min_area
        supported_process = process_like and near_supported_soma and area >= process_min_area
        if large_enough or near_main or supported_soma or supported_process:
            keep_view = keep[component_slice]
            keep_view[comp] = True
    keep = morphology.remove_small_objects(keep, min_size=spec.min_area)
    return keep.astype(bool)

def refine_fused_process_regions(
    mask: np.ndarray,
    candidate: np.ndarray,
    cellpose_mask: np.ndarray,
    dapi_proj: np.ndarray,
    struct: np.ndarray,
    spec: TestSpec,
) -> np.ndarray:
    if not spec.branch_refine or not mask.any():
        return mask.astype(bool)

    distance = ndi.distance_transform_edt(mask)
    values = struct[mask]
    evidence_cut = float(np.percentile(values, spec.branch_support_percentile))
    direct_support = morphology.binary_dilation(
        candidate & mask,
        footprint=morphology.disk(spec.branch_support_radius),
    )
    intensity_support = morphology.binary_dilation(
        mask & (struct >= evidence_cut),
        footprint=morphology.disk(1),
    )
    thin_process_corridor = mask & (distance <= spec.max_process_half_width)

    dapi_anchor = dapi_supported_anchor(dapi_proj, struct, candidate, spec)
    structural_core = struct >= full_array_percentile(struct, 74)
    cellpose_anchor = cellpose_mask & mask & structural_core
    soma_seed = dapi_anchor | cellpose_anchor
    soma_support = morphology.binary_dilation(
        soma_seed,
        footprint=morphology.disk(spec.soma_protect_radius),
    )

    refined = mask & (
        direct_support
        | intensity_support
        | thin_process_corridor
        | soma_support
    )
    refined = morphology.binary_closing(refined, footprint=morphology.disk(1))
    refined = morphology.remove_small_holes(refined, area_threshold=spec.hole_area)
    refined = morphology.remove_small_objects(refined, min_size=spec.min_area)
    return refined.astype(bool)

def refine_with_cellpose_and_dapi(
    raw_mask: np.ndarray,
    cellpose_mask: np.ndarray,
    dapi_proj: np.ndarray,
    struct: np.ndarray,
    spec: TestSpec,
) -> np.ndarray:
    candidate = cleanup_mask(raw_mask, spec)
    cp_context = morphology.binary_dilation(cellpose_mask, footprint=morphology.disk(spec.bridge_radius))
    candidate = cleanup_mask(candidate, spec)
    cp_anchor = cp_context & candidate & (
        struct >= full_array_percentile(struct, 66)
    )
    cp_anchor = morphology.remove_small_objects(cp_anchor, min_size=max(60, spec.min_area))
    dapi_anchor = dapi_supported_anchor(dapi_proj, struct, candidate, spec)
    extra_anchor = cp_anchor | dapi_anchor
    cleaned = anchor_connected_cleanup(candidate, struct, spec, extra_anchor=extra_anchor)
    cleaned = refine_fused_process_regions(
        cleaned,
        candidate,
        cellpose_mask,
        dapi_proj,
        struct,
        spec,
    )
    if spec.artifact_filter:
        cleaned = remove_isolated_artifact_fragments(cleaned, dapi_proj, struct, cellpose_mask, spec)
    return cleaned

def postprocess_without_cellpose(
    mask: np.ndarray,
    raw_mask: np.ndarray,
    dapi_proj: np.ndarray,
    struct: np.ndarray,
    spec: TestSpec,
) -> np.ndarray:
    candidate = cleanup_mask(raw_mask, spec)
    empty_cellpose = np.zeros_like(mask, dtype=bool)
    cleaned = refine_fused_process_regions(
        mask,
        candidate,
        empty_cellpose,
        dapi_proj,
        struct,
        spec,
    )
    if spec.artifact_filter:
        cleaned = remove_isolated_artifact_fragments(cleaned, dapi_proj, struct, empty_cellpose, spec)
    return cleaned

def fixed_soma_component_metrics(
    mask: np.ndarray,
    dapi_proj: np.ndarray,
    struct: np.ndarray,
    labels: np.ndarray | None = None,
    distance: np.ndarray | None = None,
) -> dict:
    if labels is None:
        labels = measure.label(mask)
    component_count = int(labels.max())
    if component_count == 0:
        return {
            "soma_supported_components": 0,
            "unanchored_components": 0,
            "unanchored_area_fraction": 0.0,
            "median_component_area_px": 0.0,
        }

    nuclei = dapi_nuclei_mask(dapi_proj, percentile_floor=85.0)
    near_nuclei = morphology.binary_dilation(nuclei, footprint=morphology.disk(4))
    if distance is None:
        distance = ndi.distance_transform_edt(mask)
    structural_core = struct >= full_array_percentile(struct, 84.0)
    fixed_anchor = mask & near_nuclei & structural_core & (distance >= 8.0)

    component_areas = np.bincount(labels.ravel(), minlength=component_count + 1)
    anchor_counts = np.bincount(labels[fixed_anchor], minlength=component_count + 1)
    supported = (component_areas >= 3000) & (anchor_counts >= 8)
    supported[0] = False
    unanchored_labels = np.flatnonzero((component_areas > 0) & ~supported)
    unanchored_labels = unanchored_labels[unanchored_labels > 0]
    unanchored_area = int(component_areas[unanchored_labels].sum())
    unanchored_fraction = unanchored_area / max(int(mask.sum()), 1)
    return {
        "soma_supported_components": int(supported.sum()),
        "unanchored_components": int(len(unanchored_labels)),
        "unanchored_area_fraction": round(unanchored_fraction, 6),
        "_raw_unanchored_area_fraction": unanchored_fraction,
        "median_component_area_px": round(float(np.median(component_areas[1:])), 3),
    }

def edge_proximity_metrics(
    mask: np.ndarray,
    dapi_proj: np.ndarray,
    struct: np.ndarray,
    spec: TestSpec,
    labels: np.ndarray | None = None,
    distance: np.ndarray | None = None,
) -> dict:
    labels, complete_labels, _ = complete_soma_component_labels(
        mask,
        dapi_proj,
        struct,
        spec,
        labels=labels,
        distance=distance,
    )
    if labels.max() == 0:
        return {
            "final_border_touching_components": 0,
            "final_complete_border_touching_components": 0,
            "final_incomplete_border_touching_components": 0,
            "edge_proximity_components": 0,
            "edge_proximity_area_fraction": 0.0,
        }

    exact_border = np.zeros_like(mask, dtype=bool)
    exact_border[[0, -1], :] = True
    exact_border[:, [0, -1]] = True
    exact_labels = np.unique(labels[exact_border])
    exact_labels = exact_labels[exact_labels > 0]
    complete_exact_labels = np.intersect1d(exact_labels, complete_labels)
    incomplete_exact_labels = np.setdiff1d(exact_labels, complete_labels)

    edge_zone = edge_zone_mask(mask.shape, spec.edge_qc_margin)
    near_labels = np.unique(labels[edge_zone])
    near_labels = near_labels[near_labels > 0]
    incomplete_near_labels = np.setdiff1d(near_labels, complete_labels)
    near_area = int(np.isin(labels, incomplete_near_labels).sum())
    edge_fraction = near_area / max(int(mask.sum()), 1)
    return {
        "final_border_touching_components": int(len(exact_labels)),
        "final_complete_border_touching_components": int(len(complete_exact_labels)),
        "final_incomplete_border_touching_components": int(len(incomplete_exact_labels)),
        "edge_proximity_components": int(len(incomplete_near_labels)),
        "edge_proximity_area_fraction": round(edge_fraction, 6),
        "_raw_edge_proximity_area_fraction": edge_fraction,
    }

def qc_metrics(mask: np.ndarray, struct: np.ndarray, dapi_proj: np.ndarray, spec: TestSpec) -> dict:
    labels = measure.label(mask)
    props = measure.regionprops(labels)
    total_signal = full_array_sum(struct)
    mask_area = int(mask.sum())
    distance = ndi.distance_transform_edt(mask) if mask_area else None
    if mask_area:
        masked_struct = struct[mask]
        in_signal = float(np.sum(masked_struct, dtype=np.float64))
        evidence_cut = float(np.percentile(masked_struct, 40))
        unsupported_wide = mask & (distance > 10) & (struct < evidence_cut)
        unsupported_wide_fraction = float(unsupported_wide.sum()) / mask_area
        structural_precision = float(np.mean(masked_struct, dtype=np.float64))
    else:
        in_signal = 0.0
        unsupported_wide_fraction = 0.0
        structural_precision = 0.0

    background_labels = measure.label(~mask)
    border_labels = np.unique(
        np.concatenate(
            [
                background_labels[0],
                background_labels[-1],
                background_labels[:, 0],
                background_labels[:, -1],
            ]
        )
    )
    internal_holes = [
        prop
        for prop in measure.regionprops(background_labels)
        if prop.label not in border_labels and prop.area >= spec.outline_hole_min_area
    ]
    structural_coverage = in_signal / total_signal if total_signal > 0 else 0.0
    mask_area_fraction = float(np.mean(mask, dtype=np.float64))
    metrics = {
        "mask_area_px": mask_area,
        "mask_area_fraction": round(mask_area_fraction, 6),
        "_raw_mask_area_fraction": mask_area_fraction,
        "connected_components": int(len(props)),
        "largest_component_px": int(max((p.area for p in props), default=0)),
        "structural_signal_coverage": round(structural_coverage, 6),
        "_raw_structural_signal_coverage": structural_coverage,
        "structural_precision": round(structural_precision, 6),
        "_raw_structural_precision": structural_precision,
        "unsupported_wide_fraction": round(unsupported_wide_fraction, 6),
        "_raw_unsupported_wide_fraction": unsupported_wide_fraction,
        "internal_holes": int(len(internal_holes)),
        "internal_hole_area_px": int(sum(prop.area for prop in internal_holes)),
    }
    metrics.update(
        fixed_soma_component_metrics(
            mask,
            dapi_proj,
            struct,
            labels=labels,
            distance=distance,
        )
    )
    metrics.update(
        edge_proximity_metrics(
            mask,
            dapi_proj,
            struct,
            spec,
            labels=labels,
            distance=distance,
        )
    )
    return metrics

def candidate_cellpose_cache_key(
    structural_channels: list[str],
    projection_key: tuple[int, int, str],
    spec: TestSpec,
) -> tuple:
    z0, z1, projection_mode = projection_key
    return (
        tuple(structural_channels),
        z0,
        z1,
        projection_mode,
        round(spec.egfp_weight, 3),
        round(spec.gfap_weight, 3),
        round(spec.smooth_sigma, 3),
        round(spec.cellpose_cellprob, 3),
        round(spec.cellpose_diameter, 3),
        spec.cellpose_max_side,
    )

def build_candidate_window_contexts(
    *,
    chosen_specs: list[TestSpec],
    structural_channels: list[str],
    dapi_stack: np.ndarray,
    structural_stacks: dict[str, np.ndarray],
    profile: np.ndarray,
    projection_cache: dict[
        tuple[int, int, str],
        tuple[np.ndarray, dict[str, np.ndarray]],
    ],
    structural_map_cache: dict[tuple, np.ndarray],
) -> list[CandidateWindowContext]:
    contexts: list[CandidateWindowContext] = []
    shared_contexts: dict[tuple, CandidateWindowContext] = {}
    for spec in chosen_specs:
        z0, z1 = z_range_from_mode(spec.z_mode, profile)
        projection_key = (z0, z1, spec.projection)
        if projection_key not in projection_cache:
            dapi_projection = project(dapi_stack, z0, z1, spec.projection)
            structural_projections = {
                channel: project(stack, z0, z1, spec.projection)
                for channel, stack in structural_stacks.items()
            }
            dapi_projection.setflags(write=False)
            for projection in structural_projections.values():
                projection.setflags(write=False)
            projection_cache[projection_key] = (
                dapi_projection,
                structural_projections,
            )
        dapi_projection, structural_projections = projection_cache[projection_key]

        structural_key = (
            *projection_key,
            round(spec.egfp_weight, 6),
            round(spec.gfap_weight, 6),
            round(spec.smooth_sigma, 6),
        )
        if structural_key not in structural_map_cache:
            cached_struct = structural_map(structural_projections, spec)
            cached_struct.setflags(write=False)
            structural_map_cache[structural_key] = cached_struct
        struct = structural_map_cache[structural_key]

        context_key = (
            structural_key,
            bool(spec.cellpose),
            round(spec.cellpose_cellprob, 3),
            round(spec.cellpose_diameter, 3),
            spec.cellpose_max_side,
        )
        context = shared_contexts.get(context_key)
        if context is None:
            if spec.cellpose:
                cellpose_mask, cellpose_note = run_cellpose_mask(
                    struct,
                    spec,
                    candidate_cellpose_cache_key(
                        structural_channels,
                        projection_key,
                        spec,
                    ),
                )
                immutable_cellpose = np.asarray(
                    cellpose_mask,
                    dtype=bool,
                ).copy()
            else:
                immutable_cellpose = np.zeros_like(struct, dtype=bool)
                cellpose_note = "cellpose_disabled"
            immutable_cellpose.setflags(write=False)
            context = CandidateWindowContext(
                projection_key=projection_key,
                structural_key=structural_key,
                dapi_projection=dapi_projection,
                structural_projections=structural_projections,
                structural_map=struct,
                cellpose_mask=immutable_cellpose,
                cellpose_note=cellpose_note,
            )
            shared_contexts[context_key] = context
        contexts.append(context)
    return contexts

def precompute_candidate_top_hat(
    context: CandidateWindowContext,
    representative: TestSpec,
) -> None:
    threshold_mask(
        context.structural_map,
        context.structural_projections,
        representative,
    )

def precompute_candidate_dapi(
    context: CandidateWindowContext,
    percentile_floor: float | None,
) -> None:
    dapi_nuclei_mask(
        context.dapi_projection,
        percentile_floor=percentile_floor,
    )

def precompute_candidate_branches(
    context: CandidateWindowContext,
    input_dir: Path,
    structural_channels: list[str],
    background_sigma: float,
) -> None:
    fine_branch_features(
        context.structural_projections,
        cache_key=(
            str(input_dir),
            tuple(structural_channels),
            *context.projection_key,
        ),
        background_sigma=background_sigma,
    )

def precompute_distribution_models(context: CandidateWindowContext) -> None:
    try:
        get_log1p_gmm_threshold(context.structural_map)
    except ValueError:
        # Distributional-threshold candidates consume the cached failure as a
        # fail-closed QC result.
        pass

def candidate_precompute_jobs(
    *,
    context_groups: list[tuple[CandidateWindowContext, list[TestSpec]]],
    input_dir: Path,
    structural_channels: list[str],
) -> list[tuple[int, object, tuple]]:
    jobs: list[tuple[int, object, tuple]] = []
    for context, specs in context_groups:
        if any(spec.method == "log1p_gmm" for spec in specs):
            jobs.append((0, precompute_distribution_models, (context,)))
        branch_sigmas = sorted(
            {
                float(spec.fine_branch_background_sigma)
                for spec in specs
                if spec.fine_branch_recovery
            }
        )
        for background_sigma in branch_sigmas:
            jobs.append(
                (
                    0,
                    precompute_candidate_branches,
                    (context, input_dir, structural_channels, background_sigma),
                )
            )
        top_hat_specs = [spec for spec in specs if spec.method == "top_hat_union"]
        if top_hat_specs:
            jobs.append(
                (
                    1,
                    precompute_candidate_top_hat,
                    (context, top_hat_specs[0]),
                )
            )
        jobs.append((2, precompute_candidate_dapi, (context, None)))
        jobs.append((2, precompute_candidate_dapi, (context, 85.0)))
    return sorted(jobs, key=lambda item: item[0])

def candidate_base_cache_key(struct: np.ndarray, spec: TestSpec) -> tuple:
    base_spec = asdict(spec)
    base_spec.pop("name", None)
    base_spec.pop("z_mode", None)
    for field_name in list(base_spec):
        if field_name.startswith("fine_branch_"):
            base_spec.pop(field_name)
    return (
        "candidate_base",
        array_identity_key(struct),
        tuple(sorted(base_spec.items())),
    )

def evaluate_ihc_candidate(
    *,
    candidate_number: int,
    candidate_count: int,
    spec: TestSpec,
    input_dir: Path,
    structural_channels: list[str],
    dapi_stack: np.ndarray,
    structural_stacks: dict[str, np.ndarray],
    profile: np.ndarray,
    projection_cache: dict[tuple[int, int, str], tuple[np.ndarray, dict[str, np.ndarray]]],
    structural_map_cache: dict[tuple, np.ndarray],
    emit_progress: bool = True,
    window_context: CandidateWindowContext | None = None,
) -> tuple[np.ndarray, dict, tuple[int, int, str]]:
    z0, z1 = z_range_from_mode(spec.z_mode, profile)
    projection_key = (z0, z1, spec.projection)
    if window_context is None:
        if projection_key not in projection_cache:
            projection_cache[projection_key] = (
                project(dapi_stack, z0, z1, spec.projection),
                {
                    channel: project(stack, z0, z1, spec.projection)
                    for channel, stack in structural_stacks.items()
                },
            )
        d_proj, structural_projections = projection_cache[projection_key]
    else:
        if window_context.projection_key != projection_key:
            raise AssertionError("Candidate context projection key mismatch")
        d_proj = window_context.dapi_projection
        structural_projections = window_context.structural_projections
    structural_key = (
        *projection_key,
        round(spec.egfp_weight, 6),
        round(spec.gfap_weight, 6),
        round(spec.smooth_sigma, 6),
    )
    if window_context is None:
        if structural_key not in structural_map_cache:
            cached_struct = structural_map(structural_projections, spec)
            cached_struct.setflags(write=False)
            structural_map_cache[structural_key] = cached_struct
        struct = structural_map_cache[structural_key]
    else:
        if window_context.structural_key != structural_key:
            raise AssertionError("Candidate context structural key mismatch")
        struct = window_context.structural_map
    normalized_weights = active_channel_weights(structural_projections, spec)
    base_key = candidate_base_cache_key(struct, spec)
    with _CACHE_LOCK:
        base_result = _CANDIDATE_BASE_CACHE.get(base_key)
        base_lock = _CANDIDATE_BASE_LOCKS.setdefault(base_key, threading.Lock())
    if base_result is None:
        with base_lock:
            with _CACHE_LOCK:
                base_result = _CANDIDATE_BASE_CACHE.get(base_key)
            if base_result is None:
                cellpose_mask = np.zeros_like(struct, dtype=bool)
                try:
                    if spec.cellpose:
                        raw_mask = threshold_mask(struct, structural_projections, spec)
                        if window_context is None:
                            cellpose_mask, cellpose_note = run_cellpose_mask(
                                struct,
                                spec,
                                candidate_cellpose_cache_key(
                                    structural_channels,
                                    projection_key,
                                    spec,
                                ),
                            )
                        else:
                            cellpose_mask = window_context.cellpose_mask
                            cellpose_note = window_context.cellpose_note
                        cellpose_mask = cleanup_mask(cellpose_mask, spec)
                        if cellpose_mask.sum() == 0:
                            if spec.cleanup_mode == "basic":
                                mask = cleanup_mask(raw_mask, spec)
                            else:
                                mask = anchor_connected_cleanup(raw_mask, struct, spec)
                            mask = postprocess_without_cellpose(
                                mask,
                                raw_mask,
                                d_proj,
                                struct,
                                spec,
                            )
                            method_used = f"{cellpose_note}_fallback_{spec.method}"
                        else:
                            mask = refine_with_cellpose_and_dapi(
                                raw_mask,
                                cellpose_mask,
                                d_proj,
                                struct,
                                spec,
                            )
                            method_used = (
                                f"{cellpose_note}+{spec.method}+"
                                "cellpose_anchor_only+dapi_anchor"
                            )
                    else:
                        raw_mask = threshold_mask(struct, structural_projections, spec)
                        if spec.cleanup_mode == "basic":
                            mask = cleanup_mask(raw_mask, spec)
                        else:
                            mask = anchor_connected_cleanup(raw_mask, struct, spec)
                        mask = postprocess_without_cellpose(
                            mask,
                            raw_mask,
                            d_proj,
                            struct,
                            spec,
                        )
                        method_used = spec.method
                    error = ""
                except Exception as exc:
                    try:
                        raw_mask = threshold_mask(
                            struct,
                            structural_projections,
                            spec,
                        )
                    except Exception as threshold_exc:
                        if spec.method != "log1p_gmm":
                            raise
                        raw_mask = np.zeros_like(struct, dtype=bool)
                        mask = raw_mask.copy()
                        method_used = "log1p_gmm_qc_failed"
                        error = repr(threshold_exc)
                    else:
                        if spec.cleanup_mode == "basic":
                            mask = cleanup_mask(raw_mask, spec)
                        else:
                            mask = anchor_connected_cleanup(raw_mask, struct, spec)
                        mask = postprocess_without_cellpose(
                            mask,
                            raw_mask,
                            d_proj,
                            struct,
                            spec,
                        )
                        method_used = f"fallback_{spec.method}"
                        error = repr(exc)
                cached_mask = mask.astype(bool, copy=True)
                cached_cellpose = cellpose_mask.astype(bool, copy=True)
                cached_mask.setflags(write=False)
                cached_cellpose.setflags(write=False)
                base_result = CandidateBaseResult(
                    mask=cached_mask,
                    cellpose_mask=cached_cellpose,
                    method_used=method_used,
                    error=error,
                )
                with _CACHE_LOCK:
                    _CANDIDATE_BASE_CACHE[base_key] = base_result
    mask = base_result.mask.copy()
    cellpose_mask = base_result.cellpose_mask
    method_used = base_result.method_used
    error = base_result.error

    branch_base_mask = mask
    branch_metrics = empty_branch_recovery_metrics()
    if spec.fine_branch_recovery:
        try:
            mask, branch_metrics = recover_anchor_connected_fine_processes(
                mask,
                structural_projections,
                d_proj,
                struct,
                cellpose_mask,
                spec,
                cache_key=(str(input_dir), tuple(structural_channels), *projection_key),
            )
            method_used = f"{method_used}+anchored_fine_branch_recovery"
            if spec.fine_branch_evidence_mode == "channel_consensus":
                method_used = f"{method_used}+channel_consensus_guarded"
            elif spec.fine_branch_evidence_mode == "topology_continuity":
                method_used = f"{method_used}+topology_continuity_guarded"
        except Exception as exc:
            detail = f"fine_branch_recovery={exc!r}"
            error = f"{error}; {detail}" if error else detail
            method_used = f"{method_used}+fine_branch_recovery_failed"

    border_metrics = empty_border_exclusion_metrics()
    if spec.exclude_border_components:
        mask, border_metrics = exclude_incomplete_border_components(mask, d_proj, struct, spec)
        method_used = f"{method_used}+border_soma_complete"

    retained_added = mask & ~branch_base_mask
    retained_added_px = int(retained_added.sum())
    final_mask_area_px = int(mask.sum())
    branch_metrics["fine_branch_retained_px"] = retained_added_px
    branch_metrics["fine_branch_retained_fraction"] = round(
        float(retained_added_px) / max(final_mask_area_px, 1),
        6,
    )
    if spec.branch_refine:
        method_used = f"{method_used}+branch_gap_refine"
    if spec.require_soma_anchor:
        method_used = f"{method_used}+soma_connected_filter"

    z_activity_mean = float(np.mean(profile[z0 : z1 + 1], dtype=np.float64))
    z_activity_integral = float(np.sum(profile[z0 : z1 + 1], dtype=np.float64))
    row = {
        "candidate": candidate_number,
        "pipeline_name": PIPELINE_NAME,
        "name": spec.name,
        "candidate_profile": candidate_profile_name(spec),
        "candidate_family": candidate_profile_family(spec),
        "candidate_module": candidate_module_name(spec),
        "z_mode": spec.z_mode,
        "z_start_0based": z0,
        "z_end_0based_inclusive": z1,
        "z_start_1based": z0 + 1,
        "z_end_1based_inclusive": z1 + 1,
        "z_slice_count": z1 - z0 + 1,
        "z_activity_mean": round(z_activity_mean, 6),
        "_raw_z_activity_mean": z_activity_mean,
        "z_activity_integral": round(z_activity_integral, 6),
        "_raw_z_activity_integral": z_activity_integral,
        "projection": spec.projection,
        "method_requested": "cellpose_cpsam_v2" if spec.cellpose else spec.method,
        "method_used": method_used,
        "egfp_weight_effective": normalized_weights.get("eGFP", 0.0),
        "gfap_weight_effective": normalized_weights.get("GFAP", 0.0),
        "threshold_scale": spec.threshold_scale,
        "min_area": spec.min_area,
        "close_radius": spec.close_radius,
        "dilate_radius": spec.dilate_radius,
        "cellpose": bool(spec.cellpose),
        "cellpose_cellprob": spec.cellpose_cellprob,
        "cellpose_diameter": spec.cellpose_diameter,
        "cellpose_max_side": spec.cellpose_max_side,
        "fine_branch_detail_percentile": spec.fine_branch_detail_percentile,
        "fine_branch_intensity_percentile": spec.fine_branch_intensity_percentile,
        "fine_branch_min_area": spec.fine_branch_min_area,
        "fine_branch_min_major_axis": spec.fine_branch_min_major_axis,
        "fine_branch_min_eccentricity": spec.fine_branch_min_eccentricity,
        "fine_branch_gap_radius": spec.fine_branch_gap_radius,
        "fine_branch_evidence_mode": spec.fine_branch_evidence_mode,
        "border_margin": spec.border_margin,
        "edge_qc_margin": spec.edge_qc_margin,
        "border_complete_soma_margin": spec.border_complete_soma_margin,
        "border_complete_min_area_ratio": spec.border_complete_min_area_ratio,
        "border_complete_min_interior_fraction": spec.border_complete_min_interior_fraction,
        "error": error,
    }
    if spec.method == "log1p_gmm":
        row.update(
            distributional_threshold_diagnostics(
                struct,
                structural_projections,
                spec,
            )
        )
    row.update(branch_metrics)
    row.update(border_metrics)
    row.update(qc_metrics(mask, struct, d_proj, spec))
    if emit_progress:
        print(
            f"candidate {candidate_number:02d}/{candidate_count}: "
            f"z={z0 + 1}-{z1 + 1}, area={row['mask_area_px']}, "
            f"components={row['connected_components']}, soma={row['soma_supported_components']}, "
            f"fine_added={row['fine_branch_retained_px']}, method={method_used}",
            flush=True,
        )
    return mask, row, projection_key


def relabel_compartment_triplet(
    whole_labels: np.ndarray,
    soma_labels: np.ndarray,
    process_labels: np.ndarray,
    retained_ids: list[int],
) -> tuple[np.ndarray, np.ndarray, np.ndarray, dict[int, int]]:
    """Relabel retained Whole/Soma/Processes IDs atomically and contiguously."""

    maximum_id = max(
        int(whole_labels.max()),
        int(soma_labels.max()),
        int(process_labels.max()),
    )
    lookup = np.zeros(maximum_id + 1, dtype=np.uint16)
    mapping: dict[int, int] = {
        int(old_id): int(new_id)
        for new_id, old_id in enumerate(retained_ids, start=1)
    }
    for old_id, new_id in mapping.items():
        lookup[old_id] = new_id
    out_whole = lookup[whole_labels]
    out_soma = lookup[soma_labels]
    out_process = lookup[process_labels]
    return out_whole, out_soma, out_process, mapping

def circular_mask(
    shape: tuple[int, int],
    center_y: int,
    center_x: int,
    radius: int,
) -> np.ndarray:
    yy, xx = np.ogrid[: shape[0], : shape[1]]
    return (yy - center_y) ** 2 + (xx - center_x) ** 2 <= radius**2

def skeleton_topology(mask: np.ndarray) -> tuple[int, int, int]:
    skeleton = morphology.skeletonize(mask.astype(bool))
    if not skeleton.any():
        return 0, 0, 0
    neighbors = ndi.convolve(
        skeleton.astype(np.uint8),
        np.ones((3, 3), dtype=np.uint8),
        mode="constant",
        cval=0,
    ) - skeleton.astype(np.uint8)
    endpoints = int((skeleton & (neighbors == 1)).sum())
    branch_clusters = int(
        measure.label(skeleton & (neighbors >= 3), connectivity=2).max()
    )
    return int(skeleton.sum()), endpoints, branch_clusters

def robust_reference(values: list[float], index: int) -> tuple[float, float]:
    reference = np.asarray(
        [value for position, value in enumerate(values) if position != index and np.isfinite(value)],
        dtype=np.float64,
    )
    if reference.size == 0:
        return 0.0, 1.0
    median = float(np.median(reference))
    mad_scale = 1.4826 * float(np.median(np.abs(reference - median)))
    q25, q75 = np.percentile(reference, [25.0, 75.0])
    iqr_scale = float(q75 - q25) / 1.349
    scale = max(mad_scale, iqr_scale, 1e-6)
    return median, scale

def filter_morphology_outlier_instances(
    whole_labels: np.ndarray,
    soma_labels: np.ndarray,
    process_labels: np.ndarray,
    struct: np.ndarray,
    mean_pixel_um: float,
    pixel_area_um2: float,
    per_cell: list[dict],
    instance_metrics: dict,
    config: CompartmentConfig,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, list[dict], dict]:
    """Remove only whole-ID morphology outliers supported by multiple independent cues."""

    original_count = int(whole_labels.max())
    empty_metrics = {
        "enabled": bool(config.morphology_outlier_filter_enabled),
        "reference_count": original_count,
        "pre_filter_roi_count": original_count,
        "post_filter_roi_count": original_count,
        "removed_count": 0,
        "removed_area_px": 0,
        "removed_area_fraction": 0.0,
        "removed_original_ids": [],
        "flagged_original_ids": [],
        "id_mapping": {str(index): index for index in range(1, original_count + 1)},
        "details": [],
    }
    if not config.morphology_outlier_filter_enabled or original_count == 0:
        return whole_labels, soma_labels, process_labels, per_cell, empty_metrics

    per_cell_by_id = {int(row["astrocyte_id"]): row for row in per_cell}
    split_child_ids = {
        int(astrocyte_id)
        for detail in instance_metrics.get("split_components", [])
        for astrocyte_id in detail.get("new_astrocyte_ids", [])
    }
    properties = {int(prop.label): prop for prop in measure.regionprops(whole_labels)}
    feature_rows: list[dict] = []
    for astrocyte_id in range(1, original_count + 1):
        prop = properties[astrocyte_id]
        min_row, min_col, max_row, max_col = prop.bbox
        crop = np.s_[min_row:max_row, min_col:max_col]
        component = whole_labels[crop] == astrocyte_id
        local_process = process_labels[crop] == astrocyte_id
        distance_um = ndi.distance_transform_edt(
            component,
            sampling=(mean_pixel_um, mean_pixel_um),
        )
        skeleton_px, endpoint_count, branchpoint_count = skeleton_topology(component)
        process_component_count = int(measure.label(local_process, connectivity=2).max())
        axis_ratio = float(prop.major_axis_length) / max(float(prop.minor_axis_length), 1e-6)
        edge_touch = bool(
            min_row == 0
            or min_col == 0
            or max_row == whole_labels.shape[0]
            or max_col == whole_labels.shape[1]
        )
        local_struct = struct[crop]
        ring = morphology.binary_dilation(component, footprint=morphology.disk(5)) & ~component
        structural_contrast = float(np.median(local_struct[component])) - (
            float(np.median(local_struct[ring])) if ring.any() else 0.0
        )
        cell = per_cell_by_id[astrocyte_id]
        feature_rows.append(
            {
                "original_astrocyte_id": astrocyte_id,
                "area_px": int(prop.area),
                "area_um2": float(prop.area) * pixel_area_um2,
                "core_radius_um": float(np.percentile(distance_um[component], 95.0)),
                "axis_ratio": axis_ratio,
                "eccentricity": float(prop.eccentricity),
                "solidity": float(prop.solidity),
                "skeleton_length_um": skeleton_px * mean_pixel_um,
                "endpoint_count": endpoint_count,
                "branchpoint_count": branchpoint_count,
                "process_component_count": process_component_count,
                "structural_contrast": structural_contrast,
                "soma_anchor_count": int(cell["soma_anchor_count"]),
                "nucleus_score": float(cell["nucleus_score"]),
                "process_fraction": float(cell["process_fraction"]),
                "edge_touch": edge_touch,
                "accepted_split_child": astrocyte_id in split_child_ids,
            }
        )

    # Border-incomplete cells were already removed upstream; preserved border cells
    # remain valid peers, but are still protected from morphology-only deletion.
    reference_indices = list(range(len(feature_rows)))
    reference_count = len(reference_indices)
    metric_values = {
        "log_area": [math.log(max(row["area_um2"], 1e-6)) for row in feature_rows],
        "log_core": [math.log(max(row["core_radius_um"], 1e-6)) for row in feature_rows],
        "log_skeleton": [
            math.log(max(row["skeleton_length_um"], 1e-6)) for row in feature_rows
        ],
        "branchpoints": [float(row["branchpoint_count"]) for row in feature_rows],
        "endpoints": [float(row["endpoint_count"]) for row in feature_rows],
        "solidity": [float(row["solidity"]) for row in feature_rows],
        "axis_ratio": [float(row["axis_ratio"]) for row in feature_rows],
        "structural_contrast": [
            float(row["structural_contrast"]) for row in feature_rows
        ],
    }
    reference_medians = {
        key: float(np.median([values[index] for index in reference_indices]))
        if reference_indices
        else 0.0
        for key, values in metric_values.items()
    }
    process_fraction_median = float(
        np.median([feature_rows[index]["process_fraction"] for index in reference_indices])
    ) if reference_indices else 0.0
    branchpoint_median = reference_medians["branchpoints"]
    ramified_field = bool(process_fraction_median >= 0.60 and branchpoint_median >= 3.0)

    proposed: list[tuple[int, float, str]] = []
    flagged_ids: list[int] = []
    for index, row in enumerate(feature_rows):
        z_scores: dict[str, float] = {}
        for key, values in metric_values.items():
            allowed_reference = [values[position] for position in reference_indices]
            if index in reference_indices:
                local_index = reference_indices.index(index)
                median, scale = robust_reference(allowed_reference, local_index)
            else:
                median = float(np.median(allowed_reference)) if allowed_reference else 0.0
                mad = (
                    1.4826
                    * float(np.median(np.abs(np.asarray(allowed_reference) - median)))
                    if allowed_reference
                    else 1.0
                )
                scale = max(mad, 1e-6)
            z_scores[key] = (values[index] - median) / scale

        median_area_um2 = math.exp(reference_medians["log_area"])
        median_core_um = math.exp(reference_medians["log_core"])
        median_skeleton_um = math.exp(reference_medians["log_skeleton"])
        no_anchor = row["soma_anchor_count"] == 0
        votes = {
            "small_area": bool(
                z_scores["log_area"] <= -config.morphology_outlier_robust_z
                and row["area_um2"] <= 0.45 * median_area_um2
            ),
            "thin_core": bool(
                z_scores["log_core"] <= -config.morphology_outlier_robust_z
                and row["core_radius_um"] <= 0.65 * median_core_um
            ),
            "short_skeleton": bool(
                z_scores["log_skeleton"] <= -config.morphology_outlier_robust_z
                and row["skeleton_length_um"] <= 0.45 * median_skeleton_um
            ),
            "few_branches": bool(
                z_scores["branchpoints"] <= -config.morphology_outlier_robust_z
                and row["branchpoint_count"] <= max(1, int(0.25 * branchpoint_median))
            ),
            "few_endpoints": bool(
                z_scores["endpoints"] <= -config.morphology_outlier_robust_z
                and row["endpoint_count"] <= 2
            ),
            "high_solidity": bool(
                z_scores["solidity"] >= config.morphology_outlier_robust_z
                and row["solidity"] >= 0.82
            ),
            "high_axis_ratio": bool(
                z_scores["axis_ratio"] >= config.morphology_outlier_robust_z
                and row["axis_ratio"] >= config.morphology_fragment_min_axis_ratio
            ),
            "low_structural_contrast": bool(
                z_scores["structural_contrast"] <= -config.morphology_outlier_robust_z
                and row["structural_contrast"] <= 0.0
            ),
        }
        consensus = int(sum(votes.values()))
        absolute_fragment = bool(
            no_anchor
            and row["axis_ratio"] >= config.morphology_fragment_min_axis_ratio
            and row["branchpoint_count"] <= config.morphology_fragment_max_branchpoints
            and row["core_radius_um"] <= config.morphology_fragment_max_core_radius_um
        )
        relative_elongated_fragment = bool(
            reference_count >= config.morphology_outlier_min_reference_count
            and row["axis_ratio"] >= config.morphology_fragment_min_axis_ratio
            and row["area_um2"] <= 0.35 * median_area_um2
            and row["skeleton_length_um"] <= 0.35 * median_skeleton_um
            and row["core_radius_um"] <= 1.10 * median_core_um
        )
        population_outlier = bool(
            reference_count >= config.morphology_outlier_min_reference_count
            and no_anchor
            and consensus >= config.morphology_outlier_min_consensus
        )
        compact_cues = {
            "relative_small_area": bool(row["area_um2"] <= 0.45 * median_area_um2),
            "relative_short_skeleton": bool(
                row["skeleton_length_um"] <= 0.45 * median_skeleton_um
            ),
            "relative_few_branches": bool(
                row["branchpoint_count"] <= max(1, int(0.45 * branchpoint_median))
            ),
            "relative_low_process_fraction": bool(
                row["process_fraction"] <= min(0.55, 0.75 * process_fraction_median)
            ),
            "relative_high_solidity": bool(
                row["solidity"] >= max(0.50, 1.25 * reference_medians["solidity"])
            ),
        }
        compact_consensus = int(sum(compact_cues.values()))
        compact_outlier = bool(
            reference_count >= config.morphology_outlier_min_reference_count
            and ramified_field
            and compact_cues["relative_high_solidity"]
            and compact_consensus >= config.morphology_outlier_min_consensus + 1
        )
        protected = bool(row["edge_touch"] or row["accepted_split_child"])
        reason = ""
        if relative_elongated_fragment:
            reason = "small_elongated_morphology_outlier"
        elif absolute_fragment:
            reason = "unanchored_thin_fragment"
        elif population_outlier:
            reason = "unanchored_multimetric_outlier"
        elif compact_outlier:
            reason = "compact_multimetric_outlier"
        if consensus > 0 or compact_consensus > 0 or absolute_fragment:
            flagged_ids.append(int(row["original_astrocyte_id"]))
        severity = float(consensus + compact_consensus) + max(
            abs(min(z_scores.values(), default=0.0)),
            abs(max(z_scores.values(), default=0.0)),
        ) / 10.0
        if reason and not protected:
            proposed.append((int(row["original_astrocyte_id"]), severity, reason))
        row.update(
            {
                "robust_z": {key: round(value, 4) for key, value in z_scores.items()},
                "outlier_votes": [key for key, value in votes.items() if value],
                "outlier_consensus": consensus,
                "compact_outlier_cues": [
                    key for key, value in compact_cues.items() if value
                ],
                "compact_outlier_consensus": compact_consensus,
                "outlier_reason": reason,
                "outlier_protected": protected,
            }
        )

    max_remove_count = max(1, int(math.floor(0.25 * original_count)))
    proposed.sort(key=lambda item: item[1], reverse=True)
    removed_ids = {item[0] for item in proposed[:max_remove_count]}
    removed_area_px = int(np.isin(whole_labels, list(removed_ids)).sum())
    if (
        original_count - len(removed_ids) < 5
        or removed_area_px > 0.25 * int((whole_labels > 0).sum())
    ):
        removed_ids.clear()
        removed_area_px = 0

    retained_ids = [
        astrocyte_id
        for astrocyte_id in range(1, original_count + 1)
        if astrocyte_id not in removed_ids
    ]
    id_mapping = {old_id: new_id for new_id, old_id in enumerate(retained_ids, start=1)}
    filtered_whole = np.zeros_like(whole_labels, dtype=np.uint16)
    filtered_soma = np.zeros_like(soma_labels, dtype=np.uint16)
    filtered_process = np.zeros_like(process_labels, dtype=np.uint16)
    filtered_per_cell: list[dict] = []
    for old_id, new_id in id_mapping.items():
        filtered_whole[whole_labels == old_id] = new_id
        filtered_soma[soma_labels == old_id] = new_id
        filtered_process[process_labels == old_id] = new_id
        updated = dict(per_cell_by_id[old_id])
        updated["original_astrocyte_id"] = old_id
        updated["astrocyte_id"] = new_id
        feature = next(
            row for row in feature_rows if row["original_astrocyte_id"] == old_id
        )
        updated["morphology_qc"] = feature
        filtered_per_cell.append(updated)

    removed_details = [
        {
            **next(row for row in feature_rows if row["original_astrocyte_id"] == old_id),
            "reason": next(item[2] for item in proposed if item[0] == old_id),
        }
        for old_id in sorted(removed_ids)
    ]
    metrics = {
        "enabled": True,
        "reference_count": reference_count,
        "pre_filter_roi_count": original_count,
        "post_filter_roi_count": len(retained_ids),
        "removed_count": len(removed_ids),
        "removed_area_px": removed_area_px,
        "removed_area_fraction": round(
            removed_area_px / max(int((whole_labels > 0).sum()), 1),
            6,
        ),
        "removed_original_ids": sorted(removed_ids),
        "flagged_original_ids": sorted(set(flagged_ids)),
        "id_mapping": {str(old): new for old, new in id_mapping.items()},
        "ramified_field": ramified_field,
        "reference_medians": {
            "area_um2": round(math.exp(reference_medians["log_area"]), 6),
            "core_radius_um": round(math.exp(reference_medians["log_core"]), 6),
            "skeleton_length_um": round(
                math.exp(reference_medians["log_skeleton"]),
                6,
            ),
            "branchpoints": round(branchpoint_median, 6),
            "process_fraction": round(process_fraction_median, 6),
        },
        "details": removed_details,
        "all_features": feature_rows,
    }
    return (
        filtered_whole,
        filtered_soma,
        filtered_process,
        filtered_per_cell,
        metrics,
    )

def filter_instances_by_valid_soma(
    whole_labels: np.ndarray,
    soma_labels: np.ndarray,
    process_labels: np.ndarray,
    per_cell: list[dict],
    profile: str,
    unresolved_multi_soma_ids: set[int] | None = None,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, list[dict], dict]:
    """Keep only cell IDs with exactly one connected, trusted Soma."""

    if profile not in {"mature", "neonatal"}:
        raise ValueError(f"Unknown Soma-gate profile: {profile}")
    if whole_labels.shape != soma_labels.shape or whole_labels.shape != process_labels.shape:
        raise ValueError("Valid-Soma gate received mismatched label geometry")
    original_ids = sorted(int(value) for value in np.unique(whole_labels) if int(value) > 0)
    expected_ids = list(range(1, int(whole_labels.max()) + 1))
    if original_ids != expected_ids:
        raise ValueError(
            "Valid-Soma gate requires contiguous pre-gate Whole IDs: "
            f"observed={original_ids}"
        )
    per_cell_by_id = {int(row["astrocyte_id"]): row for row in per_cell}
    if set(per_cell_by_id) != set(original_ids):
        raise ValueError(
            "Valid-Soma gate per-cell rows do not match Whole IDs: "
            f"whole={original_ids}, rows={sorted(per_cell_by_id)}"
        )

    retained_ids: list[int] = []
    removed_details: list[dict] = []
    unresolved_ids = set(unresolved_multi_soma_ids or set())
    if not unresolved_ids.issubset(set(original_ids)):
        raise ValueError(
            "Disqualified split IDs are absent from Whole labels: "
            f"{sorted(unresolved_ids - set(original_ids))}"
        )
    for astrocyte_id in original_ids:
        row = per_cell_by_id[astrocyte_id]
        anchor_count = int(row.get("soma_anchor_count", 0))
        whole_area_px = int((whole_labels == astrocyte_id).sum())
        soma_area_px = int((soma_labels == astrocyte_id).sum())
        process_area_px = int((process_labels == astrocyte_id).sum())
        soma_component_count = int(
            measure.label(soma_labels == astrocyte_id, connectivity=2).max()
        )
        reason = ""
        if astrocyte_id in unresolved_ids:
            reason = "unresolved_multi_soma_instance_split"
        elif anchor_count == 0 or soma_area_px == 0:
            reason = "no_valid_soma"
        elif anchor_count > 1:
            reason = "multiple_valid_somata_after_instance_split"
        elif soma_component_count != 1:
            reason = "disconnected_soma_geometry"
        elif process_area_px == 0:
            reason = "empty_processes_compartment"
        if reason:
            removed_details.append(
                {
                    "pre_gate_astrocyte_id": astrocyte_id,
                    "reason": reason,
                    "soma_anchor_count": anchor_count,
                    "soma_component_count": soma_component_count,
                    "whole_area_px": whole_area_px,
                    "soma_area_px": soma_area_px,
                    "process_area_px": process_area_px,
                }
            )
        else:
            retained_ids.append(astrocyte_id)

    if not retained_ids:
        raise RuntimeError(
            f"No {profile} Whole Astrocyte instance retained exactly one valid Soma; "
            "the run was stopped before measurement"
        )

    id_mapping = {
        old_id: new_id for new_id, old_id in enumerate(retained_ids, start=1)
    }
    filtered_whole = np.zeros_like(whole_labels, dtype=np.uint16)
    filtered_soma = np.zeros_like(soma_labels, dtype=np.uint16)
    filtered_process = np.zeros_like(process_labels, dtype=np.uint16)
    filtered_per_cell: list[dict] = []
    for old_id, new_id in id_mapping.items():
        filtered_whole[whole_labels == old_id] = new_id
        filtered_soma[soma_labels == old_id] = new_id
        filtered_process[process_labels == old_id] = new_id
        updated = dict(per_cell_by_id[old_id])
        updated.setdefault("original_astrocyte_id", old_id)
        updated["pre_soma_gate_astrocyte_id"] = old_id
        updated["astrocyte_id"] = new_id
        filtered_per_cell.append(updated)

    final_ids = set(range(1, len(retained_ids) + 1))
    observed = {
        key: set(int(value) for value in np.unique(labels) if int(value) > 0)
        for key, labels in (
            ("whole", filtered_whole),
            ("soma", filtered_soma),
            ("processes", filtered_process),
        )
    }
    if any(ids != final_ids for ids in observed.values()):
        raise RuntimeError(
            "Valid-Soma gate failed to preserve synchronized compartment IDs: "
            f"{observed}"
        )
    if any(int(row["soma_anchor_count"]) != 1 for row in filtered_per_cell):
        raise RuntimeError("A retained cell does not have exactly one Soma anchor")

    removed_ids = [row["pre_gate_astrocyte_id"] for row in removed_details]
    removed_area_px = int(np.isin(whole_labels, removed_ids).sum())
    metrics = {
        "enabled": True,
        "profile": profile,
        "criterion": "exactly_one_valid_soma_per_final_astrocyte",
        "pre_gate_roi_count": len(original_ids),
        "post_gate_roi_count": len(retained_ids),
        "removed_count": len(removed_ids),
        "removed_area_px": removed_area_px,
        "removed_area_fraction": round(
            removed_area_px / max(int((whole_labels > 0).sum()), 1),
            6,
        ),
        "removed_pre_gate_ids": removed_ids,
        "unresolved_multi_soma_pre_gate_ids": sorted(unresolved_ids),
        "retained_pre_gate_ids": retained_ids,
        "id_mapping": {str(old_id): new_id for old_id, new_id in id_mapping.items()},
        "details": removed_details,
    }
    return (
        filtered_whole,
        filtered_soma,
        filtered_process,
        filtered_per_cell,
        metrics,
    )


def retain_primary_anchor_extent(
    selected_extent: np.ndarray,
    selected_core: np.ndarray,
    component: np.ndarray,
) -> tuple[np.ndarray, int, int]:
    """Keep the single DAPI-supported extent body belonging to one Soma anchor."""

    extent = (selected_extent | selected_core) & component
    extent_labels = measure.label(extent, connectivity=2)
    component_count = int(extent_labels.max())
    if component_count <= 1:
        return extent.astype(bool), 0, 0

    areas = np.bincount(
        extent_labels.ravel(),
        minlength=component_count + 1,
    )
    core_counts = np.bincount(
        extent_labels[selected_core & component],
        minlength=component_count + 1,
    )
    primary_label = max(
        range(1, component_count + 1),
        key=lambda label_id: (int(core_counts[label_id]), int(areas[label_id])),
    )
    retained = extent_labels == primary_label
    removed_px = int(extent.sum() - retained.sum())
    return retained, component_count - 1, removed_px

def score_nuclei_for_component(
    component: np.ndarray,
    nearest_nucleus_labels: np.ndarray,
    nucleus_distance: np.ndarray,
    distance: np.ndarray,
    struct: np.ndarray,
    cellpose_mask: np.ndarray,
    link_radius_px: int,
    ambiguity_delta: float,
) -> tuple[list[dict], bool]:
    association_zone = component & (nucleus_distance <= link_radius_px)
    candidate_ids = np.unique(nearest_nucleus_labels[association_zone])
    candidate_ids = candidate_ids[candidate_ids > 0]
    if candidate_ids.size == 0:
        return [], False

    core_reference = max(float(np.percentile(distance[component], 99.0)), 1.0)
    scored: list[dict] = []
    for nucleus_id in candidate_ids:
        near_nucleus = (
            component
            & (nearest_nucleus_labels == int(nucleus_id))
            & (nucleus_distance <= link_radius_px)
        )
        if not near_nucleus.any():
            continue
        thickness_support = min(
            1.0,
            float(np.percentile(distance[near_nucleus], 90.0)) / core_reference,
        )
        structural_support = float(np.percentile(struct[near_nucleus], 75.0))
        overlap_fraction = float(
            (component & (nearest_nucleus_labels == int(nucleus_id)) & (nucleus_distance == 0)).sum()
        ) / max(int(((nearest_nucleus_labels == int(nucleus_id)) & (nucleus_distance == 0)).sum()), 1)
        model_support = float(cellpose_mask[near_nucleus].mean()) if near_nucleus.any() else 0.0
        score = (
            0.46 * thickness_support
            + 0.31 * structural_support
            + 0.18 * min(1.0, overlap_fraction * 2.0)
            + 0.05 * model_support
        )
        nucleus_pixels = (nearest_nucleus_labels == int(nucleus_id)) & (nucleus_distance == 0)
        nucleus_coords = np.argwhere(nucleus_pixels)
        if nucleus_coords.size == 0:
            continue
        center_y, center_x = nucleus_coords.mean(axis=0)
        scored.append(
            {
                "nucleus_id": int(nucleus_id),
                "score": float(score),
                "thickness_support": float(thickness_support),
                "structural_support": float(structural_support),
                "overlap_fraction": float(overlap_fraction),
                "model_support": float(model_support),
                "center_y": float(center_y),
                "center_x": float(center_x),
            }
        )
    if not scored:
        return [], False
    scored.sort(key=lambda row: row["score"], reverse=True)
    ambiguous = len(scored) > 1 and scored[0]["score"] - scored[1]["score"] < ambiguity_delta
    return scored, ambiguous

def select_soma_anchor_groups(
    scored_nuclei: list[dict],
    mean_pixel_um: float,
    config: CompartmentConfig,
) -> list[dict]:
    """Keep spatially distinct, high-confidence soma anchors and merge nearby DAPI fragments."""

    if not scored_nuclei:
        return []
    min_separation_px = config.soma_anchor_min_separation_um / mean_pixel_um
    top_score = float(scored_nuclei[0]["score"])
    groups: list[dict] = []
    for index, candidate in enumerate(scored_nuclei):
        nearest_group = None
        nearest_distance = math.inf
        for group in groups:
            distance_px = math.hypot(
                candidate["center_y"] - group["center_y"],
                candidate["center_x"] - group["center_x"],
            )
            if distance_px < nearest_distance:
                nearest_group = group
                nearest_distance = distance_px
        if nearest_group is not None and nearest_distance < min_separation_px:
            nearest_group["nucleus_ids"].append(candidate["nucleus_id"])
            nearest_group["member_scores"].append(candidate["score"])
            continue

        is_primary = index == 0
        passes_primary_threshold = (
            candidate["score"] >= config.primary_anchor_min_score
            and candidate["thickness_support"] >= config.primary_anchor_min_thickness_support
            and candidate["structural_support"] >= config.primary_anchor_min_structural_support
            and (
                candidate["overlap_fraction"] >= config.primary_anchor_min_overlap_fraction
                or candidate["model_support"] >= config.primary_anchor_min_model_support
            )
        )
        has_local_support = (
            candidate["thickness_support"] >= config.multi_anchor_min_thickness_support
            and candidate["structural_support"] >= config.multi_anchor_min_structural_support
            and (
                candidate["overlap_fraction"] >= config.multi_anchor_min_overlap_fraction
                or candidate["model_support"] >= config.multi_anchor_min_model_support
            )
        )
        passes_secondary_threshold = (
            candidate["score"] >= config.multi_anchor_min_score
            and candidate["score"] >= top_score - config.multi_anchor_max_score_delta
            and has_local_support
        )
        if is_primary and not passes_primary_threshold:
            continue
        if not is_primary and not passes_secondary_threshold:
            continue
        if len(groups) >= config.max_soma_anchors_per_whole_roi:
            continue
        groups.append(
            {
                "nucleus_ids": [candidate["nucleus_id"]],
                "member_scores": [candidate["score"]],
                "score": candidate["score"],
                "center_y": candidate["center_y"],
                "center_x": candidate["center_x"],
            }
        )
    return groups

def select_validated_soma_anchor_groups(
    scored_nuclei: list[dict],
    component: np.ndarray,
    local_nuclei_labels: np.ndarray,
    local_grouped_extent_labels: np.ndarray,
    validated_group_by_id: dict[int, dict],
    minimum_overlap_px: int,
    assigned_group_ids: set[int] | None = None,
) -> list[dict]:
    """Use each independently accepted 3D nucleus group as a required Soma anchor."""

    scored_by_id = {
        int(row["nucleus_id"]): row for row in scored_nuclei
    }
    present_group_ids = np.unique(local_grouped_extent_labels[component])
    groups: list[dict] = []
    for group_id_value in present_group_ids:
        group_id = int(group_id_value)
        if group_id <= 0 or group_id not in validated_group_by_id:
            continue
        if assigned_group_ids is not None and group_id not in assigned_group_ids:
            continue
        group = validated_group_by_id[group_id]
        if not bool(group["accepted"]):
            continue
        extent_overlap = component & (local_grouped_extent_labels == group_id)
        overlap_px = int(extent_overlap.sum())
        if overlap_px < minimum_overlap_px:
            continue
        object_ids = tuple(int(value) for value in group["object_ids"])
        nucleus_ids = [
            object_id
            for object_id in object_ids
            if np.any(local_nuclei_labels == object_id)
        ]
        if not nucleus_ids:
            continue
        members = [
            scored_by_id[object_id]
            for object_id in nucleus_ids
            if object_id in scored_by_id
        ]
        nucleus_pixels = component & np.isin(local_nuclei_labels, nucleus_ids)
        center_pixels = nucleus_pixels if nucleus_pixels.any() else extent_overlap
        center_y, center_x = np.argwhere(center_pixels).mean(axis=0)
        score = max(
            (float(row["score"]) for row in members),
            default=float(group["enclosure_score"]),
        )
        groups.append(
            {
                "nucleus_ids": nucleus_ids,
                "member_scores": [float(row["score"]) for row in members],
                "score": score,
                "center_y": float(center_y),
                "center_x": float(center_x),
                "validated_group_id": group_id,
                "validated_object_ids": list(object_ids),
                "extent_overlap_px": overlap_px,
                "source": "independently_accepted_3d_nucleus_group",
            }
        )
    groups.sort(
        key=lambda row: (
            -int(row["extent_overlap_px"]),
            -float(row["score"]),
            int(row["validated_group_id"]),
        )
    )
    return groups

def restore_low_support_branch_gaps(
    whole_mask: np.ndarray,
    dapi_projection: np.ndarray,
    struct: np.ndarray,
    pixel_width_um: float,
    pixel_height_um: float,
    config: CompartmentConfig,
    nuclei_mask: np.ndarray | None = None,
) -> tuple[np.ndarray, dict]:
    """Restore dark internal valleys without shrinking edges or breaking connectivity."""

    input_mask = whole_mask.astype(bool)
    empty_metrics = {
        "enabled": bool(config.branch_gap_restore_enabled),
        "input_area_px": int(input_mask.sum()),
        "output_area_px": int(input_mask.sum()),
        "removed_px": 0,
        "removed_fraction": 0.0,
        "accepted_gap_count": 0,
        "rejected_disconnect_count": 0,
        "structural_cut": 0.0,
        "removed_structural_mean": 0.0,
    }
    if not config.branch_gap_restore_enabled or not input_mask.any():
        return input_mask, empty_metrics

    structural_cut = float(
        np.percentile(struct[input_mask], config.branch_gap_low_percentile)
    )
    distance_um = ndi.distance_transform_edt(
        input_mask,
        sampling=(pixel_height_um, pixel_width_um),
    )
    nuclei = (
        nuclei_mask.astype(bool, copy=False)
        if nuclei_mask is not None
        else dapi_nuclei_mask(
            dapi_projection,
            percentile_floor=config.dapi_percentile_floor,
        )
    )
    if nuclei.any():
        nucleus_distance_um = ndi.distance_transform_edt(
            ~nuclei,
            sampling=(pixel_height_um, pixel_width_um),
        )
        nucleus_protection = nucleus_distance_um <= config.branch_gap_nucleus_protect_um
    else:
        nucleus_protection = np.zeros_like(input_mask, dtype=bool)

    gap_candidates = (
        input_mask
        & (distance_um >= config.branch_gap_min_depth_um)
        & (struct < structural_cut)
        & ~nucleus_protection
    )
    component_labels = measure.label(input_mask, connectivity=2)
    output = input_mask.copy()
    accepted_gap_count = 0
    rejected_disconnect_count = 0
    pixel_area_um2 = pixel_width_um * pixel_height_um
    mean_pixel_um = math.sqrt(pixel_area_um2)

    for component_id in range(1, int(component_labels.max()) + 1):
        component = component_labels == component_id
        candidate_labels = measure.label(gap_candidates & component, connectivity=2)
        properties = sorted(
            measure.regionprops(candidate_labels),
            key=lambda item: item.area,
            reverse=True,
        )
        for prop in properties:
            area_um2 = float(prop.area) * pixel_area_um2
            major_axis_um = float(prop.major_axis_length) * mean_pixel_um
            gap_like = (
                area_um2 >= config.branch_gap_min_area_um2
                and (
                    float(prop.eccentricity) >= config.branch_gap_min_eccentricity
                    or major_axis_um >= config.branch_gap_min_major_axis_um
                )
            )
            if not gap_like:
                continue
            gap = candidate_labels == prop.label
            trial_component = output & component & ~gap
            if int(measure.label(trial_component, connectivity=2).max()) != 1:
                rejected_disconnect_count += 1
                continue
            output[gap] = False
            accepted_gap_count += 1

    if int(measure.label(output, connectivity=2).max()) != int(component_labels.max()):
        raise RuntimeError("Branch-gap restoration changed the Whole component count")
    removed = input_mask & ~output
    removed_px = int(removed.sum())
    return output, {
        "enabled": True,
        "input_area_px": int(input_mask.sum()),
        "output_area_px": int(output.sum()),
        "removed_px": removed_px,
        "removed_fraction": round(removed_px / max(int(input_mask.sum()), 1), 6),
        "accepted_gap_count": accepted_gap_count,
        "rejected_disconnect_count": rejected_disconnect_count,
        "structural_cut": round(structural_cut, 6),
        "removed_structural_mean": round(
            float(struct[removed].mean()) if removed.any() else 0.0,
            6,
        ),
    }

def split_touching_whole_instances(
    whole_mask: np.ndarray,
    nuclei_labels: np.ndarray,
    nearest_nucleus_labels: np.ndarray,
    nucleus_distance: np.ndarray,
    struct: np.ndarray,
    cellpose_mask: np.ndarray,
    mean_pixel_um: float,
    pixel_area_um2: float,
    link_radius_px: int,
    config: CompartmentConfig,
) -> tuple[np.ndarray, dict]:
    """Split only high-confidence multi-soma components without changing Whole pixels."""

    base_labels = measure.label(whole_mask, connectivity=2).astype(np.uint16)
    base_count = int(base_labels.max())
    empty_metrics = {
        "base_connected_component_count": base_count,
        "final_instance_count": base_count,
        "split_component_count": 0,
        "split_added_roi_count": 0,
        "split_components": [],
        "split_rejected_count": 0,
    }
    if not config.instance_split_enabled or base_count < 1:
        return base_labels, empty_metrics

    output = np.zeros_like(base_labels, dtype=np.uint16)
    next_id = 1
    split_details: list[dict] = []
    rejected_count = 0
    padding = link_radius_px + 4

    for prop in measure.regionprops(base_labels):
        min_row, min_col, max_row, max_col = prop.bbox
        row0 = max(0, min_row - padding)
        col0 = max(0, min_col - padding)
        row1 = min(base_labels.shape[0], max_row + padding)
        col1 = min(base_labels.shape[1], max_col + padding)
        crop = np.s_[row0:row1, col0:col1]
        component = base_labels[crop] == prop.label
        local_struct = struct[crop]
        local_cellpose = cellpose_mask[crop]
        local_nuclei_labels = nuclei_labels[crop]
        local_nearest_labels = nearest_nucleus_labels[crop]
        local_nucleus_distance = nucleus_distance[crop]
        distance = ndi.distance_transform_edt(component)
        scored_nuclei, _ = score_nuclei_for_component(
            component,
            local_nearest_labels,
            local_nucleus_distance,
            distance,
            local_struct,
            local_cellpose,
            link_radius_px,
            config.ambiguity_score_delta,
        )
        anchor_groups = select_soma_anchor_groups(scored_nuclei, mean_pixel_um, config)
        scored_by_id = {row["nucleus_id"]: row for row in scored_nuclei}
        strict_groups: list[dict] = []
        for group in anchor_groups:
            representative = scored_by_id.get(group["nucleus_ids"][0])
            if representative is None:
                continue
            has_channel_support = (
                representative["overlap_fraction"] >= config.multi_anchor_min_overlap_fraction
                or representative["model_support"] >= config.multi_anchor_min_model_support
            )
            if (
                group["score"] >= config.instance_split_min_anchor_score
                and representative["thickness_support"]
                >= config.multi_anchor_min_thickness_support
                and representative["structural_support"]
                >= config.multi_anchor_min_structural_support
                and has_channel_support
            ):
                strict_groups.append(group)

        split_reason = "not_two_strict_anchors"
        partition = None
        anchor_separation_um = 0.0
        neck_core_ratio = math.inf
        child_areas: list[int] = []
        if len(strict_groups) == 2 and len(strict_groups) <= config.instance_split_max_markers:
            anchor_separation_um = mean_pixel_um * math.hypot(
                strict_groups[0]["center_y"] - strict_groups[1]["center_y"],
                strict_groups[0]["center_x"] - strict_groups[1]["center_x"],
            )
            if anchor_separation_um >= config.instance_split_min_anchor_separation_um:
                markers = np.zeros_like(component, dtype=np.int32)
                core_peaks: list[float] = []
                marker_masks: list[np.ndarray] = []
                distance_scale = max(float(np.percentile(distance[component], 99.0)), 1.0)
                for marker_id, group in enumerate(strict_groups, start=1):
                    selected_nucleus = np.isin(local_nuclei_labels, group["nucleus_ids"])
                    selected_distance = ndi.distance_transform_edt(~selected_nucleus)
                    search_region = component & (selected_distance <= link_radius_px)
                    if not search_region.any():
                        break
                    seed_score = (
                        0.72 * np.clip(distance / distance_scale, 0, 1)
                        + 0.23 * local_struct
                        + 0.05 * local_cellpose.astype(np.float32)
                    )
                    seed_score = np.where(search_region, seed_score, -np.inf)
                    seed_y, seed_x = np.unravel_index(
                        int(np.argmax(seed_score)),
                        seed_score.shape,
                    )
                    marker_radius_px = max(1, int(round(0.25 / mean_pixel_um)))
                    marker_mask = component & circular_mask(
                        component.shape,
                        seed_y,
                        seed_x,
                        marker_radius_px,
                    )
                    if not marker_mask.any() or np.any(markers[marker_mask] > 0):
                        break
                    markers[marker_mask] = marker_id
                    marker_masks.append(marker_mask)
                    core_neighborhood = component & circular_mask(
                        component.shape,
                        seed_y,
                        seed_x,
                        max(link_radius_px, int(round(0.75 / mean_pixel_um))),
                    )
                    core_peaks.append(
                        max(float(np.percentile(distance[core_neighborhood], 90.0)), 1.0)
                    )

                if len(marker_masks) == 2:
                    distance_cost = 1.0 - np.clip(distance / distance_scale, 0, 1)
                    structural_cost = 1.0 - filters.gaussian(
                        local_struct,
                        sigma=1.0,
                        preserve_range=True,
                    )
                    elevation = 0.72 * distance_cost + 0.28 * structural_cost
                    candidate_partition = segmentation.watershed(
                        elevation,
                        markers=markers,
                        mask=component,
                        watershed_line=False,
                        connectivity=np.ones((3, 3), dtype=bool),
                    ).astype(np.uint16)
                    child_areas = [
                        int((candidate_partition == marker_id).sum())
                        for marker_id in (1, 2)
                    ]
                    minimum_child_area = max(
                        int(round(config.instance_split_min_child_area_um2 / pixel_area_um2)),
                        int(round(config.instance_split_min_child_fraction * int(component.sum()))),
                    )
                    boundary = np.zeros_like(component, dtype=bool)
                    vertical = (
                        (candidate_partition[1:] > 0)
                        & (candidate_partition[:-1] > 0)
                        & (candidate_partition[1:] != candidate_partition[:-1])
                    )
                    horizontal = (
                        (candidate_partition[:, 1:] > 0)
                        & (candidate_partition[:, :-1] > 0)
                        & (candidate_partition[:, 1:] != candidate_partition[:, :-1])
                    )
                    boundary[1:] |= vertical
                    boundary[:-1] |= vertical
                    boundary[:, 1:] |= horizontal
                    boundary[:, :-1] |= horizontal
                    if boundary.any():
                        neck_core_ratio = float(np.percentile(distance[boundary], 75.0)) / max(
                            min(core_peaks),
                            1.0,
                        )
                    children_connected = all(
                        int(measure.label(candidate_partition == marker_id, connectivity=2).max()) == 1
                        for marker_id in (1, 2)
                    )
                    markers_retained = all(
                        bool(np.all(candidate_partition[marker_masks[index]] == index + 1))
                        for index in range(2)
                    )
                    if min(child_areas) < minimum_child_area:
                        split_reason = "child_too_small"
                    elif not children_connected or not markers_retained:
                        split_reason = "invalid_partition"
                    elif neck_core_ratio > config.instance_split_max_neck_core_ratio:
                        split_reason = "neck_too_thick"
                    else:
                        split_reason = "accepted"
                        partition = candidate_partition
                else:
                    split_reason = "marker_build_failed"
            else:
                split_reason = "anchors_too_close"

        output_view = output[crop]
        if partition is None:
            output_view[component] = next_id
            next_id += 1
            if len(strict_groups) >= 2:
                rejected_count += 1
            continue

        assigned_ids: list[int] = []
        for marker_id in range(1, int(partition.max()) + 1):
            output_view[partition == marker_id] = next_id
            assigned_ids.append(next_id)
            next_id += 1
        split_details.append(
            {
                "base_component_id": int(prop.label),
                "new_astrocyte_ids": assigned_ids,
                "anchor_scores": [round(float(group["score"]), 6) for group in strict_groups],
                "anchor_separation_um": round(anchor_separation_um, 6),
                "child_areas_px": child_areas,
                "neck_core_ratio": round(neck_core_ratio, 6),
                "reason": split_reason,
            }
        )

    if not np.array_equal(output > 0, whole_mask.astype(bool)):
        raise RuntimeError("Instance splitting changed the Whole Astrocyte pixel union")
    final_count = int(output.max())
    return output, {
        "base_connected_component_count": base_count,
        "final_instance_count": final_count,
        "split_component_count": len(split_details),
        "split_added_roi_count": final_count - base_count,
        "split_components": split_details,
        "split_rejected_count": rejected_count,
    }

def split_touching_whole_instances_multi(
    whole_mask: np.ndarray,
    nuclei_labels: np.ndarray,
    nearest_nucleus_labels: np.ndarray,
    nucleus_distance: np.ndarray,
    struct: np.ndarray,
    cellpose_mask: np.ndarray,
    mean_pixel_um: float,
    pixel_area_um2: float,
    link_radius_px: int,
    config: CompartmentConfig,
) -> tuple[np.ndarray, dict]:
    """Neonatal multi-center partition; preserve every Whole pixel and create no seam."""

    base_labels = measure.label(whole_mask, connectivity=2).astype(np.uint16)
    base_count = int(base_labels.max())
    empty_metrics = {
        "base_connected_component_count": base_count,
        "final_instance_count": base_count,
        "split_component_count": 0,
        "split_added_roi_count": 0,
        "split_components": [],
        "split_rejected_count": 0,
        "component_decisions": [],
    }
    if not config.instance_split_enabled or base_count < 1:
        return base_labels, empty_metrics

    output = np.zeros_like(base_labels, dtype=np.uint16)
    next_id = 1
    split_details: list[dict] = []
    component_decisions: list[dict] = []
    rejected_count = 0
    padding = link_radius_px + 4
    max_markers = max(2, int(config.instance_split_max_markers))

    for prop in measure.regionprops(base_labels):
        min_row, min_col, max_row, max_col = prop.bbox
        row0 = max(0, min_row - padding)
        col0 = max(0, min_col - padding)
        row1 = min(base_labels.shape[0], max_row + padding)
        col1 = min(base_labels.shape[1], max_col + padding)
        crop = np.s_[row0:row1, col0:col1]
        component = base_labels[crop] == prop.label
        local_struct = struct[crop]
        local_cellpose = cellpose_mask[crop]
        local_nuclei_labels = nuclei_labels[crop]
        local_nearest_labels = nearest_nucleus_labels[crop]
        local_nucleus_distance = nucleus_distance[crop]
        distance = ndi.distance_transform_edt(component)
        scored_nuclei, _ = score_nuclei_for_component(
            component,
            local_nearest_labels,
            local_nucleus_distance,
            distance,
            local_struct,
            local_cellpose,
            link_radius_px,
            config.ambiguity_score_delta,
        )
        anchor_groups = select_soma_anchor_groups(scored_nuclei, mean_pixel_um, config)
        scored_by_id = {row["nucleus_id"]: row for row in scored_nuclei}
        strict_groups: list[dict] = []
        for group in anchor_groups:
            representative = scored_by_id.get(group["nucleus_ids"][0])
            if representative is None:
                continue
            has_channel_support = (
                representative["overlap_fraction"] >= config.multi_anchor_min_overlap_fraction
                or representative["model_support"] >= config.multi_anchor_min_model_support
            )
            if (
                group["score"] >= config.instance_split_min_anchor_score
                and representative["thickness_support"]
                >= config.multi_anchor_min_thickness_support
                and representative["structural_support"]
                >= config.multi_anchor_min_structural_support
                and has_channel_support
            ):
                strict_groups.append(group)

        selected_groups: list[dict] = []
        for group in strict_groups:
            separated = all(
                mean_pixel_um
                * math.hypot(
                    group["center_y"] - accepted["center_y"],
                    group["center_x"] - accepted["center_x"],
                )
                >= config.instance_split_min_anchor_separation_um
                for accepted in selected_groups
            )
            if separated:
                selected_groups.append(group)
            if len(selected_groups) >= max_markers:
                break

        split_reason = "not_enough_strict_anchors"
        partition = None
        minimum_anchor_separation_um = 0.0
        neck_core_ratio = math.inf
        boundary_structural_ratio = math.inf
        child_areas: list[int] = []
        minimum_child_area = 0
        if len(selected_groups) >= 2:
            pairwise_separations = [
                mean_pixel_um
                * math.hypot(
                    selected_groups[left]["center_y"] - selected_groups[right]["center_y"],
                    selected_groups[left]["center_x"] - selected_groups[right]["center_x"],
                )
                for left in range(len(selected_groups))
                for right in range(left + 1, len(selected_groups))
            ]
            minimum_anchor_separation_um = min(pairwise_separations)
            markers = np.zeros_like(component, dtype=np.int32)
            marker_masks: list[np.ndarray] = []
            core_peaks: list[float] = []
            core_structural_supports: list[float] = []
            distance_scale = max(float(np.percentile(distance[component], 99.0)), 1.0)
            for marker_id, group in enumerate(selected_groups, start=1):
                selected_nucleus = np.isin(local_nuclei_labels, group["nucleus_ids"])
                selected_distance = ndi.distance_transform_edt(~selected_nucleus)
                search_region = component & (selected_distance <= link_radius_px)
                if not search_region.any():
                    break
                seed_score = (
                    0.72 * np.clip(distance / distance_scale, 0, 1)
                    + 0.23 * local_struct
                    + 0.05 * local_cellpose.astype(np.float32)
                )
                seed_score = np.where(search_region, seed_score, -np.inf)
                seed_y, seed_x = np.unravel_index(int(np.argmax(seed_score)), seed_score.shape)
                marker_radius_px = max(1, int(round(0.25 / mean_pixel_um)))
                marker_mask = component & circular_mask(
                    component.shape,
                    seed_y,
                    seed_x,
                    marker_radius_px,
                )
                if not marker_mask.any() or np.any(markers[marker_mask] > 0):
                    break
                markers[marker_mask] = marker_id
                marker_masks.append(marker_mask)
                core_neighborhood = component & circular_mask(
                    component.shape,
                    seed_y,
                    seed_x,
                    max(link_radius_px, int(round(0.75 / mean_pixel_um))),
                )
                core_peaks.append(
                    max(float(np.percentile(distance[core_neighborhood], 90.0)), 1.0)
                )
                core_structural_supports.append(
                    max(float(np.percentile(local_struct[core_neighborhood], 75.0)), 1e-6)
                )

            marker_count = len(selected_groups)
            if len(marker_masks) == marker_count:
                distance_cost = 1.0 - np.clip(distance / distance_scale, 0, 1)
                structural_cost = 1.0 - filters.gaussian(
                    local_struct,
                    sigma=1.0,
                    preserve_range=True,
                )
                elevation = 0.72 * distance_cost + 0.28 * structural_cost
                candidate_partition = segmentation.watershed(
                    elevation,
                    markers=markers,
                    mask=component,
                    watershed_line=False,
                    connectivity=np.ones((3, 3), dtype=bool),
                ).astype(np.uint16)
                marker_ids = list(range(1, marker_count + 1))
                child_areas = [
                    int((candidate_partition == marker_id).sum())
                    for marker_id in marker_ids
                ]
                minimum_child_area = max(
                    int(round(config.instance_split_min_child_area_um2 / pixel_area_um2)),
                    int(round(config.instance_split_min_child_fraction * int(component.sum()))),
                )
                boundary = np.zeros_like(component, dtype=bool)
                vertical = (
                    (candidate_partition[1:] > 0)
                    & (candidate_partition[:-1] > 0)
                    & (candidate_partition[1:] != candidate_partition[:-1])
                )
                horizontal = (
                    (candidate_partition[:, 1:] > 0)
                    & (candidate_partition[:, :-1] > 0)
                    & (candidate_partition[:, 1:] != candidate_partition[:, :-1])
                )
                boundary[1:] |= vertical
                boundary[:-1] |= vertical
                boundary[:, 1:] |= horizontal
                boundary[:, :-1] |= horizontal
                if boundary.any():
                    neck_core_ratio = float(np.percentile(distance[boundary], 75.0)) / max(
                        min(core_peaks),
                        1.0,
                    )
                    boundary_structural_ratio = float(np.median(local_struct[boundary])) / max(
                        min(core_structural_supports),
                        1e-6,
                    )
                children_connected = all(
                    int(measure.label(candidate_partition == marker_id, connectivity=2).max()) == 1
                    for marker_id in marker_ids
                )
                markers_retained = all(
                    bool(np.all(candidate_partition[marker_masks[index]] == index + 1))
                    for index in range(marker_count)
                )
                if min(child_areas) < minimum_child_area:
                    split_reason = "child_too_small"
                elif not children_connected or not markers_retained:
                    split_reason = "invalid_partition"
                elif not (
                    neck_core_ratio <= config.instance_split_strict_neck_core_ratio
                    or (
                        neck_core_ratio <= config.instance_split_max_neck_core_ratio
                        and boundary_structural_ratio
                        <= config.instance_split_max_boundary_structural_ratio
                    )
                ):
                    split_reason = "neck_or_boundary_too_supported"
                else:
                    split_reason = "accepted"
                    partition = candidate_partition
            else:
                split_reason = "marker_build_failed"

        output_view = output[crop]
        if partition is None:
            assigned_id = next_id
            output_view[component] = assigned_id
            next_id += 1
            if len(strict_groups) >= 2:
                rejected_count += 1
            component_decisions.append(
                {
                    "base_component_id": int(prop.label),
                    "output_astrocyte_ids": [assigned_id],
                    "strict_anchor_count": len(strict_groups),
                    "selected_anchor_count": len(selected_groups),
                    "split_required": len(strict_groups) >= 2,
                    "split_accepted": False,
                    "reason": split_reason,
                }
            )
            continue

        assigned_ids: list[int] = []
        for marker_id in range(1, int(partition.max()) + 1):
            output_view[partition == marker_id] = next_id
            assigned_ids.append(next_id)
            next_id += 1
        split_details.append(
            {
                "base_component_id": int(prop.label),
                "new_astrocyte_ids": assigned_ids,
                "anchor_scores": [
                    round(float(group["score"]), 6) for group in selected_groups
                ],
                "anchor_separation_um": round(minimum_anchor_separation_um, 6),
                "child_areas_px": child_areas,
                "neck_core_ratio": round(neck_core_ratio, 6),
                "boundary_structural_ratio": round(boundary_structural_ratio, 6),
                "reason": split_reason,
            }
        )
        component_decisions.append(
            {
                "base_component_id": int(prop.label),
                "output_astrocyte_ids": assigned_ids,
                "strict_anchor_count": len(strict_groups),
                "selected_anchor_count": len(selected_groups),
                "split_required": True,
                "split_accepted": True,
                "reason": split_reason,
                "child_areas_px": child_areas,
                "minimum_child_area_px": int(minimum_child_area),
            }
        )

    if not np.array_equal(output > 0, whole_mask.astype(bool)):
        raise RuntimeError("Neonatal instance splitting changed the Whole Astrocyte pixel union")
    final_count = int(output.max())
    return output, {
        "base_connected_component_count": base_count,
        "final_instance_count": final_count,
        "split_component_count": len(split_details),
        "split_added_roi_count": final_count - base_count,
        "split_components": split_details,
        "split_rejected_count": rejected_count,
        "component_decisions": component_decisions,
    }


def group_inventory_nucleus_objects(
    inventory: ValidatedNucleusAnchors,
    pixel_width_um: float,
    pixel_height_um: float,
    pixel_depth_um: float,
    config: NucleusOwnershipConfig,
    prefer_canonical: bool = False,
) -> list[dict]:
    """Group only geometrically overlapping threshold fragments of one 3D nucleus."""

    if prefer_canonical and inventory.nucleus_instance_records:
        return [
            {
                "group_id": int(row["instance_id"]),
                "object_ids": (int(row["instance_id"]),),
                "source_object_ids": tuple(row["source_object_ids"]),
                "accepted": bool(row["accepted"]),
                "independently_accepted": bool(row["accepted"]),
                "accepted_volume_gate_passed": bool(
                    float(row["volume_um3"]) >= config.accepted_min_volume_um3
                ),
                "identity_status": str(row["identity_status"]),
                "volume_um3": float(row["volume_um3"]),
                "enclosure_score": float(row["enclosure_score"]),
                "center_z": float(row["center_z"]),
                "center_y": float(row["center_y"]),
                "center_x": float(row["center_x"]),
                "z_min_0based": int(row["z_min_0based"]),
                "z_max_0based_inclusive": int(row["z_max_0based_inclusive"]),
                "extent_component_2d_ids": (
                    int(row["extent_component_2d_id"]),
                ),
            }
            for row in inventory.nucleus_instance_records
            if bool(row["dapi_valid"])
        ]

    records = [dict(row) for row in inventory.object_records if bool(row["dapi_valid"])]
    if not records:
        return []
    parent = {int(row["object_id"]): int(row["object_id"]) for row in records}

    def find(object_id: int) -> int:
        while parent[object_id] != object_id:
            parent[object_id] = parent[parent[object_id]]
            object_id = parent[object_id]
        return object_id

    def union(left_id: int, right_id: int) -> None:
        left_root = find(left_id)
        right_root = find(right_id)
        if left_root != right_root:
            parent[max(left_root, right_root)] = min(left_root, right_root)

    for left_index, left in enumerate(records):
        left_extent_component = int(left["extent_component_2d_id"])
        if left_extent_component <= 0:
            continue
        for right in records[left_index + 1 :]:
            if int(right["extent_component_2d_id"]) != left_extent_component:
                continue
            z_overlap = min(
                int(left["z_max_0based_inclusive"]),
                int(right["z_max_0based_inclusive"]),
            ) - max(
                int(left["z_min_0based"]),
                int(right["z_min_0based"]),
            ) + 1
            if z_overlap <= 0:
                continue
            delta_z_um = (
                float(left["center_z"]) - float(right["center_z"])
            ) * pixel_depth_um
            delta_y_um = (
                float(left["center_y"]) - float(right["center_y"])
            ) * pixel_height_um
            delta_x_um = (
                float(left["center_x"]) - float(right["center_x"])
            ) * pixel_width_um
            center_distance_um = math.sqrt(
                delta_z_um**2 + delta_y_um**2 + delta_x_um**2
            )
            left_radius_um = (
                3.0 * float(left["volume_um3"]) / (4.0 * math.pi)
            ) ** (1.0 / 3.0)
            right_radius_um = (
                3.0 * float(right["volume_um3"]) / (4.0 * math.pi)
            ) ** (1.0 / 3.0)
            if center_distance_um <= config.fragment_radius_sum_factor * (
                left_radius_um + right_radius_um
            ):
                union(int(left["object_id"]), int(right["object_id"]))

    grouped: dict[int, list[dict]] = {}
    for record in records:
        grouped.setdefault(find(int(record["object_id"])), []).append(record)
    output: list[dict] = []
    for group_id, members in sorted(grouped.items()):
        total_volume = float(sum(float(row["volume_um3"]) for row in members))
        independently_accepted = any(bool(row["accepted"]) for row in members)
        ownership_accepted = bool(
            independently_accepted
            and total_volume >= config.accepted_min_volume_um3
        )
        weights = np.asarray(
            [max(float(row["volume_um3"]), 1e-9) for row in members],
            dtype=np.float64,
        )
        output.append(
            {
                "group_id": int(group_id),
                "object_ids": tuple(sorted(int(row["object_id"]) for row in members)),
                "accepted": ownership_accepted,
                "independently_accepted": independently_accepted,
                "accepted_volume_gate_passed": bool(
                    total_volume >= config.accepted_min_volume_um3
                ),
                "volume_um3": total_volume,
                "enclosure_score": max(float(row["enclosure_score"]) for row in members),
                "center_z": float(
                    np.average([float(row["center_z"]) for row in members], weights=weights)
                ),
                "center_y": float(
                    np.average([float(row["center_y"]) for row in members], weights=weights)
                ),
                "center_x": float(
                    np.average([float(row["center_x"]) for row in members], weights=weights)
                ),
                "z_min_0based": min(int(row["z_min_0based"]) for row in members),
                "z_max_0based_inclusive": max(
                    int(row["z_max_0based_inclusive"]) for row in members
                ),
                "extent_component_2d_ids": tuple(
                    sorted(
                        {
                            int(row["extent_component_2d_id"])
                            for row in members
                            if int(row["extent_component_2d_id"]) > 0
                        }
                    )
                ),
            }
        )
    return output

def nucleus_group_marker(
    component: np.ndarray,
    extent_mask: np.ndarray,
    local_struct: np.ndarray,
    distance_inside_um: np.ndarray,
    pixel_width_um: float,
    pixel_height_um: float,
    config: NucleusOwnershipConfig,
) -> np.ndarray:
    distance_to_extent_um = ndi.distance_transform_edt(
        ~extent_mask,
        sampling=(pixel_height_um, pixel_width_um),
    )
    search = component & (distance_to_extent_um <= config.marker_search_um)
    if not search.any():
        return np.zeros_like(component, dtype=bool)
    distance_scale = max(float(np.percentile(distance_inside_um[component], 99.0)), 1e-6)
    proximity = np.exp(
        -np.square(distance_to_extent_um / max(config.marker_search_um, 1e-6))
    )
    score = (
        0.58 * np.clip(distance_inside_um / distance_scale, 0.0, 1.0)
        + 0.27 * local_struct
        + 0.15 * proximity
    )
    score = np.where(search, score, -np.inf)
    seed_y, seed_x = np.unravel_index(int(np.argmax(score)), score.shape)
    radius_y = max(1, int(round(config.marker_radius_um / pixel_height_um)))
    radius_x = max(1, int(round(config.marker_radius_um / pixel_width_um)))
    yy, xx = np.ogrid[: component.shape[0], : component.shape[1]]
    marker = component & (
        ((yy - seed_y) / radius_y) ** 2 + ((xx - seed_x) / radius_x) ** 2 <= 1.0
    )
    if not marker.any():
        marker[seed_y, seed_x] = True
    return marker

def apply_nucleus_ownership_guard(
    instance_labels: np.ndarray,
    struct: np.ndarray,
    inventory: ValidatedNucleusAnchors | None,
    pixel_width_um: float,
    pixel_height_um: float,
    pixel_depth_um: float | None,
    profile: str,
    config: NucleusOwnershipConfig | None = None,
    prefer_canonical: bool = False,
) -> tuple[np.ndarray, dict]:
    """Split validated owners and remove territory assigned to a foreign 3D soma."""

    cfg = config or NucleusOwnershipConfig()
    empty_metrics = {
        "enabled": inventory is not None,
        "profile": profile,
        "method": (
            "object-preserving 3D owner partition with Z-supported foreign-soma barrier"
        ),
        "evaluated_component_count": int(instance_labels.max()),
        "conflict_component_count": 0,
        "split_component_count": 0,
        "foreign_soma_pruned_component_count": 0,
        "fail_closed_component_count": 0,
        "removed_area_px": 0,
        "input_to_output_ids": {},
        "decisions": [],
        "config": asdict(cfg),
    }
    if (
        inventory is None
        or pixel_depth_um is None
        or pixel_depth_um <= 0
        or inventory.object_extent_labels_2d is None
        or not inventory.object_records
    ):
        return instance_labels, empty_metrics

    extent_labels = np.asarray(inventory.object_extent_labels_2d, dtype=np.uint32)
    if extent_labels.shape != instance_labels.shape:
        raise ValueError("3D nucleus inventory and Whole instance labels do not match")
    groups = group_inventory_nucleus_objects(
        inventory,
        pixel_width_um,
        pixel_height_um,
        pixel_depth_um,
        cfg,
        prefer_canonical=prefer_canonical,
    )
    if not groups:
        return instance_labels, empty_metrics

    object_to_group = np.zeros(int(extent_labels.max()) + 1, dtype=np.uint32)
    for group in groups:
        for object_id in group["object_ids"]:
            object_to_group[int(object_id)] = int(group["group_id"])
    grouped_extent_labels = object_to_group[extent_labels]
    group_extent_areas = np.bincount(
        grouped_extent_labels.ravel(),
        minlength=len(object_to_group),
    )
    instance_stride = int(instance_labels.max()) + 1
    group_instance_counts = np.bincount(
        (
            grouped_extent_labels.astype(np.int64) * instance_stride
            + instance_labels.astype(np.int64)
        ).ravel(),
        minlength=len(object_to_group) * instance_stride,
    ).reshape(len(object_to_group), instance_stride)

    pixel_area_um2 = pixel_width_um * pixel_height_um
    owner_min_overlap_px = max(1, int(math.ceil(cfg.owner_min_overlap_um2 / pixel_area_um2)))
    foreign_min_overlap_px = max(
        1,
        int(math.ceil(cfg.foreign_min_overlap_um2 / pixel_area_um2)),
    )
    minimum_child_area_px = max(
        1,
        int(math.ceil(cfg.minimum_child_area_um2 / pixel_area_um2)),
    )
    output = np.zeros_like(instance_labels, dtype=np.uint16)
    next_id = 1
    decisions: list[dict] = []
    split_count = 0
    pruned_count = 0
    fail_closed_count = 0
    removed_area_px = 0
    input_to_output_ids: dict[int, list[int]] = {}

    for prop in measure.regionprops(instance_labels):
        min_row, min_col, max_row, max_col = prop.bbox
        crop = np.s_[min_row:max_row, min_col:max_col]
        component = instance_labels[crop] == int(prop.label)
        local_struct = struct[crop]
        local_grouped_extent_labels = grouped_extent_labels[crop]
        component_area = int(component.sum())
        associations: list[dict] = []
        for group in groups:
            group_id = int(group["group_id"])
            extent_area = int(group_extent_areas[group_id])
            overlap = int(group_instance_counts[group_id, int(prop.label)])
            if overlap == 0 or extent_area == 0:
                continue
            extent = local_grouped_extent_labels == group_id
            total_labelled_overlap = int(group_instance_counts[group_id, 1:].sum())
            dominance = overlap / max(total_labelled_overlap, 1)
            associations.append(
                {
                    **group,
                    "extent": extent,
                    "extent_area_px": extent_area,
                    "overlap_px": overlap,
                    "overlap_fraction": overlap / extent_area,
                    "component_dominance": dominance,
                }
            )

        owner_candidates = [
            row
            for row in associations
            if bool(row["accepted"]) and int(row["overlap_px"]) >= owner_min_overlap_px
        ]
        ambiguous_nuclear_envelopes = [
            row
            for row in associations
            if str(row.get("identity_status", "resolved")) == "ambiguous"
            and int(row["overlap_px"]) >= owner_min_overlap_px
            and float(row["overlap_fraction"]) >= 0.30
        ]
        if ambiguous_nuclear_envelopes:
            input_to_output_ids[int(prop.label)] = []
            removed_area_px += component_area
            fail_closed_count += 1
            empty_metrics["conflict_component_count"] += 1
            decisions.append(
                {
                    "input_instance_id": int(prop.label),
                    "owner_group_id": 0,
                    "foreign_groups": [],
                    "ambiguous_group_ids": [
                        int(row["group_id"]) for row in ambiguous_nuclear_envelopes
                    ],
                    "status": "fail_closed_ambiguous_canonical_nuclear_envelope",
                    "output_instance_ids": [],
                    "removed_area_px": component_area,
                }
            )
            continue
        if not owner_candidates:
            output[crop][component] = next_id
            input_to_output_ids[int(prop.label)] = [next_id]
            next_id += 1
            continue
        owner = max(
            owner_candidates,
            key=lambda row: (
                int(row["overlap_px"]),
                float(row["volume_um3"]),
                float(row["enclosure_score"]),
            ),
        )
        foreign: list[dict] = []
        rejected_candidates: list[dict] = []
        for row in associations:
            if int(row["group_id"]) == int(owner["group_id"]):
                continue
            if bool(row["accepted"]):
                accepted_checks = {
                    "independently_accepted_3d_nucleus": True,
                    "volume": (
                        float(row["volume_um3"]) >= cfg.accepted_min_volume_um3
                    ),
                    "owner_overlap": (
                        int(row["overlap_px"]) >= owner_min_overlap_px
                    ),
                    "extent_overlap_fraction": (
                        float(row["overlap_fraction"])
                        >= cfg.accepted_min_extent_overlap_fraction
                    ),
                }
                if all(accepted_checks.values()):
                    foreign.append(row)
                else:
                    rejected_candidates.append(
                        {
                            "group_id": int(row["group_id"]),
                            "object_ids": list(row["object_ids"]),
                            "checks": accepted_checks,
                        }
                    )
                continue
            minimum_volume = (
                cfg.unowned_min_volume_um3
            )
            checks = {
                "volume": float(row["volume_um3"]) >= minimum_volume,
                "absolute_overlap": int(row["overlap_px"]) >= foreign_min_overlap_px,
                "overlap_fraction": (
                    float(row["overlap_fraction"]) >= cfg.foreign_min_overlap_fraction
                ),
                "owner_overlap_ratio": (
                    int(row["overlap_px"])
                    >= cfg.foreign_min_owner_overlap_ratio * int(owner["overlap_px"])
                ),
                "component_dominance": (
                    float(row["component_dominance"])
                    >= cfg.foreign_min_component_dominance
                ),
                "z_supported_enclosure": (
                    float(row["enclosure_score"])
                    >= cfg.unowned_min_enclosure_score
                ),
            }
            if all(checks.values()):
                foreign.append(row)
            else:
                rejected_candidates.append(
                    {
                        "group_id": int(row["group_id"]),
                        "object_ids": list(row["object_ids"]),
                        "checks": checks,
                    }
                )
        if not foreign:
            output[crop][component] = next_id
            input_to_output_ids[int(prop.label)] = [next_id]
            next_id += 1
            continue

        distance_inside_um = ndi.distance_transform_edt(
            component,
            sampling=(pixel_height_um, pixel_width_um),
        )
        owner_marker = nucleus_group_marker(
            component,
            owner["extent"],
            local_struct,
            distance_inside_um,
            pixel_width_um,
            pixel_height_um,
            cfg,
        )
        accepted_foreign = [row for row in foreign if bool(row["accepted"])]
        unowned_foreign = [row for row in foreign if not bool(row["accepted"])]
        decision = {
            "input_instance_id": int(prop.label),
            "owner_group_id": int(owner["group_id"]),
            "owner_object_ids": list(owner["object_ids"]),
            "foreign_groups": [
                {
                    "group_id": int(row["group_id"]),
                    "object_ids": list(row["object_ids"]),
                    "accepted": bool(row["accepted"]),
                    "volume_um3": float(row["volume_um3"]),
                    "overlap_px": int(row["overlap_px"]),
                    "overlap_fraction": float(row["overlap_fraction"]),
                    "component_dominance": float(row["component_dominance"]),
                    "enclosure_score": float(row["enclosure_score"]),
                }
                for row in foreign
            ],
            "rejected_candidates": rejected_candidates,
            "status": "pending",
            "output_instance_ids": [],
        }
        empty_metrics["conflict_component_count"] += 1

        if not owner_marker.any():
            input_to_output_ids[int(prop.label)] = []
            removed_area_px += component_area
            fail_closed_count += 1
            decision["status"] = "fail_closed_owner_marker_unavailable"
            decisions.append(decision)
            continue

        if not accepted_foreign:
            foreign_extent = np.logical_or.reduce(
                [row["extent"] for row in unowned_foreign]
            )
            distance_to_foreign_um = ndi.distance_transform_edt(
                ~foreign_extent,
                sampling=(pixel_height_um, pixel_width_um),
            )
            structural_cut = float(np.percentile(local_struct[component], 60.0))
            foreign_barrier = (
                component
                & (distance_to_foreign_um <= cfg.unowned_barrier_radius_um)
                & (
                    (distance_inside_um >= cfg.unowned_barrier_inner_width_um)
                    | (local_struct >= structural_cut)
                )
            )
            allowed = component & ~foreign_barrier
            owner_seed = owner_marker & allowed
            if owner_seed.any():
                owner_child = ndi.binary_propagation(
                    owner_seed,
                    structure=np.ones((3, 3), dtype=bool),
                    mask=allowed,
                ).astype(bool)
            else:
                owner_child = np.zeros_like(component, dtype=bool)
            minimum_owner_area = max(
                minimum_child_area_px,
                int(math.ceil(cfg.minimum_owner_child_fraction * component_area)),
            )
            if int(owner_child.sum()) < minimum_owner_area:
                input_to_output_ids[int(prop.label)] = []
                removed_area_px += component_area
                fail_closed_count += 1
                decision["status"] = "fail_closed_foreign_barrier_invalid_owner"
                decisions.append(decision)
                continue
            output_view = output[crop]
            output_view[owner_child] = next_id
            decision["output_instance_ids"] = [next_id]
            input_to_output_ids[int(prop.label)] = [next_id]
            next_id += 1
            removed = component_area - int(owner_child.sum())
            removed_area_px += removed
            pruned_count += 1
            decision["status"] = "foreign_soma_pruned"
            decision["removed_area_px"] = removed
            decisions.append(decision)
            continue

        partition_groups = [owner, *accepted_foreign]
        markers = np.zeros_like(component, dtype=np.int32)
        marker_masks: list[np.ndarray] = []
        marker_failed = False
        for marker_id, group in enumerate(partition_groups, start=1):
            marker = nucleus_group_marker(
                component,
                group["extent"],
                local_struct,
                distance_inside_um,
                pixel_width_um,
                pixel_height_um,
                cfg,
            )
            if not marker.any() or np.any(markers[marker] > 0):
                marker_failed = True
                break
            markers[marker] = marker_id
            marker_masks.append(marker)
        if marker_failed:
            input_to_output_ids[int(prop.label)] = []
            removed_area_px += component_area
            fail_closed_count += 1
            decision["status"] = "fail_closed_marker_build_failed"
            decisions.append(decision)
            continue

        partition_mask = component.copy()
        if unowned_foreign:
            unowned_extent = np.logical_or.reduce(
                [row["extent"] for row in unowned_foreign]
            )
            unowned_distance_um = ndi.distance_transform_edt(
                ~unowned_extent,
                sampling=(pixel_height_um, pixel_width_um),
            )
            partition_mask &= (
                unowned_distance_um > cfg.multi_owner_unowned_exclusion_um
            )
            if any(not np.all(partition_mask[marker]) for marker in marker_masks):
                input_to_output_ids[int(prop.label)] = []
                removed_area_px += component_area
                fail_closed_count += 1
                decision["status"] = "fail_closed_unowned_barrier_hit_owner"
                decisions.append(decision)
                continue

        distance_scale = max(float(np.percentile(distance_inside_um[component], 99.0)), 1e-6)
        distance_cost = 1.0 - np.clip(distance_inside_um / distance_scale, 0.0, 1.0)
        structural_cost = 1.0 - filters.gaussian(
            local_struct,
            sigma=1.0,
            preserve_range=True,
        )
        elevation = 0.72 * distance_cost + 0.28 * structural_cost
        partition = segmentation.watershed(
            elevation,
            markers=markers,
            mask=partition_mask,
            watershed_line=False,
            connectivity=np.ones((3, 3), dtype=bool),
        ).astype(np.uint16)
        child_areas = [
            int((partition == marker_id).sum())
            for marker_id in range(1, len(partition_groups) + 1)
        ]
        minimum_areas = [
            max(
                minimum_child_area_px,
                int(
                    math.ceil(
                        (
                            cfg.minimum_owner_child_fraction
                            if marker_id == 1
                            else cfg.minimum_accepted_child_fraction
                        )
                        * component_area
                    )
                ),
            )
            for marker_id in range(1, len(partition_groups) + 1)
        ]
        connected = all(
            int(measure.label(partition == marker_id, connectivity=2).max()) == 1
            for marker_id in range(1, len(partition_groups) + 1)
        )
        retained = all(
            bool(np.all(partition[marker_masks[index]] == index + 1))
            for index in range(len(marker_masks))
        )
        boundary = np.zeros_like(component, dtype=bool)
        vertical = (
            (partition[1:] > 0)
            & (partition[:-1] > 0)
            & (partition[1:] != partition[:-1])
        )
        horizontal = (
            (partition[:, 1:] > 0)
            & (partition[:, :-1] > 0)
            & (partition[:, 1:] != partition[:, :-1])
        )
        boundary[1:] |= vertical
        boundary[:-1] |= vertical
        boundary[:, 1:] |= horizontal
        boundary[:, :-1] |= horizontal
        core_peaks = [
            max(float(np.percentile(distance_inside_um[marker], 90.0)), 1e-6)
            for marker in marker_masks
        ]
        core_structural = [
            max(float(np.percentile(local_struct[marker], 75.0)), 1e-6)
            for marker in marker_masks
        ]
        boundary_core_ratio = (
            float(np.percentile(distance_inside_um[boundary], 75.0))
            / max(min(core_peaks), 1e-6)
            if boundary.any()
            else math.inf
        )
        boundary_structural_ratio = (
            float(np.median(local_struct[boundary])) / max(min(core_structural), 1e-6)
            if boundary.any()
            else math.inf
        )
        quality_passed = bool(
            all(area >= floor for area, floor in zip(child_areas, minimum_areas))
            and connected
            and retained
            and (
                boundary_core_ratio <= cfg.maximum_boundary_core_ratio
                or boundary_structural_ratio <= cfg.maximum_boundary_structural_ratio
            )
        )
        decision["child_areas_px"] = child_areas
        decision["minimum_child_areas_px"] = minimum_areas
        decision["boundary_core_ratio"] = float(boundary_core_ratio)
        decision["boundary_structural_ratio"] = float(boundary_structural_ratio)
        if not quality_passed:
            undersized_foreign = any(
                child_areas[index] < minimum_areas[index]
                for index in range(1, len(partition_groups))
            )
            if prefer_canonical and undersized_foreign:
                output[crop][component] = next_id
                input_to_output_ids[int(prop.label)] = [next_id]
                decision["status"] = (
                    "retained_peripheral_nucleus_without_minimum_cell_territory"
                )
                decision["output_instance_ids"] = [next_id]
                next_id += 1
                decisions.append(decision)
                continue
            input_to_output_ids[int(prop.label)] = []
            removed_area_px += component_area
            fail_closed_count += 1
            decision["status"] = "fail_closed_ambiguous_multi_owner_partition"
            decisions.append(decision)
            continue

        output_view = output[crop]
        assigned_ids: list[int] = []
        for marker_id in range(1, len(partition_groups) + 1):
            output_view[partition == marker_id] = next_id
            assigned_ids.append(next_id)
            next_id += 1
        dropped = int(component.sum()) - int((partition > 0).sum())
        removed_area_px += dropped
        split_count += 1
        decision["status"] = "accepted_multi_owner_split"
        decision["output_instance_ids"] = assigned_ids
        input_to_output_ids[int(prop.label)] = assigned_ids
        decision["removed_area_px"] = dropped
        decisions.append(decision)

    final_mask = output > 0
    if np.any(final_mask & ~(instance_labels > 0)):
        raise RuntimeError("Nucleus ownership guard expanded the frozen Whole geometry")
    empty_metrics.update(
        {
            "conflict_component_count": len(decisions),
            "split_component_count": split_count,
            "foreign_soma_pruned_component_count": pruned_count,
            "fail_closed_component_count": fail_closed_count,
            "removed_area_px": removed_area_px,
            "removed_area_fraction": removed_area_px
            / max(int((instance_labels > 0).sum()), 1),
            "final_instance_count": int(output.max()),
            "input_to_output_ids": {
                str(input_id): output_ids
                for input_id, output_ids in input_to_output_ids.items()
            },
            "decisions": decisions,
        }
    )
    return output, empty_metrics

def inventory_group_geometry(
    inventory: ValidatedNucleusAnchors | None,
    pixel_width_um: float,
    pixel_height_um: float,
    pixel_depth_um: float | None,
    config: NucleusOwnershipConfig | None = None,
) -> tuple[list[dict], np.ndarray | None]:
    """Return stable nucleus groups and their 2D extent IDs."""

    if inventory is None or pixel_depth_um is None or pixel_depth_um <= 0:
        return [], None
    cfg = config or NucleusOwnershipConfig()
    if (
        inventory.nucleus_instance_extent_labels_2d is not None
        and inventory.nucleus_instance_records
    ):
        labels = np.asarray(
            inventory.nucleus_instance_extent_labels_2d,
            dtype=np.uint32,
        )
        groups = []
        for row in inventory.nucleus_instance_records:
            if not bool(row["dapi_valid"]):
                continue
            instance_id = int(row["instance_id"])
            groups.append(
                {
                    "group_id": instance_id,
                    "object_ids": (instance_id,),
                    "source_object_ids": tuple(row["source_object_ids"]),
                    "accepted": bool(row["accepted"]),
                    "identity_status": str(row["identity_status"]),
                    "volume_um3": float(row["volume_um3"]),
                    "enclosure_score": float(row["enclosure_score"]),
                    "center_z": float(row["center_z"]),
                    "center_y": float(row["center_y"]),
                    "center_x": float(row["center_x"]),
                    "z_min_0based": int(row["z_min_0based"]),
                    "z_max_0based_inclusive": int(row["z_max_0based_inclusive"]),
                    "dapi_low_threshold": (
                        float(row["dapi_low_threshold"])
                        if row.get("dapi_low_threshold") is not None
                        else None
                    ),
                    "extent_component_2d_ids": (
                        int(row["extent_component_2d_id"]),
                    ),
                }
            )
        valid_ids = np.asarray(
            [int(group["group_id"]) for group in groups],
            dtype=np.uint32,
        )
        return groups, np.where(np.isin(labels, valid_ids), labels, 0).astype(np.uint32)
    if inventory.object_extent_labels_2d is None or not inventory.object_records:
        return [], None
    groups = group_inventory_nucleus_objects(
        inventory,
        pixel_width_um,
        pixel_height_um,
        pixel_depth_um,
        cfg,
    )
    object_labels = np.asarray(inventory.object_extent_labels_2d, dtype=np.uint32)
    object_to_group = np.zeros(int(object_labels.max()) + 1, dtype=np.uint32)
    for group in groups:
        for object_id in group["object_ids"]:
            object_to_group[int(object_id)] = int(group["group_id"])
    return groups, object_to_group[object_labels]

def owner_group_for_soma(
    soma_mask: np.ndarray,
    groups: list[dict],
    grouped_extent_labels: np.ndarray,
) -> dict | None:
    """Choose one nucleus envelope by Soma overlap without intensity-peak voting."""

    overlaps: list[tuple[int, bool, float, dict]] = []
    for group in groups:
        group_id = int(group["group_id"])
        overlap = int((soma_mask & (grouped_extent_labels == group_id)).sum())
        if overlap <= 0:
            continue
        overlaps.append(
            (
                overlap,
                bool(group["accepted"]),
                float(group["volume_um3"]),
                group,
            )
        )
    if not overlaps:
        return None
    accepted = [row for row in overlaps if row[1]]
    pool = accepted or overlaps
    return max(pool, key=lambda row: (row[0], row[2]))[3]


def apply_canonical_identity_reconciliation(
    whole_labels: np.ndarray,
    soma_labels: np.ndarray,
    process_labels: np.ndarray,
    struct: np.ndarray,
    inventory: ValidatedNucleusAnchors | None,
    pixel_width_um: float,
    pixel_height_um: float,
    pixel_depth_um: float | None,
    profile: str,
    config: CanonicalIdentityConfig | None = None,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, dict]:
    """Correct only explicit nucleus/ROI identity conflicts on validated compartment labels."""

    cfg = config or CanonicalIdentityConfig()
    metrics = {
        "enabled": bool(cfg.enabled and inventory is not None),
        "profile": profile,
        "method": (
            "local canonical identity reconciliation: same canonical nucleus merges "
            "adjacent IDs; multiple accepted nuclei trigger quality-gated partition"
        ),
        "pre_roi_count": int(whole_labels.max()),
        "post_roi_count": int(whole_labels.max()),
        "merge_count": 0,
        "split_count": 0,
        "fail_closed_count": 0,
        "changed_input_ids": [],
        "merge_decisions": [],
        "partition_decisions": [],
        "final_lineage": {},
        "config": asdict(cfg),
    }
    if (
        not metrics["enabled"]
        or pixel_depth_um is None
        or pixel_depth_um <= 0
        or inventory is None
        or inventory.nucleus_instance_extent_labels_2d is None
        or not inventory.nucleus_instance_records
    ):
        return whole_labels, soma_labels, process_labels, metrics

    canonical_records = [
        dict(row)
        for row in inventory.nucleus_instance_records
        if bool(row["accepted"])
        and bool(row["dapi_valid"])
        and str(row["identity_status"]) == "resolved"
    ]
    if not canonical_records:
        return whole_labels, soma_labels, process_labels, metrics
    satellite_aliases: dict[int, int] = {}
    for left_index, left in enumerate(canonical_records):
        for right in canonical_records[left_index + 1 :]:
            smaller, larger = sorted(
                (left, right),
                key=lambda row: float(row["volume_um3"]),
            )
            smaller_sources = set(int(value) for value in smaller["source_object_ids"])
            larger_sources = set(int(value) for value in larger["source_object_ids"])
            if not smaller_sources or not smaller_sources < larger_sources:
                continue
            volume_ratio = float(smaller["volume_um3"]) / max(
                float(larger["volume_um3"]),
                1e-9,
            )
            if volume_ratio > cfg.satellite_max_volume_ratio:
                continue
            z_overlap = max(
                0,
                min(
                    int(smaller["z_max_0based_inclusive"]),
                    int(larger["z_max_0based_inclusive"]),
                )
                - max(
                    int(smaller["z_min_0based"]),
                    int(larger["z_min_0based"]),
                )
                + 1,
            )
            smaller_z_span = (
                int(smaller["z_max_0based_inclusive"])
                - int(smaller["z_min_0based"])
                + 1
            )
            if z_overlap / max(smaller_z_span, 1) < cfg.satellite_min_z_overlap_fraction:
                continue
            delta_z = (
                float(smaller["center_z"]) - float(larger["center_z"])
            ) * pixel_depth_um
            delta_y = (
                float(smaller["center_y"]) - float(larger["center_y"])
            ) * pixel_height_um
            delta_x = (
                float(smaller["center_x"]) - float(larger["center_x"])
            ) * pixel_width_um
            center_distance_um = math.sqrt(delta_z**2 + delta_y**2 + delta_x**2)
            smaller_radius_um = (
                3.0 * float(smaller["volume_um3"]) / (4.0 * math.pi)
            ) ** (1.0 / 3.0)
            larger_radius_um = (
                3.0 * float(larger["volume_um3"]) / (4.0 * math.pi)
            ) ** (1.0 / 3.0)
            if center_distance_um > cfg.satellite_max_radius_sum_factor * (
                smaller_radius_um + larger_radius_um
            ):
                continue
            satellite_aliases[int(smaller["instance_id"])] = int(
                larger["instance_id"]
            )
    if satellite_aliases:
        def alias_root(instance_id: int) -> int:
            while instance_id in satellite_aliases:
                instance_id = satellite_aliases[instance_id]
            return instance_id

        satellite_aliases = {
            instance_id: alias_root(target_id)
            for instance_id, target_id in satellite_aliases.items()
        }
        canonical_records = [
            row
            for row in canonical_records
            if int(row["instance_id"]) not in satellite_aliases
        ]
    metrics["canonical_satellite_collapses"] = {
        str(instance_id): target_id
        for instance_id, target_id in sorted(satellite_aliases.items())
    }
    canonical_ids = np.asarray(
        [int(row["instance_id"]) for row in canonical_records],
        dtype=np.uint32,
    )
    canonical_extent_labels = np.asarray(
        inventory.nucleus_instance_extent_labels_2d,
        dtype=np.uint32,
    )
    canonical_core_labels = np.asarray(
        inventory.nucleus_instance_core_labels_2d,
        dtype=np.uint32,
    )
    if satellite_aliases:
        for source_id, target_id in satellite_aliases.items():
            canonical_extent_labels[canonical_extent_labels == source_id] = target_id
            canonical_core_labels[canonical_core_labels == source_id] = target_id
    canonical_extent_labels = np.where(
        np.isin(canonical_extent_labels, canonical_ids),
        canonical_extent_labels,
        0,
    ).astype(np.uint32)
    canonical_core_labels = np.where(
        np.isin(canonical_core_labels, canonical_ids),
        canonical_core_labels,
        0,
    ).astype(np.uint32)

    original_ids = list(range(1, int(whole_labels.max()) + 1))
    parent = {astrocyte_id: astrocyte_id for astrocyte_id in original_ids}

    def find(astrocyte_id: int) -> int:
        while parent[astrocyte_id] != astrocyte_id:
            parent[astrocyte_id] = parent[parent[astrocyte_id]]
            astrocyte_id = parent[astrocyte_id]
        return astrocyte_id

    def union(left_id: int, right_id: int) -> None:
        left_root = find(left_id)
        right_root = find(right_id)
        if left_root != right_root:
            parent[max(left_root, right_root)] = min(left_root, right_root)

    pixel_area_um2 = pixel_width_um * pixel_height_um
    minimum_overlap_px = max(
        1,
        int(math.ceil(cfg.minimum_extent_overlap_um2 / pixel_area_um2)),
    )
    for canonical_id in canonical_ids:
        extent = canonical_extent_labels == int(canonical_id)
        extent_area = int(extent.sum())
        if extent_area == 0:
            continue
        overlaps = np.bincount(
            whole_labels[extent],
            minlength=int(whole_labels.max()) + 1,
        )
        source_ids = [
            astrocyte_id
            for astrocyte_id in original_ids
            if int(overlaps[astrocyte_id]) >= minimum_overlap_px
            and float(overlaps[astrocyte_id]) / extent_area
            >= cfg.minimum_extent_overlap_fraction
        ]
        if len(source_ids) < 2 or len(source_ids) > cfg.maximum_merge_source_count:
            continue
        accepted_pairs = []
        for left_index, left_id in enumerate(source_ids):
            left = whole_labels == left_id
            contact_distance = ndi.distance_transform_edt(
                ~left,
                sampling=(pixel_height_um, pixel_width_um),
            )
            for right_id in source_ids[left_index + 1 :]:
                right = whole_labels == right_id
                if not np.any(right & (contact_distance <= cfg.merge_contact_distance_um)):
                    continue
                union(left_id, right_id)
                accepted_pairs.append([left_id, right_id])
        if accepted_pairs:
            metrics["merge_decisions"].append(
                {
                    "canonical_nucleus_id": int(canonical_id),
                    "source_astrocyte_ids": source_ids,
                    "accepted_pairs": accepted_pairs,
                    "extent_overlap_px": {
                        str(value): int(overlaps[value]) for value in source_ids
                    },
                }
            )

    root_to_sources: dict[int, list[int]] = {}
    for astrocyte_id in original_ids:
        root_to_sources.setdefault(find(astrocyte_id), []).append(astrocyte_id)
    merged_labels = np.zeros_like(whole_labels, dtype=np.uint16)
    merged_soma = np.zeros_like(soma_labels, dtype=np.uint16)
    merged_sources: dict[int, list[int]] = {}
    for merged_id, root in enumerate(sorted(root_to_sources), start=1):
        sources = sorted(root_to_sources[root])
        merged_sources[merged_id] = sources
        source_mask = np.isin(whole_labels, sources)
        merged_labels[source_mask] = merged_id
        merged_soma[np.isin(soma_labels, sources)] = merged_id
    metrics["merge_count"] = sum(len(values) > 1 for values in merged_sources.values())

    accepted_core = canonical_core_labels > 0
    accepted_extent = canonical_extent_labels > 0
    canonical_inventory = ValidatedNucleusAnchors(
        accepted_core_mask_2d=accepted_core,
        accepted_extent_mask_2d=accepted_extent,
        metrics={"status": "canonical_identity_reconciliation"},
        object_core_labels_2d=canonical_core_labels,
        object_extent_labels_2d=canonical_extent_labels,
        dapi_valid_object_ids=tuple(int(value) for value in canonical_ids),
        accepted_object_ids=tuple(int(value) for value in canonical_ids),
        object_records=tuple(canonical_records),
        nucleus_instance_core_labels_2d=canonical_core_labels,
        nucleus_instance_extent_labels_2d=canonical_extent_labels,
        accepted_instance_ids=tuple(int(value) for value in canonical_ids),
        nucleus_instance_records=tuple(canonical_records),
    )
    ownership_cfg = replace(
        NucleusOwnershipConfig(),
        accepted_min_extent_overlap_fraction=cfg.split_minimum_extent_overlap_fraction,
        unowned_min_volume_um3=math.inf,
        unowned_min_enclosure_score=math.inf,
    )
    reconciled_whole, partition_metrics = apply_nucleus_ownership_guard(
        merged_labels,
        struct,
        canonical_inventory,
        pixel_width_um,
        pixel_height_um,
        pixel_depth_um,
        profile=f"{profile}_canonical_reconciliation",
        config=ownership_cfg,
        prefer_canonical=True,
    )
    metrics["partition_decisions"] = partition_metrics.get("decisions", [])
    metrics["split_count"] = int(partition_metrics.get("split_component_count", 0))
    metrics["fail_closed_count"] = int(
        partition_metrics.get("fail_closed_component_count", 0)
    )
    input_to_output_ids = {
        int(input_id): [int(value) for value in output_ids]
        for input_id, output_ids in partition_metrics.get(
            "input_to_output_ids", {}
        ).items()
    }

    output_soma = np.zeros_like(reconciled_whole, dtype=np.uint16)
    output_process = np.zeros_like(reconciled_whole, dtype=np.uint16)
    final_lineage: dict[int, dict] = {}
    changed_merged_ids = {
        merged_id
        for merged_id, sources in merged_sources.items()
        if len(sources) > 1 or len(input_to_output_ids.get(merged_id, [])) != 1
    }
    for final_id in range(1, int(reconciled_whole.max()) + 1):
        component = reconciled_whole == final_id
        source_merged_ids = [
            merged_id
            for merged_id, output_ids in input_to_output_ids.items()
            if final_id in output_ids
        ]
        if not source_merged_ids:
            source_merged_ids = sorted(
                int(value)
                for value in np.unique(merged_labels[component])
                if int(value) > 0
            )
        source_ids = sorted(
            {
                source_id
                for merged_id in source_merged_ids
                for source_id in merged_sources.get(merged_id, [])
            }
        )
        soma = component & np.isin(soma_labels, source_ids)
        canonical_overlaps = np.bincount(
            canonical_extent_labels[component],
            minlength=int(canonical_extent_labels.max()) + 1,
        )
        owner_id = int(np.argmax(canonical_overlaps[1:]) + 1) if canonical_overlaps[1:].any() else 0
        was_split = any(
            len(input_to_output_ids.get(merged_id, [])) > 1
            for merged_id in source_merged_ids
        )
        if was_split and owner_id > 0:
            soma |= component & (canonical_extent_labels == owner_id)
        if not soma.any() and owner_id > 0:
            soma = component & (canonical_extent_labels == owner_id)
        if not soma.any():
            continue
        output_soma[soma] = final_id
        output_process[component & ~soma] = final_id
        final_lineage[final_id] = {
            "source_astrocyte_ids": source_ids,
            "source_merged_ids": source_merged_ids,
            "canonical_owner_id": owner_id,
            "identity_changed": bool(
                any(value in changed_merged_ids for value in source_merged_ids)
            ),
        }

    retained_ids = sorted(final_lineage)
    if not retained_ids:
        raise RuntimeError("Canonical identity reconciliation removed every Astrocyte ROI")
    output_whole, output_soma, output_process, final_mapping = (
        relabel_compartment_triplet(
            reconciled_whole,
            output_soma,
            output_process,
            retained_ids,
        )
    )
    metrics["final_lineage"] = {
        str(final_mapping[old_id]): row for old_id, row in final_lineage.items()
    }
    metrics["changed_input_ids"] = sorted(
        {
            source_id
            for row in metrics["final_lineage"].values()
            if bool(row["identity_changed"])
            for source_id in row["source_astrocyte_ids"]
        }
    )
    metrics["post_roi_count"] = int(output_whole.max())
    if np.any((output_whole > 0) & ~(whole_labels > 0)):
        raise RuntimeError("Canonical identity reconciliation expanded frozen Whole geometry")
    if np.any((output_soma > 0) & (output_process > 0)):
        raise RuntimeError("Canonical identity reconciliation overlapped compartments")
    if not np.array_equal(
        output_whole > 0,
        (output_soma > 0) | (output_process > 0),
    ):
        raise RuntimeError("Canonical identity reconciliation broke the partition")
    return output_whole, output_soma, output_process, metrics

def apply_axial_truncation_guard(
    whole_labels: np.ndarray,
    soma_labels: np.ndarray,
    process_labels: np.ndarray,
    inventory: ValidatedNucleusAnchors | None,
    context: Neonatal3DContext | None,
    pixel_width_um: float,
    pixel_height_um: float,
    config: AxialTruncationConfig | None = None,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, dict]:
    """Delete cells whose owner nucleus is demonstrably truncated by selected Z faces."""

    cfg = config or AxialTruncationConfig()
    metrics = {
        "enabled": bool(cfg.enabled and inventory is not None and context is not None),
        "method": (
            "selected-Z cuboid guard using owner-envelope boundary contact and "
            "connected DAPI continuation in raw guard slices"
        ),
        "evaluated_cell_count": int(whole_labels.max()),
        "removed_cell_count": 0,
        "removed_pre_guard_ids": [],
        "id_mapping": {},
        "decisions": [],
        "config": asdict(cfg),
    }
    if not metrics["enabled"]:
        return whole_labels, soma_labels, process_labels, metrics
    assert context is not None
    groups, grouped_extents = inventory_group_geometry(
        inventory,
        pixel_width_um,
        pixel_height_um,
        context.pixel_depth_um,
    )
    if not groups or grouped_extents is None:
        return whole_labels, soma_labels, process_labels, metrics

    z0 = int(context.z_start_0based)
    z1 = int(context.z_end_0based_inclusive)
    stack_depth = int(context.dapi_stack.shape[0])
    band_slices = max(1, int(math.ceil(cfg.boundary_band_um / context.pixel_depth_um)))
    guard_slices = max(1, int(math.ceil(cfg.guard_depth_um / context.pixel_depth_um)))
    pixel_area_um2 = pixel_width_um * pixel_height_um
    voxel_volume_um3 = pixel_area_um2 * context.pixel_depth_um
    per_object_rows = {
        int(row["object_id_3d"]): row
        for row in inventory.metrics.get("per_nucleus", [])
    }

    owners: dict[int, dict] = {}
    owner_volumes: list[float] = []
    for astrocyte_id in range(1, int(whole_labels.max()) + 1):
        owner = owner_group_for_soma(
            soma_labels == astrocyte_id,
            groups,
            grouped_extents,
        )
        if owner is not None:
            owners[astrocyte_id] = owner
            owner_volumes.append(float(owner["volume_um3"]))
    reference_volume = (
        float(np.median(owner_volumes))
        if len(owner_volumes) >= cfg.min_reference_nuclei
        else 0.0
    )

    removed_ids: set[int] = set()
    for astrocyte_id in range(1, int(whole_labels.max()) + 1):
        owner = owners.get(astrocyte_id)
        decision = {
            "pre_guard_astrocyte_id": astrocyte_id,
            "owner_group_id": int(owner["group_id"]) if owner else 0,
            "status": "retained",
            "faces": [],
        }
        if owner is None:
            decision["status"] = "not_evaluated_owner_unavailable"
            metrics["decisions"].append(decision)
            continue
        group_id = int(owner["group_id"])
        extent = grouped_extents == group_id
        if not extent.any():
            decision["status"] = "not_evaluated_extent_unavailable"
            metrics["decisions"].append(decision)
            continue
        rows, cols = np.nonzero(extent)
        halo_y = max(1, int(math.ceil(0.45 / pixel_height_um)))
        halo_x = max(1, int(math.ceil(0.45 / pixel_width_um)))
        row0 = max(0, int(rows.min()) - halo_y)
        row1 = min(extent.shape[0], int(rows.max()) + halo_y + 1)
        col0 = max(0, int(cols.min()) - halo_x)
        col1 = min(extent.shape[1], int(cols.max()) + halo_x + 1)
        local_extent = extent[row0:row1, col0:col1]
        support_distance = ndi.distance_transform_edt(
            ~local_extent,
            sampling=(pixel_height_um, pixel_width_um),
        )
        support = support_distance <= 0.45
        thresholds = [
            float(per_object_rows[object_id]["dapi_low_threshold"])
            for object_id in owner["object_ids"]
            if object_id in per_object_rows
            and per_object_rows[object_id].get("dapi_low_threshold") is not None
        ]
        if not thresholds and owner.get("dapi_low_threshold") is not None:
            thresholds = [float(owner["dapi_low_threshold"])]
        if not thresholds:
            decision["status"] = "not_evaluated_threshold_unavailable"
            metrics["decisions"].append(decision)
            continue
        threshold = float(np.median(thresholds))
        inside_volume = float(owner["volume_um3"])
        relative_volume = (
            inside_volume / reference_volume if reference_volume > 0 else math.nan
        )
        faces = []
        if z0 > 0 and int(owner["z_min_0based"]) <= z0 + band_slices - 1:
            faces.append(("front", max(0, z0 - guard_slices), z0 + band_slices, z0))
        if z1 < stack_depth - 1 and int(owner["z_max_0based_inclusive"]) >= z1 - band_slices + 1:
            faces.append(("back", z1 - band_slices + 1, min(stack_depth, z1 + guard_slices + 1), z1 + 1))

        reject = False
        for face_name, slab_start, slab_end, outside_boundary in faces:
            slab = context.dapi_stack[
                slab_start:slab_end,
                row0:row1,
                col0:col1,
            ].astype(np.float32, copy=False)
            binary = (slab >= threshold) & support[None, :, :]
            labelled = measure.label(binary, connectivity=3)
            split_index = outside_boundary - slab_start
            if face_name == "front":
                outside_selector = np.arange(binary.shape[0]) < split_index
                boundary_indices = range(split_index, min(binary.shape[0], split_index + band_slices))
            else:
                outside_selector = np.arange(binary.shape[0]) >= split_index
                boundary_indices = range(max(0, split_index - band_slices), split_index)
            boundary_ids: set[int] = set()
            for index in boundary_indices:
                boundary_ids.update(int(value) for value in np.unique(labelled[index]) if value > 0)
            outside_counts = {
                component_id: int(
                    ((labelled == component_id) & outside_selector[:, None, None]).sum()
                )
                for component_id in boundary_ids
            }
            connected_outside_voxels = max(outside_counts.values(), default=0)
            outside_volume = connected_outside_voxels * voxel_volume_um3
            outside_ratio = outside_volume / max(inside_volume, 1e-9)
            inside_fraction = inside_volume / max(inside_volume + outside_volume, 1e-9)
            boundary_area_px = 0
            for index in boundary_indices:
                boundary_area_px = max(
                    boundary_area_px,
                    int(np.isin(labelled[index], list(boundary_ids)).sum()),
                )
            boundary_area_ratio = (
                boundary_area_px * pixel_area_um2
                / max(int(extent.sum()) * pixel_area_um2, 1e-9)
            )
            face_decision = {
                "face": face_name,
                "outside_volume_um3": float(outside_volume),
                "outside_to_inside_ratio": float(outside_ratio),
                "inside_volume_fraction": float(inside_fraction),
                "boundary_area_ratio": float(boundary_area_ratio),
            }
            continuation = bool(
                outside_volume >= cfg.min_outside_continuation_um3
                and outside_ratio >= cfg.min_outside_to_inside_ratio
                and boundary_area_ratio >= cfg.min_boundary_area_ratio
            )
            small_or_incomplete = bool(
                inside_fraction < cfg.min_inside_volume_fraction
                or float(owner["z_max_0based_inclusive"] - owner["z_min_0based"] + 1)
                * context.pixel_depth_um
                < cfg.min_inside_z_span_um
                or (
                    np.isfinite(relative_volume)
                    and relative_volume < cfg.minimum_relative_volume
                )
            )
            face_decision["continuation_confirmed"] = continuation
            face_decision["small_or_incomplete"] = small_or_incomplete
            decision["faces"].append(face_decision)
            reject |= continuation and small_or_incomplete
        decision["inside_volume_um3"] = inside_volume
        decision["relative_owner_volume"] = float(relative_volume)
        if reject:
            removed_ids.add(astrocyte_id)
            decision["status"] = "removed_confirmed_axial_truncation"
        metrics["decisions"].append(decision)

    retained_ids = [
        astrocyte_id
        for astrocyte_id in range(1, int(whole_labels.max()) + 1)
        if astrocyte_id not in removed_ids
    ]
    if not retained_ids:
        raise RuntimeError("Axial truncation guard would remove every Astrocyte ROI")
    outputs = relabel_compartment_triplet(
        whole_labels,
        soma_labels,
        process_labels,
        retained_ids,
    )
    out_whole, out_soma, out_process, mapping = outputs
    metrics["removed_cell_count"] = len(removed_ids)
    metrics["removed_pre_guard_ids"] = sorted(removed_ids)
    metrics["id_mapping"] = {str(key): value for key, value in mapping.items()}
    return out_whole, out_soma, out_process, metrics

def apply_projected_foreign_soma_guard(
    whole_labels: np.ndarray,
    soma_labels: np.ndarray,
    process_labels: np.ndarray,
    inventory: ValidatedNucleusAnchors | None,
    pixel_width_um: float,
    pixel_height_um: float,
    pixel_depth_um: float | None,
    config: NucleusOwnershipConfig | None = None,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, dict]:
    """Remove only true distal tips ending inside another cell's DAPI projection."""

    cfg = config or NucleusOwnershipConfig()
    metrics = {
        "enabled": bool(cfg.projection_occlusion_enabled and inventory is not None),
        "method": (
            "topology-gated distal-tip exclusion at the exact foreign DAPI projection "
            "boundary; no fixed terminal length, halo, or ROI expansion"
        ),
        "evaluated_cell_count": int(whole_labels.max()),
        "changed_cell_count": 0,
        "removed_area_px": 0,
        "true_tip_count": 0,
        "terminal_overlap_component_count": 0,
        "preserved_pass_through_component_count": 0,
        "connectivity_rollback_count": 0,
        "decisions": [],
        "config": asdict(cfg),
    }
    groups, grouped_extents = inventory_group_geometry(
        inventory,
        pixel_width_um,
        pixel_height_um,
        pixel_depth_um,
        cfg,
    )
    if not metrics["enabled"] or not groups or grouped_extents is None:
        return whole_labels, soma_labels, process_labels, metrics
    pixel_area_um2 = pixel_width_um * pixel_height_um
    output_whole = np.zeros_like(whole_labels, dtype=np.uint16)
    output_soma = soma_labels.copy().astype(np.uint16)
    output_process = np.zeros_like(process_labels, dtype=np.uint16)

    eligible_groups = [
        group
        for group in groups
        if float(group["volume_um3"]) >= cfg.projection_min_foreign_volume_um3
        and (
            int(group["z_max_0based_inclusive"]) - int(group["z_min_0based"]) + 1
        ) * float(pixel_depth_um)
        >= cfg.projection_min_foreign_z_span_um
        and int((grouped_extents == int(group["group_id"])).sum()) * pixel_area_um2
        >= cfg.projection_min_foreign_extent_um2
    ]
    eligible_extents = {
        int(group["group_id"]): grouped_extents == int(group["group_id"])
        for group in eligible_groups
    }
    neighbor_kernel = np.ones((3, 3), dtype=np.uint8)
    neighbor_kernel[1, 1] = 0
    connectivity = np.ones((3, 3), dtype=np.uint8)
    image_edge = np.zeros_like(whole_labels, dtype=bool)
    image_edge[[0, -1], :] = True
    image_edge[:, [0, -1]] = True

    for astrocyte_id in range(1, int(whole_labels.max()) + 1):
        component = whole_labels == astrocyte_id
        soma = soma_labels == astrocyte_id
        process = process_labels == astrocyte_id
        owner = owner_group_for_soma(soma, groups, grouped_extents)
        owner_group_id = int(owner["group_id"]) if owner is not None else 0
        foreign_ids: list[int] = []
        removed_foreign_ids: list[int] = []
        remove_mask = np.zeros_like(component, dtype=bool)

        component_labels = measure.label(component, connectivity=2)
        soma_root_ids = np.unique(component_labels[soma])
        soma_root_ids = soma_root_ids[soma_root_ids > 0]
        soma_rooted_whole = np.isin(component_labels, soma_root_ids)
        skeleton = morphology.skeletonize(soma_rooted_whole)
        neighbor_count = ndi.convolve(
            skeleton.astype(np.uint8),
            neighbor_kernel,
            mode="constant",
            cval=0,
        )
        soma_adjacent = morphology.binary_dilation(
            soma,
            footprint=np.ones((3, 3), dtype=bool),
        )
        true_tips = (
            skeleton
            & process
            & (neighbor_count == 1)
            & ~soma_adjacent
            & ~image_edge
        )
        true_tip_count = int(true_tips.sum())
        metrics["true_tip_count"] += true_tip_count
        terminal_components = 0
        preserved_pass_through = 0

        for group_id, extent in eligible_extents.items():
            if group_id == owner_group_id:
                continue
            overlap = process & extent
            if not np.any(overlap):
                continue
            foreign_ids.append(group_id)
            overlap_labels, overlap_count = ndi.label(
                overlap,
                structure=connectivity,
            )
            tip_labels = np.unique(overlap_labels[true_tips & extent])
            tip_labels = tip_labels[tip_labels > 0]
            for overlap_id in tip_labels:
                overlap_component = overlap_labels == int(overlap_id)
                skeleton_inside = skeleton & overlap_component
                if not np.any(skeleton_inside):
                    continue
                skeleton_entry_pixels = (
                    ndi.binary_dilation(skeleton_inside, structure=connectivity)
                    & skeleton
                    & process
                    & ~extent
                )
                _, entry_count = ndi.label(
                    skeleton_entry_pixels,
                    structure=connectivity,
                )
                contains_branchpoint = bool(
                    np.any(skeleton_inside & (neighbor_count >= 3))
                )
                if entry_count != 1 or contains_branchpoint:
                    preserved_pass_through += 1
                    continue
                remove_mask |= overlap_component
                terminal_components += 1
                removed_foreign_ids.append(group_id)

        retained = soma | (process & ~remove_mask)
        before_component_count = int(measure.label(component, connectivity=2).max())
        after_component_count = int(measure.label(retained, connectivity=2).max())
        connectivity_rollback = after_component_count > before_component_count
        if connectivity_rollback:
            retained = component
            remove_mask.fill(False)
            terminal_components = 0
            removed_foreign_ids = []
            metrics["connectivity_rollback_count"] += 1

        removed = int(component.sum() - retained.sum())
        output_whole[retained] = astrocyte_id
        output_soma[soma] = astrocyte_id
        output_process[retained & ~soma] = astrocyte_id
        metrics["decisions"].append(
            {
                "astrocyte_id": astrocyte_id,
                "owner_group_id": owner_group_id,
                "foreign_group_ids": foreign_ids,
                "removed_foreign_group_ids": sorted(set(removed_foreign_ids)),
                "true_tip_count": true_tip_count,
                "terminal_overlap_component_count": terminal_components,
                "preserved_pass_through_component_count": preserved_pass_through,
                "connectivity_rollback": connectivity_rollback,
                "removed_area_px": removed,
                "status": (
                    "terminal_projection_overlap_pruned"
                    if removed
                    else (
                        "connectivity_rollback"
                        if connectivity_rollback
                        else "unchanged"
                    )
                ),
            }
        )
        metrics["changed_cell_count"] += int(removed > 0)
        metrics["removed_area_px"] += removed
        metrics["terminal_overlap_component_count"] += terminal_components
        metrics["preserved_pass_through_component_count"] += preserved_pass_through
    if np.any((output_whole > 0) & ~(whole_labels > 0)):
        raise RuntimeError("Projected foreign-soma guard expanded Whole geometry")
    if np.any((output_soma > 0) & ~(output_whole > 0)):
        raise RuntimeError("Projected foreign-soma guard removed assigned Soma pixels")
    if np.any((output_soma > 0) & (output_process > 0)):
        raise RuntimeError("Projected foreign-soma guard overlapped Soma and Processes")
    if not np.array_equal(output_whole > 0, (output_soma > 0) | (output_process > 0)):
        raise RuntimeError("Projected foreign-soma guard broke the compartment partition")
    return output_whole, output_soma, output_process, metrics

def _complete_soma_within_whole_owner_extent(
    whole_labels: np.ndarray,
    soma_labels: np.ndarray,
    process_labels: np.ndarray,
    inventory: ValidatedNucleusAnchors | None,
    pixel_width_um: float,
    pixel_height_um: float,
    pixel_depth_um: float | None,
    config: SomaNuclearCompletionConfig | None = None,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, dict]:
    """Fill only small missing parts of an already assigned 3D nuclear envelope."""

    cfg = config or SomaNuclearCompletionConfig()
    metrics = {
        "enabled": bool(cfg.enabled and inventory is not None),
        "method": (
            "owner-only canonical nuclear-envelope completion inside frozen Whole; "
            "Processes are recomputed as Whole minus Soma"
        ),
        "evaluated_cell_count": int(whole_labels.max()),
        "changed_cell_count": 0,
        "added_soma_px": 0,
        "decisions": [],
        "config": asdict(cfg),
    }
    groups, grouped_extents = inventory_group_geometry(
        inventory,
        pixel_width_um,
        pixel_height_um,
        pixel_depth_um,
    )
    if not metrics["enabled"] or not groups or grouped_extents is None:
        return whole_labels, soma_labels, process_labels, metrics

    pixel_area_um2 = pixel_width_um * pixel_height_um
    minimum_owner_overlap_px = max(
        1,
        int(math.ceil(cfg.minimum_owner_overlap_um2 / pixel_area_um2)),
    )
    output_soma = soma_labels.copy().astype(np.uint16)
    for astrocyte_id in range(1, int(whole_labels.max()) + 1):
        whole = whole_labels == astrocyte_id
        soma = output_soma == astrocyte_id
        owner = owner_group_for_soma(soma, groups, grouped_extents)
        decision = {
            "astrocyte_id": astrocyte_id,
            "owner_group_id": int(owner["group_id"]) if owner is not None else 0,
            "status": "unchanged",
            "owner_extent_inside_whole_px": 0,
            "existing_extent_coverage": 0.0,
            "added_soma_px": 0,
        }
        if owner is None or str(owner.get("identity_status", "")) != "resolved":
            decision["status"] = "skipped_no_resolved_owner"
            metrics["decisions"].append(decision)
            continue
        owner_extent = grouped_extents == int(owner["group_id"])
        inside = owner_extent & whole
        inside_area = int(inside.sum())
        covered_area = int((inside & soma).sum())
        missing = inside & ~soma
        missing_area = int(missing.sum())
        coverage = covered_area / max(inside_area, 1)
        decision["owner_extent_inside_whole_px"] = inside_area
        decision["existing_extent_coverage"] = coverage
        decision["added_soma_px"] = missing_area
        if inside_area < minimum_owner_overlap_px or covered_area < minimum_owner_overlap_px:
            decision["status"] = "skipped_insufficient_owner_overlap"
        elif coverage < cfg.minimum_existing_extent_coverage:
            decision["status"] = "skipped_extent_not_already_soma_owned"
        elif missing_area > cfg.maximum_added_fraction_of_existing_soma * max(
            int(soma.sum()),
            1,
        ):
            decision["status"] = "skipped_completion_too_large"
        elif missing_area:
            output_soma[missing] = astrocyte_id
            decision["status"] = "completed_owner_nuclear_extent"
            metrics["changed_cell_count"] += 1
            metrics["added_soma_px"] += missing_area
        metrics["decisions"].append(decision)

    output_process = np.where(
        (whole_labels > 0) & (output_soma == 0),
        whole_labels,
        0,
    ).astype(np.uint16)
    if not np.array_equal(whole_labels > 0, (output_soma > 0) | (output_process > 0)):
        raise RuntimeError("Soma nuclear-envelope completion broke the compartment partition")
    if np.any((output_soma > 0) & (output_process > 0)):
        raise RuntimeError("Soma nuclear-envelope completion overlapped Soma and Processes")
    if np.any((output_soma > 0) & ~(whole_labels > 0)):
        raise RuntimeError("Soma nuclear-envelope completion expanded outside Whole")
    return whole_labels, output_soma, output_process, metrics

def resolve_canonical_owner_assignments(
    identity_metrics: dict,
    axial_metrics: dict,
    inventory: ValidatedNucleusAnchors | None,
    final_roi_count: int,
) -> tuple[dict[int, dict], dict[int, list[str]]]:
    """Resolve one canonical, axially retained 3D owner for each final cell."""

    assignments: dict[int, dict] = {}
    failures: dict[int, list[str]] = {}

    def reject(astrocyte_id: int, reason: str) -> None:
        failures.setdefault(int(astrocyte_id), []).append(str(reason))

    if inventory is None:
        for astrocyte_id in range(1, int(final_roi_count) + 1):
            reject(astrocyte_id, "owner_inventory_missing")
        return assignments, failures

    records = {
        int(row["instance_id"]): row
        for row in inventory.nucleus_instance_records
    }
    accepted_ids = {
        int(value) for value in inventory.accepted_instance_ids
    }
    identity_lineage = {
        int(key): value
        for key, value in identity_metrics.get("final_lineage", {}).items()
    }
    axial_mapping = {
        int(old_id): int(new_id)
        for old_id, new_id in axial_metrics.get("id_mapping", {}).items()
    }
    inverse_axial_mapping: dict[int, int] = {}
    duplicate_final_ids: set[int] = set()
    for pre_axial_id, final_id in axial_mapping.items():
        if final_id in inverse_axial_mapping:
            duplicate_final_ids.add(final_id)
        else:
            inverse_axial_mapping[final_id] = pre_axial_id
    axial_decisions = {
        int(row["pre_guard_astrocyte_id"]): row
        for row in axial_metrics.get("decisions", [])
        if int(row.get("pre_guard_astrocyte_id", 0)) > 0
    }
    lineage_owner_claims: dict[int, list[int]] = {}
    for astrocyte_id in range(1, int(final_roi_count) + 1):
        pre_axial_id = inverse_axial_mapping.get(astrocyte_id)
        lineage = (
            identity_lineage.get(pre_axial_id)
            if pre_axial_id is not None
            else None
        )
        owner_id = (
            int(lineage.get("canonical_owner_id", 0))
            if lineage is not None
            else 0
        )
        if owner_id > 0:
            lineage_owner_claims.setdefault(owner_id, []).append(
                astrocyte_id
            )
    duplicate_owner_claim_cells = {
        astrocyte_id
        for cell_ids in lineage_owner_claims.values()
        if len(cell_ids) > 1
        for astrocyte_id in cell_ids
    }

    for astrocyte_id in range(1, int(final_roi_count) + 1):
        if astrocyte_id in duplicate_final_ids:
            reject(astrocyte_id, "axial_mapping_not_unique")
            continue
        pre_axial_id = inverse_axial_mapping.get(astrocyte_id)
        if pre_axial_id is None:
            reject(astrocyte_id, "axial_mapping_missing")
            continue
        lineage = identity_lineage.get(pre_axial_id)
        if lineage is None:
            reject(astrocyte_id, "owner_lineage_missing")
            continue
        owner_id = int(lineage.get("canonical_owner_id", 0))
        if owner_id <= 0:
            reject(astrocyte_id, "canonical_owner_missing")
            continue
        if astrocyte_id in duplicate_owner_claim_cells:
            reject(astrocyte_id, "owner_assignment_not_unique")
            continue
        record = records.get(owner_id)
        if record is None:
            reject(astrocyte_id, "owner_record_missing")
            continue
        if not bool(record.get("dapi_valid", False)):
            reject(astrocyte_id, "owner_not_dapi_valid")
            continue
        if str(record.get("identity_status", "")) != "resolved":
            reject(astrocyte_id, "owner_not_resolved")
            continue
        if not bool(record.get("accepted", False)):
            reject(astrocyte_id, "owner_not_accepted")
            continue
        if owner_id not in accepted_ids:
            reject(astrocyte_id, "owner_not_in_accepted_ids")
            continue
        axial_decision = axial_decisions.get(pre_axial_id)
        if axial_decision is None:
            reject(astrocyte_id, "axial_decision_missing")
            continue
        axial_owner_id = int(axial_decision.get("owner_group_id", 0))
        if axial_owner_id != owner_id:
            reject(astrocyte_id, "axial_owner_mismatch")
            continue
        if str(axial_decision.get("status", "")) != "retained":
            reject(astrocyte_id, "axial_owner_not_retained")
            continue
        assignments[astrocyte_id] = {
            "astrocyte_id": astrocyte_id,
            "pre_axial_id": pre_axial_id,
            "owner_id": owner_id,
            "lineage_owner_id": owner_id,
            "axial_owner_id": axial_owner_id,
            "owner_record": record,
            "identity_changed": bool(
                lineage.get("identity_changed", False)
            ),
        }

    return assignments, failures

def complete_soma_to_owner_nuclear_extent(
    whole_labels: np.ndarray,
    soma_labels: np.ndarray,
    process_labels: np.ndarray,
    inventory: ValidatedNucleusAnchors | None,
    pixel_width_um: float,
    pixel_height_um: float,
    pixel_depth_um: float | None,
    owner_assignments: dict[int, dict],
    owner_assignment_failures: dict[int, list[str]],
    pre_finalization_whole_union: np.ndarray,
    config: SomaNuclearCompletionConfig | None = None,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, dict, np.ndarray]:
    """Complete only an unambiguous canonical owner extent, atomically."""

    cfg = config or SomaNuclearCompletionConfig()
    if (
        whole_labels.shape != soma_labels.shape
        or whole_labels.shape != process_labels.shape
    ):
        raise ValueError("Canonical Owner Nuclear-Extent Completion compartment label shapes do not match")
    frozen_pre_finalization_whole = np.asarray(
        pre_finalization_whole_union,
        dtype=bool,
    )
    if frozen_pre_finalization_whole.shape != whole_labels.shape:
        raise ValueError(
            "Canonical Owner Nuclear-Extent Completion pre-finalization Whole shape "
            "does not match compartment labels"
        )
    original_ids = set(int(value) for value in np.unique(whole_labels))
    metrics = {
        "enabled": bool(cfg.enabled and inventory is not None),
        "method": (
            "exact canonical owner-nucleus extent completion after identity and "
            "axial consensus; no dilation, closing, hull, or intensity voting"
        ),
        "pre_finalization_whole_area_px": int(
            frozen_pre_finalization_whole.sum()
        ),
        "pre_canonical_owner_extent_completion_whole_area_px": int(
            (whole_labels > 0).sum()
        ),
        "post_canonical_owner_extent_completion_whole_area_px": int(
            (whole_labels > 0).sum()
        ),
        "pre_finalization_whole_to_final_removed_px": 0,
        "pre_finalization_whole_to_final_approved_added_px": 0,
        "evaluated_cell_count": int(whole_labels.max()),
        "eligible_cell_count": 0,
        "changed_cell_count": 0,
        "no_op_cell_count": 0,
        "fail_closed_cell_count": 0,
        "inside_added_soma_px": 0,
        "outside_added_whole_px": 0,
        "outside_added_soma_px": 0,
        "changed_cell_ids": [],
        "outside_changed_cell_ids": [],
        "no_op_cell_ids": [],
        "fail_closed_cell_ids": [],
        "decisions": [],
        "config": asdict(cfg),
    }
    approved_outside = np.zeros(whole_labels.shape, dtype=bool)
    if not metrics["enabled"]:
        for astrocyte_id in range(1, int(whole_labels.max()) + 1):
            metrics["decisions"].append(
                {
                    "astrocyte_id": astrocyte_id,
                    "status": "fail_closed_owner_assignment_missing",
                    "fail_reasons": ["owner_inventory_missing"],
                }
            )
            metrics["fail_closed_cell_ids"].append(astrocyte_id)
        metrics["fail_closed_cell_count"] = len(
            metrics["fail_closed_cell_ids"]
        )
        return (
            whole_labels,
            soma_labels,
            process_labels,
            metrics,
            approved_outside,
        )

    groups, grouped_extents = inventory_group_geometry(
        inventory,
        pixel_width_um,
        pixel_height_um,
        pixel_depth_um,
    )
    if grouped_extents is None:
        owner_assignment_failures = {
            astrocyte_id: ["owner_extent_inventory_missing"]
            for astrocyte_id in range(1, int(whole_labels.max()) + 1)
        }
        owner_assignments = {}
    group_by_id = {
        int(group["group_id"]): group for group in groups
    }
    pixel_area_um2 = float(pixel_width_um * pixel_height_um)
    minimum_owner_overlap_px = max(
        1,
        int(math.ceil(cfg.minimum_owner_overlap_um2 / pixel_area_um2)),
    )
    plans: dict[int, dict] = {}
    decisions: dict[int, dict] = {}

    def failure_status(reasons: list[str]) -> str:
        if "owner_assignment_not_unique" in reasons:
            return "fail_closed_owner_assignment_not_unique"
        if "axial_owner_mismatch" in reasons:
            return "fail_closed_axial_owner_mismatch"
        if any(reason.startswith("axial_") for reason in reasons):
            return "fail_closed_axial_status_unverified"
        if any(
            reason in {
                "owner_not_accepted",
                "owner_not_resolved",
                "owner_not_dapi_valid",
                "owner_not_in_accepted_ids",
                "owner_record_missing",
            }
            for reason in reasons
        ):
            return "fail_closed_owner_not_accepted_resolved"
        return "fail_closed_owner_assignment_missing"

    for astrocyte_id in range(1, int(whole_labels.max()) + 1):
        assignment = owner_assignments.get(astrocyte_id)
        assignment_reasons = list(
            owner_assignment_failures.get(astrocyte_id, [])
        )
        decision = {
            "astrocyte_id": astrocyte_id,
            "pre_axial_id": int(
                assignment.get("pre_axial_id", 0)
                if assignment is not None
                else 0
            ),
            "lineage_owner_id": int(
                assignment.get("lineage_owner_id", 0)
                if assignment is not None
                else 0
            ),
            "axial_owner_id": int(
                assignment.get("axial_owner_id", 0)
                if assignment is not None
                else 0
            ),
            "owner_group_id": int(
                assignment.get("owner_id", 0)
                if assignment is not None
                else 0
            ),
            "owner_assignment_unique": bool(
                assignment is not None and not assignment_reasons
            ),
            "accepted_resolved_soma_candidate_ids": [],
            "resolved_nucleus_ids_in_whole": [],
            "foreign_resolved_nucleus_ids": [],
            "incidental_resolved_nucleus_ids_in_whole": [],
            "foreign_resolved_nucleus_evidence": [],
            "owner_extent_total_px": 0,
            "owner_extent_inside_current_whole_px": 0,
            "owner_extent_current_background_px": 0,
            "owner_extent_outside_pre_finalization_whole_px": 0,
            "prior_guard_removed_owner_px": 0,
            "existing_soma_owner_overlap_px": 0,
            "missing_inside_px": 0,
            "foreign_whole_overlap_px": 0,
            "foreign_soma_overlap_px": 0,
            "owner_extent_component_count": 0,
            "owner_anchor_component_count": 0,
            "selected_owner_extent_component_count": 0,
            "selected_owner_extent_component_ids": [],
            "selected_owner_extent_component_px": 0,
            "selected_owner_core_overlap_px": 0,
            "selected_owner_soma_overlap_px": 0,
            "selected_owner_core_component_count": 0,
            "selected_owner_substantial_core_component_count": 0,
            "selected_owner_substantial_core_component_ids": [],
            "selected_owner_core_component_details": [],
            "ignored_owner_extent_component_count": 0,
            "ignored_owner_extent_px": 0,
            "owner_extent_component_details": [],
            "owner_extent_xy_edge_touch": False,
            "approved_inside_soma_px": 0,
            "approved_outside_whole_px": 0,
            "whole_delta_px": 0,
            "soma_delta_px": 0,
            "process_delta_px": 0,
            "approved": False,
            "approved_owner_extent_added_px": 0,
            "status": "unchanged",
            "fail_reasons": assignment_reasons,
        }
        decisions[astrocyte_id] = decision
        if assignment is None or assignment_reasons:
            decision["status"] = failure_status(assignment_reasons)
            continue

        owner_id = int(assignment["owner_id"])
        owner_group = group_by_id.get(owner_id)
        if owner_group is None:
            decision["fail_reasons"].append("owner_extent_record_missing")
            decision["status"] = "fail_closed_owner_assignment_missing"
            continue
        full_owner_extent = np.asarray(
            grouped_extents == owner_id,
            dtype=bool,
        )
        if (
            inventory.nucleus_instance_core_labels_2d is not None
            and np.asarray(
                inventory.nucleus_instance_core_labels_2d
            ).shape
            == full_owner_extent.shape
        ):
            owner_core = np.asarray(
                inventory.nucleus_instance_core_labels_2d == owner_id,
                dtype=bool,
            )
        else:
            owner_core = np.zeros_like(full_owner_extent, dtype=bool)
        own_whole = whole_labels == astrocyte_id
        own_soma = soma_labels == astrocyte_id
        component_labels = measure.label(
            full_owner_extent,
            connectivity=2,
        )
        component_count = int(component_labels.max())
        owner_soma_distance_um = ndi.distance_transform_edt(
            ~own_soma,
            sampling=(pixel_height_um, pixel_width_um),
        )
        component_details = []
        owner_anchor_component_ids = []
        for component_id in range(1, component_count + 1):
            component = component_labels == component_id
            component_area = int(component.sum())
            core_overlap = int((component & owner_core).sum())
            soma_overlap = int((component & own_soma).sum())
            whole_overlap = int((component & own_whole).sum())
            outside_pre_finalization_whole = int(
                (component & ~frozen_pre_finalization_whole).sum()
            )
            minimum_soma_distance_um = (
                float(owner_soma_distance_um[component].min())
                if component_area > 0
                else math.inf
            )
            core_soma_anchor = bool(
                core_overlap > 0
                and soma_overlap >= minimum_owner_overlap_px
            )
            if core_soma_anchor:
                owner_anchor_component_ids.append(component_id)
            rows, cols = np.nonzero(component)
            component_details.append(
                {
                    "component_id": component_id,
                    "area_px": component_area,
                    "core_overlap_px": core_overlap,
                    "soma_overlap_px": soma_overlap,
                    "whole_overlap_px": whole_overlap,
                    "outside_pre_finalization_whole_px": (
                        outside_pre_finalization_whole
                    ),
                    "minimum_soma_distance_um": float(
                        minimum_soma_distance_um
                    ),
                    "core_soma_anchor": core_soma_anchor,
                    "bbox_yx": [
                        int(rows.min()),
                        int(cols.min()),
                        int(rows.max()) + 1,
                        int(cols.max()) + 1,
                    ],
                }
            )
        if len(owner_anchor_component_ids) == 1:
            selected_component_ids = list(owner_anchor_component_ids)
            owner_extent = (
                component_labels == selected_component_ids[0]
            )
        else:
            selected_component_ids = []
            owner_extent = full_owner_extent
        ignored_component_ids = [
            component_id
            for component_id in range(1, component_count + 1)
            if component_id not in selected_component_ids
        ]
        ignored_owner_extent_px = int(
            np.isin(component_labels, ignored_component_ids).sum()
        )
        selected_core_component_details = []
        selected_substantial_core_component_ids = []
        selected_core_component_count = 0
        if selected_component_ids:
            selected_core = owner_core & owner_extent
            selected_core_labels = measure.label(
                selected_core,
                connectivity=2,
            )
            selected_core_component_count = int(
                selected_core_labels.max()
            )
            for core_component_id in range(
                1,
                selected_core_component_count + 1,
            ):
                core_component = (
                    selected_core_labels == core_component_id
                )
                core_component_area = int(core_component.sum())
                substantial = bool(
                    core_component_area >= minimum_owner_overlap_px
                )
                if substantial:
                    selected_substantial_core_component_ids.append(
                        core_component_id
                    )
                rows, cols = np.nonzero(core_component)
                selected_core_component_details.append(
                    {
                        "component_id": core_component_id,
                        "area_px": core_component_area,
                        "substantial": substantial,
                        "bbox_yx": [
                            int(rows.min()),
                            int(cols.min()),
                            int(rows.max()) + 1,
                            int(cols.max()) + 1,
                        ],
                    }
                )
        accepted_resolved_candidates = []
        resolved_nuclei_in_whole = []
        foreign_resolved_nucleus_ids = []
        incidental_resolved_nucleus_ids = []
        foreign_resolved_nucleus_evidence = []
        owner_distance_um = ndi.distance_transform_edt(
            ~owner_extent,
            sampling=(pixel_height_um, pixel_width_um),
        )
        local_domain = own_soma | owner_extent
        local_domain_distance_um = ndi.distance_transform_edt(
            ~local_domain,
            sampling=(pixel_height_um, pixel_width_um),
        )
        minimum_foreign_overlap_px = max(
            1,
            int(
                math.ceil(
                    cfg.minimum_foreign_overlap_um2 / pixel_area_um2
                )
            ),
        )
        for group in groups:
            group_id = int(group["group_id"])
            group_extent = np.asarray(
                grouped_extents == group_id,
                dtype=bool,
            )
            whole_overlap = int((own_whole & group_extent).sum())
            soma_overlap = int((own_soma & group_extent).sum())
            is_resolved = (
                str(group.get("identity_status", "")) == "resolved"
            )
            if is_resolved and whole_overlap > 0:
                resolved_nuclei_in_whole.append(group_id)
            if (
                bool(group.get("accepted", False))
                and is_resolved
                and soma_overlap >= minimum_owner_overlap_px
            ):
                accepted_resolved_candidates.append(group_id)
            if (
                group_id == owner_id
                or not is_resolved
                or whole_overlap <= 0
            ):
                continue
            extent_area = int(group_extent.sum())
            whole_overlap_fraction = (
                whole_overlap / max(extent_area, 1)
            )
            minimum_owner_distance_um = (
                float(owner_distance_um[group_extent].min())
                if extent_area > 0
                else math.inf
            )
            minimum_local_domain_distance_um = (
                float(local_domain_distance_um[group_extent].min())
                if extent_area > 0
                else math.inf
            )
            meaningful_soma_overlap = (
                soma_overlap >= minimum_foreign_overlap_px
            )
            meaningful_whole_overlap = (
                whole_overlap >= minimum_foreign_overlap_px
                and whole_overlap_fraction
                >= cfg.minimum_foreign_overlap_fraction
            )
            local_owner_contact = (
                whole_overlap >= minimum_foreign_overlap_px
                and minimum_local_domain_distance_um
                <= cfg.maximum_local_foreign_distance_um
            )
            veto = bool(
                meaningful_soma_overlap
                or local_owner_contact
            )
            foreign_resolved_nucleus_evidence.append(
                {
                    "nucleus_id": group_id,
                    "accepted": bool(group.get("accepted", False)),
                    "whole_overlap_px": whole_overlap,
                    "soma_overlap_px": soma_overlap,
                    "whole_overlap_fraction_of_extent": float(
                        whole_overlap_fraction
                    ),
                    "minimum_owner_extent_distance_um": float(
                        minimum_owner_distance_um
                    ),
                    "minimum_local_domain_distance_um": float(
                        minimum_local_domain_distance_um
                    ),
                    "meaningful_soma_overlap": bool(
                        meaningful_soma_overlap
                    ),
                    "meaningful_whole_overlap": bool(
                        meaningful_whole_overlap
                    ),
                    "local_owner_contact": bool(local_owner_contact),
                    "veto": veto,
                }
            )
            if veto:
                foreign_resolved_nucleus_ids.append(group_id)
            else:
                incidental_resolved_nucleus_ids.append(group_id)
        decision["accepted_resolved_soma_candidate_ids"] = sorted(
            accepted_resolved_candidates
        )
        decision["resolved_nucleus_ids_in_whole"] = sorted(
            resolved_nuclei_in_whole
        )
        decision["foreign_resolved_nucleus_ids"] = sorted(
            foreign_resolved_nucleus_ids
        )
        decision["incidental_resolved_nucleus_ids_in_whole"] = sorted(
            incidental_resolved_nucleus_ids
        )
        decision["foreign_resolved_nucleus_evidence"] = sorted(
            foreign_resolved_nucleus_evidence,
            key=lambda row: int(row["nucleus_id"]),
        )

        inside_current_whole = owner_extent & own_whole
        current_background = owner_extent & (whole_labels == 0)
        prior_guard_removed = current_background & frozen_pre_finalization_whole
        novel_outside_pre_finalization_whole = (
            current_background & ~frozen_pre_finalization_whole
        )
        foreign_whole = (
            owner_extent
            & (whole_labels > 0)
            & (whole_labels != astrocyte_id)
        )
        foreign_soma = (
            owner_extent
            & (soma_labels > 0)
            & (soma_labels != astrocyte_id)
        )
        missing_inside = inside_current_whole & ~own_soma
        existing_overlap = int((inside_current_whole & own_soma).sum())
        edge_touch = bool(
            owner_extent[0, :].any()
            or owner_extent[-1, :].any()
            or owner_extent[:, 0].any()
            or owner_extent[:, -1].any()
        )
        decision.update(
            {
                "owner_extent_total_px": int(
                    full_owner_extent.sum()
                ),
                "owner_extent_inside_current_whole_px": int(
                    inside_current_whole.sum()
                ),
                "owner_extent_current_background_px": int(
                    current_background.sum()
                ),
                "owner_extent_outside_pre_finalization_whole_px": int(
                    novel_outside_pre_finalization_whole.sum()
                ),
                "prior_guard_removed_owner_px": int(
                    prior_guard_removed.sum()
                ),
                "existing_soma_owner_overlap_px": existing_overlap,
                "missing_inside_px": int(missing_inside.sum()),
                "foreign_whole_overlap_px": int(foreign_whole.sum()),
                "foreign_soma_overlap_px": int(foreign_soma.sum()),
                "owner_extent_component_count": component_count,
                "owner_anchor_component_count": int(
                    len(owner_anchor_component_ids)
                ),
                "selected_owner_extent_component_count": len(
                    selected_component_ids
                ),
                "selected_owner_extent_component_ids": (
                    selected_component_ids
                ),
                "selected_owner_extent_component_px": int(
                    owner_extent.sum()
                    if selected_component_ids
                    else 0
                ),
                "selected_owner_core_overlap_px": int(
                    (owner_extent & owner_core).sum()
                    if selected_component_ids
                    else 0
                ),
                "selected_owner_soma_overlap_px": int(
                    (owner_extent & own_soma).sum()
                    if selected_component_ids
                    else 0
                ),
                "selected_owner_core_component_count": int(
                    selected_core_component_count
                ),
                "selected_owner_substantial_core_component_count": len(
                    selected_substantial_core_component_ids
                ),
                "selected_owner_substantial_core_component_ids": (
                    selected_substantial_core_component_ids
                ),
                "selected_owner_core_component_details": (
                    selected_core_component_details
                ),
                "ignored_owner_extent_component_count": len(
                    ignored_component_ids
                ),
                "ignored_owner_extent_px": ignored_owner_extent_px,
                "owner_extent_component_details": component_details,
                "owner_extent_xy_edge_touch": edge_touch,
            }
        )
        fail_reasons: list[str] = []
        if owner_extent.sum() == 0:
            fail_reasons.append("owner_extent_missing")
        if len(owner_anchor_component_ids) == 0:
            fail_reasons.append(
                "owner_extent_no_core_soma_anchored_component"
            )
        elif len(owner_anchor_component_ids) > 1:
            fail_reasons.append(
                "owner_extent_multiple_core_soma_anchored_components"
            )
        if (
            len(owner_anchor_component_ids) == 1
            and len(selected_substantial_core_component_ids) == 0
        ):
            fail_reasons.append(
                "owner_extent_no_substantial_core_component"
            )
        elif (
            len(owner_anchor_component_ids) == 1
            and len(selected_substantial_core_component_ids) > 1
        ):
            fail_reasons.append(
                "owner_extent_multiple_substantial_core_components"
            )
        if accepted_resolved_candidates != [owner_id]:
            fail_reasons.append("owner_assignment_not_unique")
        if decision["foreign_resolved_nucleus_ids"]:
            fail_reasons.append("multiple_resolved_nuclei_in_whole")
        if existing_overlap < minimum_owner_overlap_px:
            fail_reasons.append("insufficient_soma_owner_anchor")
        if edge_touch:
            fail_reasons.append("owner_extent_xy_edge_touch")
        if foreign_whole.any() or foreign_soma.any():
            fail_reasons.append("competing_whole_or_soma")
        if prior_guard_removed.any():
            fail_reasons.append("prior_guard_removed_owner_extent")
        if fail_reasons:
            decision["fail_reasons"].extend(fail_reasons)
            if (
                "owner_assignment_not_unique" in fail_reasons
                or "multiple_resolved_nuclei_in_whole" in fail_reasons
            ):
                decision["status"] = (
                    "fail_closed_owner_assignment_not_unique"
                )
            elif (
                "owner_extent_no_core_soma_anchored_component"
                in fail_reasons
            ):
                decision["status"] = (
                    "fail_closed_owner_component_anchor_missing"
                )
            elif (
                "owner_extent_multiple_core_soma_anchored_components"
                in fail_reasons
            ):
                decision["status"] = (
                    "fail_closed_owner_component_ambiguous"
                )
            elif (
                "owner_extent_no_substantial_core_component"
                in fail_reasons
                or "owner_extent_multiple_substantial_core_components"
                in fail_reasons
            ):
                decision["status"] = (
                    "fail_closed_owner_core_ambiguous"
                )
            elif "insufficient_soma_owner_anchor" in fail_reasons:
                decision["status"] = (
                    "fail_closed_insufficient_soma_anchor"
                )
            elif "owner_extent_xy_edge_touch" in fail_reasons:
                decision["status"] = "fail_closed_xy_edge_touch"
            elif "competing_whole_or_soma" in fail_reasons:
                decision["status"] = "fail_closed_competing_whole"
            elif "prior_guard_removed_owner_extent" in fail_reasons:
                decision["status"] = (
                    "fail_closed_prior_guard_removed_extent"
                )
            else:
                decision["status"] = "fail_closed_owner_assignment_missing"
            continue
        plans[astrocyte_id] = {
            "inside": missing_inside,
            "outside": novel_outside_pre_finalization_whole,
        }
        metrics["eligible_cell_count"] += 1

    planned_owners = sorted(plans)
    conflict_ids: set[int] = set()
    for position, first_id in enumerate(planned_owners):
        first = plans[first_id]["outside"]
        for second_id in planned_owners[position + 1 :]:
            if np.any(first & plans[second_id]["outside"]):
                conflict_ids.update((first_id, second_id))
    for astrocyte_id in sorted(conflict_ids):
        plans.pop(astrocyte_id, None)
        decision = decisions[astrocyte_id]
        decision["fail_reasons"].append("planned_outside_overlap")
        decision["status"] = "fail_closed_planned_outside_conflict"

    output_whole = whole_labels.copy()
    output_soma = soma_labels.copy()
    for astrocyte_id, plan in sorted(plans.items()):
        inside = plan["inside"]
        outside = plan["outside"]
        output_soma[inside] = astrocyte_id
        output_whole[outside] = astrocyte_id
        output_soma[outside] = astrocyte_id
        approved_outside |= outside
        inside_count = int(inside.sum())
        outside_count = int(outside.sum())
        decision = decisions[astrocyte_id]
        decision["approved_inside_soma_px"] = inside_count
        decision["approved_outside_whole_px"] = outside_count
        decision["whole_delta_px"] = outside_count
        decision["soma_delta_px"] = inside_count + outside_count
        decision["process_delta_px"] = -inside_count
        decision["approved"] = bool(inside_count or outside_count)
        decision["approved_owner_extent_added_px"] = (
            inside_count + outside_count
        )
        if inside_count or outside_count:
            decision["status"] = (
                "completed_inside_and_exact_outside_pre_finalization_whole_owner_extent"
                if outside_count
                else "completed_inside_owner_extent"
            )
            metrics["changed_cell_ids"].append(astrocyte_id)
            if outside_count:
                metrics["outside_changed_cell_ids"].append(astrocyte_id)
            metrics["inside_added_soma_px"] += inside_count
            metrics["outside_added_whole_px"] += outside_count
            metrics["outside_added_soma_px"] += outside_count
        else:
            decision["status"] = "no_op_owner_already_complete"
            metrics["no_op_cell_ids"].append(astrocyte_id)

    for astrocyte_id, decision in sorted(decisions.items()):
        if str(decision["status"]).startswith("fail_closed_"):
            metrics["fail_closed_cell_ids"].append(astrocyte_id)
        elif astrocyte_id not in plans:
            metrics["no_op_cell_ids"].append(astrocyte_id)
        metrics["decisions"].append(decision)
    metrics["changed_cell_ids"] = sorted(set(metrics["changed_cell_ids"]))
    metrics["outside_changed_cell_ids"] = sorted(
        set(metrics["outside_changed_cell_ids"])
    )
    metrics["no_op_cell_ids"] = sorted(set(metrics["no_op_cell_ids"]))
    metrics["fail_closed_cell_ids"] = sorted(
        set(metrics["fail_closed_cell_ids"])
    )
    metrics["changed_cell_count"] = len(metrics["changed_cell_ids"])
    metrics["no_op_cell_count"] = len(metrics["no_op_cell_ids"])
    metrics["fail_closed_cell_count"] = len(
        metrics["fail_closed_cell_ids"]
    )

    output_process = np.where(
        (output_whole > 0) & (output_soma == 0),
        output_whole,
        0,
    ).astype(process_labels.dtype, copy=False)
    if output_whole.dtype != whole_labels.dtype:
        raise RuntimeError(
            "Canonical Owner Nuclear-Extent Completion changed the Whole label dtype"
        )
    if output_soma.dtype != soma_labels.dtype:
        raise RuntimeError(
            "Canonical Owner Nuclear-Extent Completion changed the Soma label dtype"
        )
    if set(int(value) for value in np.unique(output_whole)) != original_ids:
        raise RuntimeError(
            "Canonical Owner Nuclear-Extent Completion changed the Whole ID set"
        )
    if np.any((output_soma > 0) & (output_process > 0)):
        raise RuntimeError(
            "Canonical Owner Nuclear-Extent Completion overlapped Soma and Processes"
        )
    if not np.array_equal(
        output_whole > 0,
        (output_soma > 0) | (output_process > 0),
    ):
        raise RuntimeError(
            "Canonical Owner Nuclear-Extent Completion broke Whole = Soma union Processes"
        )
    if not np.array_equal(
        output_soma[output_soma > 0],
        output_whole[output_soma > 0],
    ):
        raise RuntimeError(
            "Canonical Owner Nuclear-Extent Completion changed a Soma pixel to a foreign ID"
        )
    if not np.array_equal(
        output_process[output_process > 0],
        output_whole[output_process > 0],
    ):
        raise RuntimeError(
            "Canonical Owner Nuclear-Extent Completion changed a Processes pixel to a foreign ID"
        )
    added_whole = (output_whole > 0) & ~(whole_labels > 0)
    if not np.array_equal(added_whole, approved_outside):
        raise RuntimeError(
            "Canonical Owner Nuclear-Extent Completion Whole expansion differs from approved mask"
        )
    metrics["post_canonical_owner_extent_completion_whole_area_px"] = int(
        (output_whole > 0).sum()
    )
    metrics["pre_finalization_whole_to_final_removed_px"] = int(
        (frozen_pre_finalization_whole & ~(output_whole > 0)).sum()
    )
    metrics["pre_finalization_whole_to_final_approved_added_px"] = int(
        approved_outside.sum()
    )
    metrics["approved_owner_extent_added_px"] = int(
        metrics["inside_added_soma_px"]
        + metrics["outside_added_soma_px"]
    )
    return (
        output_whole,
        output_soma,
        output_process,
        metrics,
        approved_outside,
    )

def reconcile_same_id_disconnected_soma(
    whole_labels: np.ndarray,
    soma_labels: np.ndarray,
    process_labels: np.ndarray,
    inventory: ValidatedNucleusAnchors | None,
    identity_metrics: dict,
    axial_metrics: dict,
    profile: str,
    pixel_width_um: float,
    pixel_height_um: float,
    pixel_depth_um: float | None,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, dict]:
    """Join only short, same-owner Soma islands inside one frozen Whole ID."""

    if (
        whole_labels.shape != soma_labels.shape
        or whole_labels.shape != process_labels.shape
    ):
        raise ValueError(
            "Same-ID Soma Island Reconciliation compartment label shapes do not match"
        )
    frozen_whole_labels = whole_labels.copy()
    ownership_cfg = NucleusOwnershipConfig()
    compartment_cfg = compartment_config_for_profile(profile)
    pixel_area_um2 = float(pixel_width_um * pixel_height_um)
    minimum_owner_overlap_px = max(
        1,
        int(
            math.ceil(
                ownership_cfg.owner_min_overlap_um2 / pixel_area_um2
            )
        ),
    )
    maximum_absolute_added_px = max(
        1,
        int(
            math.floor(
                compartment_cfg.min_soma_area_um2
                * SomaNuclearCompletionConfig().maximum_added_fraction_of_existing_soma
                / pixel_area_um2
            )
        ),
    )
    bridge_radius_y = max(
        1,
        int(
            round(
                compartment_cfg.dapi_extent_closing_um
                / max(float(pixel_height_um), 1e-9)
            )
        ),
    )
    bridge_radius_x = max(
        1,
        int(
            round(
                compartment_cfg.dapi_extent_closing_um
                / max(float(pixel_width_um), 1e-9)
            )
        ),
    )
    bridge_structure = cv2.getStructuringElement(
        cv2.MORPH_ELLIPSE,
        (2 * bridge_radius_x + 1, 2 * bridge_radius_y + 1),
    ).astype(bool)
    metrics = {
        "enabled": bool(
            inventory is not None
            and pixel_depth_um is not None
            and pixel_depth_um > 0
        ),
        "method": (
            "identity-gated, calibration-aware same-owner island bridging; "
            "Whole is frozen and only same-ID Processes may become Soma"
        ),
        "disconnected_before_ids": [],
        "bridged_ids": [],
        "rejected_identity_split_ids": [],
        "rejected_multiple_owner_ids": [],
        "rejected_local_foreign_ids": [],
        "rejected_ambiguous_owner_ids": [],
        "rejected_gap_ids": [],
        "rejected_path_ids": [],
        "added_soma_px": 0,
        "removed_process_px": 0,
        "approved_process_to_soma_px": 0,
        "canonical_satellite_aliases": {},
        "canonical_satellite_alias_details": [],
        "decisions": [],
        "derived_thresholds": {
            "fragment_bridge_max_gap_um": float(
                ownership_cfg.fragment_bridge_max_gap_um
            ),
            "owner_min_overlap_px": minimum_owner_overlap_px,
            "owner_halo_um": float(
                compartment_cfg.soma_trusted_core_nucleus_margin_um
            ),
            "bridge_radius_y_px": bridge_radius_y,
            "bridge_radius_x_px": bridge_radius_x,
            "maximum_absolute_added_px": maximum_absolute_added_px,
            "maximum_relative_added_fraction": float(
                SomaNuclearCompletionConfig().maximum_added_fraction_of_existing_soma
            ),
            "minimum_process_fraction": float(
                compartment_cfg.min_process_fraction
            ),
        },
    }
    if profile == "neonatal":
        metrics["method"] += (
            "; safe merge-only two-island pairs may use the exact owner "
            "extent inside their pair convex hull"
        )
        metrics["owner_convex_hull_ids"] = []
        metrics["owner_convex_hull_added_px"] = 0
        metrics["derived_thresholds"].update(
            {
                "owner_convex_hull_profile": "neonatal",
                "owner_convex_hull_component_count": 2,
                "owner_convex_hull_relative_cap_only": True,
            }
        )
    approved_labels = np.zeros(whole_labels.shape, dtype=np.uint16)
    output_soma = soma_labels.copy()
    if inventory is not None:
        groups, grouped_extents = inventory_group_geometry(
            inventory,
            pixel_width_um,
            pixel_height_um,
            pixel_depth_um,
        )
    else:
        groups, grouped_extents = [], None
    group_by_id = {
        int(group["group_id"]): group for group in groups
    }
    records_by_id = (
        {
            int(row["instance_id"]): row
            for row in inventory.nucleus_instance_records
        }
        if inventory is not None
        else {}
    )
    canonical_cfg = CanonicalIdentityConfig()
    satellite_aliases: dict[int, int] = {}
    satellite_alias_details: list[dict] = []
    for satellite_id, satellite in sorted(records_by_id.items()):
        if pixel_depth_um is None or pixel_depth_um <= 0:
            break
        if (
            bool(satellite.get("accepted", False))
            or not bool(satellite.get("dapi_valid", False))
            or str(satellite.get("identity_status", "")) != "resolved"
            or bool(
                satellite.get("resolution_diagnostics", {}).get(
                    "split_accepted",
                    False,
                )
            )
        ):
            continue
        satellite_sources = {
            int(value) for value in satellite.get("source_object_ids", ())
        }
        satellite_volume = float(satellite.get("volume_um3", 0.0))
        if not satellite_sources or satellite_volume <= 0:
            continue
        candidates: list[tuple[int, dict]] = []
        for owner_candidate_id, owner_candidate in sorted(
            records_by_id.items()
        ):
            if (
                owner_candidate_id == satellite_id
                or not bool(owner_candidate.get("accepted", False))
                or not bool(owner_candidate.get("dapi_valid", False))
                or str(owner_candidate.get("identity_status", ""))
                != "resolved"
            ):
                continue
            owner_sources = {
                int(value)
                for value in owner_candidate.get("source_object_ids", ())
            }
            if not satellite_sources < owner_sources:
                continue
            owner_volume = float(owner_candidate.get("volume_um3", 0.0))
            if owner_volume <= 0:
                continue
            volume_ratio = satellite_volume / owner_volume
            if volume_ratio > canonical_cfg.satellite_max_volume_ratio:
                continue
            z_overlap = max(
                0,
                min(
                    int(satellite["z_max_0based_inclusive"]),
                    int(owner_candidate["z_max_0based_inclusive"]),
                )
                - max(
                    int(satellite["z_min_0based"]),
                    int(owner_candidate["z_min_0based"]),
                )
                + 1,
            )
            satellite_z_span = (
                int(satellite["z_max_0based_inclusive"])
                - int(satellite["z_min_0based"])
                + 1
            )
            z_overlap_fraction = z_overlap / max(satellite_z_span, 1)
            if (
                z_overlap_fraction
                < canonical_cfg.satellite_min_z_overlap_fraction
            ):
                continue
            delta_z_um = (
                float(satellite["center_z"])
                - float(owner_candidate["center_z"])
            ) * float(pixel_depth_um)
            delta_y_um = (
                float(satellite["center_y"])
                - float(owner_candidate["center_y"])
            ) * float(pixel_height_um)
            delta_x_um = (
                float(satellite["center_x"])
                - float(owner_candidate["center_x"])
            ) * float(pixel_width_um)
            center_distance_um = math.sqrt(
                delta_z_um**2 + delta_y_um**2 + delta_x_um**2
            )
            satellite_radius_um = (
                3.0 * satellite_volume / (4.0 * math.pi)
            ) ** (1.0 / 3.0)
            owner_radius_um = (
                3.0 * owner_volume / (4.0 * math.pi)
            ) ** (1.0 / 3.0)
            radius_limit_um = (
                canonical_cfg.satellite_max_radius_sum_factor
                * (satellite_radius_um + owner_radius_um)
            )
            if center_distance_um > radius_limit_um:
                continue
            candidates.append(
                (
                    owner_candidate_id,
                    {
                        "satellite_id": satellite_id,
                        "owner_id": owner_candidate_id,
                        "satellite_source_object_ids": sorted(
                            satellite_sources
                        ),
                        "owner_source_object_ids": sorted(owner_sources),
                        "volume_ratio": float(volume_ratio),
                        "z_overlap_fraction": float(z_overlap_fraction),
                        "center_distance_um": float(center_distance_um),
                        "radius_limit_um": float(radius_limit_um),
                    },
                )
            )
        if len(candidates) == 1:
            owner_candidate_id, detail = candidates[0]
            satellite_aliases[satellite_id] = owner_candidate_id
            satellite_alias_details.append(detail)
    if grouped_extents is not None and satellite_aliases:
        effective_grouped_extents = np.asarray(
            grouped_extents,
            dtype=np.uint32,
        ).copy()
        for satellite_id, owner_id in sorted(satellite_aliases.items()):
            effective_grouped_extents[
                effective_grouped_extents == satellite_id
            ] = owner_id
        grouped_extents = effective_grouped_extents
    metrics["canonical_satellite_aliases"] = {
        str(satellite_id): owner_id
        for satellite_id, owner_id in sorted(satellite_aliases.items())
    }
    metrics["canonical_satellite_alias_details"] = satellite_alias_details
    identity_lineage = {
        int(key): value
        for key, value in identity_metrics.get("final_lineage", {}).items()
    }
    merge_decisions = [
        value
        for value in identity_metrics.get("merge_decisions", [])
        if isinstance(value, dict)
    ]
    merged_id_claim_counts: dict[int, int] = {}
    for lineage in identity_lineage.values():
        for merged_id in {
            int(value)
            for value in lineage.get("source_merged_ids", [])
        }:
            merged_id_claim_counts[merged_id] = (
                merged_id_claim_counts.get(merged_id, 0) + 1
            )
    axial_mapping = {
        int(old_id): int(new_id)
        for old_id, new_id in axial_metrics.get("id_mapping", {}).items()
    }
    inverse_axial_mapping = {
        final_id: old_id for old_id, final_id in axial_mapping.items()
    }
    retained_owner_claim_counts: dict[int, int] = {}
    for pre_axial_id in axial_mapping:
        lineage = identity_lineage.get(pre_axial_id)
        if lineage is None:
            continue
        owner_id = int(lineage.get("canonical_owner_id", 0))
        if owner_id > 0:
            retained_owner_claim_counts[owner_id] = (
                retained_owner_claim_counts.get(owner_id, 0) + 1
            )

    def is_accepted_merge_only(lineage: dict) -> bool:
        source_ids = sorted(
            {
                int(value)
                for value in lineage.get("source_astrocyte_ids", [])
            }
        )
        source_merged_ids = sorted(
            {
                int(value)
                for value in lineage.get("source_merged_ids", [])
            }
        )
        owner_id = int(lineage.get("canonical_owner_id", 0))
        if (
            len(source_ids) < 2
            or len(source_merged_ids) != 1
            or owner_id <= 0
            or merged_id_claim_counts.get(source_merged_ids[0], 0) != 1
        ):
            return False
        source_set = set(source_ids)
        for merge_decision in merge_decisions:
            if int(merge_decision.get("canonical_nucleus_id", 0)) != owner_id:
                continue
            decision_sources = {
                int(value)
                for value in merge_decision.get(
                    "source_astrocyte_ids",
                    [],
                )
            }
            if decision_sources != source_set:
                continue
            parent = {source_id: source_id for source_id in source_ids}

            def find(source_id: int) -> int:
                while parent[source_id] != source_id:
                    parent[source_id] = parent[parent[source_id]]
                    source_id = parent[source_id]
                return source_id

            for pair in merge_decision.get("accepted_pairs", []):
                if not isinstance(pair, (list, tuple)) or len(pair) != 2:
                    continue
                left_id, right_id = (int(pair[0]), int(pair[1]))
                if left_id not in source_set or right_id not in source_set:
                    continue
                left_root = find(left_id)
                right_root = find(right_id)
                if left_root != right_root:
                    parent[right_root] = left_root
            if len({find(source_id) for source_id in source_ids}) == 1:
                return True
        return False

    def component_count(mask: np.ndarray) -> int:
        return int(measure.label(mask, connectivity=2).max())

    def closest_masks(
        source: np.ndarray,
        target: np.ndarray,
    ) -> tuple[float, tuple[int, int], tuple[int, int]]:
        distance, nearest = ndi.distance_transform_edt(
            ~source,
            sampling=(float(pixel_height_um), float(pixel_width_um)),
            return_indices=True,
        )
        target_coordinates = np.argwhere(target)
        target_distances = distance[target]
        minimum_distance = float(np.min(target_distances))
        tied_positions = np.flatnonzero(
            np.isclose(
                target_distances,
                minimum_distance,
                rtol=0.0,
                atol=1e-12,
            )
        )
        source_center = np.mean(np.argwhere(source), axis=0)
        target_center = np.mean(target_coordinates, axis=0)
        tied_rows = []
        for position in tied_positions:
            target_y = int(target_coordinates[position, 0])
            target_x = int(target_coordinates[position, 1])
            source_y = int(nearest[0, target_y, target_x])
            source_x = int(nearest[1, target_y, target_x])
            center_cost = float(
                (target_y - target_center[0]) ** 2
                + (target_x - target_center[1]) ** 2
                + (source_y - source_center[0]) ** 2
                + (source_x - source_center[1]) ** 2
            )
            tied_rows.append(
                (
                    center_cost,
                    target_y,
                    target_x,
                    source_y,
                    source_x,
                )
            )
        _, target_y, target_x, source_y, source_x = min(tied_rows)
        center_distance = minimum_distance
        contact_step = math.hypot(
            float(pixel_height_um) if target_y != source_y else 0.0,
            float(pixel_width_um) if target_x != source_x else 0.0,
        )
        effective_gap = max(0.0, center_distance - contact_step)
        return (
            effective_gap,
            (source_y, source_x),
            (target_y, target_x),
        )

    for astrocyte_id in range(1, int(whole_labels.max()) + 1):
        own_whole = whole_labels == astrocyte_id
        original_soma = soma_labels == astrocyte_id
        pre_components = component_count(original_soma)
        decision = {
            "astrocyte_id": astrocyte_id,
            "pre_component_count": pre_components,
            "post_component_count": pre_components,
            "owner_group_id": 0,
            "resolved_nucleus_ids_in_whole": [],
            "foreign_resolved_nucleus_ids_in_whole": [],
            "local_foreign_veto_ids": [],
            "gap_um": [],
            "added_soma_px": 0,
            "approved": False,
            "approved_process_to_soma_px": 0,
            "process_fraction_before": float(
                (process_labels == astrocyte_id).sum()
                / max(int(own_whole.sum()), 1)
            ),
            "process_fraction_after": float(
                (process_labels == astrocyte_id).sum()
                / max(int(own_whole.sum()), 1)
            ),
            "status": (
                "unchanged_connected"
                if pre_components <= 1
                else "skipped_no_owner"
            ),
        }
        if profile == "neonatal":
            decision["bridge_method"] = "none"
            decision["owner_overlap_px_by_component"] = []
        if pre_components <= 1:
            metrics["decisions"].append(decision)
            continue
        metrics["disconnected_before_ids"].append(astrocyte_id)

        pre_axial_id = inverse_axial_mapping.get(astrocyte_id)
        lineage = (
            identity_lineage.get(pre_axial_id)
            if pre_axial_id is not None
            else None
        )
        if lineage is None:
            metrics["decisions"].append(decision)
            continue
        identity_changed = bool(lineage.get("identity_changed", False))
        accepted_merge_only = (
            identity_changed and is_accepted_merge_only(lineage)
        )
        if identity_changed and not accepted_merge_only:
            decision["status"] = "skipped_identity_changed"
            metrics["rejected_identity_split_ids"].append(astrocyte_id)
            metrics["decisions"].append(decision)
            continue
        owner_id = int(lineage.get("canonical_owner_id", 0))
        decision["owner_group_id"] = owner_id
        if (
            accepted_merge_only
            and retained_owner_claim_counts.get(owner_id, 0) != 1
        ):
            decision["status"] = "skipped_ambiguous_owner"
            metrics["rejected_ambiguous_owner_ids"].append(astrocyte_id)
            metrics["decisions"].append(decision)
            continue
        owner_record = records_by_id.get(owner_id)
        if owner_record is None or grouped_extents is None:
            metrics["decisions"].append(decision)
            continue
        if str(owner_record.get("identity_status", "")) != "resolved":
            decision["status"] = "skipped_ambiguous_owner"
            metrics["rejected_ambiguous_owner_ids"].append(astrocyte_id)
            metrics["decisions"].append(decision)
            continue
        if (
            not bool(owner_record.get("dapi_valid", False))
            or not bool(owner_record.get("accepted", False))
            or owner_id not in group_by_id
        ):
            metrics["decisions"].append(decision)
            continue

        owner_extent = np.asarray(
            grouped_extents == owner_id,
            dtype=bool,
        )
        labeled_soma = measure.label(original_soma, connectivity=2)
        owner_overlap_px_by_component = [
            int(
                (
                    (labeled_soma == component_id)
                    & owner_extent
                ).sum()
            )
            for component_id in range(1, pre_components + 1)
        ]
        if profile == "neonatal":
            decision["owner_overlap_px_by_component"] = (
                owner_overlap_px_by_component
            )
        every_island_supported = all(
            overlap_px > 0
            for overlap_px in owner_overlap_px_by_component
        )
        if not every_island_supported:
            decision["status"] = "skipped_island_not_owner_supported"
            metrics["decisions"].append(decision)
            continue
        use_owner_convex_hull = bool(
            profile == "neonatal"
            and accepted_merge_only
            and pre_components == 2
        )
        if (
            use_owner_convex_hull
            and any(
                overlap_px < minimum_owner_overlap_px
                for overlap_px in owner_overlap_px_by_component
            )
        ):
            decision["status"] = "skipped_island_not_owner_supported"
            metrics["rejected_path_ids"].append(astrocyte_id)
            metrics["decisions"].append(decision)
            continue

        resolved_ids = []
        for group_id, record in sorted(records_by_id.items()):
            if (
                not bool(record.get("dapi_valid", False))
                or str(record.get("identity_status", "")) != "resolved"
            ):
                continue
            overlap = int(
                (
                    own_whole
                    & np.asarray(grouped_extents == group_id, dtype=bool)
                ).sum()
            )
            if overlap >= minimum_owner_overlap_px:
                resolved_ids.append(group_id)
        decision["resolved_nucleus_ids_in_whole"] = resolved_ids
        decision["foreign_resolved_nucleus_ids_in_whole"] = [
            group_id for group_id in resolved_ids if group_id != owner_id
        ]
        if owner_id not in resolved_ids:
            decision["status"] = "skipped_owner_not_resolved_in_whole"
            metrics["decisions"].append(decision)
            continue

        if (
            decision["process_fraction_before"]
            < compartment_cfg.min_process_fraction
        ):
            decision["status"] = "skipped_process_fraction"
            metrics["decisions"].append(decision)
            continue

        owner_distance = ndi.distance_transform_edt(
            ~owner_extent,
            sampling=(float(pixel_height_um), float(pixel_width_um)),
        )
        allowed = (
            own_whole
            & (
                owner_distance
                <= float(
                    compartment_cfg.soma_trusted_core_nucleus_margin_um
                )
            )
        )
        trial = original_soma.copy()
        bridge = np.zeros_like(trial)
        failure_status: str | None = None
        gap_values: list[float] = []
        while component_count(trial) > 1:
            labels = measure.label(trial, connectivity=2)
            properties = list(measure.regionprops(labels))
            root_property = max(
                properties,
                key=lambda row: (int(row.area), -int(row.label)),
            )
            root = labels == int(root_property.label)
            closest: tuple[
                float,
                int,
                tuple[int, int],
                tuple[int, int],
            ] | None = None
            for prop in properties:
                if int(prop.label) == int(root_property.label):
                    continue
                candidate = labels == int(prop.label)
                gap, source_point, target_point = closest_masks(
                    root,
                    candidate,
                )
                row = (
                    gap,
                    int(prop.label),
                    source_point,
                    target_point,
                )
                if closest is None or row < closest:
                    closest = row
            if closest is None:
                failure_status = "skipped_no_safe_path"
                break
            gap, target_label, source_point, target_point = closest
            selected_target = labels == int(target_label)
            gap_values.append(float(gap))
            if gap > ownership_cfg.fragment_bridge_max_gap_um + 1e-12:
                failure_status = "skipped_gap_too_large"
                break
            target_foreign_ids = []
            for group_id in sorted(
                int(value)
                for value in np.unique(grouped_extents[selected_target])
                if int(value) > 0 and int(value) != owner_id
            ):
                target_overlap_px = int(
                    (
                        selected_target
                        & np.asarray(grouped_extents == group_id, dtype=bool)
                    ).sum()
                )
                record = records_by_id.get(group_id)
                incidental_unaccepted_resolved_overlap = bool(
                    record is not None
                    and not bool(record.get("accepted", False))
                    and bool(record.get("dapi_valid", False))
                    and str(record.get("identity_status", "")) == "resolved"
                    and target_overlap_px < minimum_owner_overlap_px
                )
                if not incidental_unaccepted_resolved_overlap:
                    target_foreign_ids.append(group_id)
            if target_foreign_ids:
                decision["local_foreign_veto_ids"] = sorted(
                    set(decision["local_foreign_veto_ids"])
                    | set(target_foreign_ids)
                )
                failure_status = "skipped_foreign_nucleus_near_connection"
                break
            if use_owner_convex_hull:
                pair_hull = morphology.convex_hull_image(original_soma)
                eligible_hull = (
                    pair_hull
                    & own_whole
                    & owner_extent
                    & (process_labels == astrocyte_id)
                    & ~original_soma
                )
                eligible_labels = measure.label(
                    eligible_hull,
                    connectivity=2,
                )
                joining_components = []
                for component_id in range(
                    1,
                    int(eligible_labels.max()) + 1,
                ):
                    component = eligible_labels == component_id
                    if component_count(original_soma | component) == 1:
                        joining_components.append(component)
                if len(joining_components) != 1:
                    failure_status = "skipped_no_safe_path"
                    break
                footprint = joining_components[0]
                if profile == "neonatal":
                    decision["bridge_method"] = (
                        "accepted_merge_owner_convex_hull"
                    )
            else:
                line = np.zeros_like(trial, dtype=np.uint8)
                cv2.line(
                    line,
                    (int(source_point[1]), int(source_point[0])),
                    (int(target_point[1]), int(target_point[0])),
                    color=1,
                    thickness=1,
                    lineType=cv2.LINE_8,
                )
                footprint = ndi.binary_dilation(
                    line.astype(bool),
                    structure=bridge_structure,
                )
                if profile == "neonatal":
                    decision["bridge_method"] = (
                        "calibration_aware_shortest_line"
                    )
            distance_to_connection_um = ndi.distance_transform_edt(
                ~footprint,
                sampling=(
                    float(pixel_height_um),
                    float(pixel_width_um),
                ),
            )
            local_foreign_ids = sorted(
                int(value)
                for value in np.unique(
                    grouped_extents[
                        distance_to_connection_um
                        <= ownership_cfg.unowned_barrier_radius_um
                    ]
                )
                if int(value) > 0 and int(value) != owner_id
            )
            if local_foreign_ids:
                decision["local_foreign_veto_ids"] = sorted(
                    set(decision["local_foreign_veto_ids"])
                    | set(local_foreign_ids)
                )
                failure_status = "skipped_foreign_nucleus_near_connection"
                break
            if use_owner_convex_hull:
                proposed = footprint
            else:
                if np.any(footprint & ~(allowed | trial)):
                    failure_status = "skipped_no_safe_path"
                    break
                proposed = footprint & allowed & ~trial
            if not proposed.any():
                failure_status = "skipped_no_safe_path"
                break
            before_count = component_count(trial)
            trial |= proposed
            bridge |= proposed
            if component_count(trial) >= before_count:
                failure_status = "skipped_no_safe_path"
                break
        decision["gap_um"] = [round(value, 9) for value in gap_values]
        if failure_status is not None:
            decision["status"] = failure_status
            if failure_status == "skipped_gap_too_large":
                metrics["rejected_gap_ids"].append(astrocyte_id)
            elif failure_status == "skipped_foreign_nucleus_near_connection":
                metrics["rejected_multiple_owner_ids"].append(astrocyte_id)
                metrics["rejected_local_foreign_ids"].append(astrocyte_id)
            else:
                metrics["rejected_path_ids"].append(astrocyte_id)
            metrics["decisions"].append(decision)
            continue
        if component_count(trial) != 1:
            decision["status"] = "skipped_no_safe_path"
            metrics["rejected_path_ids"].append(astrocyte_id)
            metrics["decisions"].append(decision)
            continue

        added_count = int(bridge.sum())
        maximum_relative_added_px = int(
            math.floor(
                SomaNuclearCompletionConfig().maximum_added_fraction_of_existing_soma
                * max(int(original_soma.sum()), 1)
            )
        )
        maximum_added_px = (
            max(1, maximum_relative_added_px)
            if use_owner_convex_hull
            else min(
                maximum_absolute_added_px,
                max(1, maximum_relative_added_px),
            )
        )
        if profile == "neonatal":
            decision["maximum_added_px"] = maximum_added_px
        if added_count > maximum_added_px:
            decision["status"] = "skipped_bridge_too_large"
            metrics["decisions"].append(decision)
            continue
        process_after = int(
            (process_labels == astrocyte_id).sum()
        ) - added_count
        process_fraction_after = process_after / max(
            int(own_whole.sum()),
            1,
        )
        decision["process_fraction_after"] = float(process_fraction_after)
        if process_fraction_after < compartment_cfg.min_process_fraction:
            decision["status"] = "skipped_process_fraction"
            metrics["decisions"].append(decision)
            continue
        if np.any(bridge & (process_labels != astrocyte_id)):
            raise RuntimeError(
                "Same-ID Soma Island Reconciliation attempted to convert pixels "
                "outside same-ID Processes"
            )
        output_soma[bridge] = astrocyte_id
        approved_labels[bridge] = astrocyte_id
        decision["post_component_count"] = 1
        decision["added_soma_px"] = added_count
        decision["approved"] = True
        decision["approved_process_to_soma_px"] = added_count
        decision["status"] = "bridged_same_owner_islands"
        metrics["bridged_ids"].append(astrocyte_id)
        metrics["added_soma_px"] += added_count
        metrics["removed_process_px"] += added_count
        metrics["approved_process_to_soma_px"] += added_count
        if use_owner_convex_hull:
            metrics["owner_convex_hull_ids"].append(astrocyte_id)
            metrics["owner_convex_hull_added_px"] += added_count
        metrics["decisions"].append(decision)

    output_process = np.where(
        (whole_labels > 0) & (output_soma == 0),
        whole_labels,
        0,
    ).astype(process_labels.dtype, copy=False)
    if not np.array_equal(whole_labels, frozen_whole_labels):
        raise RuntimeError(
            "Same-ID Soma Island Reconciliation changed Whole labels"
        )
    if np.any((output_soma > 0) & (output_process > 0)):
        raise RuntimeError(
            "Same-ID Soma Island Reconciliation overlapped Soma and Processes"
        )
    if not np.array_equal(
        whole_labels > 0,
        (output_soma > 0) | (output_process > 0),
    ):
        raise RuntimeError(
            "Same-ID Soma Island Reconciliation broke Whole = Soma union Processes"
        )
    if np.any((output_soma > 0) & (output_soma != whole_labels)):
        raise RuntimeError(
            "Same-ID Soma Island Reconciliation created a foreign Soma label"
        )
    if np.any(
        (approved_labels > 0)
        & (process_labels != approved_labels)
    ):
        raise RuntimeError(
            "Same-ID Soma Island Reconciliation approval map is not same-ID Processes"
        )
    metrics["disconnected_before_ids"] = sorted(
        metrics["disconnected_before_ids"]
    )
    metrics["bridged_ids"] = sorted(metrics["bridged_ids"])
    if profile == "neonatal":
        metrics["owner_convex_hull_ids"] = sorted(
            metrics["owner_convex_hull_ids"]
        )
    return whole_labels, output_soma, output_process, metrics

def finalize_compartment_geometry_and_metrics(
    whole_labels: np.ndarray,
    soma_labels: np.ndarray,
    process_labels: np.ndarray,
    metrics: dict,
    inventory: ValidatedNucleusAnchors | None,
    context: Neonatal3DContext | None,
    struct: np.ndarray,
    profile: str,
    pixel_width_um: float,
    pixel_height_um: float,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, dict]:
    """Apply post-selection guards and refresh synchronized cell metrics."""

    pre_finalization_whole_union = (whole_labels > 0).copy()
    pre_finalization_whole_union.flags.writeable = False
    whole_labels, soma_labels, process_labels, identity_metrics = (
        apply_canonical_identity_reconciliation(
            whole_labels,
            soma_labels,
            process_labels,
            struct,
            inventory,
            pixel_width_um,
            pixel_height_um,
            context.pixel_depth_um if context is not None else None,
            profile,
        )
    )
    whole_labels, soma_labels, process_labels, axial_metrics = (
        apply_axial_truncation_guard(
            whole_labels,
            soma_labels,
            process_labels,
            inventory,
            context,
            pixel_width_um,
            pixel_height_um,
        )
    )
    if profile == "mature":
        whole_labels, soma_labels, process_labels, projection_metrics = (
            apply_projected_foreign_soma_guard(
                whole_labels,
                soma_labels,
                process_labels,
                inventory,
                pixel_width_um,
                pixel_height_um,
                context.pixel_depth_um if context is not None else None,
            )
        )
    else:
        projection_metrics = {
            "enabled": False,
            "status": "skipped_for_dense_neonatal_projection",
            "method": (
                "projection foreign-soma exclusion is restricted to mature samples; "
                "neonatal identity uses explicit canonical nucleus reconciliation"
            ),
            "evaluated_cell_count": int(whole_labels.max()),
            "changed_cell_count": 0,
            "removed_area_px": 0,
            "decisions": [],
        }
    # Preserve the within-Whole owner-extent completion result first. Canonical
    # owner completion then adds only canonical-owner pixels, so cells outside
    # the eligibility gates remain unchanged.
    whole_labels, soma_labels, process_labels, within_whole_completion_metrics = (
        _complete_soma_within_whole_owner_extent(
            whole_labels,
            soma_labels,
            process_labels,
            inventory,
            pixel_width_um,
            pixel_height_um,
            context.pixel_depth_um if context is not None else None,
        )
    )
    owner_assignments, owner_assignment_failures = (
        resolve_canonical_owner_assignments(
            identity_metrics,
            axial_metrics,
            inventory,
            int(whole_labels.max()),
        )
    )
    pre_canonical_owner_extent_completion_union = (whole_labels > 0).copy()
    pre_canonical_owner_extent_completion_soma_labels = soma_labels.copy()
    (
        whole_labels,
        soma_labels,
        process_labels,
        soma_completion_metrics,
        approved_outside_pre_finalization_whole_mask,
    ) = (
        complete_soma_to_owner_nuclear_extent(
            whole_labels,
            soma_labels,
            process_labels,
            inventory,
            pixel_width_um,
            pixel_height_um,
            context.pixel_depth_um if context is not None else None,
            owner_assignments,
            owner_assignment_failures,
            pre_finalization_whole_union,
        )
    )
    canonical_owner_extent_approved_owner_extent_labels = np.where(
        (pre_canonical_owner_extent_completion_soma_labels == 0)
        & (soma_labels > 0),
        soma_labels,
        0,
    ).astype(np.uint16)
    if (
        int((canonical_owner_extent_approved_owner_extent_labels > 0).sum())
        != int(
            soma_completion_metrics[
                "approved_owner_extent_added_px"
            ]
        )
    ):
        raise RuntimeError(
            "Canonical Owner Nuclear-Extent Completion approval labels differ "
            "from recorded approved pixels"
        )
    pre_same_id_soma_reconciliation_whole_labels = whole_labels.copy()
    pre_same_id_soma_reconciliation_soma_labels = soma_labels.copy()
    pre_same_id_soma_reconciliation_process_labels = process_labels.copy()
    (
        whole_labels,
        soma_labels,
        process_labels,
        soma_reconciliation_metrics,
    ) = reconcile_same_id_disconnected_soma(
        whole_labels,
        soma_labels,
        process_labels,
        inventory,
        identity_metrics,
        axial_metrics,
        profile,
        pixel_width_um,
        pixel_height_um,
        context.pixel_depth_um if context is not None else None,
    )
    same_id_soma_reconciliation_approved_process_to_soma_labels = np.where(
        (pre_same_id_soma_reconciliation_soma_labels == 0)
        & (soma_labels > 0),
        soma_labels,
        0,
    ).astype(np.uint16)
    if not np.array_equal(
        whole_labels,
        pre_same_id_soma_reconciliation_whole_labels,
    ):
        raise RuntimeError(
            "Same-ID Soma Island Reconciliation changed Whole labels"
        )
    if not np.array_equal(
        same_id_soma_reconciliation_approved_process_to_soma_labels,
        np.where(
            (pre_same_id_soma_reconciliation_process_labels > 0)
            & (process_labels == 0),
            pre_same_id_soma_reconciliation_process_labels,
            0,
        ).astype(np.uint16),
    ):
        raise RuntimeError(
            "Same-ID Soma Island Reconciliation Soma additions differ from "
            "Processes removals"
        )
    if (
        int(
            (
                same_id_soma_reconciliation_approved_process_to_soma_labels > 0
            ).sum()
        )
        != int(
            soma_reconciliation_metrics[
                "approved_process_to_soma_px"
            ]
        )
    ):
        raise RuntimeError(
            "Same-ID Soma Island Reconciliation approval labels differ from "
            "recorded approved pixels"
        )
    if not np.array_equal(
        (whole_labels > 0) & ~pre_canonical_owner_extent_completion_union,
        approved_outside_pre_finalization_whole_mask,
    ):
        raise RuntimeError(
            "Canonical Owner Nuclear-Extent Completion final Whole expansion "
            "differs from its approved mask"
        )
    if not np.array_equal(
        (whole_labels > 0) & ~pre_finalization_whole_union,
        approved_outside_pre_finalization_whole_mask,
    ):
        raise RuntimeError(
            "Final Whole expansion outside the pre-finalization Whole is not "
            "exclusively from Canonical Owner Nuclear-Extent Completion"
        )
    if np.any((soma_labels > 0) & (process_labels > 0)):
        raise RuntimeError("Final geometry overlapped Soma and Processes")
    if not np.array_equal(
        whole_labels > 0,
        (soma_labels > 0) | (process_labels > 0),
    ):
        raise RuntimeError("Final geometry violates Whole = Soma union Processes")

    original_rows = {
        int(row["astrocyte_id"]): dict(row) for row in metrics.get("per_cell", [])
    }
    axial_mapping = {
        int(old_id): int(new_id)
        for old_id, new_id in axial_metrics.get("id_mapping", {}).items()
    }
    inverse_axial_mapping = {new_id: old_id for old_id, new_id in axial_mapping.items()}
    identity_lineage = {
        int(final_id): row
        for final_id, row in identity_metrics.get("final_lineage", {}).items()
    }
    refreshed_rows = []
    for astrocyte_id in range(1, int(whole_labels.max()) + 1):
        pre_axial_id = inverse_axial_mapping.get(astrocyte_id, astrocyte_id)
        lineage = identity_lineage.get(
            pre_axial_id,
            {"source_astrocyte_ids": [pre_axial_id]},
        )
        source_ids = [int(value) for value in lineage["source_astrocyte_ids"]]
        source_row = next(
            (original_rows[value] for value in source_ids if value in original_rows),
            {"astrocyte_id": astrocyte_id},
        )
        row = dict(source_row)
        row["astrocyte_id"] = astrocyte_id
        row["source_astrocyte_ids_before_identity_reconciliation"] = source_ids
        row["canonical_nucleus_id"] = int(lineage.get("canonical_owner_id", 0))
        row["identity_reconciled"] = bool(lineage.get("identity_changed", False))
        whole_area = int((whole_labels == astrocyte_id).sum())
        soma_area = int((soma_labels == astrocyte_id).sum())
        process_area = int((process_labels == astrocyte_id).sum())
        row["whole_area_px"] = whole_area
        row["soma_area_px"] = soma_area
        row["process_area_px"] = process_area
        row["soma_fraction"] = soma_area / max(whole_area, 1)
        row["process_fraction"] = process_area / max(whole_area, 1)
        refreshed_rows.append(row)
    metrics["per_cell"] = refreshed_rows
    metrics["canonical_identity_reconciliation"] = identity_metrics
    metrics[
        "within_whole_soma_nuclear_extent_completion"
    ] = within_whole_completion_metrics
    metrics["soma_nuclear_extent_completion"] = soma_completion_metrics
    metrics[
        "_canonical_owner_extent_pre_finalization_whole_union_mask"
    ] = pre_finalization_whole_union
    metrics[
        "_canonical_owner_extent_approved_outside_pre_finalization_whole_mask"
    ] = approved_outside_pre_finalization_whole_mask
    metrics[
        "_canonical_owner_extent_approved_owner_extent_labels"
    ] = canonical_owner_extent_approved_owner_extent_labels
    if (
        inventory is not None
        and inventory.nucleus_instance_core_labels_2d is not None
        and inventory.nucleus_instance_extent_labels_2d is not None
    ):
        metrics[
            "_canonical_nucleus_instance_core_labels_2d"
        ] = np.asarray(
            inventory.nucleus_instance_core_labels_2d,
            dtype=np.uint32,
        )
        metrics[
            "_canonical_nucleus_instance_extent_labels_2d"
        ] = np.asarray(
            inventory.nucleus_instance_extent_labels_2d,
            dtype=np.uint32,
        )
    metrics[
        "same_id_disconnected_soma_reconciliation"
    ] = soma_reconciliation_metrics
    metrics[
        "_same_id_soma_reconciliation_approved_process_to_soma_labels"
    ] = same_id_soma_reconciliation_approved_process_to_soma_labels
    metrics["axial_truncation_guard"] = axial_metrics
    metrics["projected_foreign_soma_guard"] = projection_metrics
    metrics["roi_count"] = int(whole_labels.max())
    metrics["whole_area_px"] = int((whole_labels > 0).sum())
    metrics["soma_area_px"] = int((soma_labels > 0).sum())
    metrics["process_area_px"] = int((process_labels > 0).sum())
    metrics["soma_area_fraction"] = (
        metrics["soma_area_px"] / max(metrics["whole_area_px"], 1)
    )
    metrics["process_area_fraction"] = (
        metrics["process_area_px"] / max(metrics["whole_area_px"], 1)
    )
    return whole_labels, soma_labels, process_labels, metrics


def prune_soma_to_trusted_core_shell(
    soma: np.ndarray,
    trusted_core: np.ndarray,
    anchor_seeds: np.ndarray,
    pixel_width_um: float,
    pixel_height_um: float,
    max_shell_um: float,
    min_soma_area_px: int,
) -> tuple[np.ndarray, bool]:
    """Conservatively remove thin Soma extensions without adding any pixels."""

    if not soma.any() or not trusted_core.any() or not anchor_seeds.any():
        return soma, False
    trusted_core = trusted_core & soma
    anchor_seeds = anchor_seeds & soma
    if not trusted_core.any() or not anchor_seeds.any():
        return soma, False
    distance_um = ndi.distance_transform_edt(
        ~trusted_core,
        sampling=(pixel_height_um, pixel_width_um),
    )
    allowed = soma & (distance_um <= max_shell_um)
    allowed |= anchor_seeds
    pruned = ndi.binary_propagation(
        anchor_seeds,
        structure=np.ones((3, 3), dtype=bool),
        mask=allowed,
    ).astype(bool)
    if not np.all(pruned[anchor_seeds]):
        return soma, False
    if int(pruned.sum()) < min_soma_area_px:
        return soma, False
    if np.any(pruned & ~soma):
        raise RuntimeError("Soma core-shell pruning added pixels outside the original Soma")
    return pruned, bool(np.any(soma & ~pruned))

def compartment_config_for_profile(profile: str) -> CompartmentConfig:
    if profile == "mature":
        return CompartmentConfig()
    if profile != "neonatal":
        raise ValueError(f"Unknown astrocyte profile: {profile}")
    return replace(
        CompartmentConfig(),
        soma_zone_max_um=4.60,
        soma_zone_scale_process_rich=1.95,
        soma_zone_scale_compact=2.45,
        thickness_fraction_process_rich=0.40,
        thickness_fraction_compact=0.30,
        structural_percentile_process_rich=62.0,
        structural_percentile_compact=50.0,
        fallback_soma_radius_um=1.45,
        max_soma_fraction=0.64,
        primary_anchor_min_score=0.60,
        primary_anchor_min_thickness_support=0.56,
        primary_anchor_min_structural_support=0.32,
        primary_anchor_min_overlap_fraction=0.32,
        multi_anchor_min_score=0.58,
        multi_anchor_max_score_delta=0.18,
        multi_anchor_min_thickness_support=0.60,
        multi_anchor_min_structural_support=0.33,
        multi_anchor_min_overlap_fraction=0.48,
        soma_anchor_min_separation_um=3.8,
        soma_part_max_axis_ratio=4.50,
        soma_core_shell_max_um=0.75,
        soma_trusted_core_radius_scale=1.45,
        soma_trusted_core_max_um=3.20,
        soma_trusted_core_nucleus_margin_um=0.60,
        soma_nucleus_shape_preserving=True,
        instance_split_min_anchor_score=0.66,
        instance_split_min_anchor_separation_um=4.0,
        instance_split_min_child_area_um2=12.0,
        instance_split_min_child_fraction=0.10,
        instance_split_max_neck_core_ratio=0.82,
        instance_split_max_boundary_structural_ratio=0.72,
        instance_split_max_markers=4,
        instance_split_strategy="neonatal_multi",
    )

def split_astrocyte_compartments(
    whole_mask: np.ndarray,
    dapi_projection: np.ndarray,
    struct: np.ndarray,
    cellpose_mask: np.ndarray,
    pixel_width_um: float,
    pixel_height_um: float,
    config: CompartmentConfig | None = None,
    validated_anchors: ValidatedNucleusAnchors | None = None,
    ownership_inventory: ValidatedNucleusAnchors | None = None,
    ownership_pixel_depth_um: float | None = None,
    ownership_profile: str = "mature",
) -> tuple[np.ndarray, np.ndarray, np.ndarray, dict]:
    """Partition each Whole ROI into its soma union and exact process complement."""

    cfg = config or CompartmentConfig()
    if not whole_mask.any():
        raise ValueError("Cannot split compartments because the Whole ROI mask is empty")

    mean_pixel_um = max(1e-4, math.sqrt(pixel_width_um * pixel_height_um))
    pixel_area_um2 = pixel_width_um * pixel_height_um
    nuclei, nuclei_extent, dapi_norm, dapi_extent_metrics = dapi_nuclei_core_and_extent(
        dapi_projection,
        mean_pixel_um,
        cfg,
    )
    preserved_object_nuclei_labels = None
    if validated_anchors is not None:
        accepted_core = np.asarray(
            validated_anchors.accepted_core_mask_2d,
            dtype=bool,
        )
        accepted_extent = np.asarray(
            validated_anchors.accepted_extent_mask_2d,
            dtype=bool,
        )
        if accepted_core.shape != nuclei.shape or accepted_extent.shape != nuclei.shape:
            raise ValueError(
                "Validated neonatal nucleus masks do not match the 2D compartment geometry"
            )
        pre_validation_core_count = int(measure.label(nuclei, connectivity=2).max())
        nuclei &= accepted_core
        nuclei_extent &= accepted_extent
        nuclei_extent |= nuclei
        dapi_extent_metrics["pre_3d_validation_core_count"] = pre_validation_core_count
        dapi_extent_metrics["post_3d_validation_core_count"] = int(
            measure.label(nuclei, connectivity=2).max()
        )
        dapi_extent_metrics["strict_core_px_after_3d_validation"] = int(nuclei.sum())
        dapi_extent_metrics["extent_px_after_3d_validation"] = int(nuclei_extent.sum())
        if validated_anchors.object_core_labels_2d is not None:
            object_core_labels = np.asarray(
                validated_anchors.object_core_labels_2d,
                dtype=np.uint32,
            )
            preserved_object_nuclei_labels = np.where(
                np.isin(object_core_labels, validated_anchors.accepted_object_ids),
                object_core_labels,
                0,
            ).astype(np.uint32)
            dapi_extent_metrics["preserved_3d_object_id_count"] = len(
                np.unique(preserved_object_nuclei_labels)
            ) - int(np.any(preserved_object_nuclei_labels == 0))
    refined_whole_mask, branch_gap_metrics = restore_low_support_branch_gaps(
        whole_mask,
        dapi_projection,
        struct,
        pixel_width_um,
        pixel_height_um,
        cfg,
        nuclei_mask=nuclei,
    )
    link_radius_px = max(3, int(round(cfg.nucleus_link_um / mean_pixel_um)))
    min_zone_px = max(link_radius_px + 2, int(round(cfg.soma_zone_min_um / mean_pixel_um)))
    max_zone_px = max(min_zone_px, int(round(cfg.soma_zone_max_um / mean_pixel_um)))
    fallback_radius_px = max(4, int(round(cfg.fallback_soma_radius_um / mean_pixel_um)))
    min_soma_area_px = max(40, int(round(cfg.min_soma_area_um2 / pixel_area_um2)))

    nuclei_labels = (
        preserved_object_nuclei_labels
        if preserved_object_nuclei_labels is not None
        else measure.label(nuclei, connectivity=2)
    )
    validated_grouped_extent_labels = None
    validated_group_extent_areas = None
    validated_group_by_id: dict[int, dict] = {}
    validated_anchor_minimum_overlap_px = 1
    if (
        validated_anchors is not None
        and ownership_pixel_depth_um is not None
        and ownership_pixel_depth_um > 0
        and validated_anchors.object_extent_labels_2d is not None
    ):
        ownership_config = NucleusOwnershipConfig()
        validated_groups = group_inventory_nucleus_objects(
            validated_anchors,
            pixel_width_um,
            pixel_height_um,
            ownership_pixel_depth_um,
            ownership_config,
        )
        object_extent_labels = np.asarray(
            validated_anchors.object_extent_labels_2d,
            dtype=np.uint32,
        )
        object_to_group = np.zeros(
            int(object_extent_labels.max()) + 1,
            dtype=np.uint32,
        )
        for group in validated_groups:
            group_id = int(group["group_id"])
            validated_group_by_id[group_id] = group
            for object_id in group["object_ids"]:
                object_to_group[int(object_id)] = group_id
        validated_grouped_extent_labels = object_to_group[object_extent_labels]
        validated_group_extent_areas = np.bincount(
            validated_grouped_extent_labels.ravel(),
            minlength=len(object_to_group),
        )
        validated_anchor_minimum_overlap_px = max(
            1,
            int(
                math.ceil(
                    ownership_config.owner_min_overlap_um2 / pixel_area_um2
                )
            ),
        )
    if nuclei.any():
        nucleus_distance, nearest_indices = ndi.distance_transform_edt(
            ~nuclei,
            return_indices=True,
        )
        nearest_nucleus_labels = nuclei_labels[nearest_indices[0], nearest_indices[1]]
    else:
        nucleus_distance = np.full(nuclei.shape, np.inf, dtype=np.float32)
        nearest_nucleus_labels = np.zeros(nuclei.shape, dtype=np.int32)
    if cfg.instance_split_strategy == "pairwise_soma_anchor_split":
        instance_splitter = split_touching_whole_instances
    elif cfg.instance_split_strategy == "neonatal_multi":
        instance_splitter = split_touching_whole_instances_multi
    else:
        raise ValueError(f"Unknown instance split strategy: {cfg.instance_split_strategy}")
    labels, instance_metrics = instance_splitter(
        refined_whole_mask,
        nuclei_labels,
        nearest_nucleus_labels,
        nucleus_distance,
        struct,
        cellpose_mask,
        mean_pixel_um,
        pixel_area_um2,
        link_radius_px,
        cfg,
    )
    labels, nucleus_ownership_metrics = apply_nucleus_ownership_guard(
        labels,
        struct,
        ownership_inventory,
        pixel_width_um,
        pixel_height_um,
        ownership_pixel_depth_um,
        ownership_profile,
    )
    validated_groups_by_instance_id: dict[int, set[int]] = {}
    if validated_grouped_extent_labels is not None:
        explicit_group_owner: dict[int, int] = {}
        for decision in nucleus_ownership_metrics.get("decisions", []):
            output_ids = [
                int(value) for value in decision.get("output_instance_ids", [])
            ]
            if not output_ids:
                continue
            owner_group_id = int(decision.get("owner_group_id", 0))
            if owner_group_id > 0:
                explicit_group_owner[owner_group_id] = output_ids[0]
            accepted_foreign_group_ids = [
                int(row["group_id"])
                for row in decision.get("foreign_groups", [])
                if bool(row.get("accepted"))
            ]
            for group_id, output_id in zip(
                accepted_foreign_group_ids,
                output_ids[1:],
            ):
                explicit_group_owner[group_id] = output_id
        for group_id, group in validated_group_by_id.items():
            if not bool(group["accepted"]):
                continue
            if group_id in explicit_group_owner:
                validated_groups_by_instance_id.setdefault(
                    explicit_group_owner[group_id],
                    set(),
                ).add(int(group_id))
                continue
            overlapping_labels = labels[validated_grouped_extent_labels == group_id]
            overlapping_labels = overlapping_labels[overlapping_labels > 0]
            if overlapping_labels.size == 0:
                continue
            label_ids, label_counts = np.unique(
                overlapping_labels,
                return_counts=True,
            )
            winner_index = int(np.argmax(label_counts))
            if int(label_counts[winner_index]) < validated_anchor_minimum_overlap_px:
                continue
            group_extent_area_px = int(validated_group_extent_areas[group_id])
            if (
                int(label_counts[winner_index]) / max(group_extent_area_px, 1)
                < ownership_config.accepted_min_extent_overlap_fraction
            ):
                continue
            owner_label = int(label_ids[winner_index])
            validated_groups_by_instance_id.setdefault(owner_label, set()).add(
                int(group_id)
            )
    ownership_id_mapping = {
        int(old_id): [int(value) for value in new_ids]
        for old_id, new_ids in nucleus_ownership_metrics.get(
            "input_to_output_ids",
            {},
        ).items()
    }
    if ownership_id_mapping:
        for detail in instance_metrics.get("split_components", []):
            pre_guard_ids = [int(value) for value in detail.get("new_astrocyte_ids", [])]
            detail["pre_ownership_guard_new_astrocyte_ids"] = pre_guard_ids
            detail["new_astrocyte_ids"] = [
                mapped_id
                for old_id in pre_guard_ids
                for mapped_id in ownership_id_mapping.get(old_id, [])
            ]
        for decision in instance_metrics.get("component_decisions", []):
            pre_guard_ids = [
                int(value) for value in decision.get("output_astrocyte_ids", [])
            ]
            decision["pre_ownership_guard_output_astrocyte_ids"] = pre_guard_ids
            decision["output_astrocyte_ids"] = [
                mapped_id
                for old_id in pre_guard_ids
                for mapped_id in ownership_id_mapping.get(old_id, [])
            ]
    roi_count = int(labels.max())
    soma_labels = np.zeros_like(labels, dtype=np.uint16)
    process_labels = np.zeros_like(labels, dtype=np.uint16)
    per_cell: list[dict] = []
    fallback_count = 0
    ambiguous_count = 0
    no_dapi_count = 0
    multi_soma_roi_count = 0
    total_soma_anchor_count = 0
    rejected_soma_anchor_count = 0
    dapi_extent_satellite_component_count = 0
    dapi_extent_satellite_px = 0
    component_properties = {prop.label: prop for prop in measure.regionprops(labels)}
    crop_padding = max_zone_px + link_radius_px + 4

    for astrocyte_id in range(1, roi_count + 1):
        prop = component_properties[astrocyte_id]
        min_row, min_col, max_row, max_col = prop.bbox
        row0 = max(0, min_row - crop_padding)
        col0 = max(0, min_col - crop_padding)
        row1 = min(labels.shape[0], max_row + crop_padding)
        col1 = min(labels.shape[1], max_col + crop_padding)
        crop = np.s_[row0:row1, col0:col1]
        component = labels[crop] == astrocyte_id
        local_struct = struct[crop]
        local_cellpose = cellpose_mask[crop]
        local_nuclei_labels = nuclei_labels[crop]
        local_nuclei_extent = nuclei_extent[crop]
        local_dapi_norm = dapi_norm[crop]
        local_nearest_labels = nearest_nucleus_labels[crop]
        local_nucleus_distance = nucleus_distance[crop]
        local_validated_grouped_extent_labels = (
            validated_grouped_extent_labels[crop]
            if validated_grouped_extent_labels is not None
            else None
        )
        component_area = int(component.sum())
        distance = ndi.distance_transform_edt(component)
        scored_nuclei, ambiguous = score_nuclei_for_component(
            component,
            local_nearest_labels,
            local_nucleus_distance,
            distance,
            local_struct,
            local_cellpose,
            link_radius_px,
            cfg.ambiguity_score_delta,
        )
        if local_validated_grouped_extent_labels is not None:
            anchor_groups = select_validated_soma_anchor_groups(
                scored_nuclei,
                component,
                local_nuclei_labels,
                local_validated_grouped_extent_labels,
                validated_group_by_id,
                validated_anchor_minimum_overlap_px,
                validated_groups_by_instance_id.get(astrocyte_id, set()),
            )
        else:
            anchor_groups = select_soma_anchor_groups(
                scored_nuclei,
                mean_pixel_um,
                cfg,
            )
        nucleus_candidates = len(scored_nuclei)
        nucleus_score = float(scored_nuclei[0]["score"]) if scored_nuclei else 0.0
        ambiguous_count += int(ambiguous)

        soma_parts: list[np.ndarray] = []
        anchor_details: list[dict] = []
        fallback_used = False
        rejected_anchor_count = 0
        distance_scale = max(float(np.percentile(distance[component], 99.0)), 1.0)

        for anchor in anchor_groups:
            nucleus_ids = anchor["nucleus_ids"]
            if nucleus_ids:
                selected_nucleus = np.isin(local_nuclei_labels, nucleus_ids)
                selected_voronoi = np.isin(local_nearest_labels, nucleus_ids)
                selected_nucleus_distance_um = ndi.distance_transform_edt(
                    ~selected_nucleus,
                    sampling=(pixel_height_um, pixel_width_um),
                )
                selected_nucleus_extent = (
                    local_nuclei_extent
                    & selected_voronoi
                    & (selected_nucleus_distance_um <= cfg.dapi_extent_max_expand_um)
                )
                selected_nucleus_extent |= selected_nucleus
                selected_nucleus_extent = morphology.binary_closing(
                    selected_nucleus_extent,
                    footprint=morphology.disk(1),
                )
                (
                    selected_nucleus_extent,
                    removed_extent_components,
                    removed_extent_px,
                ) = retain_primary_anchor_extent(
                    selected_nucleus_extent,
                    selected_nucleus,
                    component,
                )
                dapi_extent_satellite_component_count += removed_extent_components
                dapi_extent_satellite_px += removed_extent_px
                selected_nucleus_distance = ndi.distance_transform_edt(
                    ~selected_nucleus_extent
                )
                selected_nucleus_extent_distance_um = ndi.distance_transform_edt(
                    ~selected_nucleus_extent,
                    sampling=(pixel_height_um, pixel_width_um),
                )
                search_region = component & (selected_nucleus_distance <= link_radius_px)
            else:
                selected_nucleus = np.zeros_like(component, dtype=bool)
                selected_nucleus_extent = np.zeros_like(component, dtype=bool)
                selected_nucleus_distance = np.full(component.shape, np.inf, dtype=np.float32)
                selected_nucleus_extent_distance_um = np.full(
                    component.shape,
                    np.inf,
                    dtype=np.float32,
                )
                search_region = component.copy()
            if not search_region.any():
                search_region = component.copy()

            nucleus_proximity = (
                np.exp(-np.square(selected_nucleus_distance / max(link_radius_px, 1)))
                if nucleus_ids
                else np.zeros_like(distance, dtype=np.float32)
            )
            seed_score = (
                0.64 * np.clip(distance / distance_scale, 0, 1)
                + 0.23 * local_struct
                + 0.05 * local_cellpose.astype(np.float32)
                + 0.08 * nucleus_proximity
            )
            seed_score = np.where(search_region, seed_score, -np.inf)
            seed_y, seed_x = np.unravel_index(int(np.argmax(seed_score)), seed_score.shape)
            seed_point = np.zeros_like(component, dtype=bool)
            seed_point[seed_y, seed_x] = True

            core_neighborhood_radius = max(
                link_radius_px,
                int(round(0.75 / mean_pixel_um)),
            )
            core_neighborhood = component & circular_mask(
                component.shape,
                seed_y,
                seed_x,
                core_neighborhood_radius,
            )
            core_peak_px = max(
                float(np.percentile(distance[core_neighborhood], 90.0)),
                0.55 / mean_pixel_um,
            )
            thin_cut = max(1.5, 0.42 * core_peak_px)
            thin_fraction = float((distance[component] <= thin_cut).mean())
            process_richness = float(np.clip((thin_fraction - 0.20) / 0.55, 0, 1))

            zone_scale = (
                cfg.soma_zone_scale_compact
                + process_richness
                * (cfg.soma_zone_scale_process_rich - cfg.soma_zone_scale_compact)
            )
            zone_radius_px = int(np.clip(round(zone_scale * core_peak_px), min_zone_px, max_zone_px))
            thickness_fraction = (
                cfg.thickness_fraction_compact
                + process_richness
                * (cfg.thickness_fraction_process_rich - cfg.thickness_fraction_compact)
            )
            structural_percentile = (
                cfg.structural_percentile_compact
                + process_richness
                * (cfg.structural_percentile_process_rich - cfg.structural_percentile_compact)
            )
            thickness_cut = max(0.32 / mean_pixel_um, core_peak_px * thickness_fraction)
            structural_cut = float(np.percentile(local_struct[component], structural_percentile))

            if nucleus_ids:
                soma_zone = selected_nucleus_distance <= zone_radius_px
            else:
                soma_zone = circular_mask(
                    component.shape,
                    seed_y,
                    seed_x,
                    zone_radius_px,
                )
            secondary_thickness_cut = max(1.5, 0.14 * core_peak_px)
            soma_domain = component & soma_zone & (
                (distance >= thickness_cut)
                | ((local_struct >= structural_cut) & (distance >= secondary_thickness_cut))
            )
            seed_radius_px = max(2, min(fallback_radius_px, int(round(0.35 * core_peak_px))))
            soma_seed = component & circular_mask(
                component.shape,
                seed_y,
                seed_x,
                seed_radius_px,
            )
            required_nucleus = component & selected_nucleus_extent
            soma_domain |= soma_seed | required_nucleus
            soma_seed |= required_nucleus
            soma_part = ndi.binary_propagation(
                soma_seed,
                structure=np.ones((3, 3), dtype=bool),
                mask=soma_domain,
            ).astype(bool)
            soma_part = morphology.binary_closing(soma_part, footprint=morphology.disk(2)) & component
            soma_part = morphology.remove_small_holes(
                soma_part,
                area_threshold=max(16, int(round(0.35 / pixel_area_um2))),
            ) & component

            part_fallback = False
            soma_fraction = float(soma_part.sum()) / max(component_area, 1)
            if int(soma_part.sum()) < min_soma_area_px or soma_fraction > cfg.max_soma_fraction:
                part_fallback = True
                if nucleus_ids and cfg.soma_nucleus_shape_preserving:
                    fallback_zone = (
                        selected_nucleus_extent_distance_um
                        <= cfg.fallback_soma_radius_um
                    )
                    fallback_domain = component & fallback_zone & (
                        (distance >= secondary_thickness_cut)
                        | (local_struct >= structural_cut)
                    )
                else:
                    fallback_zone = circular_mask(
                        component.shape,
                        seed_y,
                        seed_x,
                        fallback_radius_px,
                    )
                    fallback_domain = component & fallback_zone
                fallback_domain |= component & selected_nucleus_extent
                fallback_soma = ndi.binary_propagation(
                    seed_point | required_nucleus,
                    structure=np.ones((3, 3), dtype=bool),
                    mask=fallback_domain,
                ).astype(bool)
                fallback_soma = morphology.binary_closing(
                    fallback_soma,
                    footprint=morphology.disk(2),
                ) & component
                if fallback_soma.any():
                    soma_part = fallback_soma
            part_properties = measure.regionprops(measure.label(soma_part, connectivity=2))
            if part_properties:
                soma_property = max(part_properties, key=lambda item: item.area)
                axis_ratio = float(soma_property.major_axis_length) / max(
                    float(soma_property.minor_axis_length),
                    1e-6,
                )
            else:
                axis_ratio = math.inf
            core_radius_um = core_peak_px * mean_pixel_um
            if (
                not soma_part.any()
                or core_radius_um < cfg.soma_part_min_core_radius_um
                or axis_ratio > cfg.soma_part_max_axis_ratio
            ):
                rejected_anchor_count += 1
                continue
            fallback_used |= part_fallback
            soma_parts.append(soma_part)
            trusted_core_radius_um = float(
                np.clip(
                    cfg.soma_trusted_core_radius_scale * core_peak_px * mean_pixel_um,
                    cfg.soma_trusted_core_min_um,
                    cfg.soma_trusted_core_max_um,
                )
            )
            trusted_core_radius_px = max(
                seed_radius_px,
                int(round(trusted_core_radius_um / mean_pixel_um)),
            )
            if nucleus_ids and cfg.soma_nucleus_shape_preserving:
                trusted_core_zone = (
                    selected_nucleus_extent_distance_um <= trusted_core_radius_um
                )
            else:
                trusted_core_zone = circular_mask(
                    component.shape,
                    seed_y,
                    seed_x,
                    trusted_core_radius_px,
                )
            nucleus_trusted_zone = (
                selected_nucleus_extent_distance_um
                <= cfg.soma_trusted_core_nucleus_margin_um
            )
            trusted_core = soma_part & trusted_core_zone & nucleus_trusted_zone & (
                (distance >= thickness_cut) | soma_seed
            )
            trusted_core |= required_nucleus
            trusted_core |= seed_point
            anchor_details.append(
                {
                    "seed_y": seed_y,
                    "seed_x": seed_x,
                    "selected_nucleus": selected_nucleus,
                    "selected_nucleus_extent": required_nucleus,
                    "required_nucleus_px": int(required_nucleus.sum()),
                    "required_nucleus_mean": round(
                        float(local_dapi_norm[required_nucleus].mean())
                        if required_nucleus.any()
                        else 0.0,
                        6,
                    ),
                    "core_peak_px": core_peak_px,
                    "thin_fraction": thin_fraction,
                    "process_richness": process_richness,
                    "zone_radius_px": zone_radius_px,
                    "score": float(anchor["score"]),
                    "axis_ratio": axis_ratio,
                    "seed_point": seed_point,
                    "trusted_core": trusted_core,
                    "trusted_core_radius_um": trusted_core_radius_um,
                    "anchor_source": anchor.get("source", "2d_scored_nucleus"),
                    "validated_group_id": anchor.get("validated_group_id"),
                }
            )

        soma_anchor_count = len(soma_parts)
        no_dapi_count += int(soma_anchor_count == 0)
        multi_soma_roi_count += int(soma_anchor_count > 1)
        total_soma_anchor_count += soma_anchor_count
        rejected_soma_anchor_count += rejected_anchor_count
        soma = (
            np.logical_or.reduce(soma_parts) & component
            if soma_parts
            else np.zeros_like(component, dtype=bool)
        )
        process = component & ~soma
        if soma.any() and float(process.sum()) / max(component_area, 1) < cfg.min_process_fraction:
            fallback_used = True
            radial_limit = max(4, int(round(fallback_radius_px * 0.85)))
            restricted = np.zeros_like(component, dtype=bool)
            for detail in anchor_details:
                restricted |= component & circular_mask(
                    component.shape,
                    detail["seed_y"],
                    detail["seed_x"],
                    radial_limit,
                )
                restricted |= detail["selected_nucleus_extent"]
            if restricted.any():
                soma = restricted
                process = component & ~soma

        soma_area_before_core_shell_px = int(soma.sum())
        soma_core_shell_applied = False
        if soma.any() and anchor_details:
            trusted_core_union = np.logical_or.reduce(
                [detail["trusted_core"] for detail in anchor_details]
            ) & soma
            anchor_seed_union = np.logical_or.reduce(
                [
                    detail["seed_point"] | detail["selected_nucleus_extent"]
                    for detail in anchor_details
                ]
            ) & soma
            soma, soma_core_shell_applied = prune_soma_to_trusted_core_shell(
                soma,
                trusted_core_union,
                anchor_seed_union,
                pixel_width_um,
                pixel_height_um,
                cfg.soma_core_shell_max_um,
                min_soma_area_px,
            )
            process = component & ~soma
            required_nucleus_union = np.logical_or.reduce(
                [detail["selected_nucleus_extent"] for detail in anchor_details]
            ) & component
            missing_required_nucleus_px = int((required_nucleus_union & ~soma).sum())
            if missing_required_nucleus_px:
                raise RuntimeError(
                    f"Astrocyte_{astrocyte_id:03d} Soma lost "
                    f"{missing_required_nucleus_px} protected DAPI pixels"
                )
        else:
            required_nucleus_union = np.zeros_like(component, dtype=bool)
        soma_core_shell_removed_px = soma_area_before_core_shell_px - int(soma.sum())

        if not process.any():
            raise RuntimeError(
                f"Astrocyte_{astrocyte_id:03d} could not retain a non-empty Processes compartment"
            )
        fallback_count += int(fallback_used)
        soma_view = soma_labels[crop]
        process_view = process_labels[crop]
        soma_view[soma] = astrocyte_id
        process_view[process] = astrocyte_id
        per_cell.append(
            {
                "astrocyte_id": astrocyte_id,
                "whole_area_px": component_area,
                "soma_area_px": int(soma.sum()),
                "process_area_px": int(process.sum()),
                "soma_fraction": round(float(soma.sum()) / component_area, 6),
                "process_fraction": round(float(process.sum()) / component_area, 6),
                "process_richness": round(
                    float(np.mean([detail["process_richness"] for detail in anchor_details]))
                    if anchor_details else 1.0,
                    6,
                ),
                "thin_fraction": round(
                    float(np.mean([detail["thin_fraction"] for detail in anchor_details]))
                    if anchor_details else 1.0,
                    6,
                ),
                "core_peak_px": round(
                    max((detail["core_peak_px"] for detail in anchor_details), default=0.0),
                    3,
                ),
                "soma_zone_radius_px": max(
                    (detail["zone_radius_px"] for detail in anchor_details),
                    default=0,
                ),
                "nucleus_candidates": nucleus_candidates,
                "nucleus_score": round(nucleus_score, 6),
                "nucleus_ambiguous": bool(ambiguous),
                "soma_anchor_count": soma_anchor_count,
                "soma_anchor_scores": [
                    round(float(detail["score"]), 6) for detail in anchor_details
                ],
                "soma_anchor_sources": [
                    str(detail["anchor_source"]) for detail in anchor_details
                ],
                "validated_soma_group_ids": [
                    int(detail["validated_group_id"])
                    for detail in anchor_details
                    if detail["validated_group_id"] is not None
                ],
                "rejected_soma_anchor_count": rejected_anchor_count,
                "fallback_used": bool(fallback_used),
                "soma_area_before_core_shell_px": soma_area_before_core_shell_px,
                "soma_core_shell_removed_px": soma_core_shell_removed_px,
                "soma_core_shell_applied": bool(soma_core_shell_applied),
                "required_nucleus_px": int(required_nucleus_union.sum()),
                "required_nucleus_coverage": round(
                    float((required_nucleus_union & soma).sum())
                    / max(int(required_nucleus_union.sum()), 1),
                    6,
                ),
            }
        )

    pre_filter_roi_count = int(labels.max())
    labels, soma_labels, process_labels, per_cell, morphology_filter_metrics = (
        filter_morphology_outlier_instances(
            labels,
            soma_labels,
            process_labels,
            struct,
            mean_pixel_um,
            pixel_area_um2,
            per_cell,
            instance_metrics,
            cfg,
        )
    )
    roi_count = int(labels.max())
    id_mapping = {
        int(old_id): int(new_id)
        for old_id, new_id in morphology_filter_metrics["id_mapping"].items()
    }
    for detail in instance_metrics.get("split_components", []):
        original_ids = [int(value) for value in detail["new_astrocyte_ids"]]
        detail["pre_filter_new_astrocyte_ids"] = original_ids
        detail["new_astrocyte_ids"] = [
            id_mapping[value] for value in original_ids if value in id_mapping
        ]
    instance_metrics["pre_morphology_filter_instance_count"] = pre_filter_roi_count
    instance_metrics["final_instance_count"] = roi_count

    final_whole_mask = labels > 0
    soma_mask = soma_labels > 0
    process_mask = process_labels > 0
    overlap_px = int((soma_mask & process_mask).sum())
    gap_px = int((final_whole_mask & ~(soma_mask | process_mask)).sum())
    outside_px = int(((soma_mask | process_mask) & ~final_whole_mask).sum())
    if overlap_px or gap_px or outside_px:
        raise RuntimeError(
            "Compartment partition invariant failed: "
            f"overlap={overlap_px}, gap={gap_px}, outside={outside_px}"
        )

    whole_area_px = int(final_whole_mask.sum())
    soma_area_px = int(soma_mask.sum())
    process_area_px = int(process_mask.sum())
    metrics = {
        "method": "connectivity-preserving low-support branch-gap restoration + high-confidence DAPI/structural marker-controlled instance partition + single-body assigned-DAPI extent protection + local thickness/core-shell Soma + conservative multimetric whole-ID morphology filtering; Processes=Whole-Soma; no trusted Soma is forced",
        "adaptation": "continuous per-cell morphology adaptation; no animal-age label inferred",
        "config": asdict(cfg),
        "pixel_width_um": pixel_width_um,
        "pixel_height_um": pixel_height_um,
        "roi_count": roi_count,
        "whole_area_px": whole_area_px,
        "soma_area_px": soma_area_px,
        "process_area_px": process_area_px,
        "soma_area_fraction": round(soma_area_px / whole_area_px, 6),
        "process_area_fraction": round(process_area_px / whole_area_px, 6),
        "fallback_soma_count": int(sum(bool(row["fallback_used"]) for row in per_cell)),
        "ambiguous_nucleus_count": int(
            sum(bool(row["nucleus_ambiguous"]) for row in per_cell)
        ),
        "no_dapi_anchor_count": int(
            sum(int(row["soma_anchor_count"]) == 0 for row in per_cell)
        ),
        "total_soma_anchor_count": int(
            sum(int(row["soma_anchor_count"]) for row in per_cell)
        ),
        "multi_soma_whole_roi_count": int(
            sum(int(row["soma_anchor_count"]) > 1 for row in per_cell)
        ),
        "rejected_soma_anchor_count": int(
            sum(int(row["rejected_soma_anchor_count"]) for row in per_cell)
        ),
        "instance_split": instance_metrics,
        "nucleus_ownership_guard": nucleus_ownership_metrics,
        "branch_gap_restoration": branch_gap_metrics,
        "dapi_extent": dapi_extent_metrics,
        "dapi_extent_satellite_components_removed": int(
            dapi_extent_satellite_component_count
        ),
        "dapi_extent_satellite_px_removed": int(dapi_extent_satellite_px),
        "morphology_filter": morphology_filter_metrics,
        "soma_core_shell_removed_px": int(
            sum(row["soma_core_shell_removed_px"] for row in per_cell)
        ),
        "soma_core_shell_applied_roi_count": int(
            sum(bool(row["soma_core_shell_applied"]) for row in per_cell)
        ),
        "partition_overlap_px": overlap_px,
        "partition_gap_px": gap_px,
        "partition_outside_whole_px": outside_px,
        "per_cell": per_cell,
    }
    return labels, soma_labels, process_labels, metrics


def _linear_evidence(value: float, low: float, high: float) -> float:
    if high <= low:
        raise ValueError("Age-profile evidence bounds must increase")
    return float(np.clip((value - low) / (high - low), 0.0, 1.0))

def classify_age_profile(
    whole_mask: np.ndarray,
    dapi_projection: np.ndarray,
    struct: np.ndarray,
    pixel_width_um: float,
    pixel_height_um: float,
) -> AgeProfileDecision:
    """Return a deterministic binary morphology profile from structural channels only."""

    mask = whole_mask.astype(bool, copy=False)
    labels = measure.label(mask, connectivity=2)
    pixel_area_um2 = pixel_width_um * pixel_height_um
    mean_pixel_um = math.sqrt(pixel_area_um2)
    all_props = sorted(measure.regionprops(labels), key=lambda prop: prop.area, reverse=True)
    minimum_area_px = max(32, int(round(12.0 / pixel_area_um2)))
    area_eligible = [prop for prop in all_props if int(prop.area) >= minimum_area_px]
    interior_eligible = [
        prop
        for prop in area_eligible
        if prop.bbox[0] > 0
        and prop.bbox[1] > 0
        and prop.bbox[2] < labels.shape[0]
        and prop.bbox[3] < labels.shape[1]
    ]
    eligible = interior_eligible if len(interior_eligible) >= 3 else area_eligible
    if not eligible:
        eligible = all_props
    if not eligible:
        raise ValueError("Cannot classify age profile from an empty Whole mask")
    eligible = eligible[:24]
    eligible_ids = {int(prop.label) for prop in eligible}

    nuclei = dapi_nuclei_mask(dapi_projection, percentile_floor=85.0)
    nuclei_labels = measure.label(nuclei, connectivity=2)
    nucleus_counts = {label_id: 0 for label_id in eligible_ids}
    if mask.any():
        distance_to_whole_um, nearest_indices = ndi.distance_transform_edt(
            ~mask,
            sampling=(pixel_height_um, pixel_width_um),
            return_indices=True,
        )
        nearest_whole_labels = labels[nearest_indices[0], nearest_indices[1]]
        for nucleus_prop in measure.regionprops(nuclei_labels):
            coords = nucleus_prop.coords
            overlapping = labels[coords[:, 0], coords[:, 1]]
            overlapping = overlapping[overlapping > 0]
            assigned = 0
            if overlapping.size:
                ids, counts = np.unique(overlapping, return_counts=True)
                assigned = int(ids[int(np.argmax(counts))])
            else:
                cy = int(np.clip(round(nucleus_prop.centroid[0]), 0, labels.shape[0] - 1))
                cx = int(np.clip(round(nucleus_prop.centroid[1]), 0, labels.shape[1] - 1))
                if distance_to_whole_um[cy, cx] <= 1.2:
                    assigned = int(nearest_whole_labels[cy, cx])
            if assigned in nucleus_counts:
                nucleus_counts[assigned] += 1

    widths_um: list[float] = []
    thin_fractions: list[float] = []
    solidities: list[float] = []
    axis_ratios: list[float] = []
    branch_densities: list[float] = []
    endpoint_densities: list[float] = []
    for prop in eligible:
        min_row, min_col, max_row, max_col = prop.bbox
        component = labels[min_row:max_row, min_col:max_col] == prop.label
        distance_um = ndi.distance_transform_edt(
            component,
            sampling=(pixel_height_um, pixel_width_um),
        )
        skeleton_px, endpoint_count, branchpoint_count = skeleton_topology(component)
        skeleton_length_um = max(skeleton_px * mean_pixel_um, mean_pixel_um)
        area_um2 = float(prop.area) * pixel_area_um2
        widths_um.append(area_um2 / skeleton_length_um)
        thin_fractions.append(float((distance_um[component] <= 0.70).mean()))
        solidities.append(float(prop.solidity))
        axis_ratios.append(
            float(prop.major_axis_length) / max(float(prop.minor_axis_length), 1e-6)
        )
        branch_densities.append(10.0 * branchpoint_count / skeleton_length_um)
        endpoint_densities.append(10.0 * endpoint_count / skeleton_length_um)

    median_width_um = float(np.median(widths_um))
    median_thin_fraction = float(np.median(thin_fractions))
    median_solidity = float(np.median(solidities))
    median_axis_ratio = float(np.median(axis_ratios))
    median_branch_density = float(np.median(branch_densities))
    median_endpoint_density = float(np.median(endpoint_densities))
    multi_nucleus_fraction = float(
        np.mean([nucleus_counts[int(prop.label)] >= 2 for prop in eligible])
    )

    evidence = {
        "broad_structure": _linear_evidence(median_width_um, 0.95, 2.20),
        "solid_structure": _linear_evidence(median_solidity, 0.24, 0.54),
        "limited_thin_arbor": 1.0
        - _linear_evidence(median_thin_fraction, 0.62, 0.88),
        "limited_branching": 1.0
        - _linear_evidence(median_branch_density, 0.14, 0.50),
        "polarized_shape": _linear_evidence(median_axis_ratio, 1.70, 3.20),
        "multi_nucleus_overlap": _linear_evidence(multi_nucleus_fraction, 0.05, 0.30),
    }
    neonatal_score = (
        0.26 * evidence["broad_structure"]
        + 0.20 * evidence["solid_structure"]
        + 0.18 * evidence["limited_thin_arbor"]
        + 0.16 * evidence["limited_branching"]
        + 0.10 * evidence["polarized_shape"]
        + 0.10 * evidence["multi_nucleus_overlap"]
    )
    threshold = AGE_PROFILE_THRESHOLD
    profile = "neonatal" if neonatal_score >= threshold else "mature"
    features: dict[str, float | int] = {
        "component_count": len(eligible),
        "edge_components_excluded": max(0, len(area_eligible) - len(eligible)),
        "median_width_um": round(median_width_um, 6),
        "median_thin_fraction": round(median_thin_fraction, 6),
        "median_solidity": round(median_solidity, 6),
        "median_axis_ratio": round(median_axis_ratio, 6),
        "median_branchpoints_per_10um": round(median_branch_density, 6),
        "median_endpoints_per_10um": round(median_endpoint_density, 6),
        "multi_nucleus_component_fraction": round(multi_nucleus_fraction, 6),
        **{f"evidence_{key}": round(value, 6) for key, value in evidence.items()},
    }
    return AgeProfileDecision(
        profile=profile,
        source="morphology_classifier",
        neonatal_score=round(float(neonatal_score), 6),
        threshold=threshold,
        confidence_margin=round(abs(float(neonatal_score) - threshold), 6),
        tagged_files=(),
        features=features,
    )

def split_astrocyte_compartments_for_profile(
    whole_mask: np.ndarray,
    dapi_projection: np.ndarray,
    struct: np.ndarray,
    cellpose_mask: np.ndarray,
    pixel_width_um: float,
    pixel_height_um: float,
    profile: str,
    neonatal_3d_context: Neonatal3DContext | None = None,
    dapi_fragment_workload_diagnostic_path: Path | None = None,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, dict]:
    """Apply shared 3D ownership before profile-specific Soma/Processes rules."""

    mature_config = compartment_config_for_profile("mature")
    ownership_inventory = None
    if neonatal_3d_context is not None:
        ownership_started = time.perf_counter()
        ownership_inventory = build_dapi_object_inventory_3d(
            whole_mask,
            dapi_projection,
            neonatal_3d_context,
            pixel_width_um,
            pixel_height_um,
            mature_config,
            max_workers=_EFFECTIVE_DAPI_INVENTORY_CPU_WORKERS,
            workload_diagnostic_path=dapi_fragment_workload_diagnostic_path,
        )
        nucleus_inventory_metrics = ownership_inventory.metrics
        print(
            "3D DAPI nucleus ownership complete | "
            f"elapsed={time.perf_counter() - ownership_started:.3f} s; "
            "partitioning Whole/Soma/Processes...",
            flush=True,
        )
    else:
        nucleus_inventory_metrics = {
            "status": "not_run_structural_stack_or_Z_calibration_unavailable",
            "method": "object-preserving calibrated 3D DAPI inventory",
            "measurement_channel_used": False,
            "candidate_count": 0,
            "dapi_valid_count": 0,
            "accepted_count": 0,
            "rejected_count": 0,
            "per_nucleus": [],
        }
    if profile == "mature":
        base_partition_started = time.perf_counter()
        labels, soma_labels, process_labels, metrics = split_astrocyte_compartments(
            whole_mask,
            dapi_projection,
            struct,
            cellpose_mask,
            pixel_width_um,
            pixel_height_um,
            config=mature_config,
            ownership_inventory=ownership_inventory,
            ownership_pixel_depth_um=(
                neonatal_3d_context.pixel_depth_um
                if neonatal_3d_context is not None
                else None
            ),
            ownership_profile="mature",
        )
        pre_gate_labels = labels.copy()
        pre_gate_roi_count = int(labels.max())
        (
            labels,
            soma_labels,
            process_labels,
            gated_per_cell,
            soma_identity_gate,
        ) = filter_instances_by_valid_soma(
            labels,
            soma_labels,
            process_labels,
            metrics["per_cell"],
            profile="mature",
        )
        retained_pre_gate_ids = {
            int(value) for value in soma_identity_gate["retained_pre_gate_ids"]
        }
        expected_final_mask = np.isin(pre_gate_labels, list(retained_pre_gate_ids))
        final_whole_mask = labels > 0
        if not np.array_equal(final_whole_mask, expected_final_mask):
            raise RuntimeError(
                "Mature valid-Soma gate changed retained geometry instead of deleting "
                "complete cell IDs"
            )
        if int(final_whole_mask.sum()) + int(soma_identity_gate["removed_area_px"]) != int(
            (pre_gate_labels > 0).sum()
        ):
            raise RuntimeError("Mature valid-Soma gate removed partial cell geometry")

        id_mapping = {
            int(old_id): int(new_id)
            for old_id, new_id in soma_identity_gate["id_mapping"].items()
        }
        for detail in metrics["instance_split"].get("split_components", []):
            pre_gate_ids = [int(value) for value in detail.get("new_astrocyte_ids", [])]
            detail["pre_soma_gate_new_astrocyte_ids"] = pre_gate_ids
            detail["removed_pre_soma_gate_ids"] = [
                value for value in pre_gate_ids if value not in id_mapping
            ]
            detail["new_astrocyte_ids"] = [
                id_mapping[value] for value in pre_gate_ids if value in id_mapping
            ]
        metrics["instance_split"]["pre_soma_gate_instance_count"] = pre_gate_roi_count
        metrics["instance_split"]["final_instance_count"] = int(labels.max())
        metrics["soma_identity_gate"] = soma_identity_gate
        metrics["per_cell"] = gated_per_cell
        metrics["roi_count"] = int(labels.max())
        metrics["whole_area_px"] = int(final_whole_mask.sum())
        metrics["soma_area_px"] = int((soma_labels > 0).sum())
        metrics["process_area_px"] = int((process_labels > 0).sum())
        metrics["soma_area_fraction"] = round(
            metrics["soma_area_px"] / metrics["whole_area_px"],
            6,
        )
        metrics["process_area_fraction"] = round(
            metrics["process_area_px"] / metrics["whole_area_px"],
            6,
        )
        metrics["fallback_soma_count"] = int(
            sum(bool(row["fallback_used"]) for row in gated_per_cell)
        )
        metrics["ambiguous_nucleus_count"] = int(
            sum(bool(row["nucleus_ambiguous"]) for row in gated_per_cell)
        )
        metrics["no_dapi_anchor_count"] = 0
        metrics["total_soma_anchor_count"] = len(gated_per_cell)
        metrics["multi_soma_whole_roi_count"] = 0
        metrics["method"] += (
            "; mature whole-ID valid-Soma gate with synchronized relabeling"
        )
        metrics["nucleus_3d_inventory"] = nucleus_inventory_metrics
        print(
            "Base compartment partition complete | profile=mature | "
            f"elapsed={time.perf_counter() - base_partition_started:.3f} s; "
            "finalizing identity and Soma safeguards...",
            flush=True,
        )
        finalization_started = time.perf_counter()
        labels, soma_labels, process_labels, metrics = (
            finalize_compartment_geometry_and_metrics(
                labels,
                soma_labels,
                process_labels,
                metrics,
                ownership_inventory,
                neonatal_3d_context,
                struct,
                "mature",
                pixel_width_um,
                pixel_height_um,
            )
        )
        print(
            "Compartment finalization complete | profile=mature | "
            f"elapsed={time.perf_counter() - finalization_started:.3f} s",
            flush=True,
        )
        return labels, soma_labels, process_labels, metrics
    if profile != "neonatal":
        raise ValueError(f"Unknown astrocyte profile: {profile}")

    base_partition_started = time.perf_counter()
    shared_whole_labels, _, _, shared_metrics = split_astrocyte_compartments(
        whole_mask,
        dapi_projection,
        struct,
        cellpose_mask,
        pixel_width_um,
        pixel_height_um,
        config=mature_config,
        ownership_inventory=ownership_inventory,
        ownership_pixel_depth_um=(
            neonatal_3d_context.pixel_depth_um
            if neonatal_3d_context is not None
            else None
        ),
        ownership_profile="neonatal_shared_whole",
    )
    frozen_whole_mask = shared_whole_labels > 0
    neonatal_config = replace(
        compartment_config_for_profile("neonatal"),
        branch_gap_restore_enabled=False,
        morphology_outlier_filter_enabled=False,
        instance_split_min_anchor_separation_um=4.0,
    )
    validated_anchors = ownership_inventory
    neonatal_3d_metrics = nucleus_inventory_metrics
    labels, soma_labels, process_labels, metrics = split_astrocyte_compartments(
        frozen_whole_mask,
        dapi_projection,
        struct,
        cellpose_mask,
        pixel_width_um,
        pixel_height_um,
        config=neonatal_config,
        validated_anchors=validated_anchors,
        ownership_inventory=ownership_inventory,
        ownership_pixel_depth_um=(
            neonatal_3d_context.pixel_depth_um
            if neonatal_3d_context is not None
            else None
        ),
        ownership_profile="neonatal",
    )
    if np.any((labels > 0) & ~frozen_whole_mask):
        raise RuntimeError(
            "Neonatal repartition expanded the frozen shared Whole pixel union before "
            "the valid-Soma cell gate"
        )
    pre_gate_roi_count = int(labels.max())
    pre_gate_labels = labels.copy()
    unresolved_multi_soma_ids = {
        int(astrocyte_id)
        for decision in metrics["instance_split"].get("component_decisions", [])
        if bool(decision.get("split_required"))
        and not bool(decision.get("split_accepted"))
        for astrocyte_id in decision.get("output_astrocyte_ids", [])
    }
    (
        labels,
        soma_labels,
        process_labels,
        gated_per_cell,
        soma_identity_gate,
    ) = filter_instances_by_valid_soma(
        labels,
        soma_labels,
        process_labels,
        metrics["per_cell"],
        profile="neonatal",
        unresolved_multi_soma_ids=unresolved_multi_soma_ids,
    )
    final_whole_mask = labels > 0
    if np.any(final_whole_mask & ~frozen_whole_mask):
        raise RuntimeError("Neonatal valid-Soma gate expanded the frozen Whole pixel union")
    retained_pre_gate_ids = {
        int(value) for value in soma_identity_gate["retained_pre_gate_ids"]
    }
    expected_final_mask = np.isin(
        pre_gate_labels,
        list(retained_pre_gate_ids),
    )
    if not np.array_equal(final_whole_mask, expected_final_mask):
        raise RuntimeError(
            "Neonatal valid-Soma gate changed retained cell geometry instead of only relabeling"
        )
    ownership_removed_area_px = int(
        metrics.get("nucleus_ownership_guard", {}).get("removed_area_px", 0)
    )
    if (
        int(final_whole_mask.sum())
        + int(soma_identity_gate["removed_area_px"])
        + ownership_removed_area_px
        != int(frozen_whole_mask.sum())
    ):
        raise RuntimeError(
            "Neonatal ownership and valid-Soma gates do not account for the frozen "
            "Whole geometry exactly"
        )
    del pre_gate_labels, expected_final_mask

    id_mapping = {
        int(old_id): int(new_id)
        for old_id, new_id in soma_identity_gate["id_mapping"].items()
    }
    for detail in metrics["instance_split"].get("split_components", []):
        pre_gate_ids = [int(value) for value in detail.get("new_astrocyte_ids", [])]
        pre_gate_areas = [int(value) for value in detail.get("child_areas_px", [])]
        detail["pre_soma_gate_new_astrocyte_ids"] = pre_gate_ids
        detail["pre_soma_gate_child_areas_px"] = pre_gate_areas
        detail["removed_pre_soma_gate_ids"] = [
            value for value in pre_gate_ids if value not in id_mapping
        ]
        detail["new_astrocyte_ids"] = [
            id_mapping[value] for value in pre_gate_ids if value in id_mapping
        ]
        detail["retained_child_areas_px"] = [
            area
            for value, area in zip(pre_gate_ids, pre_gate_areas)
            if value in id_mapping
        ]
    for decision in metrics["instance_split"].get("component_decisions", []):
        pre_gate_ids = [
            int(value) for value in decision.get("output_astrocyte_ids", [])
        ]
        decision["pre_soma_gate_output_astrocyte_ids"] = pre_gate_ids
        decision["output_astrocyte_ids"] = [
            id_mapping[value] for value in pre_gate_ids if value in id_mapping
        ]
    metrics["instance_split"]["pre_soma_gate_instance_count"] = pre_gate_roi_count
    metrics["instance_split"]["post_soma_gate_instance_count"] = int(labels.max())
    metrics["soma_identity_gate"] = soma_identity_gate
    metrics["per_cell"] = gated_per_cell
    metrics["roi_count"] = int(labels.max())
    metrics["whole_area_px"] = int((labels > 0).sum())
    metrics["soma_area_px"] = int((soma_labels > 0).sum())
    metrics["process_area_px"] = int((process_labels > 0).sum())
    metrics["soma_area_fraction"] = round(
        metrics["soma_area_px"] / metrics["whole_area_px"], 6
    )
    metrics["process_area_fraction"] = round(
        metrics["process_area_px"] / metrics["whole_area_px"], 6
    )
    metrics["fallback_soma_count"] = int(
        sum(bool(row["fallback_used"]) for row in gated_per_cell)
    )
    metrics["ambiguous_nucleus_count"] = int(
        sum(bool(row["nucleus_ambiguous"]) for row in gated_per_cell)
    )
    metrics["no_dapi_anchor_count"] = 0
    metrics["total_soma_anchor_count"] = len(gated_per_cell)
    metrics["multi_soma_whole_roi_count"] = 0
    metrics["rejected_soma_anchor_count"] = int(
        sum(int(row["rejected_soma_anchor_count"]) for row in gated_per_cell)
    )
    metrics["soma_core_shell_removed_px"] = int(
        sum(int(row["soma_core_shell_removed_px"]) for row in gated_per_cell)
    )
    metrics["soma_core_shell_applied_roi_count"] = int(
        sum(bool(row["soma_core_shell_applied"]) for row in gated_per_cell)
    )
    overlap_px = int(((soma_labels > 0) & (process_labels > 0)).sum())
    gap_px = int(
        ((labels > 0) & ~((soma_labels > 0) | (process_labels > 0))).sum()
    )
    outside_px = int(
        (((soma_labels > 0) | (process_labels > 0)) & ~(labels > 0)).sum()
    )
    if overlap_px or gap_px or outside_px:
        raise RuntimeError(
            "Post-gate compartment partition invariant failed: "
            f"overlap={overlap_px}, gap={gap_px}, outside={outside_px}"
        )
    metrics["partition_overlap_px"] = overlap_px
    metrics["partition_gap_px"] = gap_px
    metrics["partition_outside_whole_px"] = outside_px
    metrics["shared_whole_baseline"] = {
        "method": (
            "Frozen Whole geometry/filter with shared object-preserving 3D nucleus "
            "ownership refinement"
        ),
        "roi_count_before_neonatal_repartition": int(shared_whole_labels.max()),
        "whole_area_px": int(frozen_whole_mask.sum()),
        "branch_gap_restoration": shared_metrics["branch_gap_restoration"],
        "instance_split": shared_metrics["instance_split"],
        "nucleus_ownership_guard": shared_metrics["nucleus_ownership_guard"],
        "morphology_filter": shared_metrics["morphology_filter"],
    }
    metrics["neonatal_3d_validation"] = neonatal_3d_metrics
    shared_rows = {
        int(row["astrocyte_id"]): row for row in shared_metrics["per_cell"]
    }
    for row in metrics["per_cell"]:
        astrocyte_id = int(row["astrocyte_id"])
        final_mask = labels == astrocyte_id
        source_ids = sorted(
            int(value)
            for value in np.unique(shared_whole_labels[final_mask])
            if int(value) > 0
        )
        row["shared_whole_ids"] = source_ids
        row["process_component_count"] = int(
            measure.label(process_labels == astrocyte_id, connectivity=2).max()
        )
        if len(source_ids) == 1 and source_ids[0] in shared_rows:
            row["shared_morphology_qc"] = shared_rows[source_ids[0]].get(
                "morphology_qc", {}
            )
    validation_method = (
        "object-preserving calibrated 3D DAPI/"
        f"{neonatal_3d_context.structural_channel} nucleus ownership and anchor gate + "
        if neonatal_3d_context is not None
        else "2D DAPI anchors because calibrated 3D ownership was unavailable + "
    )
    metrics["method"] = (
        "shared frozen Whole geometry/filter with 3D nucleus ownership refinement + "
        + validation_method
        + "neonatal multi-center ID partition + neonatal local thickness/core-shell Soma + "
        "whole-ID valid-Soma gate and synchronized relabeling; "
        "Processes=Whole-Soma"
    )
    print(
        "Base compartment partition complete | profile=neonatal | "
        f"elapsed={time.perf_counter() - base_partition_started:.3f} s; "
        "finalizing identity and Soma safeguards...",
        flush=True,
    )
    finalization_started = time.perf_counter()
    labels, soma_labels, process_labels, metrics = (
        finalize_compartment_geometry_and_metrics(
            labels,
            soma_labels,
            process_labels,
            metrics,
            ownership_inventory,
            neonatal_3d_context,
            struct,
            "neonatal",
            pixel_width_um,
            pixel_height_um,
        )
    )
    print(
        "Compartment finalization complete | profile=neonatal | "
        f"elapsed={time.perf_counter() - finalization_started:.3f} s",
        flush=True,
    )
    return labels, soma_labels, process_labels, metrics


def rank_candidates(masks: list[np.ndarray], rows: list[dict]) -> int:
    count = len(masks)
    if count == 0:
        raise ValueError("No candidate masks to rank")
    if count == 1:
        rows[0].update({"mean_candidate_iou": 1.0, "auto_selection_score": 1.0, "auto_selected": True})
        return 0

    mean_iou = np.zeros(count, dtype=np.float64)
    mask_areas = np.asarray(
        [np.count_nonzero(mask) for mask in masks],
        dtype=np.int64,
    )
    intersection_buffer = np.empty_like(masks[0], dtype=bool)
    for left in range(count):
        for right in range(left + 1, count):
            np.logical_and(
                masks[left],
                masks[right],
                out=intersection_buffer,
            )
            intersection = int(np.count_nonzero(intersection_buffer))
            union = int(mask_areas[left] + mask_areas[right] - intersection)
            iou = intersection / union if union else 1.0
            mean_iou[left] += iou
            mean_iou[right] += iou
    mean_iou /= count - 1

    def rank01(values: list[float], higher_is_better: bool = True) -> np.ndarray:
        array = np.asarray(values, dtype=np.float64)
        ranks = np.asarray([(array < value).mean() + 0.5 * (array == value).mean() for value in array])
        return ranks if higher_is_better else 1.0 - ranks

    def raw_metric(row: dict, key: str) -> float:
        return float(row.get(f"_raw_{key}", row[key]))

    coverage_rank = rank01([raw_metric(row, "structural_signal_coverage") for row in rows])
    precision_rank = rank01([raw_metric(row, "structural_precision") for row in rows])
    unsupported_rank = rank01(
        [raw_metric(row, "unsupported_wide_fraction") for row in rows],
        higher_is_better=False,
    )
    soma_counts = np.asarray(
        [float(row["soma_supported_components"]) for row in rows],
        dtype=np.float64,
    )
    soma_consensus_cap = float(np.percentile(soma_counts, 75))
    soma_rank = rank01(np.minimum(soma_counts, soma_consensus_cap).tolist())
    unanchored_rank = rank01(
        [raw_metric(row, "unanchored_area_fraction") for row in rows],
        higher_is_better=False,
    )
    z_activity_rank = rank01([raw_metric(row, "z_activity_mean") for row in rows])
    edge_rank = rank01(
        [raw_metric(row, "edge_proximity_area_fraction") for row in rows],
        higher_is_better=False,
    )
    border_burden_rank = rank01(
        [raw_metric(row, "border_removed_area_fraction") for row in rows],
        higher_is_better=False,
    )
    preserved_border_rank = rank01(
        [raw_metric(row, "border_preserved_complete_area_fraction") for row in rows],
        higher_is_better=False,
    )
    score = (
        0.27 * mean_iou
        + 0.10 * coverage_rank
        + 0.15 * precision_rank
        + 0.09 * unsupported_rank
        + 0.20 * soma_rank
        + 0.08 * unanchored_rank
        + 0.04 * z_activity_rank
        + 0.02 * edge_rank
        + 0.01 * border_burden_rank
        + 0.04 * preserved_border_rank
    )
    best = int(np.argmax(score))
    for index, row in enumerate(rows):
        row.update(
            {
                "mean_candidate_iou": round(float(mean_iou[index]), 6),
                "_raw_mean_candidate_iou": float(mean_iou[index]),
                "auto_selection_score": round(float(score[index]), 6),
                "_raw_auto_selection_score": float(score[index]),
                "auto_selected": index == best,
            }
        )
    return best

def rank_production_candidates(masks: list[np.ndarray], rows: list[dict]) -> int:
    eligible = [index for index, row in enumerate(rows) if not row.get("error")]
    if not eligible:
        errors = "\n".join(
            f"candidate {row.get('candidate', index + 1)}: {row.get('error')}"
            for index, row in enumerate(rows)
        )
        raise RuntimeError(f"All ROI candidates failed or used an exception fallback:\n{errors}")
    for index, row in enumerate(rows):
        if index not in eligible:
            row.update(
                {
                    "mean_candidate_iou": 0.0,
                    "auto_selection_score": -1.0,
                    "auto_selected": False,
                    "selection_eligible": False,
                }
            )
    eligible_masks = [masks[index] for index in eligible]
    eligible_rows = [rows[index] for index in eligible]
    best_eligible_position = rank_candidates(eligible_masks, eligible_rows)
    for row in eligible_rows:
        row["selection_eligible"] = True
    return eligible[best_eligible_position]

def weighted_rank01(
    values: list[float],
    weights: np.ndarray,
    *,
    higher_is_better: bool = True,
) -> np.ndarray:
    array = np.asarray(values, dtype=np.float64)
    weights = np.asarray(weights, dtype=np.float64)
    if array.shape != weights.shape or np.any(weights < 0) or not np.any(weights > 0):
        raise ValueError("Invalid candidate ranking weights")
    weights /= float(weights.sum())
    ranks = np.asarray(
        [
            float(weights[array < value].sum())
            + 0.5 * float(weights[array == value].sum())
            for value in array
        ],
        dtype=np.float64,
    )
    return ranks if higher_is_better else 1.0 - ranks

def weighted_percentile(
    values: np.ndarray,
    weights: np.ndarray,
    percentile: float,
) -> float:
    ordered_values, inverse = np.unique(
        np.asarray(values, dtype=np.float64),
        return_inverse=True,
    )
    ordered_weights = np.zeros(len(ordered_values), dtype=np.float64)
    np.add.at(ordered_weights, inverse, np.asarray(weights, dtype=np.float64))
    cumulative = np.cumsum(ordered_weights, dtype=np.float64)
    target = float(np.clip(percentile / 100.0, 0.0, 1.0)) * float(
        ordered_weights.sum(dtype=np.float64)
    )
    position = int(np.searchsorted(cumulative, target, side="left"))
    return float(ordered_values[min(position, len(ordered_values) - 1)])

def cluster_near_duplicate_candidates(
    indices: np.ndarray,
    pairwise_iou: np.ndarray,
) -> list[np.ndarray]:
    parent = {int(index): int(index) for index in indices}

    def find(value: int) -> int:
        while parent[value] != value:
            parent[value] = parent[parent[value]]
            value = parent[value]
        return value

    def union(left: int, right: int) -> None:
        left_root = find(left)
        right_root = find(right)
        if left_root != right_root:
            parent[max(left_root, right_root)] = min(left_root, right_root)

    integer_indices = [int(index) for index in indices]
    for left_position, left in enumerate(integer_indices):
        for right in integer_indices[left_position + 1 :]:
            if pairwise_iou[left, right] >= NEAR_DUPLICATE_CANDIDATE_IOU:
                union(left, right)
    clusters: dict[int, list[int]] = {}
    for index in integer_indices:
        clusters.setdefault(find(index), []).append(index)
    return [
        np.asarray(clusters[root], dtype=np.int64)
        for root in sorted(clusters)
    ]

def family_z_balance_structure(
    rows: list[dict],
    pairwise_iou: np.ndarray,
) -> tuple[dict[tuple[str, str], list[np.ndarray]], np.ndarray]:
    cells: dict[tuple[str, str], list[int]] = {}
    for index, row in enumerate(rows):
        key = (str(row["candidate_family"]), str(row["z_mode"]))
        cells.setdefault(key, []).append(index)
    clustered_cells: dict[tuple[str, str], list[np.ndarray]] = {}
    candidate_weights = np.zeros(len(rows), dtype=np.float64)
    cell_weight = 1.0 / max(len(cells), 1)
    for cell_key in sorted(cells):
        clusters = cluster_near_duplicate_candidates(
            np.asarray(cells[cell_key], dtype=np.int64),
            pairwise_iou,
        )
        clustered_cells[cell_key] = clusters
        cluster_weight = cell_weight / max(len(clusters), 1)
        for cluster in clusters:
            # A near-duplicate cluster contributes one fixed vote.  Its
            # lowest-index member is the deterministic representative, so
            # appending an exact or near duplicate cannot change the metric
            # reference distribution or amplify that candidate family.
            candidate_weights[int(cluster[0])] = cluster_weight
    candidate_weights /= float(candidate_weights.sum(dtype=np.float64))
    return clustered_cells, candidate_weights

def rank_candidates_family_balanced(
    masks: list[np.ndarray],
    rows: list[dict],
    *,
    expected_families: tuple[str, ...] = EXPECTED_CANDIDATE_FAMILIES,
) -> int:
    count = len(masks)
    if count == 0:
        raise ValueError("No candidate masks to rank")
    if count == 1:
        rows[0].update(
            {
                "mean_candidate_iou": 1.0,
                "_raw_mean_candidate_iou": 1.0,
                "auto_selection_score": 1.0,
                "_raw_auto_selection_score": 1.0,
                "auto_selected": True,
            }
        )
        return 0

    pairwise_iou = np.eye(count, dtype=np.float64)
    mask_areas = np.asarray(
        [np.count_nonzero(mask) for mask in masks],
        dtype=np.int64,
    )
    intersection_buffer = np.empty_like(masks[0], dtype=bool)
    for left in range(count):
        for right in range(left + 1, count):
            np.logical_and(
                masks[left],
                masks[right],
                out=intersection_buffer,
            )
            intersection = int(np.count_nonzero(intersection_buffer))
            union = int(mask_areas[left] + mask_areas[right] - intersection)
            iou = intersection / union if union else 1.0
            pairwise_iou[left, right] = iou
            pairwise_iou[right, left] = iou

    clustered_cells, candidate_weights = family_z_balance_structure(
        rows,
        pairwise_iou,
    )
    observed_families = {str(row["candidate_family"]) for row in rows}
    observed_z_modes = {str(row["z_mode"]) for row in rows}
    if observed_families != set(expected_families):
        raise RuntimeError(
            "Family-Z candidate balance is missing a predefined candidate family"
        )
    expected_cell_count = len(expected_families) * EXPECTED_Z_INTERVAL_COUNT
    if (
        len(observed_z_modes) != EXPECTED_Z_INTERVAL_COUNT
        or len(clustered_cells) != expected_cell_count
    ):
        raise RuntimeError(
            f"Family-Z candidate balance requires all {expected_cell_count} "
            "predefined cells"
        )
    mean_iou = np.zeros(count, dtype=np.float64)
    for index in range(count):
        cell_means: list[float] = []
        for clusters in clustered_cells.values():
            cluster_means: list[float] = []
            for cluster in clusters:
                if np.any(cluster == index):
                    continue
                cluster_means.append(
                    float(pairwise_iou[index, int(cluster[0])])
                )
            if cluster_means:
                cell_means.append(float(np.mean(cluster_means, dtype=np.float64)))
        mean_iou[index] = float(np.mean(cell_means, dtype=np.float64))

    def raw_metric(row: dict, key: str) -> float:
        return float(row.get(f"_raw_{key}", row[key]))

    coverage_rank = weighted_rank01(
        [raw_metric(row, "structural_signal_coverage") for row in rows],
        candidate_weights,
    )
    precision_rank = weighted_rank01(
        [raw_metric(row, "structural_precision") for row in rows],
        candidate_weights,
    )
    unsupported_rank = weighted_rank01(
        [raw_metric(row, "unsupported_wide_fraction") for row in rows],
        candidate_weights,
        higher_is_better=False,
    )
    soma_counts = np.asarray(
        [float(row["soma_supported_components"]) for row in rows],
        dtype=np.float64,
    )
    soma_consensus_cap = weighted_percentile(
        soma_counts,
        candidate_weights,
        75.0,
    )
    soma_rank = weighted_rank01(
        np.minimum(soma_counts, soma_consensus_cap).tolist(),
        candidate_weights,
    )
    unanchored_rank = weighted_rank01(
        [raw_metric(row, "unanchored_area_fraction") for row in rows],
        candidate_weights,
        higher_is_better=False,
    )
    z_activity_rank = weighted_rank01(
        [raw_metric(row, "z_activity_mean") for row in rows],
        candidate_weights,
    )
    edge_rank = weighted_rank01(
        [raw_metric(row, "edge_proximity_area_fraction") for row in rows],
        candidate_weights,
        higher_is_better=False,
    )
    border_burden_rank = weighted_rank01(
        [raw_metric(row, "border_removed_area_fraction") for row in rows],
        candidate_weights,
        higher_is_better=False,
    )
    preserved_border_rank = weighted_rank01(
        [raw_metric(row, "border_preserved_complete_area_fraction") for row in rows],
        candidate_weights,
        higher_is_better=False,
    )
    score = (
        0.27 * mean_iou
        + 0.10 * coverage_rank
        + 0.15 * precision_rank
        + 0.09 * unsupported_rank
        + 0.20 * soma_rank
        + 0.08 * unanchored_rank
        + 0.04 * z_activity_rank
        + 0.02 * edge_rank
        + 0.01 * border_burden_rank
        + 0.04 * preserved_border_rank
    )
    best = int(np.argmax(score))
    for index, row in enumerate(rows):
        row.update(
            {
                "mean_candidate_iou": round(float(mean_iou[index]), 6),
                "_raw_mean_candidate_iou": float(mean_iou[index]),
                "auto_selection_score": round(float(score[index]), 6),
                "_raw_auto_selection_score": float(score[index]),
                "auto_selected": index == best,
            }
        )
    return best

def challenger_dominates_incumbent(
    challenger: dict,
    incumbent: dict,
) -> tuple[bool, dict[str, object]]:
    def raw(row: dict, key: str) -> float:
        return float(row.get(f"_raw_{key}", row[key]))

    higher_metrics = (
        "structural_signal_coverage",
        "structural_precision",
    )
    lower_metrics = (
        "unsupported_wide_fraction",
        "unanchored_area_fraction",
        "edge_proximity_area_fraction",
        "border_removed_area_fraction",
        "border_preserved_complete_area_fraction",
    )
    comparisons: dict[str, bool] = {
        "same_z_interval": (
            int(challenger["z_start_0based"])
            == int(incumbent["z_start_0based"])
            and int(challenger["z_end_0based_inclusive"])
            == int(incumbent["z_end_0based_inclusive"])
            and str(challenger["projection"]) == str(incumbent["projection"])
        )
    }
    for key in higher_metrics:
        comparisons[f"{key}_not_lower"] = raw(challenger, key) >= raw(
            incumbent, key
        )
    for key in lower_metrics:
        comparisons[f"{key}_not_higher"] = raw(challenger, key) <= raw(
            incumbent, key
        )
    comparisons["soma_supported_components_equal"] = int(
        challenger["soma_supported_components"]
    ) == int(incumbent["soma_supported_components"])
    comparisons["incomplete_border_components_not_higher"] = int(
        challenger["final_incomplete_border_touching_components"]
    ) <= int(incumbent["final_incomplete_border_touching_components"])
    comparisons["family_balanced_score_margin"] = float(
        challenger["_raw_auto_selection_score"]
    ) >= (
        float(incumbent["_raw_auto_selection_score"])
        + CHALLENGER_MIN_SCORE_MARGIN
    )
    passed = bool(all(comparisons.values()))
    return passed, {
        "passed": passed,
        "comparisons": comparisons,
    }

def rank_pre_distribution_baseline_candidates(
    masks: list[np.ndarray],
    rows: list[dict],
    *,
    morphology_baseline_count: int = 30,
) -> tuple[int, dict[str, object]]:
    if len(masks) != len(rows):
        raise ValueError("Candidate mask and row counts do not match")
    if morphology_baseline_count <= 0 or morphology_baseline_count > len(masks):
        raise ValueError("Invalid morphology-baseline candidate count")
    morphology_baseline_errors = [
        index + 1
        for index, row in enumerate(rows[:morphology_baseline_count])
        if row.get("error")
    ]
    if morphology_baseline_errors:
        raise RuntimeError(
            "Frozen morphology baseline is incomplete; failed candidates: "
            f"{morphology_baseline_errors}"
        )

    morphology_baseline_rows = [
        dict(row) for row in rows[:morphology_baseline_count]
    ]
    morphology_baseline_position = rank_production_candidates(
        masks[:morphology_baseline_count],
        morphology_baseline_rows,
    )
    for index, morphology_baseline_row in enumerate(morphology_baseline_rows):
        rows[index]["morphology_baseline_mean_candidate_iou"] = (
            morphology_baseline_row.get(
                "mean_candidate_iou"
            )
        )
        rows[index]["morphology_baseline_auto_selection_score"] = (
            morphology_baseline_row.get("auto_selection_score")
        )
        rows[index]["_raw_morphology_baseline_mean_candidate_iou"] = (
            morphology_baseline_row.get("_raw_mean_candidate_iou")
        )
        rows[index]["_raw_morphology_baseline_auto_selection_score"] = (
            morphology_baseline_row.get("_raw_auto_selection_score")
        )
        rows[index]["morphology_baseline_auto_selected"] = (
            index == morphology_baseline_position
        )

    eligible = [index for index, row in enumerate(rows) if not row.get("error")]
    if not eligible:
        raise RuntimeError(f"All {len(rows)} ROI candidates failed")
    for index, row in enumerate(rows):
        row["selection_eligible"] = index in eligible
        if index not in eligible:
            row.update(
                {
                    "mean_candidate_iou": 0.0,
                    "auto_selection_score": -1.0,
                    "auto_selected": False,
                }
            )
    if len(eligible) != len(rows):
        challenger_position = morphology_baseline_position
        chosen_position = morphology_baseline_position
        rows[morphology_baseline_position].update(
            {
                "mean_candidate_iou": morphology_baseline_rows[
                    morphology_baseline_position
                ]["mean_candidate_iou"],
                "_raw_mean_candidate_iou": morphology_baseline_rows[
                    morphology_baseline_position
                ]["_raw_mean_candidate_iou"],
                "auto_selection_score": morphology_baseline_rows[
                    morphology_baseline_position
                ]["auto_selection_score"],
                "_raw_auto_selection_score": morphology_baseline_rows[
                    morphology_baseline_position
                ]["_raw_auto_selection_score"],
            }
        )
        guard = {
            "passed": False,
            "reason": (
                "candidate_error_fail_closed_to_"
                "morphology_baseline_incumbent"
            ),
            "comparisons": {},
        }
    else:
        rank_candidates_family_balanced(
            masks,
            rows,
            expected_families=PRE_DISTRIBUTION_BASELINE_CANDIDATE_FAMILIES,
        )
        incumbent_z_key = (
            int(rows[morphology_baseline_position]["z_start_0based"]),
            int(rows[morphology_baseline_position]["z_end_0based_inclusive"]),
            str(rows[morphology_baseline_position]["projection"]),
        )
        same_z_candidates = [
            index
            for index, row in enumerate(rows)
            if (
                int(row["z_start_0based"]),
                int(row["z_end_0based_inclusive"]),
                str(row["projection"]),
            )
            == incumbent_z_key
        ]
        challenger_position = max(
            same_z_candidates,
            key=lambda index: (
                float(rows[index]["_raw_auto_selection_score"]),
                -index,
            ),
        )
        if challenger_position == morphology_baseline_position:
            chosen_position = morphology_baseline_position
            guard = {
                "passed": True,
                "reason": (
                    "family_z_balanced_rank_retained_"
                    "morphology_baseline_incumbent"
                ),
                "comparisons": {},
            }
        else:
            passed, guard = challenger_dominates_incumbent(
                rows[challenger_position],
                rows[morphology_baseline_position],
            )
            guard["reason"] = (
                "same_z_challenger_dominated_incumbent"
                if passed
                else "challenger_rejected_by_non_regression_guard"
            )
            chosen_position = (
                challenger_position if passed else morphology_baseline_position
            )

    for index, row in enumerate(rows):
        row["family_balanced_challenger"] = index == challenger_position
        row["auto_selected"] = index == chosen_position
    details = {
        "morphology_baseline_incumbent_candidate": morphology_baseline_position
        + 1,
        "family_balanced_challenger_candidate": challenger_position + 1,
        "selected_candidate": chosen_position + 1,
        "guard": guard,
    }
    rows[chosen_position].update(
        {
            "selection_guard_reason": str(guard["reason"]),
            "morphology_baseline_incumbent_candidate": (
                morphology_baseline_position + 1
            ),
            "morphology_baseline_incumbent_score": morphology_baseline_rows[
                morphology_baseline_position
            ].get("auto_selection_score"),
            "family_balanced_challenger_candidate": challenger_position + 1,
            "family_balanced_challenger_score": rows[challenger_position].get(
                "auto_selection_score"
            ),
        }
    )
    return chosen_position, details

def rank_complete_production_candidates(
    masks: list[np.ndarray],
    rows: list[dict],
) -> tuple[int, dict[str, object]]:
    """Rank all 90 candidates while preserving the 60-candidate baseline outcome."""
    if len(masks) != TOTAL_CANDIDATE_COUNT or len(rows) != TOTAL_CANDIDATE_COUNT:
        raise ValueError(
            "Complete production ranking requires exactly "
            f"{TOTAL_CANDIDATE_COUNT} candidates"
        )

    pre_distribution_baseline_rows = [
        dict(row)
        for row in rows[:PRE_DISTRIBUTION_BASELINE_CANDIDATE_COUNT]
    ]
    pre_distribution_baseline_position, pre_distribution_baseline_details = (
        rank_pre_distribution_baseline_candidates(
            masks[:PRE_DISTRIBUTION_BASELINE_CANDIDATE_COUNT],
            pre_distribution_baseline_rows,
            morphology_baseline_count=MORPHOLOGY_BASELINE_CANDIDATE_COUNT,
        )
    )
    morphology_baseline_position = int(
        pre_distribution_baseline_details[
            "morphology_baseline_incumbent_candidate"
        ]
    ) - 1
    pre_distribution_baseline_z_key = (
        int(
            pre_distribution_baseline_rows[pre_distribution_baseline_position][
                "z_start_0based"
            ]
        ),
        int(
            pre_distribution_baseline_rows[pre_distribution_baseline_position][
                "z_end_0based_inclusive"
            ]
        ),
        str(
            pre_distribution_baseline_rows[pre_distribution_baseline_position][
                "projection"
            ]
        ),
    )
    morphology_baseline_z_key = (
        int(
            pre_distribution_baseline_rows[morphology_baseline_position][
                "z_start_0based"
            ]
        ),
        int(
            pre_distribution_baseline_rows[morphology_baseline_position][
                "z_end_0based_inclusive"
            ]
        ),
        str(
            pre_distribution_baseline_rows[morphology_baseline_position][
                "projection"
            ]
        ),
    )
    if pre_distribution_baseline_z_key != morphology_baseline_z_key:
        raise AssertionError(
            "Pre-distribution baseline incumbent changed the morphology-baseline "
            "Z interval"
        )

    morphology_baseline_fields = (
        "morphology_baseline_mean_candidate_iou",
        "morphology_baseline_auto_selection_score",
        "_raw_morphology_baseline_mean_candidate_iou",
        "_raw_morphology_baseline_auto_selection_score",
        "morphology_baseline_auto_selected",
    )
    for index, row in enumerate(rows):
        row["pre_distribution_baseline_auto_selected"] = (
            index == pre_distribution_baseline_position
        )
        if index < PRE_DISTRIBUTION_BASELINE_CANDIDATE_COUNT:
            baseline_row = pre_distribution_baseline_rows[index]
            row["pre_distribution_baseline_frozen_mean_candidate_iou"] = (
                baseline_row.get("mean_candidate_iou")
            )
            row["pre_distribution_baseline_frozen_auto_selection_score"] = (
                baseline_row.get("auto_selection_score")
            )
            row["_raw_pre_distribution_baseline_frozen_mean_candidate_iou"] = (
                baseline_row.get("_raw_mean_candidate_iou")
            )
            row[
                "_raw_pre_distribution_baseline_frozen_auto_selection_score"
            ] = baseline_row.get("_raw_auto_selection_score")
            for field_name in morphology_baseline_fields:
                if field_name in baseline_row:
                    row[field_name] = baseline_row[field_name]

    candidate_errors = [
        index + 1 for index, row in enumerate(rows) if row.get("error")
    ]
    if candidate_errors:
        challenger_position = pre_distribution_baseline_position
        chosen_position = pre_distribution_baseline_position
        for index, row in enumerate(rows):
            row["selection_eligible"] = (
                index < PRE_DISTRIBUTION_BASELINE_CANDIDATE_COUNT
                and not bool(row.get("error"))
            )
            if index < PRE_DISTRIBUTION_BASELINE_CANDIDATE_COUNT:
                baseline_row = pre_distribution_baseline_rows[index]
                for field_name in (
                    "mean_candidate_iou",
                    "_raw_mean_candidate_iou",
                    "auto_selection_score",
                    "_raw_auto_selection_score",
                ):
                    if field_name in baseline_row:
                        row[field_name] = baseline_row[field_name]
            else:
                row.update(
                    {
                        "mean_candidate_iou": 0.0,
                        "_raw_mean_candidate_iou": 0.0,
                        "auto_selection_score": -1.0,
                        "_raw_auto_selection_score": -1.0,
                    }
                )
        guard: dict[str, object] = {
            "passed": False,
            "reason": (
                "candidate_error_fail_closed_to_"
                "pre_distribution_baseline_incumbent"
            ),
            "comparisons": {},
            "failed_candidates": candidate_errors,
        }
    else:
        for row in rows:
            row["selection_eligible"] = True
        rank_candidates_family_balanced(
            masks,
            rows,
            expected_families=EXPECTED_CANDIDATE_FAMILIES,
        )
        same_z_candidates = [
            index
            for index, row in enumerate(rows)
            if (
                int(row["z_start_0based"]),
                int(row["z_end_0based_inclusive"]),
                str(row["projection"]),
            )
            == pre_distribution_baseline_z_key
        ]
        challenger_position = max(
            same_z_candidates,
            key=lambda index: (
                float(rows[index]["_raw_auto_selection_score"]),
                -index,
            ),
        )
        if challenger_position == pre_distribution_baseline_position:
            chosen_position = pre_distribution_baseline_position
            guard = {
                "passed": True,
                "reason": (
                    "family_z_balanced_rank_retained_"
                    "pre_distribution_baseline_incumbent"
                ),
                "comparisons": {},
            }
        else:
            passed, guard = challenger_dominates_incumbent(
                rows[challenger_position],
                rows[pre_distribution_baseline_position],
            )
            guard["reason"] = (
                "same_z_challenger_dominated_"
                "pre_distribution_baseline_incumbent"
                if passed
                else (
                    "challenger_rejected_by_"
                    "pre_distribution_baseline_non_regression_guard"
                )
            )
            chosen_position = (
                challenger_position
                if passed
                else pre_distribution_baseline_position
            )

    for index, row in enumerate(rows):
        row["family_balanced_challenger"] = index == challenger_position
        row["auto_selected"] = index == chosen_position
    details = {
        "morphology_baseline_incumbent_candidate": morphology_baseline_position
        + 1,
        "pre_distribution_baseline_incumbent_candidate": (
            pre_distribution_baseline_position + 1
        ),
        "family_balanced_challenger_candidate": challenger_position + 1,
        "selected_candidate": chosen_position + 1,
        "pre_distribution_baseline_selection": pre_distribution_baseline_details,
        "guard": guard,
    }
    rows[chosen_position].update(
        {
            "selection_guard_reason": str(guard["reason"]),
            "morphology_baseline_incumbent_candidate": (
                morphology_baseline_position + 1
            ),
            "morphology_baseline_incumbent_score": pre_distribution_baseline_rows[
                morphology_baseline_position
            ].get(
                "morphology_baseline_auto_selection_score",
                pre_distribution_baseline_rows[
                    morphology_baseline_position
                ].get("auto_selection_score"),
            ),
            "pre_distribution_baseline_incumbent_candidate": (
                pre_distribution_baseline_position + 1
            ),
            "pre_distribution_baseline_incumbent_score_frozen": (
                pre_distribution_baseline_rows[
                    pre_distribution_baseline_position
                ].get("auto_selection_score")
            ),
            "pre_distribution_baseline_incumbent_score_complete_pool": rows[
                pre_distribution_baseline_position
            ].get(
                "auto_selection_score"
            ),
            "family_balanced_challenger_candidate": challenger_position + 1,
            "family_balanced_challenger_score": rows[challenger_position].get(
                "auto_selection_score"
            ),
        }
    )
    return chosen_position, details


def make_morphology_baseline_spec(
    name: str,
    z_mode: str = "peak_narrow",
    *,
    threshold_scale: float = 0.94,
    smooth_sigma: float = 1.0,
    min_area: int = 70,
    anchor_area: int = 1150,
    bridge_radius: int = 9,
    low_percentile: float = 78,
    seed_percentile: float = 93,
    seed_min_area: int = 140,
    max_area_fraction: float = 0.34,
    cellpose_cellprob: float = 0.0,
    cellpose_diameter: float = 70,
    dapi_support_radius: int = 16,
    outline_smooth_sigma: float = 1.7,
    artifact_min_area: int = 650,
    artifact_near_radius: int = 36,
    process_eccentricity: float = 0.70,
    process_major_axis: float = 30,
    branch_support_percentile: float = 38,
    branch_support_radius: int = 1,
    max_process_half_width: float = 9,
    soma_protect_radius: int = 21,
    outline_hole_min_area: int = 70,
    require_soma_anchor: bool = False,
    soma_anchor_radius: int = 4,
    soma_anchor_percentile: float = 84.0,
    soma_core_radius: float = 8.0,
    soma_anchor_min_pixels: int = 8,
    anchor_component_min_area: int = 3000,
    connection_radius: int = 3,
    connection_support_percentile: float = 84.0,
    fine_branch_recovery: bool = False,
    fine_branch_detail_percentile: float = 92.0,
    fine_branch_intensity_percentile: float = 68.0,
    fine_branch_min_area: int = 16,
    fine_branch_min_major_axis: float = 12.0,
    fine_branch_min_eccentricity: float = 0.62,
    fine_branch_gap_radius: int = 2,
    fine_branch_background_sigma: float = 7.0,
    fine_branch_single_channel_offset: float = 4.0,
    exclude_border_components: bool = False,
    border_margin: int = 12,
    edge_qc_margin: int = 48,
) -> TestSpec:
    return TestSpec(
        name=name,
        z_mode=z_mode,
        projection="max",
        method="top_hat_union",
        egfp_weight=0.55,
        gfap_weight=0.45,
        smooth_sigma=smooth_sigma,
        threshold_scale=threshold_scale,
        min_area=min_area,
        close_radius=1,
        dilate_radius=1,
        cleanup_mode="hybrid_reconstruct",
        anchor_area=anchor_area,
        bridge_radius=bridge_radius,
        low_percentile=low_percentile,
        seed_percentile=seed_percentile,
        seed_min_area=seed_min_area,
        max_area_fraction=max_area_fraction,
        cellpose=True,
        cellpose_cellprob=cellpose_cellprob,
        cellpose_diameter=cellpose_diameter,
        cellpose_max_side=2048,
        dapi_support_radius=dapi_support_radius,
        outline_smooth_sigma=outline_smooth_sigma,
        outline_epsilon=2.2,
        artifact_filter=True,
        artifact_min_area=artifact_min_area,
        artifact_near_radius=artifact_near_radius,
        process_eccentricity=process_eccentricity,
        process_major_axis=process_major_axis,
        branch_refine=True,
        branch_support_percentile=branch_support_percentile,
        branch_support_radius=branch_support_radius,
        max_process_half_width=max_process_half_width,
        soma_protect_radius=soma_protect_radius,
        outline_hole_min_area=outline_hole_min_area,
        require_soma_anchor=require_soma_anchor,
        soma_anchor_radius=soma_anchor_radius,
        soma_anchor_percentile=soma_anchor_percentile,
        soma_core_radius=soma_core_radius,
        soma_anchor_min_pixels=soma_anchor_min_pixels,
        anchor_component_min_area=anchor_component_min_area,
        connection_radius=connection_radius,
        connection_support_percentile=connection_support_percentile,
        fine_branch_recovery=fine_branch_recovery,
        fine_branch_detail_percentile=fine_branch_detail_percentile,
        fine_branch_intensity_percentile=fine_branch_intensity_percentile,
        fine_branch_min_area=fine_branch_min_area,
        fine_branch_min_major_axis=fine_branch_min_major_axis,
        fine_branch_min_eccentricity=fine_branch_min_eccentricity,
        fine_branch_gap_radius=fine_branch_gap_radius,
        fine_branch_background_sigma=fine_branch_background_sigma,
        fine_branch_single_channel_offset=fine_branch_single_channel_offset,
        exclude_border_components=exclude_border_components,
        border_margin=border_margin,
        edge_qc_margin=edge_qc_margin,
    )

def morphology_baseline_template_specs() -> list[TestSpec]:
    variants = [
        (
            "process",
            dict(
                threshold_scale=0.93,
                min_area=55,
                anchor_area=1150,
                bridge_radius=9,
                low_percentile=80,
                seed_min_area=130,
                max_area_fraction=0.32,
                artifact_min_area=650,
                artifact_near_radius=32,
                branch_support_percentile=44,
                max_process_half_width=8,
                soma_protect_radius=19,
                soma_anchor_percentile=82,
                soma_core_radius=7,
                soma_anchor_min_pixels=6,
                anchor_component_min_area=2800,
                connection_radius=5,
                connection_support_percentile=82,
            ),
        ),
        (
            "balanced",
            dict(
                threshold_scale=0.96,
                min_area=75,
                anchor_area=1250,
                bridge_radius=8,
                low_percentile=81,
                seed_min_area=140,
                max_area_fraction=0.30,
                artifact_min_area=750,
                artifact_near_radius=30,
                branch_support_percentile=48,
                max_process_half_width=7,
                soma_protect_radius=18,
                soma_anchor_percentile=84,
                soma_core_radius=8,
                soma_anchor_min_pixels=8,
                anchor_component_min_area=3600,
                connection_radius=3,
                connection_support_percentile=84,
            ),
        ),
        (
            "clean",
            dict(
                threshold_scale=1.00,
                min_area=95,
                anchor_area=1500,
                bridge_radius=7,
                low_percentile=83,
                seed_percentile=94,
                seed_min_area=175,
                max_area_fraction=0.28,
                artifact_min_area=950,
                artifact_near_radius=27,
                process_eccentricity=0.76,
                process_major_axis=36,
                branch_support_percentile=52,
                max_process_half_width=7,
                soma_protect_radius=18,
                soma_anchor_percentile=85,
                soma_core_radius=9,
                soma_anchor_min_pixels=10,
                anchor_component_min_area=4500,
                connection_radius=2,
                connection_support_percentile=86,
            ),
        ),
        (
            "strict",
            dict(
                threshold_scale=1.04,
                min_area=120,
                anchor_area=1800,
                bridge_radius=6,
                low_percentile=85,
                seed_percentile=95,
                seed_min_area=210,
                max_area_fraction=0.26,
                artifact_min_area=1200,
                artifact_near_radius=24,
                process_eccentricity=0.80,
                process_major_axis=42,
                branch_support_percentile=56,
                max_process_half_width=6,
                soma_protect_radius=17,
                soma_anchor_percentile=87,
                soma_core_radius=10,
                soma_anchor_min_pixels=12,
                anchor_component_min_area=6000,
                connection_radius=1,
                connection_support_percentile=88,
            ),
        ),
    ]
    candidates: list[TestSpec] = []
    for variant, parameters in variants:
        candidates.append(
            make_morphology_baseline_spec(
                f"morphology_baseline_template_auto17_{variant}",
                "auto_17",
                require_soma_anchor=True,
                **parameters,
            )
        )
    return candidates

def adaptive_z_window_widths(n_slices: int) -> list[int]:
    if n_slices < 10:
        raise ValueError(
            f"Morphology baseline requires at least 10 Z slices, got {n_slices}"
        )
    if n_slices >= 47:
        return [17, 23, 31, 39, 47]

    allowed = list(range(3, n_slices + 1, 2))
    if n_slices not in allowed:
        allowed.append(n_slices)
    targets = [
        n_slices * 0.49,
        n_slices * 0.66,
        n_slices * 0.83,
        n_slices * 0.94,
        float(n_slices),
    ]
    selected: list[int] = []
    for target in targets:
        available = [width for width in allowed if width not in selected]
        width = min(available, key=lambda value: (abs(value - target), value))
        selected.append(width)
    return sorted(selected)

def morphology_baseline_specs(n_slices: int) -> list[TestSpec]:
    """Build six morphology-baseline profiles per adaptive Z window."""
    templates = morphology_baseline_template_specs()
    profile_templates = [
        (
            templates[0],
            dict(
                variant="process",
                fine_branch_detail_percentile=92.0,
                fine_branch_intensity_percentile=68.0,
                fine_branch_min_area=16,
                fine_branch_min_major_axis=12.0,
                fine_branch_min_eccentricity=0.62,
                fine_branch_gap_radius=2,
            ),
        ),
        (
            templates[1],
            dict(
                variant="balanced",
                fine_branch_detail_percentile=93.0,
                fine_branch_intensity_percentile=70.0,
                fine_branch_min_area=18,
                fine_branch_min_major_axis=13.0,
                fine_branch_min_eccentricity=0.65,
                fine_branch_gap_radius=2,
            ),
        ),
        (
            templates[2],
            dict(
                variant="clean",
                fine_branch_detail_percentile=94.0,
                fine_branch_intensity_percentile=72.0,
                fine_branch_min_area=20,
                fine_branch_min_major_axis=15.0,
                fine_branch_min_eccentricity=0.68,
                fine_branch_gap_radius=2,
            ),
        ),
        (
            templates[3],
            dict(
                variant="strict",
                fine_branch_detail_percentile=95.0,
                fine_branch_intensity_percentile=74.0,
                fine_branch_min_area=22,
                fine_branch_min_major_axis=18.0,
                fine_branch_min_eccentricity=0.72,
                fine_branch_gap_radius=1,
            ),
        ),
        (
            templates[0],
            dict(
                variant="process_refined",
                fine_branch_detail_percentile=92.5,
                fine_branch_intensity_percentile=69.0,
                fine_branch_min_area=17,
                fine_branch_min_major_axis=12.5,
                fine_branch_min_eccentricity=0.64,
                fine_branch_gap_radius=2,
            ),
        ),
        (
            templates[2],
            dict(
                variant="clean_refined",
                fine_branch_detail_percentile=93.5,
                fine_branch_intensity_percentile=71.0,
                fine_branch_min_area=19,
                fine_branch_min_major_axis=14.0,
                fine_branch_min_eccentricity=0.67,
                fine_branch_gap_radius=2,
            ),
        ),
    ]
    candidates: list[TestSpec] = []
    for width in adaptive_z_window_widths(n_slices):
        for template, profile in profile_templates:
            variant = str(profile["variant"])
            parameters = {key: value for key, value in profile.items() if key != "variant"}
            candidates.append(
                replace(
                    template,
                    name=f"morphology_baseline_auto{width}_{variant}",
                    z_mode=f"auto_{width}",
                    fine_branch_recovery=True,
                    fine_branch_background_sigma=7.0,
                    fine_branch_single_channel_offset=4.0,
                    exclude_border_components=True,
                    border_margin=12,
                    edge_qc_margin=48,
                    preserve_complete_border_components=True,
                    border_complete_soma_margin=48,
                    border_complete_min_area_ratio=0.75,
                    border_complete_min_interior_fraction=0.75,
                    **parameters,
                )
            )
    if len(candidates) != 30:
        raise AssertionError(
            "Morphology baseline must contain exactly 30 Whole-ROI candidates, "
            f"got {len(candidates)}"
        )
    return candidates


def structural_snr_sensitivity(structural_stacks: dict[str, np.ndarray]) -> float:
    """Map structural-channel contrast/noise to a bounded branch sensitivity."""
    channel_scores: list[float] = []
    for stack in structural_stacks.values():
        sample = np.asarray(stack[:, ::8, ::8], dtype=np.float64).reshape(-1)
        if sample.size == 0:
            continue
        p50, p75, p995 = np.percentile(sample, [50.0, 75.0, 99.5])
        background = sample[sample <= p75]
        if background.size:
            background_median = float(np.median(background))
            noise = 1.4826 * float(
                np.median(np.abs(background - background_median))
            )
        else:
            noise = 0.0
        signal = max(float(p995 - p50), 0.0)
        scale = max(noise, 0.02 * signal, np.finfo(np.float64).eps)
        channel_scores.append(signal / scale)
    if not channel_scores:
        return 0.5
    robust_score = float(np.median(np.asarray(channel_scores, dtype=np.float64)))
    lower = math.log1p(3.0)
    upper = math.log1p(20.0)
    return float(np.clip((math.log1p(robust_score) - lower) / (upper - lower), 0.0, 1.0))

def morphology_with_adaptive_refinement_specs(
    n_slices: int,
    structural_stacks: dict[str, np.ndarray],
) -> list[TestSpec]:
    """Retain the validated 30 candidates and add 20 complementary profiles."""
    morphology_baseline_candidates = morphology_baseline_specs(n_slices)
    templates = morphology_baseline_template_specs()
    snr_sensitivity = structural_snr_sensitivity(structural_stacks)
    adaptive_detail = 94.5 - 2.0 * snr_sensitivity
    adaptive_intensity = 73.0 - 3.0 * snr_sensitivity
    adaptive_area = int(round(21.0 - 4.0 * snr_sensitivity))
    adaptive_major_axis = 16.0 - 3.0 * snr_sensitivity
    adaptive_eccentricity = 0.70 - 0.05 * snr_sensitivity
    adaptive_gap = 2 if snr_sensitivity >= 0.35 else 1
    added_profiles = [
        (
            templates[0],
            dict(
                variant="fine_process_guarded",
                fine_branch_detail_percentile=91.5,
                fine_branch_intensity_percentile=66.5,
                fine_branch_min_area=14,
                fine_branch_min_major_axis=10.5,
                fine_branch_min_eccentricity=0.58,
                fine_branch_gap_radius=2,
            ),
        ),
        (
            replace(
                templates[3],
                threshold_scale=1.06,
                min_area=135,
                anchor_area=1950,
                bridge_radius=5,
                low_percentile=86,
                seed_percentile=96,
                seed_min_area=230,
                max_area_fraction=0.24,
                artifact_min_area=1400,
                artifact_near_radius=22,
                branch_support_percentile=58,
                max_process_half_width=6,
                soma_anchor_percentile=88,
                soma_anchor_min_pixels=13,
                anchor_component_min_area=6500,
                connection_radius=1,
                connection_support_percentile=90,
            ),
            dict(
                variant="precision_guarded",
                fine_branch_detail_percentile=96.0,
                fine_branch_intensity_percentile=76.0,
                fine_branch_min_area=24,
                fine_branch_min_major_axis=20.0,
                fine_branch_min_eccentricity=0.75,
                fine_branch_gap_radius=1,
            ),
        ),
        (
            replace(
                templates[2],
                threshold_scale=1.02,
                min_area=105,
                anchor_area=1650,
                bridge_radius=5,
                low_percentile=84,
                seed_percentile=95,
                seed_min_area=190,
                max_area_fraction=0.27,
                artifact_min_area=1050,
                artifact_near_radius=24,
                branch_support_percentile=55,
                max_process_half_width=6,
                soma_anchor_percentile=86,
                soma_anchor_min_pixels=11,
                anchor_component_min_area=5000,
                connection_radius=1,
                connection_support_percentile=90,
            ),
            dict(
                variant="merge_resistant",
                fine_branch_detail_percentile=94.5,
                fine_branch_intensity_percentile=73.0,
                fine_branch_min_area=20,
                fine_branch_min_major_axis=16.0,
                fine_branch_min_eccentricity=0.72,
                fine_branch_gap_radius=1,
            ),
        ),
        (
            templates[1],
            dict(
                variant="structural_snr_adaptive",
                fine_branch_detail_percentile=adaptive_detail,
                fine_branch_intensity_percentile=adaptive_intensity,
                fine_branch_min_area=adaptive_area,
                fine_branch_min_major_axis=adaptive_major_axis,
                fine_branch_min_eccentricity=adaptive_eccentricity,
                fine_branch_gap_radius=adaptive_gap,
            ),
        ),
    ]
    added: list[TestSpec] = []
    for width in adaptive_z_window_widths(n_slices):
        for template, profile in added_profiles:
            variant = str(profile["variant"])
            parameters = {
                key: value for key, value in profile.items() if key != "variant"
            }
            added.append(
                replace(
                    template,
                    name=f"structural_refinement_auto{width}_{variant}",
                    z_mode=f"auto_{width}",
                    fine_branch_recovery=True,
                    fine_branch_background_sigma=7.0,
                    fine_branch_single_channel_offset=4.0,
                    exclude_border_components=True,
                    border_margin=12,
                    edge_qc_margin=48,
                    preserve_complete_border_components=True,
                    border_complete_soma_margin=48,
                    border_complete_min_area_ratio=0.75,
                    border_complete_min_interior_fraction=0.75,
                    **parameters,
                )
            )
    candidates = [*morphology_baseline_candidates, *added]
    if len(candidates) != 50:
        raise AssertionError(
            "Morphology with adaptive refinement requires exactly 50 Whole-ROI "
            f"candidates, got {len(candidates)}"
        )
    widths = adaptive_z_window_widths(n_slices)
    for width in widths:
        group = [spec for spec in candidates if spec.z_mode == f"auto_{width}"]
        if len(group) != 10:
            raise AssertionError(
                f"Z width {width} must contain exactly 10 candidate profiles"
            )
        signatures = {
            tuple(sorted({
                key: value
                for key, value in asdict(spec).items()
                if key not in {"name", "z_mode"}
            }.items()))
            for spec in group
        }
        if len(signatures) != 10:
            raise AssertionError(f"Z width {width} contains duplicate candidates")
    return candidates

def morphology_with_structural_refinement_specs(
    n_slices: int,
    structural_stacks: dict[str, np.ndarray],
) -> list[TestSpec]:
    """Retain 50 candidates and append ten orthogonal structural refinements."""
    candidates = morphology_with_adaptive_refinement_specs(
        n_slices,
        structural_stacks,
    )
    new_candidates: list[TestSpec] = []
    for width in adaptive_z_window_widths(n_slices):
        z_mode = f"auto_{width}"
        fine_process = next(
            spec
            for spec in candidates
            if spec.z_mode == z_mode
            and candidate_profile_name(spec) == "fine_process_guarded"
        )
        new_candidates.extend(
            [
                replace(
                    fine_process,
                    name=f"structural_refinement_auto{width}_channel_consensus_guarded",
                    fine_branch_evidence_mode="channel_consensus",
                    fine_branch_consensus_radius=1,
                ),
                replace(
                    fine_process,
                    name=f"structural_refinement_auto{width}_topology_continuity_guarded",
                    fine_branch_evidence_mode="topology_continuity",
                    fine_branch_topology_max_gap=4,
                    fine_branch_topology_min_skeleton=10,
                    fine_branch_topology_max_hops=3,
                ),
            ]
        )
    candidates.extend(new_candidates)
    if len(candidates) != PRE_DISTRIBUTION_BASELINE_CANDIDATE_COUNT:
        raise AssertionError(
            "Pre-distribution baseline requires exactly "
            f"{PRE_DISTRIBUTION_BASELINE_CANDIDATE_COUNT} Whole-ROI candidates, "
            f"got {len(candidates)}"
        )
    widths = adaptive_z_window_widths(n_slices)
    for width in widths:
        group = [spec for spec in candidates if spec.z_mode == f"auto_{width}"]
        if len(group) != PRE_DISTRIBUTION_BASELINE_PROFILES_PER_Z:
            raise AssertionError(
                "Z width "
                f"{width} must contain exactly "
                f"{PRE_DISTRIBUTION_BASELINE_PROFILES_PER_Z} "
                "candidate profiles"
            )
        signatures = {
            tuple(
                sorted(
                    {
                        key: value
                        for key, value in asdict(spec).items()
                        if key not in {"name", "z_mode"}
                    }.items()
                )
            )
            for spec in group
        }
        if len(signatures) != PRE_DISTRIBUTION_BASELINE_PROFILES_PER_Z:
            raise AssertionError(f"Z width {width} contains duplicate candidates")
    return candidates


def candidate_profile_name(spec: TestSpec) -> str:
    match = re.search(r"_auto\d+_(.+)$", spec.name)
    if match is None:
        raise ValueError(f"Candidate name does not encode a profile: {spec.name}")
    return match.group(1)

def candidate_profile_family(spec: TestSpec) -> str:
    profile_name = candidate_profile_name(spec)
    families = {
        "process": "process_sensitivity",
        "process_refined": "process_sensitivity",
        "fine_process_guarded": "process_sensitivity",
        "balanced": "balanced_adaptive",
        "structural_snr_adaptive": "balanced_adaptive",
        "clean": "precision",
        "clean_refined": "precision",
        "precision_guarded": "precision",
        "strict": "strict_merge",
        "merge_resistant": "strict_merge",
        "channel_consensus_guarded": "channel_consensus",
        "topology_continuity_guarded": "topology_continuity",
        "gmm_process_guarded": "distributional_threshold",
        "gmm_balanced": "distributional_threshold",
        "gmm_precision_guarded": "distributional_threshold",
        "gmm_merge_resistant": "distributional_threshold",
        "gmm_channel_consensus_guarded": "distributional_threshold",
        "gmm_topology_continuity_guarded": "distributional_threshold",
    }
    try:
        return families[profile_name]
    except KeyError as exc:
        raise ValueError(f"Unknown candidate profile family: {profile_name}") from exc

def candidate_module_name(spec: TestSpec) -> str:
    profile_name = candidate_profile_name(spec)
    if profile_name.startswith("gmm_"):
        return "distributional_threshold"
    if profile_name in {
        "process",
        "balanced",
        "clean",
        "strict",
        "process_refined",
        "clean_refined",
    }:
        return "morphology_baseline"
    return "structural_refinement"

def candidate_module_display_name(spec: TestSpec) -> str:
    return {
        "morphology_baseline": "Morphology Baseline",
        "structural_refinement": "Structural Refinement",
        "distributional_threshold": "Distributional Threshold",
    }[candidate_module_name(spec)]

def candidate_threshold_display_name(spec: TestSpec) -> str:
    return "log1p_gmm" if spec.method == "log1p_gmm" else str(spec.method)


def complete_candidate_specs(
    n_slices: int,
    structural_stacks: dict[str, np.ndarray],
) -> list[TestSpec]:
    """Retain the 60-candidate baseline and append 30 distributional challengers."""
    candidates = morphology_with_structural_refinement_specs(
        n_slices,
        structural_stacks,
    )
    added: list[TestSpec] = []
    for width in adaptive_z_window_widths(n_slices):
        z_mode = f"auto_{width}"
        profiles = {
            candidate_profile_name(spec): spec
            for spec in candidates
            if spec.z_mode == z_mode
        }
        templates = (
            (profiles["fine_process_guarded"], "gmm_process_guarded", 0.90),
            (profiles["structural_snr_adaptive"], "gmm_balanced", 0.98),
            (profiles["precision_guarded"], "gmm_precision_guarded", 1.08),
            (profiles["merge_resistant"], "gmm_merge_resistant", 1.04),
            (
                profiles["channel_consensus_guarded"],
                "gmm_channel_consensus_guarded",
                1.00,
            ),
            (
                profiles["topology_continuity_guarded"],
                "gmm_topology_continuity_guarded",
                0.94,
            ),
        )
        for template, profile_name, threshold_scale in templates:
            added.append(
                replace(
                    template,
                    name=f"distributional_threshold_auto{width}_{profile_name}",
                    method="log1p_gmm",
                    threshold_scale=threshold_scale,
                )
            )
    candidates.extend(added)
    if len(candidates) != TOTAL_CANDIDATE_COUNT:
        raise AssertionError(
            f"Complete candidate catalog requires exactly {TOTAL_CANDIDATE_COUNT} "
            "Whole-ROI candidates, "
            f"got {len(candidates)}"
        )
    widths = adaptive_z_window_widths(n_slices)
    for width in widths:
        group = [spec for spec in candidates if spec.z_mode == f"auto_{width}"]
        if len(group) != EXPECTED_PROFILES_PER_Z:
            raise AssertionError(
                f"Z width {width} must contain exactly {EXPECTED_PROFILES_PER_Z} "
                "candidate profiles"
            )
        signatures = {
            tuple(
                sorted(
                    {
                        key: value
                        for key, value in asdict(spec).items()
                        if key not in {"name", "z_mode"}
                    }.items()
                )
            )
            for spec in group
        }
        if len(signatures) != EXPECTED_PROFILES_PER_Z:
            raise AssertionError(f"Z width {width} contains duplicate candidates")
    if (
        sum(
            candidate_module_name(spec) == "distributional_threshold"
            for spec in candidates
        )
        != 30
    ):
        raise AssertionError(
            "Distributional threshold group must contain exactly 30 candidates"
        )
    return candidates


def write_analysis_report(
    path: Path,
    *,
    input_dir: Path,
    paths: dict[str, Path],
    metadata: dict[str, dict],
    structural_channels: list[str],
    measurement: str,
    best_row: dict,
    candidate_rows: list[dict],
    candidate_specs: list[TestSpec],
    roi_count: int,
    compartment_metrics: dict,
    fiji_status: str,
    fiji_details: dict | None = None,
) -> None:
    if len(candidate_specs) != len(candidate_rows):
        raise ValueError(
            "Candidate report metadata is incomplete: "
            f"{len(candidate_specs)} specifications for {len(candidate_rows)} rows"
        )
    age_profile = compartment_metrics.get("age_profile", {})
    age_score = age_profile.get("neonatal_score")
    age_margin = age_profile.get("confidence_margin")
    tagged_files = age_profile.get("tagged_files", [])
    neonatal_3d = compartment_metrics.get(
        "neonatal_3d_validation",
        compartment_metrics.get(
            "nucleus_3d_inventory",
            {
                "status": "not_run_calibrated_Z_unavailable",
                "method": "object-preserving calibrated 3D DAPI inventory",
                "measurement_channel_used": False,
                "candidate_count": 0,
                "accepted_count": 0,
                "rejected_count": 0,
                "per_nucleus": [],
            },
        ),
    )
    canonical_nuclei = neonatal_3d.get("canonical_resolution", {})
    axial_guard = compartment_metrics.get("axial_truncation_guard", {})
    projected_foreign_guard = compartment_metrics.get(
        "projected_foreign_soma_guard",
        {},
    )
    canonical_owner_extent_completion = compartment_metrics.get(
        "soma_nuclear_extent_completion",
        {},
    )
    same_id_soma_reconciliation = compartment_metrics.get(
        "same_id_disconnected_soma_reconciliation",
        {},
    )
    ownership_guard = compartment_metrics.get("nucleus_ownership_guard", {})
    shared_whole = compartment_metrics.get("shared_whole_baseline", {})
    soma_identity_gate = compartment_metrics.get(
        "soma_identity_gate",
        {
            "enabled": False,
            "pre_gate_roi_count": compartment_metrics["roi_count"],
            "post_gate_roi_count": compartment_metrics["roi_count"],
            "removed_count": 0,
            "removed_area_px": 0,
            "removed_area_fraction": 0.0,
            "removed_pre_gate_ids": [],
            "id_mapping": {},
            "details": [],
        },
    )
    shared_filter = shared_whole.get("morphology_filter", {})
    shared_gap = shared_whole.get("branch_gap_restoration", {})
    effective_filter = shared_filter or compartment_metrics["morphology_filter"]
    effective_gap = shared_gap or compartment_metrics["branch_gap_restoration"]
    candidate_times = _RUNTIME_TIMINGS["candidate_postprocess_seconds"]
    cellpose_events = _RUNTIME_TIMINGS["cellpose_inference_events"]
    assert isinstance(candidate_times, list)
    assert isinstance(cellpose_events, list)
    module_counts = {
        name: sum(candidate_module_name(spec) == name for spec in candidate_specs)
        for name in (
            "morphology_baseline",
            "structural_refinement",
            "distributional_threshold",
        )
    }
    lines = [
        "PROJECT LEAP 2D ANALYSIS REPORT",
        "===============================",
        "",
        "[Run]",
        f"Product: {PRODUCT_DISPLAY_NAME}",
        f"Scientific Core: {ANALYSIS_CORE_NAME}",
        f"Pipeline: {PIPELINE_NAME}",
        f"Generated: {time.strftime('%Y-%m-%d %H:%M:%S %Z')}",
        f"Input Folder: {input_dir}",
        "",
        "[Channels]",
    ]
    for channel in ("DAPI", "eGFP", "GFAP", *MEASUREMENT_CHANNELS):
        if channel in paths:
            lines.append(f"{channel}: {paths[channel].name}")
    lines.extend(
        [
            f"ROI Definition: DAPI + {' + '.join(structural_channels)}",
            f"Measurement: {measurement}",
            "Measurement Channel Used for ROI Definition: No",
            f"Image Geometry: ZYX {tuple(metadata['DAPI']['shape'])}, {metadata['DAPI']['dtype']}",
            "Channel Shape and Calibration Match: Yes",
            "",
            "[Candidate Architecture and Runtime]",
            "Module Counts: "
            f"Morphology Baseline={module_counts['morphology_baseline']}; "
            f"Structural Refinement={module_counts['structural_refinement']}; "
            "Distributional Threshold="
            f"{module_counts['distributional_threshold']}",
            f"Z Intervals: {EXPECTED_Z_INTERVAL_COUNT}; "
            f"Profiles per Z Interval: {EXPECTED_PROFILES_PER_Z}",
            "Shared Preparation: each Z interval generates one projection set, "
            "one Cellpose-SAM mask, and one reusable DAPI/Sato/top-hat feature context.",
            f"Cellpose Model Initialization: "
            f"{float(_RUNTIME_TIMINGS['cellpose_model_init_seconds']):.6f} s",
            f"Cellpose Inference Calls: {len(cellpose_events)}",
        ]
    )
    for index, event in enumerate(cellpose_events, start=1):
        lines.append(
            f"  Inference {index:02d}: status="
            f"{'ok' if event.get('success', False) else 'failed'}; "
            f"device={event.get('device', 'unknown')}; "
            f"max_side={event.get('max_side', 'unknown')}; "
            f"diameter={event.get('diameter', 'unknown')}; "
            f"cellprob={event.get('cellprob', 'unknown')}; "
            f"time={float(event.get('seconds', 0.0)):.6f} s"
        )
    lines.extend(
        [
            f"Candidate Stage Wall Time: "
            f"{float(_RUNTIME_TIMINGS['candidate_stage_wall_seconds']):.6f} s",
            f"Candidate Worker Compute Total: "
            f"{sum(float(value) for value in candidate_times):.6f} s",
            f"Candidate Ranking Time: "
            f"{float(_RUNTIME_TIMINGS['rank_candidates_seconds']):.6f} s",
            f"Soma/Processes and 3D Ownership Time: "
            f"{float(_RUNTIME_TIMINGS['compartment_split_seconds']):.6f} s",
            f"Fiji Startup to Review Ready: "
            + (
                f"{float(fiji_details['fiji_startup_seconds']):.6f} s"
                if fiji_details and fiji_details.get("fiji_startup_seconds") is not None
                else "Not completed"
            ),
            f"Fiji Native Measurement Time: "
            + (
                f"{float(fiji_details['measurement_seconds']):.6f} s"
                if fiji_details and fiji_details.get("measurement_seconds") is not None
                else "Not completed"
            ),
            "",
            "Candidate Inventory:",
            "ID | Module | Family | Profile | Threshold | Z | Area_px | Components | "
            "Soma_Support | Fine_Added_px | Eligible | Worker_Time_s",
        ]
    )
    for index, (spec, row) in enumerate(zip(candidate_specs, candidate_rows)):
        worker_time = (
            f"{float(candidate_times[index]):.6f}"
            if index < len(candidate_times)
            else "Not available"
        )
        lines.append(
            f"{int(row['candidate']):02d} | "
            f"{candidate_module_display_name(spec)} | "
            f"{candidate_profile_family(spec)} | "
            f"{candidate_profile_name(spec)} | "
            f"{candidate_threshold_display_name(spec)} | "
            f"{int(row['z_start_1based'])}-{int(row['z_end_1based_inclusive'])} | "
            f"{int(row['mask_area_px'])} | "
            f"{int(row['connected_components'])} | "
            f"{int(row['soma_supported_components'])} | "
            f"{int(row['fine_branch_retained_px'])} | "
            f"{bool(row.get('selection_eligible'))} | "
            f"{worker_time}"
        )
    lines.extend(
        [
            "",
            "[Selected Analysis]",
            f"Astrocyte Profile: {age_profile.get('profile', 'unknown')}",
            f"Profile Decision Source: {age_profile.get('source', 'unknown')}",
            f"Filename Profile Evidence: {', '.join(tagged_files) if tagged_files else 'None'}",
            f"Morphology Neonatal Score: {age_score if age_score is not None else 'Not run'}",
            f"Morphology Decision Threshold: {age_profile.get('threshold', AGE_PROFILE_THRESHOLD)}",
            f"Morphology Confidence Margin: {age_margin if age_margin is not None else 'Not applicable'}",
            f"Candidates: {len(candidate_rows)} evaluated; "
            f"{sum(bool(row.get('selection_eligible')) for row in candidate_rows)} eligible",
            "Candidate Design: 30 Morphology Baseline + 30 Structural Refinement + "
            "30 Distributional Threshold candidates across five Z intervals",
            "Ranking: unified 90-candidate family-by-Z balance with near-duplicate "
            "clustering and a pre-distribution-baseline same-Z non-regression guard",
            "Morphology Baseline Incumbent Candidate: "
            f"{best_row.get('morphology_baseline_incumbent_candidate', 'Not available')}",
            "Morphology Baseline Incumbent Score: "
            f"{best_row.get('morphology_baseline_incumbent_score', 'Not available')}",
            "Pre-Distribution Baseline Incumbent Candidate: "
            f"{best_row.get('pre_distribution_baseline_incumbent_candidate', 'Not available')}",
            "Pre-Distribution Baseline Incumbent Score: "
            f"{best_row.get('pre_distribution_baseline_incumbent_score_frozen', 'Not available')}",
            "Pre-Distribution Baseline Incumbent Score in Complete Pool: "
            f"{best_row.get('pre_distribution_baseline_incumbent_score_complete_pool', 'Not available')}",
            f"Family-Balanced Challenger Candidate: {best_row.get('family_balanced_challenger_candidate', 'Not available')}",
            f"Family-Balanced Challenger Score: {best_row.get('family_balanced_challenger_score', 'Not available')}",
            f"Selection Guard Decision: {best_row.get('selection_guard_reason', 'Not available')}",
            f"Best Candidate: {best_row['candidate']:02d} ({best_row['name']})",
            f"Selection Score: {best_row['auto_selection_score']}",
            f"Z Range: slices {best_row['z_start_1based']}-{best_row['z_end_1based_inclusive']} "
            "(1-based inclusive)",
            f"Projection: {str(best_row['projection']).upper()}",
            f"Automatically Selected Whole Astrocyte ROIs: {roi_count}",
            f"Selected Candidate Whole Area Before Final Refinement: "
            f"{best_row['mask_area_px']} pixels",
            f"Cellpose-SAM: {best_row['cellpose']} on {_CELLPOSE_DEVICE or 'not initialized'}",
            f"Candidate CPU Workers: requested={CANDIDATE_CPU_WORKERS}; effective={_EFFECTIVE_CANDIDATE_CPU_WORKERS}",
            f"3D DAPI Inventory CPU Workers: requested={DAPI_INVENTORY_CPU_WORKERS}; effective={_EFFECTIVE_DAPI_INVENTORY_CPU_WORKERS}",
            f"Cellpose Batch Size: requested={CELLPOSE_BATCH_SIZE}; effective={_CELLPOSE_EFFECTIVE_BATCH_SIZE if _CELLPOSE_EFFECTIVE_BATCH_SIZE is not None else 'Not run'}",
            f"Segmentation Method: {best_row['method_used']}",
            f"Distribution Model: {best_row.get('distribution_model', 'Not selected')}",
            f"Distribution Background-Peak Role: {best_row.get('distribution_background_peak_role', 'Not selected')}",
            f"Distribution Valid Channels: {best_row.get('distribution_valid_channels', 'Not selected')}",
            f"Distribution Posterior Thresholds: {best_row.get('distribution_posterior_thresholds_raw', 'Not selected')}",
            f"Distribution Background Peaks: {best_row.get('distribution_background_peaks_raw', 'Not selected')}",
            f"Distribution Component Separations: {best_row.get('distribution_ashman_separations', 'Not selected')}",
            f"Candidate Errors: {sum(bool(row['error']) for row in candidate_rows)}",
            "",
            "[Soma and Processes]",
            f"Method: {compartment_metrics['method']}",
            f"Adaptation: {compartment_metrics['adaptation']}",
            f"XY Pixel Size: {compartment_metrics['pixel_width_um']:.6f} x "
            f"{compartment_metrics['pixel_height_um']:.6f} um",
            f"Connected Whole Components Before Instance Split: "
            f"{compartment_metrics['instance_split']['base_connected_component_count']}",
            "Whole Instances Before Valid-Soma Cell Gate: "
            f"{compartment_metrics['instance_split'].get('pre_soma_gate_instance_count', compartment_metrics['instance_split']['final_instance_count'])}",
            f"Final Whole Astrocyte Instances: {compartment_metrics['roi_count']}",
            f"High-Confidence Multi-Astrocyte Components Split: "
            f"{compartment_metrics['instance_split']['split_component_count']}; "
            f"added ROIs={compartment_metrics['instance_split']['split_added_roi_count']}; "
            f"rejected ambiguous splits={compartment_metrics['instance_split']['split_rejected_count']}",
            f"Whole Geometry Path: "
            f"{'shared frozen Whole geometry with 3D ownership, then neonatal compartment repartition' if shared_whole else 'frozen mature Soma/Processes with shared 3D ownership guard'}",
            f"Shared Whole Baseline Area: "
            f"{shared_whole.get('whole_area_px', compartment_metrics['whole_area_px'])} pixels",
            f"Shared Whole Baseline Morphology Filter Removed IDs: "
            f"{shared_filter.get('removed_original_ids', compartment_metrics['morphology_filter']['removed_original_ids'])}",
            f"Shared Whole Baseline Branch-Gap Removal: "
            f"{shared_gap.get('removed_px', compartment_metrics['branch_gap_restoration']['removed_px'])} pixels",
            f"Branch-Gap Restoration: "
            f"{effective_gap['accepted_gap_count']} internal valleys; "
            f"removed={effective_gap['removed_px']} pixels "
            f"({effective_gap['removed_fraction']:.2%}); "
            f"connectivity-protected rejections="
            f"{effective_gap['rejected_disconnect_count']}",
            f"DAPI Extent Reconstruction: high={compartment_metrics['dapi_extent']['high_threshold']}; "
            f"low={compartment_metrics['dapi_extent']['low_threshold']}; "
            f"strict_core={compartment_metrics['dapi_extent']['strict_core_px']} pixels; "
            f"reconstructed_extent={compartment_metrics['dapi_extent']['reconstructed_extent_px']} pixels",
            f"DAPI Extent Satellite Cleanup: "
            f"{compartment_metrics['dapi_extent_satellite_components_removed']} components; "
            f"{compartment_metrics['dapi_extent_satellite_px_removed']} pixels reassigned to Processes",
            f"{'Shared Whole Morphology Filter' if shared_whole else 'Morphology Filter'}: "
            f"pre={effective_filter['pre_filter_roi_count']} ROIs; "
            f"post={effective_filter['post_filter_roi_count']} ROIs; "
            f"removed={effective_filter['removed_original_ids']}; "
            f"removed_area={effective_filter['removed_area_px']} pixels "
            f"({effective_filter['removed_area_fraction']:.2%})",
            "Age-Independent Valid-Soma Filtering: Whole-ID gate; retained cell boundaries unchanged",
            f"Valid-Soma Cell Gate ({soma_identity_gate.get('profile', age_profile.get('profile', 'unknown'))}): "
            f"pre={soma_identity_gate['pre_gate_roi_count']}; "
            f"post={soma_identity_gate['post_gate_roi_count']}; "
            f"removed={soma_identity_gate['removed_pre_gate_ids']}; "
            f"removed_area={soma_identity_gate['removed_area_px']} pixels "
            f"({soma_identity_gate['removed_area_fraction']:.2%})",
            f"Soma-Gate ID Mapping: {soma_identity_gate['id_mapping']}",
            f"Final Whole Area: {compartment_metrics['whole_area_px']} pixels",
            f"Final Soma Area: {compartment_metrics['soma_area_px']} pixels "
            f"({compartment_metrics['soma_area_fraction']:.2%})",
            f"Final Processes Area: {compartment_metrics['process_area_px']} pixels "
            f"({compartment_metrics['process_area_fraction']:.2%})",
            f"Fallback Soma ROIs: {compartment_metrics['fallback_soma_count']}",
            f"Ambiguous DAPI Anchors: {compartment_metrics['ambiguous_nucleus_count']}",
            f"Whole ROIs Without a Trusted Soma: {compartment_metrics['no_dapi_anchor_count']}",
            f"Selected DAPI Soma Anchors: {compartment_metrics['total_soma_anchor_count']}",
            f"Whole ROIs With Multiple Soma Anchors: "
            f"{compartment_metrics['multi_soma_whole_roi_count']}",
            f"Rejected Soma Anchors: {compartment_metrics['rejected_soma_anchor_count']}",
            f"Soma Core-Shell Pruning: {compartment_metrics['soma_core_shell_applied_roi_count']} ROIs; "
            f"{compartment_metrics['soma_core_shell_removed_px']} pixels reassigned to Processes",
            "No-Soma Rule: in both mature and neonatal profiles, a Whole ID without exactly one connected trusted Soma is removed together with its Processes.",
            "Partition QC: overlap=0 pixels; gap=0 pixels; outside Whole=0 pixels",
            "Processes ROI Semantics: one Astrocyte ID may contain multiple disconnected subregions; all are measured together as one composite ROI.",
            "",
            "Removed Morphology Outliers:",
        ]
    )
    if effective_filter["details"]:
        for row in effective_filter["details"]:
            lines.append(
                f"  Original ID {row['original_astrocyte_id']:03d}: {row['reason']}; "
                f"area={row['area_um2']:.3f} um2; core_R95={row['core_radius_um']:.3f} um; "
                f"axis_ratio={row['axis_ratio']:.3f}; branches={row['branchpoint_count']}; "
                f"anchors={row['soma_anchor_count']}; robust_votes={row['outlier_votes']}; "
                f"relative_cues={row['compact_outlier_cues']}"
            )
    else:
        lines.append("  None")
    lines.extend(["", "Valid-Soma Cell Gate Removals:"])
    if soma_identity_gate["details"]:
        for row in soma_identity_gate["details"]:
            lines.append(
                f"  Pre-gate ID {row['pre_gate_astrocyte_id']:03d}: {row['reason']}; "
                f"anchors={row['soma_anchor_count']}; soma_components={row['soma_component_count']}; "
                f"whole={row['whole_area_px']} px; "
                f"soma={row['soma_area_px']} px; processes={row['process_area_px']} px"
            )
    else:
        lines.append("  None")
    lines.extend(
        [
            "",
            "Accepted Instance Splits:",
        ]
    )
    if compartment_metrics["instance_split"]["split_components"]:
        for row in compartment_metrics["instance_split"]["split_components"]:
            lines.append(
                f"  Base {row['base_component_id']:03d}: "
                f"pre_gate_IDs={row.get('pre_soma_gate_new_astrocyte_ids', row['new_astrocyte_ids'])}; "
                f"pre_gate_child_areas={row.get('pre_soma_gate_child_areas_px', row['child_areas_px'])}; "
                f"removed_pre_gate_IDs={row.get('removed_pre_soma_gate_ids', [])}; "
                f"final_IDs={row['new_astrocyte_ids']}; "
                f"anchor_scores={row['anchor_scores']}; "
                f"separation={row['anchor_separation_um']:.3f} um; "
                f"neck/core={row['neck_core_ratio']:.3f}; "
                f"boundary/core_struct={row.get('boundary_structural_ratio', 'not used')}; "
                f"reason={row.get('reason', 'accepted')}"
            )
    else:
        lines.append("  None")
    lines.extend(
        [
            "",
            "Compartment Parameters:",
        ]
    )
    for key, value in compartment_metrics["config"].items():
        lines.append(f"  {key}: {value}")
    lines.extend(["", "Morphology Profile Features:"])
    if age_profile.get("features"):
        for key, value in age_profile["features"].items():
            lines.append(f"  {key}: {value}")
    else:
        lines.append("  Not computed because a filename profile label was present")
    lines.extend(
        [
            "",
            "[3D Nucleus Ownership QC]",
            f"Status: {neonatal_3d.get('status', 'unknown')}",
            f"Method: {neonatal_3d.get('method', 'unknown')}",
            f"Measurement Channel Used: {neonatal_3d.get('measurement_channel_used', False)}",
            f"Validation Z Range: "
            f"{neonatal_3d.get('z_start_1based', 'Not applicable')}-"
            f"{neonatal_3d.get('z_end_1based_inclusive', 'Not applicable')}",
            f"Voxel Size ZYX (um): {neonatal_3d.get('voxel_size_um', 'Not available')}",
            f"Calibration Source: {neonatal_3d.get('calibration_source', 'Not available')}",
            f"DAPI Candidates Near Frozen Whole: {neonatal_3d.get('candidate_count', 0)}",
            f"DAPI Candidates Passed 3D Gate: {neonatal_3d.get('accepted_count', 0)}",
            f"DAPI Candidates Rejected by 3D Gate: {neonatal_3d.get('rejected_count', 0)}",
            f"Accepted 2D Nucleus IDs: {neonatal_3d.get('accepted_2d_nucleus_ids', [])}",
            f"Rejected 2D Nucleus IDs: {neonatal_3d.get('rejected_2d_nucleus_ids', [])}",
            f"Ownership Conflicts: {ownership_guard.get('conflict_component_count', 0)}",
            f"Mandatory Multi-Nucleus Splits: {ownership_guard.get('split_component_count', 0)}",
            f"Foreign Soma Territories Pruned: {ownership_guard.get('foreign_soma_pruned_component_count', 0)}",
            f"Ambiguous Components Removed Fail-Closed: {ownership_guard.get('fail_closed_component_count', 0)}",
            f"Ownership Area Removed (px): {ownership_guard.get('removed_area_px', 0)}",
            f"Canonical 3D Nucleus Instances: {canonical_nuclei.get('instance_count', 0)}",
            f"Canonical Envelopes Split: {canonical_nuclei.get('connected_envelope_split_count', 0)}",
            f"Canonical Envelopes Ambiguous: {canonical_nuclei.get('ambiguous_instance_count', 0)}",
            f"Axially Truncated Cells Removed: {axial_guard.get('removed_cell_count', 0)}; "
            f"pre-guard IDs={axial_guard.get('removed_pre_guard_ids', [])}",
            f"Projected Foreign-DAPI Terminal Guard: changed "
            f"{projected_foreign_guard.get('changed_cell_count', 0)} cells; "
            f"removed {projected_foreign_guard.get('removed_area_px', 0)} px; "
            f"true tips={projected_foreign_guard.get('true_tip_count', 0)}; "
            f"terminal overlaps={projected_foreign_guard.get('terminal_overlap_component_count', 0)}; "
            f"pass-through preserved={projected_foreign_guard.get('preserved_pass_through_component_count', 0)}; "
            f"connectivity rollbacks={projected_foreign_guard.get('connectivity_rollback_count', 0)}",
            "Projection Guard Boundary: exact validated DAPI extent only; no halo, fixed terminal length, or trunk subtraction.",
            f"Canonical Owner Nuclear-Extent Completion: changed "
            f"{canonical_owner_extent_completion.get('changed_cell_count', 0)} cells; "
            "approved "
            f"{canonical_owner_extent_completion.get('approved_owner_extent_added_px', 0)} "
            "Soma px; pre-finalization-Whole-external "
            f"{canonical_owner_extent_completion.get('outside_added_whole_px', 0)} px; "
            "fail-closed IDs="
            f"{canonical_owner_extent_completion.get('fail_closed_cell_ids', [])}",
            f"Same-ID Soma Island Reconciliation: bridged IDs="
            f"{same_id_soma_reconciliation.get('bridged_ids', [])}; "
            f"approved Processes-to-Soma "
            f"{same_id_soma_reconciliation.get('approved_process_to_soma_px', 0)} px; "
            f"identity-split rejections="
            f"{same_id_soma_reconciliation.get('rejected_identity_split_ids', [])}; "
            f"multiple-nucleus rejections="
            f"{same_id_soma_reconciliation.get('rejected_multiple_owner_ids', [])}",
            "Per-Nucleus 3D QC:",
            "ID  Accepted  Score  Surface  XY_Boundary  Angular  Z_Support  Shell_Enrichment  Radial_Bands  DAPI_Z_um  Reason",
        ]
    )
    for row in neonatal_3d.get("per_nucleus", []):
        lines.append(
            f"{row['nucleus_id_2d']:03d}  {str(row['accepted']):8s}  "
            f"{row['enclosure_score']:.3f}  {row['surface_coverage']:.3f}  "
            f"{row['median_xy_boundary_coverage']:.3f}  {row['angular_coverage']:.3f}  "
            f"{row['z_support_fraction']:.3f}  {row['shell_enrichment']:.3f}  "
            f"{row['radial_band_fraction']:.3f}  {row['dapi_z_span_um']:.3f}  "
            f"{row['reason']}"
        )
    lines.extend(
        [
            "",
            "Per-Astrocyte Final Compartment QC:",
            "ID  Original_ID  Shared_Whole_IDs  Whole_px  Soma_px  Processes_px  Soma_%  Processes_%  Process_parts  DAPI_candidates  Soma_anchors  Nucleus_coverage  Flags",
        ]
    )
    for row in compartment_metrics["per_cell"]:
        flags = []
        if row["nucleus_ambiguous"]:
            flags.append("ambiguous_DAPI")
        if row["fallback_used"]:
            flags.append("fallback_soma")
        if row["nucleus_candidates"] == 0:
            flags.append("no_DAPI_anchor")
        elif row["soma_anchor_count"] == 0:
            flags.append("no_trusted_soma")
        if row["soma_anchor_count"] > 1:
            flags.append("multiple_soma_anchors")
        if row["rejected_soma_anchor_count"]:
            flags.append("rejected_soma_anchor")
        if row["soma_core_shell_applied"]:
            flags.append(f"core_shell_removed_{row['soma_core_shell_removed_px']}px")
        morphology_qc = row.get(
            "shared_morphology_qc", row.get("morphology_qc", {})
        )
        if morphology_qc.get("outlier_consensus", 0):
            flags.append(f"morphology_votes_{morphology_qc['outlier_consensus']}")
        shared_whole_ids = row.get("shared_whole_ids", [])
        shared_whole_text = (
            ",".join(str(value) for value in shared_whole_ids)
            if shared_whole_ids
            else "-"
        )
        lines.append(
            f"{row['astrocyte_id']:03d}  {row.get('original_astrocyte_id', row['astrocyte_id']):11d}  "
            f"{shared_whole_text:16s}  "
            f"{row['whole_area_px']:8d}  {row['soma_area_px']:7d}  "
            f"{row['process_area_px']:12d}  {100 * row['soma_fraction']:6.2f}  "
            f"{100 * row['process_fraction']:11.2f}  "
            f"{row.get('process_component_count', morphology_qc.get('process_component_count', 0)):13d}  "
            f"{row['nucleus_candidates']:15d}  {row['soma_anchor_count']:12d}  "
            f"{row['required_nucleus_coverage']:16.3f}  "
            f"{','.join(flags) or 'OK'}"
        )
    lines.extend(
        [
            "",
            "[Fiji]",
            f"Status: {fiji_status}",
            "Display: DAPI=blue; measurement=red; eGFP=green and GFAP=yellow when both are present; lone structural channel=green.",
            "ROI Colors: Whole=magenta; Processes=white; Soma=cyan.",
            "Fiji Windows: three Composite views and three raw grayscale measurement views; contour strokes and display-only synchronized IDs, no biological ROI fills.",
            "ROI Review: linked Delete, Soma Merge, Whole Split, and Soma Enlarge "
            "update Whole/Soma/Processes together; Revert is LIFO.",
            f"Measurement Image: raw grayscale {measurement} {str(best_row['projection']).upper()}, "
            f"slices {best_row['z_start_1based']}-{best_row['z_end_1based_inclusive']}.",
            "Measurement Order: Whole Astrocyte Cell -> Astrocyte Processes -> Astrocyte Soma.",
            "Measurement Command: Fiji ROI Manager > Measure; RGB/composite pixels are never quantified.",
            "Outputs:",
            f"  {WHOLE_OVERLAY_FILENAME}",
            f"  {PROCESS_OVERLAY_FILENAME}",
            f"  {SOMA_OVERLAY_FILENAME}",
            f"  {WORKBOOK_FILENAME}",
        ]
    )
    if fiji_details:
        lines.extend(["", "Final Fiji Results:"])
        result_sets = fiji_details.get("result_sets", {})
        for key in ("whole", "processes", "soma"):
            detail = result_sets.get(key, {})
            lines.append(
                f"  {detail.get('title', key)}: {detail.get('rows', 'unknown')} rows; "
                f"total area={detail.get('area_sum', 'unknown')}"
            )
        lines.append(
            "Final ROI Counts: "
            f"Whole={result_sets.get('whole', {}).get('rows', 'unknown')}; "
            f"Processes={result_sets.get('processes', {}).get('rows', 'unknown')}; "
            f"Soma={result_sets.get('soma', {}).get('rows', 'unknown')}"
        )
        lines.append(f"Measurement Window: {fiji_details.get('measurement_title', 'unknown')}")
        lines.append(
            f"Manual ROI Review Used: {fiji_details.get('manual_review_used', False)}; "
            f"Deleted Original IDs: {fiji_details.get('deleted_original_ids', [])}; "
            f"Reverted Actions: {fiji_details.get('reverted_actions', 0)}"
        )
        lines.append(f"Review Audit: {fiji_details.get('review_audit', [])}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


import json
import os
import signal
import subprocess
import time
from pathlib import Path


def terminate_fiji_process_group(
    process: subprocess.Popen,
    *,
    grace_seconds: float = 5.0,
) -> None:
    """Stop only the Fiji process group launched for the current analysis."""

    if process.poll() is not None:
        return
    try:
        os.killpg(process.pid, signal.SIGTERM)
    except ProcessLookupError:
        return
    except OSError:
        process.terminate()
    try:
        process.wait(timeout=max(0.1, float(grace_seconds)))
        return
    except subprocess.TimeoutExpired:
        pass
    try:
        os.killpg(process.pid, signal.SIGKILL)
    except ProcessLookupError:
        return
    except OSError:
        process.kill()
    try:
        process.wait(timeout=2.0)
    except subprocess.TimeoutExpired:
        pass


def find_fiji_launcher(explicit: Path | None = None) -> Path:
    candidates = [
        explicit.expanduser() if explicit else None,
        Path("/Applications/Fiji/fiji"),
        Path("/Applications/Fiji/Fiji.app/Contents/MacOS/fiji-macos-arm64"),
        Path("/Applications/Fiji/Fiji.app/Contents/MacOS/fiji-macos"),
        Path("/Applications/Fiji.app/Contents/MacOS/ImageJ-macosx"),
    ]
    for candidate in candidates:
        if candidate and candidate.is_file() and os.access(candidate, os.X_OK):
            return candidate.resolve()
    raise FileNotFoundError(
        "Fiji launcher not found. Expected /Applications/Fiji/fiji or an explicit --fiji-launcher."
    )

def launch_fiji_workflow(
    *,
    launcher: Path,
    run_dir: Path,
    manifest_path: Path,
    timeout_minutes: float,
) -> dict:
    script_path = run_dir / "ihc_fiji_bridge.groovy"
    ready_path = run_dir / "fiji_ready.json"
    done_path = run_dir / "fiji_done.json"
    error_path = run_dir / "fiji_error.txt"
    log_path = run_dir / "fiji_console.log"
    command = [
        str(launcher),
        "--memory",
        "16G",
        "--allow-multiple",
        "--no-splash",
        "--run",
        str(script_path),
        "manifest='" + str(manifest_path).replace("'", "\\'") + "'",
    ]
    print("Launching Fiji with 16 GB heap...", flush=True)
    log_handle = log_path.open("w", encoding="utf-8")
    try:
        process = subprocess.Popen(
            command,
            cwd=str(run_dir),
            stdout=log_handle,
            stderr=subprocess.STDOUT,
            start_new_session=True,
        )
        deadline = time.monotonic() + timeout_minutes * 60.0
        ready_announced = False
        ready_seconds: float | None = None
        started_at = time.monotonic()
        while time.monotonic() < deadline:
            if error_path.exists():
                detail = error_path.read_text(encoding="utf-8", errors="replace")
                raise RuntimeError(f"Fiji workflow failed:\n{detail}")
            if ready_path.exists() and not ready_announced:
                ready = json.loads(ready_path.read_text(encoding="utf-8"))
                print(
                    f"Fiji is ready at {ready.get('stage', 'measurement')}: "
                    f"{ready['roi_count']} Whole ROIs. "
                    "Six ROI windows are open and waiting for the review decision.",
                    flush=True,
                )
                ready_seconds = float(time.monotonic() - started_at)
                ready_announced = True
            if done_path.exists():
                details = json.loads(done_path.read_text(encoding="utf-8"))
                details["fiji_startup_seconds"] = ready_seconds
                details["fiji_total_seconds"] = float(time.monotonic() - started_at)
                return details
            if not ready_path.exists() and time.monotonic() - started_at > 8.0:
                console = log_path.read_text(encoding="utf-8", errors="replace")
                if "[ERROR]" in console or "Exception" in console:
                    raise RuntimeError(f"Fiji failed before the workflow initialized:\n{console}")
            if process.poll() is not None and not done_path.exists():
                time.sleep(1.0)
                if not done_path.exists():
                    console = log_path.read_text(encoding="utf-8", errors="replace")
                    raise RuntimeError(
                        f"Fiji exited with code {process.returncode} before completion.\n{console}"
                    )
            time.sleep(0.75)
        raise TimeoutError(
            f"Fiji display and measurement did not finish within {timeout_minutes:g} minutes. "
            f"Runtime diagnostics were kept at {run_dir}"
        )
    finally:
        log_handle.close()


def prepare_fiji_runtime(
    *,
    output_dir: Path,
    paths: dict[str, Path],
    metadata: dict[str, dict],
    structural_channels: list[str],
    measurement: str,
    best_row: dict,
    whole_labels: np.ndarray,
    soma_labels: np.ndarray,
    process_labels: np.ndarray,
    selected_projections: dict[str, np.ndarray],
    auto_continue: bool,
) -> tuple[Path, Path]:
    cache_root = Path.home() / "Library" / "Caches" / "IHC2DAnalysis"
    cache_root.mkdir(parents=True, exist_ok=True)
    now = time.time()
    for old_dir in cache_root.glob("run-*"):
        try:
            if old_dir.is_dir() and now - old_dir.stat().st_mtime > 2 * 24 * 3600:
                shutil.rmtree(old_dir)
        except OSError:
            pass
    run_dir = cache_root / f"run-{uuid.uuid4().hex}"
    run_dir.mkdir()

    roi_count = int(whole_labels.max())
    if roi_count < 1:
        raise ValueError("The selected candidate contains no ROI")
    expected_ids = set(range(1, roi_count + 1))
    observed_ids = {
        key: set(int(value) for value in np.unique(labels) if int(value) > 0)
        for key, labels in (
            ("whole", whole_labels),
            ("soma", soma_labels),
            ("processes", process_labels),
        )
    }
    if any(ids != expected_ids for ids in observed_ids.values()):
        raise ValueError(
            "Whole, Soma, and Processes label IDs must match exactly: "
            f"{observed_ids}"
        )
    if np.any((soma_labels > 0) & (whole_labels != soma_labels)):
        raise ValueError("A Soma label is outside or assigned to a different Whole Astrocyte ID")
    if np.any((process_labels > 0) & (whole_labels != process_labels)):
        raise ValueError("A Processes label is outside or assigned to a different Whole Astrocyte ID")
    partition_count = (soma_labels > 0).astype(np.uint8) + (process_labels > 0).astype(np.uint8)
    if np.any(partition_count[whole_labels > 0] != 1) or np.any(partition_count[whole_labels == 0] != 0):
        raise ValueError("Soma and Processes do not form an exact partition of the Whole labels")
    label_paths = {
        "whole": run_dir / "whole_roi_labels.tif",
        "soma": run_dir / "soma_roi_labels.tif",
        "processes": run_dir / "process_roi_labels.tif",
    }
    for key, labels in (
        ("whole", whole_labels),
        ("soma", soma_labels),
        ("processes", process_labels),
    ):
        tf.imwrite(
            label_paths[key],
            labels.astype(np.uint16, copy=False),
            photometric="minisblack",
            metadata={"axes": "YX"},
        )
    script_path = run_dir / "ihc_fiji_bridge.groovy"
    script_path.write_text(FIJI_GROOVY_SCRIPT, encoding="utf-8")
    manifest_path = run_dir / "manifest.json"
    channel_order = ["DAPI", measurement, *structural_channels]
    display_ranges = {
        channel: display_range(selected_projections[channel])
        for channel in channel_order
    }
    manifest_data = {
        "pipeline_name": PIPELINE_NAME,
        "channels": {channel: str(paths[channel]) for channel in channel_order},
        "channel_order": channel_order,
        "structural_channels": structural_channels,
        "measurement_channel": measurement,
        "expected_shape": list(metadata["DAPI"]["shape"]),
        "pixel_width_um": float(metadata["DAPI"]["pixel_width_um"]),
        "pixel_height_um": float(metadata["DAPI"]["pixel_height_um"]),
        "review_merge_max_soma_gap_um": REVIEW_MERGE_MAX_SOMA_GAP_UM,
        "z_start_1based": int(best_row["z_start_1based"]),
        "z_end_1based_inclusive": int(best_row["z_end_1based_inclusive"]),
        "projection": best_row["projection"],
        "display_ranges": display_ranges,
        "roi_count": roi_count,
        "require_complete_soma_ids": True,
        "label_mask_paths": {key: str(value) for key, value in label_paths.items()},
        "overlay_output_paths": {
            "whole": str(run_dir / "reviewed_whole_overlay.png"),
            "soma": str(run_dir / "reviewed_soma_overlay.png"),
            "processes": str(run_dir / "reviewed_processes_overlay.png"),
        },
        "ready_marker": str(run_dir / "fiji_ready.json"),
        "done_marker": str(run_dir / "fiji_done.json"),
        "error_marker": str(run_dir / "fiji_error.txt"),
        "auto_continue": bool(auto_continue),
    }
    manifest_path.write_text(json.dumps(manifest_data, indent=2), encoding="utf-8")
    return run_dir, manifest_path


def write_measurement_workbook(
    path: Path,
    *,
    fiji_details: dict,
    measurement: str,
    best_row: dict,
    age_profile: str,
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl 3.1.5 is required for the XLSX output in the formal Python environment"
        ) from exc

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.title = "IHC 2D Fluorescence Results"
    workbook.properties.creator = PIPELINE_NAME
    headers = [
        "ROI_Index",
        "Astrocyte_ID",
        "Original_Astrocyte_ID",
        "Source_Original_Astrocyte_IDs",
        "Compartment",
        "ROI_Name",
        "Label",
        "Area",
        "Mean",
        "Median",
        "Min",
        "Max",
        "IntDen",
        "RawIntDen",
        "Measurement_Channel",
        "Projection",
        "Z_Start_1Based",
        "Z_End_1Based_Inclusive",
        "Age_Profile",
        "Manual_Review_Used",
    ]
    sheet_specs = (
        ("Whole Cell", "whole"),
        ("Processes", "processes"),
        ("Soma", "soma"),
    )
    result_sets = fiji_details.get("result_sets", {})
    for sheet_name, key in sheet_specs:
        detail = result_sets.get(key, {})
        row_data = detail.get("row_data", [])
        if len(row_data) != int(detail.get("rows", -1)):
            raise RuntimeError(f"Fiji {key} row payload is incomplete: {detail}")
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="355070")
        for source in row_data:
            enriched = dict(source)
            enriched.update(
                {
                    "Measurement_Channel": measurement,
                    "Projection": str(best_row["projection"]).upper(),
                    "Z_Start_1Based": int(best_row["z_start_1based"]),
                    "Z_End_1Based_Inclusive": int(
                        best_row["z_end_1based_inclusive"]
                    ),
                    "Age_Profile": age_profile,
                    "Manual_Review_Used": bool(
                        fiji_details.get("manual_review_used", False)
                    ),
                }
            )
            sheet.append([enriched.get(header) for header in headers])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        numeric_headers = {"Area", "Mean", "Median", "Min", "Max", "IntDen", "RawIntDen"}
        for column_index, header in enumerate(headers, start=1):
            if header not in numeric_headers:
                continue
            for cell in list(sheet.columns)[column_index - 1][1:]:
                cell.number_format = "0.000000"
        widths = {
            "A": 12,
            "B": 14,
            "C": 22,
            "D": 28,
            "E": 14,
            "F": 34,
            "G": 54,
            "H": 14,
            "I": 14,
            "J": 12,
            "K": 12,
            "L": 12,
            "M": 16,
            "N": 18,
            "O": 22,
            "P": 12,
            "Q": 17,
            "R": 25,
            "S": 14,
            "T": 22,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
    audit_sheet = workbook.create_sheet("Review Audit")
    audit_headers = ["Sequence", "Action", "Source_Original_IDs", "Result_Lineage", "Reverted"]
    audit_sheet.append(audit_headers)
    for cell in audit_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="6D597A")
    for event in fiji_details.get("review_audit", []):
        audit_sheet.append(
            [
                event.get("sequence"),
                event.get("action"),
                ",".join(str(value) for value in event.get("source_ids", [])),
                ",".join(str(value) for value in event.get("result_lineage", [])),
                bool(event.get("reverted", False)),
            ]
        )
    audit_sheet.freeze_panes = "A2"
    for column, width in {"A": 12, "B": 14, "C": 28, "D": 28, "E": 12}.items():
        audit_sheet.column_dimensions[column].width = width
    workbook.save(path)

def validate_measurement_workbook(path: Path, fiji_details: dict) -> None:
    from openpyxl import load_workbook

    expected = {
        "Whole Cell": "whole",
        "Processes": "processes",
        "Soma": "soma",
    }
    with path.open("rb") as handle:
        workbook = load_workbook(handle, read_only=True, data_only=True)
        if set(workbook.sheetnames) != set(expected) | {"Review Audit"}:
            raise RuntimeError(f"Unexpected XLSX worksheets: {workbook.sheetnames}")
        observed_sequences: dict[str, tuple[list[int], list[int]]] = {}
        for sheet_name, key in expected.items():
            sheet = workbook[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                raise RuntimeError(f"XLSX worksheet {sheet_name} is empty")
            headings = list(rows[0])
            if "Median" not in headings or "P90" in headings or "P95" in headings:
                raise RuntimeError(f"XLSX measurement columns are invalid: {headings}")
            expected_rows = int(fiji_details["result_sets"][key]["rows"])
            if len(rows) - 1 != expected_rows:
                raise RuntimeError(
                    f"XLSX {sheet_name} has {len(rows) - 1} rows; expected {expected_rows}"
                )
            astrocyte_column = headings.index("Astrocyte_ID")
            original_column = headings.index("Original_Astrocyte_ID")
            current_ids = [int(row[astrocyte_column]) for row in rows[1:]]
            original_ids = [int(row[original_column]) for row in rows[1:]]
            if current_ids != list(range(1, len(current_ids) + 1)):
                raise RuntimeError(
                    f"XLSX {sheet_name} Astrocyte IDs are not contiguous: {current_ids}"
                )
            if len(set(original_ids)) != len(original_ids):
                raise RuntimeError(
                    f"XLSX {sheet_name} repeats an Original Astrocyte ID: {original_ids}"
                )
            observed_sequences[key] = (current_ids, original_ids)
        reference_sequences = observed_sequences["whole"]
        for key in ("processes", "soma"):
            if observed_sequences[key] != reference_sequences:
                raise RuntimeError(
                    "XLSX Whole, Soma, and Processes ID sequences do not match: "
                    f"{observed_sequences}"
                )
        workbook.close()


def publish_output_bundle(
    *,
    staged_files: dict[str, Path],
    final_files: dict[str, Path],
    run_dir: Path,
) -> None:
    if set(staged_files) != set(final_files):
        raise ValueError("Staged and final output keys do not match")
    token = uuid.uuid4().hex
    temporary: dict[str, Path] = {}
    backups: dict[str, Path] = {}
    existed: dict[str, bool] = {}
    for key, final_path in final_files.items():
        temporary[key] = (
            final_path.parent / f"temporary_IHC_{token}_{key}.tmp"
        )
        backups[key] = run_dir / f"previous_{key}.backup"
        existed[key] = final_path.exists()
        if existed[key]:
            shutil.copy2(final_path, backups[key])
        shutil.copy2(staged_files[key], temporary[key])
    try:
        for key in staged_files:
            os.replace(temporary[key], final_files[key])
    except Exception:
        for key, final_path in final_files.items():
            if existed[key] and backups[key].exists():
                shutil.copy2(backups[key], final_path)
            elif not existed[key]:
                final_path.unlink(missing_ok=True)
        raise
    finally:
        for path in temporary.values():
            path.unlink(missing_ok=True)


def _main_locked(argv: list[str] | None = None) -> int:
    global _CELLPOSE_EFFECTIVE_BATCH_SIZE
    global _EFFECTIVE_CANDIDATE_CPU_WORKERS
    global _EFFECTIVE_DAPI_INVENTORY_CPU_WORKERS
    global _RUN_STARTED_AT
    _RUN_STARTED_AT = time.perf_counter()
    _CELLPOSE_EFFECTIVE_BATCH_SIZE = None
    _EFFECTIVE_CANDIDATE_CPU_WORKERS = select_candidate_cpu_workers(
        CANDIDATE_CPU_WORKERS
    )
    _EFFECTIVE_DAPI_INVENTORY_CPU_WORKERS = select_candidate_cpu_workers(
        DAPI_INVENTORY_CPU_WORKERS
    )
    _RUNTIME_TIMINGS.update(
        {
            "cellpose_model_init_seconds": 0.0,
            "cellpose_inference_events": [],
            "candidate_postprocess_seconds": [],
            "candidate_total_seconds": [],
            "candidate_stage_wall_seconds": 0.0,
            "rank_candidates_seconds": 0.0,
            "compartment_split_seconds": 0.0,
        }
    )
    args = parse_args(argv)
    try:
        import openpyxl  # noqa: F401
    except ImportError as exc:
        raise RuntimeError(
            "The analysis environment requires openpyxl 3.1.5 before analysis starts"
        ) from exc
    if args.fiji_timeout_minutes <= 0:
        raise ValueError("--fiji-timeout-minutes must be positive")
    if (
        args.dapi_fragment_workload_preflight_only
        and args.dapi_fragment_workload_json is None
    ):
        raise ValueError(
            "--dapi-fragment-workload-json is required with "
            "--dapi-fragment-workload-preflight-only"
        )
    input_dir = args.input_dir.expanduser().resolve()
    output_dir = args.output_dir.expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    dapi_fragment_workload_path = (
        args.dapi_fragment_workload_json.expanduser().resolve()
        if args.dapi_fragment_workload_json is not None
        else output_dir / "IHC_2D_DAPI_Fragment_Workload_Failure.json"
    )
    whole_overlay_path = output_dir / WHOLE_OVERLAY_FILENAME
    soma_overlay_path = output_dir / SOMA_OVERLAY_FILENAME
    process_overlay_path = output_dir / PROCESS_OVERLAY_FILENAME
    report_path = output_dir / REPORT_FILENAME
    workbook_path = output_dir / WORKBOOK_FILENAME

    paths, ignored_files = discover_channel_paths(input_dir)
    if "eGFP" not in paths:
        raise ValueError(
            "The single-file fallback supports the eGFP route only. Use "
            "run_project_leap_2d.command for mature GFAP-only analysis."
        )
    filename_age_decision = detect_filename_age_profile(paths)
    structural_channels = [channel for channel in STRUCTURAL_CHANNELS if channel in paths]
    measurement = measurement_channel(paths)
    mode = channel_mode(paths)
    metadata = {channel: read_meta(path) for channel, path in paths.items()}
    reference_shape = tuple(metadata["DAPI"]["shape"])
    mismatched = {
        channel: tuple(channel_meta["shape"])
        for channel, channel_meta in metadata.items()
        if tuple(channel_meta["shape"]) != reference_shape
    }
    if mismatched:
        raise ValueError(f"Channel shapes do not match DAPI {reference_shape}: {mismatched}")
    if any(channel_meta["axes"] != "ZYX" for channel_meta in metadata.values()):
        raise ValueError("All input channels must have explicit ZYX axes")
    validate_shared_geometry(metadata)
    if any(
        metadata["DAPI"].get(key) is None
        or float(metadata["DAPI"][key]) <= 0
        for key in ("pixel_width_um", "pixel_height_um")
    ):
        raise ValueError(
            "Positive XY pixel calibration is required for physical Soma and morphology thresholds"
        )

    print_terminal_stage(
        "01 | INPUT CONTRACT",
        f"{PRODUCT_DISPLAY_NAME}\nScientific core: {ANALYSIS_CORE_NAME}",
    )
    print(f"Input: {input_dir}", flush=True)
    print(f"ROI mode: {mode}; measurement: {measurement} (excluded from ROI definition)", flush=True)
    for channel, path in paths.items():
        print(f"  {channel}: {path.name}", flush=True)
    for note in ignored_files:
        print(f"  ignored: {note}", flush=True)
    if filename_age_decision is not None:
        print(
            f"Age profile from filename: {filename_age_decision.profile} "
            f"({', '.join(filename_age_decision.tagged_files)})",
            flush=True,
        )

    print_terminal_stage(
        "02 | STACK LOADING AND ADAPTIVE Z PREPARATION",
        "Common method: load each split ZYX stack once, verify shared calibration, "
        "and build five adaptive contiguous Z intervals.",
    )
    _CELLPOSE_MASK_CACHE.clear()
    clear_candidate_computation_caches()
    dapi_stack = load_stack(paths["DAPI"])
    structural_stacks = {
        channel: load_stack(paths[channel])
        for channel in structural_channels
    }
    z_activity_profile = z_profile(structural_stacks)
    chosen_specs = complete_candidate_specs(reference_shape[0], structural_stacks)
    if args.disable_cellpose:
        chosen_specs = [replace(spec, cellpose=False) for spec in chosen_specs]

    rows: list[dict] = []
    candidate_masks: list[np.ndarray] = []
    projection_keys: list[tuple[int, int, str]] = []
    projection_cache: dict[tuple[int, int, str], tuple[np.ndarray, dict[str, np.ndarray]]] = {}
    structural_map_cache: dict[tuple, np.ndarray] = {}

    def evaluate_timed(
        position: int,
        spec: TestSpec,
        context: CandidateWindowContext,
    ):
        candidate_started = time.perf_counter()
        mask, row, projection_key = evaluate_ihc_candidate(
            candidate_number=position,
            candidate_count=len(chosen_specs),
            spec=spec,
            input_dir=input_dir,
            structural_channels=structural_channels,
            dapi_stack=dapi_stack,
            structural_stacks=structural_stacks,
            profile=z_activity_profile,
            projection_cache=projection_cache,
            structural_map_cache=structural_map_cache,
            emit_progress=False,
            window_context=context,
        )
        candidate_total = time.perf_counter() - candidate_started
        return mask, row, projection_key, candidate_total, candidate_total

    def print_candidate_row(
        row: dict,
        spec: TestSpec,
        completed_count: int,
        worker_seconds: float,
    ) -> None:
        print_terminal_event(
            f"Candidate {int(row['candidate']):02d} | "
            f"completed={completed_count:02d}/{len(chosen_specs)} | "
            f"module={candidate_module_display_name(spec)} | "
            f"family={candidate_profile_family(spec)} | "
            f"profile={candidate_profile_name(spec)} | "
            f"threshold={candidate_threshold_display_name(spec)} | "
            f"Z={int(row['z_start_1based'])}-{int(row['z_end_1based_inclusive'])} | "
            f"area={int(row['mask_area_px'])} | "
            f"components={int(row['connected_components'])} | "
            f"soma_support={int(row['soma_supported_components'])} | "
            f"fine_added={int(row['fine_branch_retained_px'])} | "
            f"worker_time={float(worker_seconds):.3f} s"
        )

    if len(chosen_specs) != TOTAL_CANDIDATE_COUNT:
        raise AssertionError(
            f"Complete candidate evaluation requires exactly {TOTAL_CANDIDATE_COUNT} candidates"
        )
    z_group_counts: dict[str, int] = {}
    for spec in chosen_specs:
        z_group_counts[spec.z_mode] = z_group_counts.get(spec.z_mode, 0) + 1
    if (
        len(z_group_counts) != EXPECTED_Z_INTERVAL_COUNT
        or set(z_group_counts.values()) != {EXPECTED_PROFILES_PER_Z}
    ):
        raise AssertionError(
            "Complete candidate evaluation requires five Z groups with exactly eighteen profiles per group"
        )
    print_terminal_stage(
        "03 | CELLPOSE-SAM GPU INFERENCE",
        "Common method: one serial Apple Metal/MPS Cellpose-SAM inference per Z "
        "interval; masks are cached and reused by all matching candidates.\n"
        f"Cellpose batch requested={CELLPOSE_BATCH_SIZE}; "
        f"candidate CPU workers requested={CANDIDATE_CPU_WORKERS}, "
        f"effective={_EFFECTIVE_CANDIDATE_CPU_WORKERS}; "
        f"3D DAPI workers requested={DAPI_INVENTORY_CPU_WORKERS}, "
        f"effective={_EFFECTIVE_DAPI_INVENTORY_CPU_WORKERS}.",
    )
    candidate_stage_started = time.perf_counter()
    candidate_contexts = build_candidate_window_contexts(
        chosen_specs=chosen_specs,
        structural_channels=structural_channels,
        dapi_stack=dapi_stack,
        structural_stacks=structural_stacks,
        profile=z_activity_profile,
        projection_cache=projection_cache,
        structural_map_cache=structural_map_cache,
    )
    context_groups: list[tuple[CandidateWindowContext, list[TestSpec]]] = []
    context_positions: dict[int, int] = {}
    for spec, context in zip(chosen_specs, candidate_contexts):
        group_position = context_positions.get(id(context))
        if group_position is None:
            context_positions[id(context)] = len(context_groups)
            context_groups.append((context, [spec]))
        else:
            context_groups[group_position][1].append(spec)

    print_terminal_stage(
        "04 | CPU CANDIDATE MORPHOLOGY",
        "Common method: reusable DAPI, Sato, top-hat, distributional-threshold, "
        "and structural features are precomputed once per complete cache key; "
        "90 immutable candidate jobs then run in the global worker pool.",
    )
    with ThreadPoolExecutor(
        max_workers=min(_EFFECTIVE_CANDIDATE_CPU_WORKERS, len(chosen_specs)),
        thread_name_prefix="ihc-candidate",
    ) as executor:
        precompute_jobs = candidate_precompute_jobs(
            context_groups=context_groups,
            input_dir=input_dir,
            structural_channels=structural_channels,
        )
        precompute_futures = [
            executor.submit(function, *job_args)
            for _priority, function, job_args in precompute_jobs
        ]
        for future in precompute_futures:
            future.result()

        primary_jobs: list[tuple[int, TestSpec, CandidateWindowContext]] = []
        reuse_jobs: list[tuple[int, TestSpec, CandidateWindowContext]] = []
        seen_base_keys: set[tuple] = set()
        for position, (spec, context) in enumerate(
            zip(chosen_specs, candidate_contexts),
            start=1,
        ):
            base_key = candidate_base_cache_key(context.structural_map, spec)
            job = (position, spec, context)
            if base_key in seen_base_keys:
                reuse_jobs.append(job)
            else:
                seen_base_keys.add(base_key)
                primary_jobs.append(job)
        def job_priority(job: tuple[int, TestSpec, CandidateWindowContext]) -> tuple:
            position, spec, _context = job
            return (
                float(spec.fine_branch_detail_percentile),
                float(spec.fine_branch_intensity_percentile),
                int(spec.fine_branch_min_area),
                position,
            )

        primary_jobs.sort(key=job_priority)
        reuse_jobs.sort(key=job_priority)
        futures_by_position = {
            position: executor.submit(evaluate_timed, position, spec, context)
            for position, spec, context in (*primary_jobs, *reuse_jobs)
        }
        future_positions = {
            future: position for position, future in futures_by_position.items()
        }
        ordered_results: list[tuple | None] = [None] * len(chosen_specs)
        candidate_errors: dict[int, BaseException] = {}
        completed_candidates = 0
        for future in as_completed(future_positions):
            position = future_positions[future]
            try:
                result = future.result()
                ordered_results[position - 1] = result
                completed_candidates += 1
                print_candidate_row(
                    result[1],
                    chosen_specs[position - 1],
                    completed_candidates,
                    float(result[3]),
                )
            except BaseException as exc:
                candidate_errors[position] = exc
                print_terminal_event(
                    f"Candidate {position:02d} failed | "
                    f"module={candidate_module_display_name(chosen_specs[position - 1])} | "
                    f"profile={candidate_profile_name(chosen_specs[position - 1])} | "
                    f"error={exc!r}"
                )
        if candidate_errors:
            first_failed = min(candidate_errors)
            raise RuntimeError(
                f"Candidate {first_failed:02d} failed during parallel morphology"
            ) from candidate_errors[first_failed]

    for result in ordered_results:
        if result is None:
            raise RuntimeError("Candidate scheduler returned an incomplete result set")
        mask, row, projection_key, candidate_total, candidate_postprocess = result
        candidate_totals = _RUNTIME_TIMINGS["candidate_total_seconds"]
        candidate_postprocess_times = _RUNTIME_TIMINGS[
            "candidate_postprocess_seconds"
        ]
        assert isinstance(candidate_totals, list)
        assert isinstance(candidate_postprocess_times, list)
        candidate_totals.append(float(candidate_total))
        candidate_postprocess_times.append(float(candidate_postprocess))
        candidate_masks.append(mask)
        rows.append(row)
        projection_keys.append(projection_key)
    _RUNTIME_TIMINGS["candidate_stage_wall_seconds"] = float(
        time.perf_counter() - candidate_stage_started
    )

    print_terminal_stage(
        "05 | BALANCED RANKING AND NON-REGRESSION GUARD",
        "Common method: unified family-by-Z ranking, near-duplicate clustering, "
        "and a same-Z challenger comparison against the pre-distribution baseline incumbent.",
    )
    ranking_started = time.perf_counter()
    best_position, selection_guard = rank_complete_production_candidates(
        candidate_masks,
        rows,
    )
    _RUNTIME_TIMINGS["rank_candidates_seconds"] = float(
        time.perf_counter() - ranking_started
    )
    best_mask = candidate_masks[best_position]
    best_row = rows[best_position]
    best_spec = chosen_specs[best_position]
    best_projection_key = projection_keys[best_position]
    dapi_projection, structural_projections = projection_cache[best_projection_key]
    candidate_component_count = int(measure.label(best_mask, connectivity=2).max())
    print(
        f"Auto-selected candidate {best_row['candidate']:02d}: "
        f"score={best_row['auto_selection_score']}, "
        f"slices={best_row['z_start_1based']}-{best_row['z_end_1based_inclusive']}, "
        f"connected_components={candidate_component_count}",
        flush=True,
    )
    print(
        "Selection non-regression guard: "
        f"morphology_baseline={selection_guard['morphology_baseline_incumbent_candidate']:02d}, "
        "pre_distribution_baseline="
        f"{selection_guard['pre_distribution_baseline_incumbent_candidate']:02d}, "
        f"challenger={selection_guard['family_balanced_challenger_candidate']:02d}, "
        f"selected={selection_guard['selected_candidate']:02d}, "
        f"reason={selection_guard['guard']['reason']}",
        flush=True,
    )

    best_struct_key = (
        *best_projection_key,
        round(best_spec.egfp_weight, 6),
        round(best_spec.gfap_weight, 6),
        round(best_spec.smooth_sigma, 6),
    )
    best_struct = structural_map_cache[best_struct_key]
    best_cellpose_mask = np.zeros_like(best_mask, dtype=bool)
    if best_spec.cellpose and not args.disable_cellpose:
        best_cellpose_key = (
            tuple(structural_channels),
            int(best_row["z_start_0based"]),
            int(best_row["z_end_0based_inclusive"]),
            best_row["projection"],
            round(best_spec.egfp_weight, 3),
            round(best_spec.gfap_weight, 3),
            round(best_spec.smooth_sigma, 3),
            round(best_spec.cellpose_cellprob, 3),
            round(best_spec.cellpose_diameter, 3),
            best_spec.cellpose_max_side,
        )
        try:
            best_cellpose_mask, _ = run_cellpose_mask(
                best_struct,
                best_spec,
                best_cellpose_key,
            )
        except Exception as exc:
            print(f"Compartment Cellpose prior unavailable; continuing without it: {exc!r}", flush=True)
            best_cellpose_mask = np.zeros_like(best_mask, dtype=bool)

    del futures_by_position, future_positions, ordered_results
    del precompute_futures, precompute_jobs
    del primary_jobs, reuse_jobs, candidate_contexts, context_groups, context_positions
    del seen_base_keys
    candidate_masks.clear()
    projection_keys.clear()
    projection_cache.clear()
    structural_map_cache.clear()
    _CELLPOSE_MASK_CACHE.clear()
    clear_candidate_computation_caches()
    gc.collect()

    print_terminal_stage(
        "06 | AGE PROFILE, 3D DAPI OWNERSHIP, AND COMPARTMENTS",
        "Common method: determine mature/neonatal profile after Whole/Z selection, "
        "validate calibrated 3D DAPI objects, reconcile identities, and enforce "
        "Whole = Soma + Processes with synchronized IDs.",
    )
    pixel_width_um = float(metadata["DAPI"]["pixel_width_um"])
    pixel_height_um = float(metadata["DAPI"]["pixel_height_um"])
    age_profile_decision = filename_age_decision or classify_age_profile(
        best_mask,
        dapi_projection,
        best_struct,
        pixel_width_um,
        pixel_height_um,
    )
    if age_profile_decision.source == "morphology_classifier":
        print(
            f"Age profile: {age_profile_decision.profile} "
            f"(morphology neonatal_score={age_profile_decision.neonatal_score:.6f}, "
            f"threshold={age_profile_decision.threshold:.2f})",
            flush=True,
        )
    else:
        print(f"Age profile: {age_profile_decision.profile} (filename)", flush=True)
    z0 = int(best_row["z_start_0based"])
    z1 = int(best_row["z_end_0based_inclusive"])
    neonatal_3d_context = None
    if structural_stacks:
        ownership_structural_channel = (
            "eGFP" if "eGFP" in structural_stacks else "GFAP"
        )
        dapi_depth_um = metadata["DAPI"].get("pixel_depth_um")
        structural_depth_um = metadata[ownership_structural_channel].get(
            "pixel_depth_um"
        )
        if dapi_depth_um is not None and structural_depth_um is not None:
            dapi_depth_um = float(dapi_depth_um)
            structural_depth_um = float(structural_depth_um)
            if not math.isclose(
                dapi_depth_um,
                structural_depth_um,
                rel_tol=1e-4,
                abs_tol=1e-6,
            ):
                raise ValueError(
                    "DAPI/structural Z calibration mismatch prevents 3D nucleus ownership: "
                    f"DAPI={dapi_depth_um}, "
                    f"{ownership_structural_channel}={structural_depth_um} um"
                )
            neonatal_3d_context = Neonatal3DContext(
                dapi_stack=dapi_stack,
                egfp_stack=structural_stacks[ownership_structural_channel],
                z_start_0based=z0,
                z_end_0based_inclusive=z1,
                pixel_depth_um=dapi_depth_um,
                calibration_source=str(metadata["DAPI"].get("pixel_depth_source") or "unknown"),
                structural_channel=ownership_structural_channel,
            )
        else:
            print(
                "3D nucleus ownership skipped: calibrated DAPI/structural Z spacing is unavailable",
                flush=True,
            )
    if args.dapi_fragment_workload_preflight_only:
        if neonatal_3d_context is None:
            raise RuntimeError(
                "DAPI fragment-workload preflight requires calibrated DAPI and structural "
                "ZYX stacks with matching positive Z spacing"
            )
        print_terminal_stage(
            "06 | DAPI 3D WORKLOAD PREFLIGHT",
            "Validation-only mode: use the production parent reconstruction path, "
            "count fragment workload, and stop before fragment evaluation, the "
            "measurement stack, or Fiji.",
        )
        preflight_started = time.perf_counter()
        preflight_inventory = build_dapi_object_inventory_3d(
            best_mask,
            dapi_projection,
            neonatal_3d_context,
            pixel_width_um,
            pixel_height_um,
            compartment_config_for_profile("mature"),
            max_workers=1,
            preflight_only=True,
            workload_diagnostic_path=dapi_fragment_workload_path,
        )
        _RUNTIME_TIMINGS["compartment_split_seconds"] = float(
            time.perf_counter() - preflight_started
        )
        workload_summary = preflight_inventory.metrics.get(
            "dapi_fragment_workload"
        )
        if not isinstance(workload_summary, dict):
            raise RuntimeError(
                "DAPI fragment-workload preflight did not return a workload summary"
            )
        workload_summary.update(
            {
                "input_dir": str(input_dir),
                "input_channels": {
                    channel: path.name for channel, path in paths.items()
                },
                "product_display_name": PRODUCT_DISPLAY_NAME,
                "analysis_core_name": ANALYSIS_CORE_NAME,
                "selected_candidate": int(best_row["candidate"]),
                "selected_age_profile": age_profile_decision.profile,
                "selected_z_range_1based": [int(z0 + 1), int(z1 + 1)],
                "measurement_stack_loaded": False,
                "fragment_evaluator_called": False,
                "fiji_launched": False,
                "production_outputs_replaced": False,
            }
        )
        atomic_write_dapi_fragment_workload_json(
            dapi_fragment_workload_path,
            workload_summary,
        )
        print(
            "DAPI fragment-workload preflight completed safely: "
            f"parents={int(workload_summary['parents_linked_to_whole'])}, "
            f"fragments={int(workload_summary['total_fragments'])}, "
            "max_parent_voxel_checks="
            f"{int(workload_summary['max_parent_voxel_comparisons']):,}",
            flush=True,
        )
        print(f"Workload JSON: {dapi_fragment_workload_path}", flush=True)
        return 0
    compartment_started = time.perf_counter()
    whole_labels, soma_labels, process_labels, compartment_metrics = (
        split_astrocyte_compartments_for_profile(
            best_mask,
            dapi_projection,
            best_struct,
            best_cellpose_mask,
            pixel_width_um,
            pixel_height_um,
            age_profile_decision.profile,
            neonatal_3d_context=neonatal_3d_context,
            dapi_fragment_workload_diagnostic_path=dapi_fragment_workload_path,
        )
    )
    _RUNTIME_TIMINGS["compartment_split_seconds"] = float(
        time.perf_counter() - compartment_started
    )
    compartment_metrics["age_profile"] = asdict(age_profile_decision)
    compartment_metrics["adaptation"] = (
        (
            "object-preserving calibrated 3D nucleus ownership guard followed by "
            "the frozen mature Soma/Processes path"
            if neonatal_3d_context is not None
            else "frozen mature compartment path without calibrated 3D ownership"
        )
        if age_profile_decision.profile == "mature"
        else (
            "shared Whole geometry followed by object-preserving calibrated 3D DAPI/"
            f"{neonatal_3d_context.structural_channel} ownership, mandatory multi-nucleus "
            "partition, neonatal shape-preserving Soma/Processes, and whole-ID valid-Soma gate"
            if neonatal_3d_context is not None
            else (
                "frozen Whole geometry followed by neonatal 2D anchors, multi-center ID, "
                "neonatal Soma/Processes profile, and whole-ID valid-Soma cell gate"
            )
        )
    )
    roi_count = int(whole_labels.max())
    print(
        "Compartments: "
        f"Whole_ROIs={roi_count} "
        f"Whole={compartment_metrics['whole_area_px']} px, "
        f"Soma={compartment_metrics['soma_area_px']} px "
        f"({compartment_metrics['soma_area_fraction']:.1%}), "
        f"Processes={compartment_metrics['process_area_px']} px "
        f"({compartment_metrics['process_area_fraction']:.1%})",
        flush=True,
    )

    print_terminal_stage(
        "07 | RAW MEASUREMENT PROJECTION",
        f"Common method: project the untouched grayscale {measurement} stack over "
        f"the selected inclusive Z range {z0 + 1}-{z1 + 1}; rendered composites "
        "remain display-only.",
    )
    measurement_projection_started = time.perf_counter()
    measurement_substack = tf.imread(str(paths[measurement]), key=range(z0, z1 + 1))
    measurement_projection = project(
        measurement_substack,
        0,
        measurement_substack.shape[0] - 1,
        best_row["projection"],
    )
    print(
        "Raw fluorescence measurement projection complete | "
        f"elapsed={time.perf_counter() - measurement_projection_started:.3f} s",
        flush=True,
    )

    if args.skip_fiji:
        debug_whole_overlay_path = output_dir / DEBUG_WHOLE_OVERLAY_FILENAME
        debug_soma_overlay_path = output_dir / DEBUG_SOMA_OVERLAY_FILENAME
        debug_process_overlay_path = output_dir / DEBUG_PROCESS_OVERLAY_FILENAME
        debug_report_path = output_dir / DEBUG_REPORT_FILENAME
        debug_state_path = output_dir / DEBUG_STATE_FILENAME
        preview_rgb = make_fiji_like_composite(
            dapi_projection,
            structural_projections,
            measurement_projection,
        )
        Image.fromarray(
            draw_smooth_label_contours(preview_rgb, whole_labels, best_spec, (255, 0, 255))
        ).save(debug_whole_overlay_path)
        Image.fromarray(
            draw_smooth_label_contours(preview_rgb, soma_labels, best_spec, (0, 255, 255))
        ).save(debug_soma_overlay_path)
        Image.fromarray(
            draw_smooth_label_contours(preview_rgb, process_labels, best_spec, (255, 255, 255))
        ).save(debug_process_overlay_path)
        np.savez_compressed(
            debug_state_path,
            whole_mask=(whole_labels > 0).astype(np.uint8),
            candidate_whole_mask=best_mask.astype(np.uint8),
            whole_labels=whole_labels.astype(np.uint16),
            soma_labels=soma_labels.astype(np.uint16),
            process_labels=process_labels.astype(np.uint16),
            canonical_owner_extent_approved_owner_extent_labels=np.asarray(
                compartment_metrics.get(
                    "_canonical_owner_extent_approved_owner_extent_labels",
                    np.zeros_like(whole_labels, dtype=np.uint16),
                ),
                dtype=np.uint16,
            ),
            same_id_soma_reconciliation_approved_process_to_soma_labels=np.asarray(
                compartment_metrics.get(
                    "_same_id_soma_reconciliation_approved_process_to_soma_labels",
                    np.zeros_like(whole_labels, dtype=np.uint16),
                ),
                dtype=np.uint16,
            ),
            canonical_nucleus_instance_core_labels_2d=np.asarray(
                compartment_metrics.get(
                    "_canonical_nucleus_instance_core_labels_2d",
                    np.zeros_like(whole_labels, dtype=np.uint32),
                ),
                dtype=np.uint32,
            ),
            canonical_nucleus_instance_extent_labels_2d=np.asarray(
                compartment_metrics.get(
                    "_canonical_nucleus_instance_extent_labels_2d",
                    np.zeros_like(whole_labels, dtype=np.uint32),
                ),
                dtype=np.uint32,
            ),
            dapi_projection=dapi_projection,
            structural_map=best_struct,
            cellpose_mask=best_cellpose_mask.astype(np.uint8),
            measurement_projection=measurement_projection,
            pixel_width_um=np.asarray(pixel_width_um),
            pixel_height_um=np.asarray(pixel_height_um),
            age_profile=np.asarray(age_profile_decision.profile),
            age_profile_source=np.asarray(age_profile_decision.source),
            age_profile_neonatal_score=np.asarray(
                np.nan
                if age_profile_decision.neonatal_score is None
                else age_profile_decision.neonatal_score
            ),
        )
        write_analysis_report(
            debug_report_path,
            input_dir=input_dir,
            paths=paths,
            metadata=metadata,
            structural_channels=structural_channels,
            measurement=measurement,
            best_row=best_row,
            candidate_rows=rows,
            candidate_specs=chosen_specs,
            roi_count=roi_count,
            compartment_metrics=compartment_metrics,
            fiji_status="Skipped by --skip-fiji (debug preview only; no measurement performed)",
        )
        print(f"Debug Whole overlay: {debug_whole_overlay_path}", flush=True)
        print(f"Debug Soma overlay: {debug_soma_overlay_path}", flush=True)
        print(f"Debug Processes overlay: {debug_process_overlay_path}", flush=True)
        print(f"Debug compartment state: {debug_state_path}", flush=True)
        print(f"Debug report: {debug_report_path}", flush=True)
        return 0

    print_terminal_stage(
        "08 | FIJI REVIEW, RAW-GRAYSCALE MEASUREMENT, AND PUBLICATION",
        "Common method: open three Composite and three raw grayscale views; "
        "optionally apply linked Delete, Soma-only proximity-chain Merge, or "
        "Revert; then measure Whole, Processes, and Soma on the raw grayscale projection.",
    )
    run_dir: Path | None = None
    try:
        fiji_runtime_started = time.perf_counter()
        selected_projections = {
            "DAPI": dapi_projection,
            measurement: measurement_projection,
            **structural_projections,
        }
        run_dir, manifest_path = prepare_fiji_runtime(
            output_dir=output_dir,
            paths=paths,
            metadata=metadata,
            structural_channels=structural_channels,
            measurement=measurement,
            best_row=best_row,
            whole_labels=whole_labels,
            soma_labels=soma_labels,
            process_labels=process_labels,
            selected_projections=selected_projections,
            auto_continue=args.fiji_auto_continue,
        )
        runtime_report = run_dir / "analysis_report_pending.txt"
        write_analysis_report(
            runtime_report,
            input_dir=input_dir,
            paths=paths,
            metadata=metadata,
            structural_channels=structural_channels,
            measurement=measurement,
            best_row=best_row,
            candidate_rows=rows,
            candidate_specs=chosen_specs,
            roi_count=roi_count,
            compartment_metrics=compartment_metrics,
            fiji_status="Pending automatic Fiji six-window display and native measurement",
        )
        launcher = find_fiji_launcher(args.fiji_launcher)

        del measurement_substack, measurement_projection, selected_projections
        del neonatal_3d_context
        del dapi_stack, structural_stacks, candidate_masks, projection_cache, structural_map_cache
        _CELLPOSE_MASK_CACHE.clear()
        clear_candidate_computation_caches()
        gc.collect()
        try:
            import torch

            if torch.backends.mps.is_available():
                torch.mps.empty_cache()
        except Exception:
            pass

        print(
            "Fiji runtime package prepared and analysis memory released | "
            f"elapsed={time.perf_counter() - fiji_runtime_started:.3f} s; "
            "launching Fiji...",
            flush=True,
        )
        fiji_details = launch_fiji_workflow(
            launcher=launcher,
            run_dir=run_dir,
            manifest_path=manifest_path,
            timeout_minutes=args.fiji_timeout_minutes,
        )
        print_terminal_event(
            "Fiji workflow returned; validating synchronized ROI IDs, measurement "
            "partitions, overlays, report, and workbook before atomic publication."
        )
        if bool(fiji_details.get("cancelled", False)):
            shutil.rmtree(run_dir, ignore_errors=True)
            print("Analysis cancelled in Fiji before measurement; no production output was replaced.", flush=True)
            return 0
        final_roi_count = int(fiji_details.get("roi_count", -1))
        if final_roi_count < 1:
            raise RuntimeError(f"Fiji returned an invalid final ROI count: {fiji_details}")
        required_headings = {
            "Label",
            "Area",
            "Mean",
            "Median",
            "Min",
            "Max",
            "IntDen",
            "ROI_Index",
            "Astrocyte_ID",
            "Original_Astrocyte_ID",
            "Source_Original_Astrocyte_IDs",
            "Compartment",
            "ROI_Name",
        }
        final_roi_ids = [int(value) for value in fiji_details.get("final_roi_ids", [])]
        if final_roi_ids != list(range(1, final_roi_count + 1)):
            raise RuntimeError(f"Fiji returned invalid final Astrocyte IDs: {final_roi_ids}")
        final_process_ids = [int(value) for value in fiji_details.get("final_process_ids", [])]
        final_soma_ids = [int(value) for value in fiji_details.get("final_soma_ids", [])]
        if final_process_ids != final_roi_ids:
            raise RuntimeError(f"Fiji Processes IDs do not match Whole IDs: {final_process_ids}")
        if final_soma_ids != final_roi_ids:
            raise RuntimeError(
                "Fiji Soma IDs do not match Whole IDs exactly: "
                f"Soma={final_soma_ids}, Whole={final_roi_ids}"
            )
        result_sets = fiji_details.get("result_sets", {})
        expected_original_ids = [
            int(value)
            for value in result_sets.get("whole", {}).get(
                "original_astrocyte_ids",
                [],
            )
        ]
        if len(expected_original_ids) != final_roi_count or len(
            set(expected_original_ids)
        ) != final_roi_count:
            raise RuntimeError(
                f"Fiji returned invalid Original Astrocyte IDs: {expected_original_ids}"
            )
        for key in ("whole", "processes", "soma"):
            detail = result_sets.get(key, {})
            expected_count = final_roi_count
            if int(detail.get("rows", -1)) != expected_count:
                raise RuntimeError(f"Fiji {key} Results row count is inconsistent: {detail}")
            result_ids = [int(value) for value in detail.get("astrocyte_ids", [])]
            expected_result_ids = final_roi_ids
            if result_ids != expected_result_ids:
                raise RuntimeError(f"Fiji {key} Results Astrocyte IDs are inconsistent: {detail}")
            result_original_ids = [
                int(value) for value in detail.get("original_astrocyte_ids", [])
            ]
            if result_original_ids != expected_original_ids:
                raise RuntimeError(
                    f"Fiji {key} Original Astrocyte IDs are inconsistent: {detail}"
                )
            observed_headings = set(detail.get("headings", []))
            if expected_count > 0 and not required_headings.issubset(observed_headings):
                raise RuntimeError(
                    f"Fiji {key} Results columns are incomplete: expected {sorted(required_headings)}, "
                    f"observed {sorted(observed_headings)}"
                )
            row_data = detail.get("row_data", [])
            if len(row_data) != expected_count:
                raise RuntimeError(f"Fiji {key} Results row payload is incomplete: {detail}")
            payload_ids = [int(row["Astrocyte_ID"]) for row in row_data]
            if payload_ids != expected_result_ids:
                raise RuntimeError(
                    f"Fiji {key} row payload Astrocyte IDs are inconsistent: {payload_ids}"
                )
            payload_original_ids = [
                int(row["Original_Astrocyte_ID"]) for row in row_data
            ]
            if payload_original_ids != expected_original_ids:
                raise RuntimeError(
                    f"Fiji {key} row payload Original IDs are inconsistent: "
                    f"{payload_original_ids}"
                )
            if any(not np.isfinite(float(row["Median"])) for row in row_data):
                raise RuntimeError(f"Fiji {key} returned a non-finite Median value")
        whole_area = float(result_sets["whole"]["area_sum"])
        partition_area = float(result_sets["processes"]["area_sum"]) + float(
            result_sets["soma"]["area_sum"]
        )
        area_tolerance = max(
            pixel_width_um * pixel_height_um * final_roi_count * 2.0,
            abs(whole_area) * 1e-6,
        )
        if abs(whole_area - partition_area) > area_tolerance:
            raise RuntimeError(
                "Fiji compartment area identity failed: "
                f"Whole={whole_area}, Processes+Soma={partition_area}"
            )
        whole_intden = float(result_sets["whole"]["integrated_density_sum"])
        partition_intden = float(result_sets["processes"]["integrated_density_sum"]) + float(
            result_sets["soma"]["integrated_density_sum"]
        )
        if abs(whole_intden - partition_intden) > max(1.0, abs(whole_intden) * 1e-6):
            raise RuntimeError(
                "Fiji compartment integrated-density identity failed: "
                f"Whole={whole_intden}, Processes+Soma={partition_intden}"
            )

        overlay_paths = fiji_details.get("overlay_paths", {})
        staged_overlays: dict[str, Path] = {}
        for key in ("whole", "processes", "soma"):
            staged_overlay = Path(str(overlay_paths.get(key, ""))).resolve()
            if staged_overlay.parent != run_dir.resolve() or not staged_overlay.is_file():
                raise RuntimeError(
                    f"Fiji returned an invalid run-specific {key} overlay path: {staged_overlay}"
                )
            with Image.open(staged_overlay) as image:
                if image.size != (reference_shape[2], reference_shape[1]) or image.mode != "RGB":
                    raise RuntimeError(
                        f"Unexpected Fiji {key} overlay format: mode={image.mode}, size={image.size}"
                    )
                image.verify()
            staged_overlays[key] = staged_overlay
        completed_report = run_dir / "analysis_report_completed.txt"
        staged_workbook = run_dir / WORKBOOK_FILENAME
        write_measurement_workbook(
            staged_workbook,
            fiji_details=fiji_details,
            measurement=measurement,
            best_row=best_row,
            age_profile=age_profile_decision.profile,
        )
        validate_measurement_workbook(staged_workbook, fiji_details)
        write_analysis_report(
            completed_report,
            input_dir=input_dir,
            paths=paths,
            metadata=metadata,
            structural_channels=structural_channels,
            measurement=measurement,
            best_row=best_row,
            candidate_rows=rows,
            candidate_specs=chosen_specs,
            roi_count=roi_count,
            compartment_metrics=compartment_metrics,
            fiji_status=(
                "Completed after the Fiji pre-measurement review decision: six ROI image windows, "
                "three native measurement Results windows, and one synchronized XLSX workbook"
            ),
            fiji_details=fiji_details,
        )
        publish_output_bundle(
            staged_files={
                "whole": staged_overlays["whole"],
                "processes": staged_overlays["processes"],
                "soma": staged_overlays["soma"],
                "report": completed_report,
                "workbook": staged_workbook,
            },
            final_files={
                "whole": whole_overlay_path,
                "processes": process_overlay_path,
                "soma": soma_overlay_path,
                "report": report_path,
                "workbook": workbook_path,
            },
            run_dir=run_dir,
        )
        print_terminal_event("Validated production bundle published atomically.")
    except Exception as exc:
        if run_dir is not None:
            try:
                failed_report = run_dir / "analysis_report_failed.txt"
                write_analysis_report(
                    failed_report,
                    input_dir=input_dir,
                    paths=paths,
                    metadata=metadata,
                    structural_channels=structural_channels,
                    measurement=measurement,
                    best_row=best_row,
                    candidate_rows=rows,
                    candidate_specs=chosen_specs,
                    roi_count=roi_count,
                    compartment_metrics=compartment_metrics,
                    fiji_status=f"Failed: {exc!r}; previous production outputs were preserved",
                )
                print(f"Failure diagnostics retained at: {run_dir}", flush=True)
            except Exception:
                pass
        raise

    shutil.rmtree(run_dir, ignore_errors=True)
    print(f"Whole overlay: {whole_overlay_path}", flush=True)
    print(f"Processes overlay: {process_overlay_path}", flush=True)
    print(f"Soma overlay: {soma_overlay_path}", flush=True)
    print(f"Report: {report_path}", flush=True)
    print(f"Workbook: {workbook_path}", flush=True)
    for key in ("whole", "processes", "soma"):
        detail = fiji_details["result_sets"][key]
        print(f"Fiji {detail['title']}: {detail['rows']} rows", flush=True)
    print_runtime_timing_summary(fiji_details)
    return 0

def main(argv: list[str] | None = None) -> int:
    try:
        with analysis_lock():
            return _main_locked(argv)
    except DapiFragmentWorkloadLimitExceeded as exc:
        print(exc.user_message(), flush=True)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
